from datetime import datetime, timedelta
import pytz

import os
from tkcalendar import DateEntry
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import StringVar, ttk

from influxdb_client import InfluxDBClient, Point, WritePrecision
from influxdb_client.client.write_api import SYNCHRONOUS

import pandas as pd
import plotly.express as px
import pytz
import csv
import datetime

import openpyxl
from datetime import datetime, timedelta

import threading
import time

from itertools import groupby
start_datetime = None
end_datetime = None

output_file = 'CELL_OUPUT.csv'

DataBP = False

token = "oKKYZLlI--VzuvTqThsHVkF-TbtwxlPKO0ntiYOJQs5liKjNzN15szgOHPzRsDybK7di5V8dUVe0lj6DgY7Qrg=="
org = "Ferry_24m"
bucket = "datalogger"
client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)


# Specify the local time zone (GMT+7 in this case)
local_tz = pytz.timezone('Asia/Bangkok')  # Replace with your actual time zone

# Get the current date and time in the local time zone
current_datetime = datetime.now(local_tz)

# Take user input for the number of days to subtract
try:
    days_to_subtract = int(input("Enter the number of days to subtract: "))
except ValueError:
    print("Invalid input. Please enter a valid number.")
    exit()

# Calculate yesterday's date and time
yesterday_datetime = current_datetime - timedelta(days=days_to_subtract )

# Set the desired time for yesterday (22:00:00.000 in local time)
desired_time1 = yesterday_datetime.replace(hour=22, minute=0, second=0, microsecond=0)

# Set another desired time (05:00:00.000 in local time)
desired_time2 = yesterday_datetime.replace(hour=5, minute=0, second=0, microsecond=0)

# Create a naive datetime object without a time zone
naive_datetime = datetime(desired_time1.year, desired_time1.month, desired_time1.day, desired_time1.hour, desired_time1.minute, desired_time1.second, desired_time1.microsecond)

# Localize the naive datetime to the specified time zone (GMT+7)
localized_desired_time = local_tz.localize(naive_datetime)

# Convert the localized time to UTC
utc_time = localized_desired_time.astimezone(pytz.utc)

# Create a naive datetime object without a time zone for desired_time2
naive_datetime2 = datetime(desired_time2.year, desired_time2.month, desired_time2.day, desired_time2.hour, desired_time2.minute, desired_time2.second, desired_time2.microsecond)

# Localize the naive datetime to the specified time zone (GMT+7)
localized_desired_time2 = local_tz.localize(naive_datetime2)

# Convert the localized time to UTC for desired_time2
utc_time2 = localized_desired_time2.astimezone(pytz.utc)

# Format both UTC datetimes as strings
formatted_utc_time1 = utc_time.strftime('%Y-%m-%d %H:%M:%S.%f')  # Include microseconds for desired_time1
formatted_utc_time2 = utc_time2.strftime('%Y-%m-%d %H:%M:%S.%f')  # Include microseconds for desired_time2

# print("Original Date and Time 1 (Local):", desired_time1)
# print("Equivalent Date and Time 1 in UTC:", formatted_utc_time1)

# print("Original Date and Time 2 (Local):", desired_time2)
# print("Equivalent Date and Time 2 in UTC:", formatted_utc_time2)

o_start_datetime = desired_time1.replace(tzinfo=None)
o_end_datetime  =  desired_time2.replace(tzinfo=None)
date_str = str(desired_time1)
date = date_str.replace("22:00:00+07:00", "")
# print(date)

# Replace ".000000" with "Z"
end_time1 = formatted_utc_time1.replace(".000000", "Z")
start_time1 = formatted_utc_time2.replace(".000000", "Z")

end_t = end_time1.replace(" ","T")
start_t = start_time1.replace(" ","T")

# print("Original Date and Time 1 (Local):", desired_time1)
# print("Equivalent Date and Time 1 in UTC:", end_t)

# print("Original Date and Time 2 (Local):", desired_time2)
# print("Equivalent Date and Time 2 in UTC:", start_t)


def Get_ferry_id():
    global ferry_ids
    ferry_ids =[] 
      
    query3  = f' from(bucket:"datalogger")\
    |> range(start:{start_t}, stop:{end_t})\
    |> filter(fn:(r) => r._measurement == "sbcu")\
    |> filter(fn:(r) => r._field == "0x180a0001_S1_BatPack_Current" )\
    |> group(columns: ["ferry_id"])\
    |> distinct(column: "ferry_id")'

    resulta = client.query_api().query(org = org,query=query3)
    ferry_ids =[]
    for table in resulta:
        for record in table.records:
            # Get the tag value from the "_measurement" column
            ferry_id = record.get_value()
            ferry_ids.append(ferry_id)    
            print(ferry_id)        
    return ferry_ids

ferries = Get_ferry_id()
# print(ferries)
headers = [
    "Date","Start Time","Stop Time","Ferry","Pack No.",
    "Charged Energy (kWh)",
    "Discharged Energy (kWh)","Start_Min","End_Max","Start_SOC", "End_SOC", "Remaining_Capacity(kWh)","Remaining_Capacity(%)"
]
headers1 = [
    
"Start_Min","End_Max","Start_SOC", "End_SOC", "Remaining_Capacity(%)","Remaining_Capacity(kWh)","Charge"
]
headers2 = [
    
"Start_Max","End_Min","Start_SOC", "End_SOC", "Remaining_Capacity(%)","Remaining_Capacity(kWh)","DisCharge"
]
target_time_zone = pytz.timezone('Asia/Bangkok')

print("Start")


def show_result_window(result,result1,the_loop,app,total_DC,total_C,column,soop,workbook,ferry_ided,sheetProcess):
    date_str = str(desired_time1)
    date = date_str.replace("22:00:00+07:00", "")
    if app < 26:
        Value_maxV = []
        timestampss = []
        Value_mV = []
        timestampsm = []    
        target_time_zone = pytz.timezone('Asia/Bangkok')


        value_C = []
        Value_V = []
        timestamps = []
        time_difference = []

        Value_maxV = []
        timestampss = []
        Value_mV = []
        for table in result5:
            for record in table.records:
                timestamp = record.get_time()
                converted_times = timestamp.astimezone(target_time_zone)
                timestampsm.append(converted_times)
                field = record.get_field()
                value = record.get_value()/1000
                Value_mV.append(value)

        for table in result4:
            for record in table.records:
                value = record.get_value()/1000
                Value_maxV.append(value)
        # Insert the result into the Text widget with row numbers
        row_number = 2
        for table in result1:
            for record in table.records:
                
                timestamp = record.get_time()
                converted_times = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                field = record.get_field()
                values = record.get_value()
                # cell = sheet2.cell(row=row_number, column=column, value=values)
                value_C.append(values)
                # if app == 25:
                    # cell = sheet2.cell(row=row_number, column=1, value=converted_times)
                row_number += 1
        
        # Insert the result into the Text widget with row numbers
        row_number1 = 1
        timestamp_1 = []
        for table in result:
            for record in table.records:

                timestamp1 = record.get_time()
                converted_time1 = timestamp1.astimezone(target_time_zone).replace(tzinfo=None)
                field1 = record.get_field()
                value1 = record.get_value()
                timestamp_1.append(timestamp1)
                if value1 is not None:
                    timestamps.append(converted_time1)

                    Value_V.append(value1)

                row_number1 += 1
        if app == 1:
            global SystemC
            rows = 2  
            system_V_Time = []
            SystemV = []
            SystemC = []
            for table in result_totalC:
                for record in table.records:   
                    timestamp = record.get_time()
                    converted_times = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                    system_V_Time.append(converted_times)
                    value = record.get_value()
                    SystemC.append(value)
                    # cell = sheet2.cell(row=rows, column = 1, value = converted_times)
                    # cell = sheet2.cell(row=rows, column = 3, value = value)
                    rows += 1
            rows = 2  
            for table in result_totalV:
                for record in table.records:
                    value = record.get_value()
                    SystemV.append(value)
                    # cell = sheet2.cell(row=rows, column = 2, value = value)
                    rows += 1
            map1 = dict(zip(system_V_Time,SystemC))
            map2 = dict(zip(system_V_Time,SystemV))
            # Initialize the result arrays
            result_values1 = []
            result_values2 = []
            # Initialize a variable to keep track of the last valid value
            last_valid_value1 = 0  # Initialize it with the desired initial value
            last_valid_value2 = 0  # Initialize it with the desired initial value

            # Iterate through all timestamps and fill missing values with 0
            added_timestamps = set()
            ino = 0
            for timestamp in timestamps:
                ditt = len(timestamps)-len(system_V_Time)
                
                if ino+1 >= len(system_V_Time) - ditt:
                    result_values1.append(map1.get(timestamp, SystemC[ino-ditt]))
                    result_values2.append(map2.get(timestamp, SystemV[ino-ditt]))
                else:                 
                    value1 = map1.get(timestamp, last_valid_value1)
                    value2 = map2.get(timestamp, last_valid_value2)
                    result_values1.append(value1)
                    result_values2.append(value2)
                    if value1 != 0:
                        last_valid_value1 = value1
                    if value2 != 0:
                        last_valid_value2 = value2

                ino += 1
            rows = 2
            for value in result_values1:
                # cell = sheet2.cell(row=rows, column = 3, value = value)
                rows += 1
            rows = 2
            for value in result_values2:
                # cell = sheet2.cell(row=rows, column = 2, value = value)
                rows += 1
        else:
    
            map1 = dict(zip(system_V_Timee,SystemCC))
            # Initialize the result arrays
            result_values1 = []
            # Initialize a variable to keep track of the last valid value
            last_valid_value1 = 0  # Initialize it with the desired initial value

            # Iterate through all timestamps and fill missing values with 0
            added_timestamps = set()
            ino = 0
            for timestamp in timestamps:
                ditt = len(timestamps)-len(system_V_Timee) 
                if ino+1 > len(SystemCC):
                    break
                elif ino+1 >= len(system_V_Timee) - ditt :
                    result_values1.append(map1.get(timestamp, SystemCC[ino-ditt]))

                else:                 
                    value1 = map1.get(timestamp, last_valid_value1)

                    result_values1.append(value1)
                    if value1 != 0:
                        last_valid_value1 = value1

                ino += 1

        time_differences = []
        for i in range(len(timestamps)):
            if i > 0 :
                difference = timestamps[i] - timestamps[i - 1]
                time_differences.append(difference.total_seconds())  # Convert to seconds
            elif i == 0:
                difference = timestamps[i + 1] - timestamps[i]
                time_differences.append(difference.total_seconds()) 
            elif i == len(timestamps):
                time_differences[i] = timestamps[i + 1] - timestamps[i]
            else:
                time_difference[i] = 0
        if app == 25:
            rowp = 2
            # cell = sheet2.cell(row=1, column=column+1, value="TimeDiff")    
            for a in time_differences:

                # cell = sheet2.cell(row=rowp, column=column+1, value=a)       
                rowp += 1
        Energy = []
        # print(len(timestamps))
        if(len(value_C) < len(Value_V)):
            for i in range(len(value_C)):
                if time_differences[i] < 300:
                    energy_value = value_C[i] * Value_V[i] * float(time_differences[i])
                    Energy.append(energy_value)
                else:
                    energy_value = value_C[i] * Value_V[i] * float(time_differences[i-1])
                    Energy.append(energy_value)
        
        else:
            for i in range(len(Value_V)):
                if time_differences[i] < 300:
                    energy_value = value_C[i] * Value_V[i] * float(time_differences[i])
                    Energy.append(energy_value)
                else:
                    energy_value = value_C[i] * Value_V[i] * float(time_differences[i-1])
                    Energy.append(energy_value)
        # print(Energy)
        row_number2 = 1
        if(len(value_C) < len(Value_V)):
            for table in result1:
                for record in table.records:
                    timestamp = record.get_time()
                    converted_times2 = timestamp.astimezone(target_time_zone)

                    row_number2 += 1 
                    
        Total_E = sum(Energy)
        P_DChar = []
        P_Char = []
        Total_P = []
        for Total in Energy:
            Total_E = Total_E+ Total
            Total_P.append(Total_E)
            if Total < 0:
                Total_P_DC = Total
                P_DChar.append(Total_P_DC)
            else:
                Total_P_C = Total
                P_Char.append(Total_P_C)

        Energy_Dhar = sum(P_DChar)/(3600*1000)
        Energy_Char = sum(P_Char)/(3600*1000)
        total_DC.append(Energy_Dhar)
        total_C.append(Energy_Char)
        date = str(date)
        modified_string = date.replace("-", "/")
        cell = sheetProcess.cell(row=the_loop, column=5, value= app+1)
        cell = sheetProcess.cell(row=the_loop, column=1, value= modified_string)
        cell = sheetProcess.cell(row=the_loop, column=4, value=ferry_ided)
        cell = sheetProcess.cell(row=the_loop, column=6, value= Energy_Char )
        cell = sheetProcess.cell(row=the_loop, column=7, value=Energy_Dhar)
        cell = sheetProcess.cell(row=the_loop, column=2, value= "5:00:00")
        cell = sheetProcess.cell(row=the_loop, column=3, value="22:00:00")  
        # cell = sheet.cell(row=the_loop, column=5, value=app + 1)
        # cell = sheet.cell(row=the_loop, column=5, value=app + 1)
        # cell = sheet.cell(row=the_loop, column=1, value= modified_string)
        # cell = sheet.cell(row=the_loop, column=4, value=ferry_ided)
        # cell = sheet200.cell(row=the_loop, column=5, value=app + 1)
        # cell = sheet200.cell(row=the_loop, column=5, value=app + 1)
        # cell = sheet200.cell(row=the_loop, column=1, value= modified_string)
        # cell = sheet200.cell(row=the_loop, column=4, value=ferry_ided)

        Data_Set = {
                    'Timestamps': timestamps,
                    'Value_V': Value_V,
                    'Value_C': value_C,
                    'Max_V' : Value_maxV,
                    'Min_V': Value_mV,
                    'Sys_C': result_values1
                    }
        Discharge_Set = {
            'Timestamps': [],
            'Voltage': [],
            'Current': [],
            'Max_V' : [],
            'Min_V': [],
            'Diff': [],
            'Group':[],
            'Time_Diff':[]
        }
        Charge_Set = {
            'Timestamps': [],
            'Voltage': [],
            'Current': [],
            'Max_V' : [],
            'Min_V': [],
            'Diff': [],
            'Group':[],
            'Time_Diff':[]
        }
        # sheet_Pack = workbook.create_sheet(title= f"Pack+{app+1}")
        start_time2 = time.time()
        for loc, Currents in enumerate(Data_Set['Sys_C']):
            
            
            # print(f"This is Location:{loc}")
            if Currents >= 0  :
                
                Discharge_Set['Timestamps'].append(Data_Set['Timestamps'][loc]) 
                # print(Data_Set['Timestamps'][loc])
                Discharge_Set['Voltage'].append(Data_Set['Value_V'][loc])
                Discharge_Set['Current'].append(Data_Set['Value_C'][loc])
                Discharge_Set['Max_V'].append(Data_Set['Max_V'][loc])
                Discharge_Set['Min_V'].append(Data_Set['Min_V'][loc])
                times = Data_Set['Timestamps'][loc]
                DC_Vo = Data_Set['Value_V'][loc]
                DC_c = Data_Set['Value_C'][loc]

            elif Currents < -5:
                
                Charge_Set['Timestamps'].append(Data_Set['Timestamps'][loc]) 
                Charge_Set['Voltage'].append(Data_Set['Value_V'][loc])
                Charge_Set['Current'].append(Data_Set['Value_C'][loc])  
                Charge_Set['Max_V'].append(Data_Set['Max_V'][loc])
                Charge_Set['Min_V'].append(Data_Set['Min_V'][loc])
                times = Data_Set['Timestamps'][loc]
                C_Vo = Data_Set['Value_V'][loc]
                C_c = Data_Set['Value_C'][loc]

            else:
                times = Data_Set['Timestamps'][loc]
                C_Vo = Data_Set['Value_V'][loc]
                C_c = Data_Set['Value_C'][loc]

        
        for loc, Currents in enumerate(Discharge_Set['Current']):
                Sample = '30s'
                
                Sample = float(Sample.replace('s', ''))
                
                
                if loc > 0:
                    difference = Discharge_Set['Timestamps'][loc] - Discharge_Set['Timestamps'][loc-1]
                    if difference.total_seconds() > 300:
                        Discharge_Set['Time_Diff'].append(Sample)
                    else:
                        Discharge_Set['Time_Diff'].append(difference.total_seconds())
                elif loc == 0:
                    difference = Sample 
                    Discharge_Set['Time_Diff'].append(difference)
                if loc == 0:
                    if difference< 300:
                        
                        Discharge_Set['Diff'].append(0)
                    else:
                        Discharge_Set['Diff'].append(1)
                else:
                    if difference.total_seconds()< 300:
                        
                        Discharge_Set['Diff'].append(0)
                    else:
                        Discharge_Set['Diff'].append(1)
                    

        for loc, Currents in enumerate(Charge_Set['Current']):

                Sample = '30s'
                Sample = float(Sample.replace('s', ''))
                if loc > 0:
                    difference = Charge_Set['Timestamps'][loc] - Charge_Set['Timestamps'][loc-1]
                    if difference.total_seconds() > 300:
                        Charge_Set['Time_Diff'].append(Sample)
                    else:
                        Charge_Set['Time_Diff'].append(difference.total_seconds())
                elif loc == 0:
                    difference = Sample
                    Charge_Set['Time_Diff'].append(difference)
                if loc == 0:
                    if difference< 300:
                        
                        Charge_Set['Diff'].append(0)
                    else:
                        Charge_Set['Diff'].append(1)
                else:
                    if difference.total_seconds()< 300:
                        
                        Charge_Set['Diff'].append(0)
                    else:
                        Charge_Set['Diff'].append(1)
        Goop = 0
        for loc,Diff in enumerate(Discharge_Set['Diff']):
            if Diff > 0: #1
                Goop = Discharge_Set['Diff'][loc] + Discharge_Set['Group'][loc-1] #1
                
            else:
                if loc > 0:
                    shit = Discharge_Set['Group'][loc-1]
                    
                    Goop = Discharge_Set['Diff'][loc] + Discharge_Set['Group'][loc-1]
                else:
                    Goop = 0

            
            # times = Discharge_Set['Timestamps'][loc]
            # C_Vo = Discharge_Set['Voltage'][loc]
            # C_c = Discharge_Set['Current'][loc]
            # Ma_V = Discharge_Set['Max_V'][loc]
            # Mi_V = Discharge_Set['Min_V'][loc]
            # ditt = Discharge_Set['Diff'][loc]
            # cell = sheet_Pack.cell(row=loc+1, column=1, value=times)
            # cell = sheet_Pack.cell(row=loc+1, column=2, value=C_Vo)
            # cell = sheet_Pack.cell(row=loc+1, column=3, value=C_c)
            # cell = sheet_Pack.cell(row=loc+1, column=4, value=Mi_V)
            # cell = sheet_Pack.cell(row=loc+1, column=5, value=Ma_V)
            # cell = sheet_Pack.cell(row=loc+1, column=6, value=ditt)
            # cell = sheet_Pack.cell(row=loc+1, column=7, value=Goop)
            Discharge_Set['Group'].append(Goop)
        
        for loc,Diff in enumerate(Charge_Set['Diff']):
            SOC_Start = 0
            SOC_End = 0
            
            if Diff > 0: #1
                Goop = Charge_Set['Diff'][loc] + Charge_Set['Group'][loc-1] #1
                
            else:
                if loc > 0:
                    shit = Charge_Set['Group'][loc-1]
                    
                    Goop = Charge_Set['Diff'][loc] + Charge_Set['Group'][loc-1]
                else:
                    Goop = 0

            
            # times = Charge_Set['Timestamps'][loc]
            # C_Vo = Charge_Set['Voltage'][loc]
            # C_c = Charge_Set['Current'][loc]
            # Ma_V = Charge_Set['Max_V'][loc]
            # Mi_V = Charge_Set['Min_V'][loc]
            # ditt = Charge_Set['Diff'][loc]
            # cell = sheet_Pack.cell(row=loc+1, column=9, value=times)
            # cell = sheet_Pack.cell(row=loc+1, column=10, value=C_Vo)
            # cell = sheet_Pack.cell(row=loc+1, column=11, value=C_c)
            # cell = sheet_Pack.cell(row=loc+1, column=12, value=Mi_V)
            # cell = sheet_Pack.cell(row=loc+1, column=13, value=Ma_V)
            # cell = sheet_Pack.cell(row=loc+1, column=14, value=ditt)
            # cell = sheet_Pack.cell(row=loc+1, column=15, value=Goop)
            Charge_Set['Group'].append(Goop)

            

        end_time2 = time.time()  # Record the end time
        execution_time2 = end_time2 - start_time2  # Calculate the execution time
        print(f"loc took {execution_time2:.6f} seconds to execute.")         
        # sheet_Pack = workbook.create_sheet(title= f"Pack+{app+1}")
        ferries = int(ferry_ided)
        # print(f"tHIS IS ferry {ferries}")
        if ferries == 2 or ferries == 18 or ferries == 22 or ferries == 21:
            def CCha():
                def SOH(input_voltage):
                    global closest_percentage
                    closest_percentage = 0
                    Volt = [3.0607,3.4149,3.4929,3.5287,3.5706,3.604,3.6294,3.6452,3.6587,3.6743,3.6951,3.7226,3.7584,3.8057,3.8619,3.9186,3.9725,4.0266,4.0828,4.141,4.1999]

                    Percentage = [0,0.05,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]

                    mapping = list(zip(Volt,Percentage))
                    # print(len(Volt))
                    # print(len(Percentage))
                    # print(len(mapping))

                    # Find the closest voltage in the list of Volt
                    closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                    # Find the index of the closest voltage
                    index = Volt.index(closest_voltage)
                    # print(index)
                    # print(Percentage[index])

                    # Check if the index is not the last index to avoid index out of range
                    if index < len(Volt) - 1:
                        if(input_voltage > closest_voltage) or index == 0  and input_voltage != 0:
                            next_voltage = Volt[index + 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")
                            # Calculate the slope using the closest_voltage and next_voltage
                            if next_voltage != closest_voltage :
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            else:
                                next_voltage = Volt[index + 2]
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)

                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif input_voltage < closest_voltage and input_voltage != 0:
                            next_voltage = Volt[index - 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                            closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif input_voltage == 0:
                            closest_percentage = 0
                        else:
                            next_voltage = Volt[index + 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                    elif index + 1 == len(Volt):

                        if input_voltage < closest_voltage and input_voltage != 0:
                            # print("This is herer 3")
                            next_voltage = Volt[index - 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                            closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif input_voltage == 0:
                            closest_percentage = 0
                        elif input_voltage > Volt[index] :

                            next_voltage = Volt[index - 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                            closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
        
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        else:
                            next_voltage = Volt[index + 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                    else:
                        print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                    return closest_percentage
                def SOH2(input_voltage):
                    global closest_percentage2
                    closest_percentage2 = 0
                    # print("This is SOH2")
                    Volt = [3.0607,3.4149,3.4929,3.5287,3.5706,3.604,3.6294,3.6452,3.6587,3.6743,3.6951,3.7226,3.7584,3.8057,3.8619,3.9186,3.9725,4.0266,4.0828,4.141,4.1999]

                    Percentage = [0,0.05,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]

                    mapping = list(zip(Volt,Percentage))
                    # print(len(Volt))
                    # print(len(Percentage))
                    # print(len(mapping))

                    # Find the closest voltage in the list of Volt
                    closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                    # Find the index of the closest voltage
                    index = Volt.index(closest_voltage)
                    # print(index)
                    # print(Percentage[index])

                    # Check if the index is not the last index to avoid index out of range
                    if index < len(Volt) - 1:
                        # Get the next voltage value
                        if(input_voltage > closest_voltage) or index == 0 and input_voltage != 0:
                            next_voltage = Volt[index + 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        elif input_voltage < closest_voltage and input_voltage != 0:
                            next_voltage = Volt[index - 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                            closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        elif input_voltage == 0:
                            closest_percentage2 = 0
                        else:
                            next_voltage = Volt[index + 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")  
                    elif index + 1 == len(Volt):

                        if input_voltage < closest_voltage and input_voltage != 0:
                            # print("This is herer 3")
                            next_voltage = Volt[index - 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                            closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        elif input_voltage == 0:
                            closest_percentage2 = 0
                        elif input_voltage > Volt[index]:
                            next_voltage = Volt[index - 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                            closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        else:
                            next_voltage = Volt[index + 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")          
                    else:
                        print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                # Find indices of 'Group1' in the 'Group' list


                # Sort the array
                sorted_array = sorted(Charge_Set['Group'])

                # Get unique numbers using set
                unique_numbers = set(Charge_Set['Group'])

                # Convert the unique numbers back to a sorted list if needed
                sorted_unique_numbers = sorted(list(unique_numbers))
                first_min_v = None
                last_max_v = None
                # Print the sorted unique numbers
                # print(sorted_unique_numbers)
                shitty = 1
                for numbers in sorted_unique_numbers:
                    group_indices = [i for i, group in enumerate(Charge_Set['Group']) if group == numbers]

                    # Check if there are 'Group1' elements in the 'Group' list
                    if group_indices:
                        # Access the first and last 'Max_V' values for 'Group1'
                        first_min_v = Charge_Set['Min_V'][group_indices[0]]
                        last_max_v = Charge_Set['Max_V'][group_indices[-1]]

                        # print(f"First Max_V for 'Group1': {first_min_v}")
                        # print(f"Last Max_V for 'Group1': {last_max_v}")
                    else:
                        first_min_v = 0
                        last_max_v = 0
                    # print(f"First Max_V for 'Group1': {first_min_v}")
                    # print(f"Last Max_V for 'Group1': {last_max_v}")
                    SOH_Goop['Goop'].append(numbers)
                    SOH_Goop['Min_V'].append(first_min_v)
                    SOH_Goop['Max_V'].append(last_max_v)
                    thread1 = threading.Thread(target=SOH, args=(first_min_v,))
                    # print(f'This is last_m:{last_max_v}')
                    thread2 = threading.Thread(target=SOH2, args=(last_max_v,))
                    thread1.start()
                    thread2.start()
                    thread1.join()
                    thread2.join()
                    # print(f'This is SOH2:{closest_percentage2}')
                    # Calculate energy for 'Group1' where 'Group' is 0
            # Find indices of 'Group1' where 'Group' is 0
                    
                    energy_sum = 0
                    save_data_start = None
                    a= None
                    # Iterate over the numerical indices of Charge_Set['Group']
                    for index, group in enumerate(Charge_Set['Group']):
                        if group == numbers:
                            voltage = Charge_Set['Voltage'][index]
                            current = Charge_Set['Current'][index]
                            time_diff = Charge_Set['Time_Diff'][index]
                            # Split the original string by space to get the time portion
                            teiam = str( Charge_Set['Timestamps'][index])
                            # print(teiam)
                            split_string = teiam.split(" ")

                            # Check if there are at least two parts (date and time)
                            if len(split_string) >= 2:
                                # Join the time portion and discard the date
                                time_portion = " ".join(split_string[1:])
                                # print(time_portion)

                            if save_data_start == None:
                            
                                save_data_start = time_portion
                                SOH_Goop['S_Time'].append(time_portion)
                                
                            save_end = time_portion
                            # print(f'This is time diff { leg}')
                            # print(index)  # Assuming 'Diff' represents time intervals
                            # print(f'This is V {voltage}')
                            # print(f'This is C {current}')
                            # print(f'This is D {time_diff}')
                            if index != 0:
                                energy =  abs(0.5*(current+previous_current))  * time_diff
                                previous_current = current
                                # print(f'This is Energy {energy}')
                                # cell = sheetPackProcess.cell(row=gginp, column=28, value= energy)
                                # gginp += 1
                            else :
                                energy = 0
                                previous_current = current
                            energy_sum += energy
                    SOH_Goop['E_Time'].append(save_end)
                    energy_sum = energy_sum* voltage
                    # print(f'Total energy for "Group1": {energy_sum} Joules')
                        

                    # print(f"Total energy for 'Group1' where 'Group' is 0: {energy_sum} Joules")
                    
                    SOC_Start = closest_percentage*100
                    SOC_End = closest_percentage2*100
                    Difference_SOC = SOC_End - SOC_Start
                    DesignCapacity_NH02 = 30
                    SOh_E = energy_sum/(3600*1000)
                    SOH_Goop['Start_SOC'].append(SOC_Start)
                    SOH_Goop['End_SOC'].append(SOC_End)
                    SOH_Goop['Charge'].append(SOh_E)
                    if Difference_SOC != 0:
                        Cal_Capacity = SOh_E/(Difference_SOC/100)
                    else:
                        Cal_Capacity = 0
                    if(Cal_Capacity != 0):
                        Remaining_Capacity = (Cal_Capacity/DesignCapacity_NH02)*100
                    elif(Cal_Capacity == None):
                        Remaining_Capacity = 0
                    else:
                        Remaining_Capacity = 0
                    SOH_Goop['Cal'].append(Remaining_Capacity)
                    SOH_Goop['SOH'].append(Cal_Capacity)
                    SOH_Goop['Cal_E'].append(SOh_E)
                    
                    # for index, header in enumerate(headers1, start=shitty):
                    #     cell = sheet8.cell(row=4, column=index, value=header)

                    # cell = sheet8.cell(row=the_loop, column=shitty, value=first_min_v)
                    # cell = sheet8.cell(row=the_loop, column=shitty+1, value=last_max_v)
                    # cell = sheet8.cell(row=the_loop, column=shitty+2, value=SOC_Start)
                    # cell = sheet8.cell(row=the_loop, column=shitty+3, value=SOC_End)
                    # cell = sheet8.cell(row=the_loop, column=shitty+4, value=Cal_Capacity)
                    # cell = sheet8.cell(row=the_loop, column=shitty+5, value=Remaining_Capacity)    
                    # cell = sheet8.cell(row=the_loop, column=shitty+6, value=SOh_E)  
                    shitty += 7
                if first_min_v == None and last_max_v == None:
                    lost_data.append(app+1)
                    lost_data3.append(app+1)
                # print(lost_data)
            def SOH_Min(input_voltage):
                    global closest_percentage
                    closest_percentage = 0
                    Volt =[2.8016,3.3466,3.4631,3.4931,3.53,3.5599,3.5862,3.6129,3.6328,3.6492,3.6685,3.6941,3.728,3.7725,3.8287,3.8883,3.946,4.0028,4.062,4.1258,4.1977]

                    Percentage = [0,0.5,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]

                
                    mapping = list(zip(Volt,Percentage))
                    # print(len(Volt))
                    # print(len(Percentage))
                    # print(len(mapping))

                    # Find the closest voltage in the list of Volt
                    closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                    # Find the index of the closest voltage
                    index = Volt.index(closest_voltage)
                    # print(index)
                    # print(Percentage[index])

                    # Check if the index is not the last index to avoid index out of range
                    if index < len(Volt) - 1:
                        if(input_voltage > closest_voltage) or index == 0  and input_voltage != 0:
                            next_voltage = Volt[index + 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")
                            # Calculate the slope using the closest_voltage and next_voltage
                            if next_voltage != closest_voltage :
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            else:
                                next_voltage = Volt[index + 2]
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                        elif (closest_voltage - input_voltage) < 0.1 :    
                            next_voltage = Volt[index + 1]
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif input_voltage < closest_voltage and input_voltage != 0:
                            next_voltage = Volt[index - 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                            closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif input_voltage == 0:
                            closest_percentage = 0
                        else:
                            next_voltage = Volt[index + 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                    elif index + 1 == len(Volt):

                        if input_voltage < closest_voltage and input_voltage != 0:
                            # print("This is herer 3")
                            next_voltage = Volt[index - 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                            closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif input_voltage == 0:
                            closest_percentage = 0
    
                        else:
                            next_voltage = Volt[index + 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")

                    else:
                        print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                    return closest_percentage
        
            def SOH2_Min(input_voltage):
                global closest_percentage2
                closest_percentage2 = 0
                # print("This is SOH2")
                Volt =[2.8016,3.3466,3.4631,3.4931,3.53,3.5599,3.5862,3.6129,3.6328,3.6492,3.6685,3.6941,3.728,3.7725,3.8287,3.8883,3.946,4.0028,4.062,4.1258,4.1977]

                Percentage = [0,0.5,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]

                mapping = list(zip(Volt,Percentage))
                # print(len(Volt))
                # print(len(Percentage))
                # print(len(mapping))

                # Find the closest voltage in the list of Volt
                closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                # Find the index of the closest voltage
                index = Volt.index(closest_voltage)
                # print(index)
                # print(Percentage[index])

                # Check if the index is not the last index to avoid index out of range
                if index < len(Volt) - 1:
                    # Get the next voltage value
                    if(input_voltage > closest_voltage) or index == 0 and input_voltage != 0:
                        next_voltage = Volt[index + 1]
                        
                        # Calculate the slope using the closest_voltage and next_voltage
                        slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                        closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                    elif (closest_voltage - input_voltage) < 0.1 :    
                        next_voltage = Volt[index + 1]
                        slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                        closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                                
                        # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                    elif input_voltage < closest_voltage and input_voltage != 0:
                        next_voltage = Volt[index - 1]
                        
                        # Calculate the slope using the closest_voltage and next_voltage
                        slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                        closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                        
                        # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                    elif input_voltage == 0:
                        closest_percentage2 = 0
                    else:
                        next_voltage = Volt[index + 1]
                        
                        # Calculate the slope using the closest_voltage and next_voltage
                        slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                        closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                        
                        # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")   
                elif index + 1 == len(Volt):

                    if input_voltage < closest_voltage and input_voltage != 0:
                        # print("This is herer 3")
                        next_voltage = Volt[index - 1]
                        # print(f"This is next{next_voltage}")
                        # print(f"This is next{closest_voltage}")        
                        # Calculate the slope using the closest_voltage and next_voltage
                        slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                        closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                        
                        # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                    elif input_voltage == 0:
                        closest_percentage2 = 0
                    elif input_voltage > Volt[index]:
                        # print(Volt[index])
                        closest_percentage2 = 0
                    else:
                        next_voltage = Volt[index + 1]
                        # print(f"This is next{next_voltage}")
                        # print(f"This is next{closest_voltage}")        
                        # Calculate the slope using the closest_voltage and next_voltage
                        slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                        closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                        
                        # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")         
                else:
                    print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
            # Find indices of 'Group1' in the 'Group' list


            # Sort the array
            sorted_array = sorted(Discharge_Set['Group'])

            # Get unique numbers using set
            unique_numbers = set(Discharge_Set['Group'])

            # Convert the unique numbers back to a sorted list if needed
            sorted_unique_numbers = sorted(list(unique_numbers))
            first_min_v = None
            last_max_v = None
            # Print the sorted unique numbers
            # print(sorted_unique_numbers)
            shitty = 1
            for numbers in sorted_unique_numbers:
                group_indices = [i for i, group in enumerate(Discharge_Set['Group']) if group == numbers]

                # Check if there are 'Group1' elements in the 'Group' list
                if group_indices:
                    # Access the first and last 'Max_V' values for 'Group1'
                    first_min_v = Discharge_Set['Max_V'][group_indices[0]]
                    last_max_v = Discharge_Set['Min_V'][group_indices[-1]]

                    # print(f"First Max_V for 'Group1': {first_min_v}")
                    # print(f"Last Max_V for 'Group1': {last_max_v}")
                else:
                    first_min_v = 0
                    last_max_v = 0
                # print(f"First Max_V for 'Group1': {first_min_v}")
                # print(f"Last Max_V for 'Group1': {last_max_v}")
                SOH_Goop_Min['Goop'].append(numbers)
                SOH_Goop_Min['Min_V'].append(first_min_v)
                SOH_Goop_Min['Max_V'].append(last_max_v)
                thread1 = threading.Thread(target=SOH_Min, args=(first_min_v,))
                # print(f'This is last_m:{last_max_v}')
                thread2 = threading.Thread(target=SOH2_Min, args=(last_max_v,))
                thread1.start()
                thread2.start()
                thread1.join()
                thread2.join()
                # print(f'This is SOH2:{closest_percentage2}')
                # Calculate energy for 'Group1' where 'Group' is 0
        # Find indices of 'Group1' where 'Group' is 0
                
                energy_sum = 0
                save_data_start = None
                a= None
                # Iterate over the numerical indices of Discharge_Set['Group']
                for index, group in enumerate(Discharge_Set['Group']):
                    if group == numbers:
                        voltage = Discharge_Set['Voltage'][index]
                        current = Discharge_Set['Current'][index]
                        time_diff = Discharge_Set['Time_Diff'][index]
                        # Split the original string by space to get the time portion
                        teiam = str( Discharge_Set['Timestamps'][index])
                        # print(teiam)
                        split_string = teiam.split(" ")

                        # Check if there are at least two parts (date and time)
                        if len(split_string) >= 2:
                            # Join the time portion and discard the date
                            time_portion = " ".join(split_string[1:])
                            # print(time_portion)

                        if save_data_start == None:
                        
                            save_data_start = time_portion
                            SOH_Goop_Min['S_Time'].append(time_portion)
                            
                        save_end = time_portion

                        # print(f'This is time diff { leg}')
                        # print(index)  # Assuming 'Diff' represents time intervals
                        # print(f'This is V {voltage}')
                        # print(f'This is C {current}')
                        # print(f'This is D {time_diff}')
                        if index != 0:
                            energy =  abs(0.5*(current+previous_current))  * time_diff
                            previous_current = current
                            # print(f'This is Energy {energy}')
                            # cell = sheetPackProcess.cell(row=gginp, column=27, value= energy)
                            # gginp += 1

                        else :
                            energy = 0
                            previous_current = current
                        energy_sum += energy
                SOH_Goop_Min['E_Time'].append(save_end)
                energy_sum = energy_sum* voltage
                    

                # print(f"Total energy for 'Group1' where 'Group' is 0: {energy_sum} Joules")
                
                SOC_Start = closest_percentage
                SOC_End = closest_percentage2
                Difference_SOC = SOC_Start - SOC_End 
                DesignCapacity_NH02 = 30
                SOh_E = energy_sum/(3600*1000)
                SOH_Goop_Min['Start_SOC'].append(SOC_Start*100)
                SOH_Goop_Min['End_SOC'].append(SOC_End*100)
                SOH_Goop_Min['Discharge'].append(SOh_E)
                if Difference_SOC != 0:
                    Cal_Capacity = abs(SOh_E)/(Difference_SOC)
                else:
                    Cal_Capacity = 0
                if(Cal_Capacity != 0):
                    Remaining_Capacity = (Cal_Capacity/DesignCapacity_NH02)*100
                elif(Cal_Capacity == None):
                    Remaining_Capacity = 0
                else:
                    Remaining_Capacity = 0
                SOH_Goop_Min['Cal'].append(Remaining_Capacity)
                SOH_Goop_Min['SOH'].append(Cal_Capacity)
                SOH_Goop_Min['Cal_E'].append(SOh_E)
                
                # for index, header in enumerate(headers2, start=shitty):
                #     cell = sheet8.cell(row=4+31, column=index, value=header)
                # cell = sheet8.cell(row=the_loop +31, column=shitty, value=first_min_v)
                # cell = sheet8.cell(row=the_loop +31, column=shitty+1, value=last_max_v)
                # cell = sheet8.cell(row=the_loop +31, column=shitty+2, value=SOC_Start)
                # cell = sheet8.cell(row=the_loop +31, column=shitty+3, value=SOC_End)
                # cell = sheet8.cell(row=the_loop +31, column=shitty+4, value=Cal_Capacity)
                # cell = sheet8.cell(row=the_loop +31, column=shitty+5, value=Remaining_Capacity)    
                # cell = sheet8.cell(row=the_loop +31, column=shitty+6, value=SOh_E)  
                shitty += 7
            if first_min_v == None and last_max_v == None:
                lost_data2.append(app+1)
                lost_data4.append(app+1)
            # print(lost_data2)
                # thread1 = threading.Thread(target=CCha)
                # thread1.start()
                # thread1.join()
            CCha()
        elif ferries == 17:

            # Define the file path to the Excel file
            file_path = os.path.join(os.path.expanduser('~'), 'Documents','Ferry_Battery_Pack_Type', 'Ferry_Bat_Pack_List.xlsx')

            # Load the workbook
            workbook = openpyxl.load_workbook(file_path)

            # Access the 'msf17' worksheet
            worksheet = workbook['MSF17']
            Bat_Type = worksheet.cell(row=app+2, column=2).value  # Assuming it's a single column
            if Bat_Type == "NH04":
                def CCha():
                    def SOH(input_voltage):
                        global closest_percentage
                        closest_percentage = 0
                        Volt = [3.0607,3.4149,3.4929,3.5287,3.5706,3.604,3.6294,3.6452,3.6587,3.6743,3.6951,3.7226,3.7584,3.8057,3.8619,3.9186,3.9725,4.0266,4.0828,4.141,4.1999]

                        Percentage = [0,0.05,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]

                        mapping = list(zip(Volt,Percentage))
                        # print(len(Volt))
                        # print(len(Percentage))
                        # print(len(mapping))

                        # Find the closest voltage in the list of Volt
                        closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                        # Find the index of the closest voltage
                        index = Volt.index(closest_voltage)
                        # print(index)
                        # print(Percentage[index])

                        # Check if the index is not the last index to avoid index out of range
                        if index < len(Volt) - 1:
                            if(input_voltage > closest_voltage) or index == 0  and input_voltage != 0:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")
                                # Calculate the slope using the closest_voltage and next_voltage
                                if next_voltage != closest_voltage :
                                    slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                    closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                else:
                                    next_voltage = Volt[index + 2]
                                    slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                    closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)

                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage < closest_voltage and input_voltage != 0:
                                next_voltage = Volt[index - 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage == 0:
                                closest_percentage = 0
                            else:
                                next_voltage = Volt[index + 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif index + 1 == len(Volt):

                            if input_voltage < closest_voltage and input_voltage != 0:
                                # print("This is herer 3")
                                next_voltage = Volt[index - 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage == 0:
                                closest_percentage = 0
                            elif input_voltage > Volt[index] :

                                next_voltage = Volt[index - 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
            
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            else:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        else:
                            print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                        return closest_percentage
                    def SOH2(input_voltage):
                        global closest_percentage2
                        closest_percentage2 = 0
                        # print("This is SOH2")
                        Volt = [3.0607,3.4149,3.4929,3.5287,3.5706,3.604,3.6294,3.6452,3.6587,3.6743,3.6951,3.7226,3.7584,3.8057,3.8619,3.9186,3.9725,4.0266,4.0828,4.141,4.1999]

                        Percentage = [0,0.05,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]

                        mapping = list(zip(Volt,Percentage))
                        # print(len(Volt))
                        # print(len(Percentage))
                        # print(len(mapping))

                        # Find the closest voltage in the list of Volt
                        closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                        # Find the index of the closest voltage
                        index = Volt.index(closest_voltage)
                        # print(index)
                        # print(Percentage[index])

                        # Check if the index is not the last index to avoid index out of range
                        if index < len(Volt) - 1:
                            # Get the next voltage value
                            if(input_voltage > closest_voltage) or index == 0 and input_voltage != 0:
                                next_voltage = Volt[index + 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                            elif input_voltage < closest_voltage and input_voltage != 0:
                                next_voltage = Volt[index - 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                                closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                            elif input_voltage == 0:
                                closest_percentage2 = 0
                            else:
                                next_voltage = Volt[index + 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")  
                        elif index + 1 == len(Volt):

                            if input_voltage < closest_voltage and input_voltage != 0:
                                # print("This is herer 3")
                                next_voltage = Volt[index - 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                                closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                            elif input_voltage == 0:
                                closest_percentage2 = 0
                            elif input_voltage > Volt[index]:
                                next_voltage = Volt[index - 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                                closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                            else:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")          
                        else:
                            print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                    # Find indices of 'Group1' in the 'Group' list


                    # Sort the array
                    sorted_array = sorted(Charge_Set['Group'])

                    # Get unique numbers using set
                    unique_numbers = set(Charge_Set['Group'])

                    # Convert the unique numbers back to a sorted list if needed
                    sorted_unique_numbers = sorted(list(unique_numbers))
                    first_min_v = None
                    last_max_v = None
                    # Print the sorted unique numbers
                    # print(sorted_unique_numbers)
                    shitty = 1
                    for numbers in sorted_unique_numbers:
                        group_indices = [i for i, group in enumerate(Charge_Set['Group']) if group == numbers]

                        # Check if there are 'Group1' elements in the 'Group' list
                        if group_indices:
                            # Access the first and last 'Max_V' values for 'Group1'
                            first_min_v = Charge_Set['Min_V'][group_indices[0]]
                            last_max_v = Charge_Set['Max_V'][group_indices[-1]]

                            # print(f"First Max_V for 'Group1': {first_min_v}")
                            # print(f"Last Max_V for 'Group1': {last_max_v}")
                        else:
                            first_min_v = 0
                            last_max_v = 0
                        # print(f"First Max_V for 'Group1': {first_min_v}")
                        # print(f"Last Max_V for 'Group1': {last_max_v}")
                        SOH_Goop['Goop'].append(numbers)
                        SOH_Goop['Min_V'].append(first_min_v)
                        SOH_Goop['Max_V'].append(last_max_v)
                        thread1 = threading.Thread(target=SOH, args=(first_min_v,))
                        # print(f'This is last_m:{last_max_v}')
                        thread2 = threading.Thread(target=SOH2, args=(last_max_v,))
                        thread1.start()
                        thread2.start()
                        thread1.join()
                        thread2.join()
                        # print(f'This is SOH2:{closest_percentage2}')
                        # Calculate energy for 'Group1' where 'Group' is 0
                # Find indices of 'Group1' where 'Group' is 0
                        
                        energy_sum = 0
                        save_data_start = None
                        a= None
                        # Iterate over the numerical indices of Charge_Set['Group']
                        for index, group in enumerate(Charge_Set['Group']):
                            if group == numbers:
                                voltage = Charge_Set['Voltage'][index]
                                current = Charge_Set['Current'][index]
                                time_diff = Charge_Set['Time_Diff'][index]
                                # Split the original string by space to get the time portion
                                teiam = str( Charge_Set['Timestamps'][index])
                                # print(teiam)
                                split_string = teiam.split(" ")

                                # Check if there are at least two parts (date and time)
                                if len(split_string) >= 2:
                                    # Join the time portion and discard the date
                                    time_portion = " ".join(split_string[1:])
                                    # print(time_portion)

                                if save_data_start == None:
                                
                                    save_data_start = time_portion
                                    SOH_Goop['S_Time'].append(time_portion)
                                    
                                save_end = time_portion
                                # print(f'This is time diff { leg}')
                                # print(index)  # Assuming 'Diff' represents time intervals
                                # print(f'This is V {voltage}')
                                # print(f'This is C {current}')
                                # print(f'This is D {time_diff}')
                                if index != 0:
                                    energy =  abs(0.5*(current+previous_current))  * time_diff
                                    previous_current = current
                                    # print(f'This is Energy {energy}')
                                    # cell = sheetPackProcess.cell(row=gginp, column=28, value= energy)
                                    # gginp += 1
                                else :
                                    energy = 0
                                    previous_current = current
                                energy_sum += energy
                        SOH_Goop['E_Time'].append(save_end)
                        energy_sum = energy_sum* voltage
                        # print(f'Total energy for "Group1": {energy_sum} Joules')
                            

                        # print(f"Total energy for 'Group1' where 'Group' is 0: {energy_sum} Joules")
                        
                        SOC_Start = closest_percentage*100
                        SOC_End = closest_percentage2*100
                        Difference_SOC = SOC_End - SOC_Start
                        DesignCapacity_NH02 = 30
                        SOh_E = energy_sum/(3600*1000)
                        SOH_Goop['Start_SOC'].append(SOC_Start)
                        SOH_Goop['End_SOC'].append(SOC_End)
                        SOH_Goop['Charge'].append(SOh_E)
                        if Difference_SOC != 0:
                            Cal_Capacity = SOh_E/(Difference_SOC/100)
                        else:
                            Cal_Capacity = 0
                        if(Cal_Capacity != 0):
                            Remaining_Capacity = (Cal_Capacity/DesignCapacity_NH02)*100
                        elif(Cal_Capacity == None):
                            Remaining_Capacity = 0
                        else:
                            Remaining_Capacity = 0
                        SOH_Goop['Cal'].append(Remaining_Capacity)
                        SOH_Goop['SOH'].append(Cal_Capacity)
                        SOH_Goop['Cal_E'].append(SOh_E)
                        
                        # for index, header in enumerate(headers1, start=shitty):
                        #     cell = sheet8.cell(row=4, column=index, value=header)

                        # cell = sheet8.cell(row=the_loop, column=shitty, value=first_min_v)
                        # cell = sheet8.cell(row=the_loop, column=shitty+1, value=last_max_v)
                        # cell = sheet8.cell(row=the_loop, column=shitty+2, value=SOC_Start)
                        # cell = sheet8.cell(row=the_loop, column=shitty+3, value=SOC_End)
                        # cell = sheet8.cell(row=the_loop, column=shitty+4, value=Cal_Capacity)
                        # cell = sheet8.cell(row=the_loop, column=shitty+5, value=Remaining_Capacity)    
                        # cell = sheet8.cell(row=the_loop, column=shitty+6, value=SOh_E)  
                        shitty += 7
                    if first_min_v == None and last_max_v == None:
                        lost_data.append(app+1)
                        lost_data3.append(app+1)
                    # print(lost_data)
                def SOH_Min(input_voltage):
                        global closest_percentage
                        closest_percentage = 0
                        Volt =[2.8016,3.3466,3.4631,3.4931,3.53,3.5599,3.5862,3.6129,3.6328,3.6492,3.6685,3.6941,3.728,3.7725,3.8287,3.8883,3.946,4.0028,4.062,4.1258,4.1977]

                        Percentage = [0,0.5,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]

                    
                        mapping = list(zip(Volt,Percentage))
                        # print(len(Volt))
                        # print(len(Percentage))
                        # print(len(mapping))

                        # Find the closest voltage in the list of Volt
                        closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                        # Find the index of the closest voltage
                        index = Volt.index(closest_voltage)
                        # print(index)
                        # print(Percentage[index])

                        # Check if the index is not the last index to avoid index out of range
                        if index < len(Volt) - 1:
                            if(input_voltage > closest_voltage) or index == 0  and input_voltage != 0:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")
                                # Calculate the slope using the closest_voltage and next_voltage
                                if next_voltage != closest_voltage :
                                    slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                    closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                else:
                                    next_voltage = Volt[index + 2]
                                    slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                    closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            elif (closest_voltage - input_voltage) < 0.1 :    
                                next_voltage = Volt[index + 1]
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage < closest_voltage and input_voltage != 0:
                                next_voltage = Volt[index - 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage == 0:
                                closest_percentage = 0
                            else:
                                next_voltage = Volt[index + 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif index + 1 == len(Volt):

                            if input_voltage < closest_voltage and input_voltage != 0:
                                # print("This is herer 3")
                                next_voltage = Volt[index - 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage == 0:
                                closest_percentage = 0
        
                            else:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")

                        else:
                            print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                        return closest_percentage
            
                def SOH2_Min(input_voltage):
                    global closest_percentage2
                    closest_percentage2 = 0
                    # print("This is SOH2")
                    Volt =[2.8016,3.3466,3.4631,3.4931,3.53,3.5599,3.5862,3.6129,3.6328,3.6492,3.6685,3.6941,3.728,3.7725,3.8287,3.8883,3.946,4.0028,4.062,4.1258,4.1977]

                    Percentage = [0,0.5,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]

                    mapping = list(zip(Volt,Percentage))
                    # print(len(Volt))
                    # print(len(Percentage))
                    # print(len(mapping))

                    # Find the closest voltage in the list of Volt
                    closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                    # Find the index of the closest voltage
                    index = Volt.index(closest_voltage)
                    # print(index)
                    # print(Percentage[index])

                    # Check if the index is not the last index to avoid index out of range
                    if index < len(Volt) - 1:
                        # Get the next voltage value
                        if(input_voltage > closest_voltage) or index == 0 and input_voltage != 0:
                            next_voltage = Volt[index + 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                        elif (closest_voltage - input_voltage) < 0.1 :    
                            next_voltage = Volt[index + 1]
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                                    
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        elif input_voltage < closest_voltage and input_voltage != 0:
                            next_voltage = Volt[index - 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                            closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        elif input_voltage == 0:
                            closest_percentage2 = 0
                        else:
                            next_voltage = Volt[index + 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")   
                    elif index + 1 == len(Volt):

                        if input_voltage < closest_voltage and input_voltage != 0:
                            # print("This is herer 3")
                            next_voltage = Volt[index - 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                            closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        elif input_voltage == 0:
                            closest_percentage2 = 0
                        elif input_voltage > Volt[index]:
                            # print(Volt[index])
                            closest_percentage2 = 0
                        else:
                            next_voltage = Volt[index + 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")         
                    else:
                        print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                # Find indices of 'Group1' in the 'Group' list


                # Sort the array
                sorted_array = sorted(Discharge_Set['Group'])

                # Get unique numbers using set
                unique_numbers = set(Discharge_Set['Group'])

                # Convert the unique numbers back to a sorted list if needed
                sorted_unique_numbers = sorted(list(unique_numbers))
                first_min_v = None
                last_max_v = None
                # Print the sorted unique numbers
                # print(sorted_unique_numbers)
                shitty = 1
                for numbers in sorted_unique_numbers:
                    group_indices = [i for i, group in enumerate(Discharge_Set['Group']) if group == numbers]

                    # Check if there are 'Group1' elements in the 'Group' list
                    if group_indices:
                        # Access the first and last 'Max_V' values for 'Group1'
                        first_min_v = Discharge_Set['Max_V'][group_indices[0]]
                        last_max_v = Discharge_Set['Min_V'][group_indices[-1]]

                        # print(f"First Max_V for 'Group1': {first_min_v}")
                        # print(f"Last Max_V for 'Group1': {last_max_v}")
                    else:
                        first_min_v = 0
                        last_max_v = 0
                    # print(f"First Max_V for 'Group1': {first_min_v}")
                    # print(f"Last Max_V for 'Group1': {last_max_v}")
                    SOH_Goop_Min['Goop'].append(numbers)
                    SOH_Goop_Min['Min_V'].append(first_min_v)
                    SOH_Goop_Min['Max_V'].append(last_max_v)
                    thread1 = threading.Thread(target=SOH_Min, args=(first_min_v,))
                    # print(f'This is last_m:{last_max_v}')
                    thread2 = threading.Thread(target=SOH2_Min, args=(last_max_v,))
                    thread1.start()
                    thread2.start()
                    thread1.join()
                    thread2.join()
                    # print(f'This is SOH2:{closest_percentage2}')
                    # Calculate energy for 'Group1' where 'Group' is 0
            # Find indices of 'Group1' where 'Group' is 0
                    
                    energy_sum = 0
                    save_data_start = None
                    a= None
                    # Iterate over the numerical indices of Discharge_Set['Group']
                    for index, group in enumerate(Discharge_Set['Group']):
                        if group == numbers:
                            voltage = Discharge_Set['Voltage'][index]
                            current = Discharge_Set['Current'][index]
                            time_diff = Discharge_Set['Time_Diff'][index]
                            # Split the original string by space to get the time portion
                            teiam = str( Discharge_Set['Timestamps'][index])
                            # print(teiam)
                            split_string = teiam.split(" ")

                            # Check if there are at least two parts (date and time)
                            if len(split_string) >= 2:
                                # Join the time portion and discard the date
                                time_portion = " ".join(split_string[1:])
                                # print(time_portion)

                            if save_data_start == None:
                            
                                save_data_start = time_portion
                                SOH_Goop_Min['S_Time'].append(time_portion)
                                
                            save_end = time_portion

                            # print(f'This is time diff { leg}')
                            # print(index)  # Assuming 'Diff' represents time intervals
                            # print(f'This is V {voltage}')
                            # print(f'This is C {current}')
                            # print(f'This is D {time_diff}')
                            if index != 0:
                                energy =  abs(0.5*(current+previous_current))  * time_diff
                                previous_current = current
                                # print(f'This is Energy {energy}')
                                # cell = sheetPackProcess.cell(row=gginp, column=27, value= energy)
                                # gginp += 1

                            else :
                                energy = 0
                                previous_current = current
                            energy_sum += energy
                    SOH_Goop_Min['E_Time'].append(save_end)
                    energy_sum = energy_sum* voltage
                        

                    # print(f"Total energy for 'Group1' where 'Group' is 0: {energy_sum} Joules")
                    
                    SOC_Start = closest_percentage
                    SOC_End = closest_percentage2
                    Difference_SOC = SOC_Start - SOC_End 
                    DesignCapacity_NH02 = 30
                    SOh_E = energy_sum/(3600*1000)
                    SOH_Goop_Min['Start_SOC'].append(SOC_Start*100)
                    SOH_Goop_Min['End_SOC'].append(SOC_End*100)
                    SOH_Goop_Min['Discharge'].append(SOh_E)
                    if Difference_SOC != 0:
                        Cal_Capacity = abs(SOh_E)/(Difference_SOC)
                    else:
                        Cal_Capacity = 0
                    if(Cal_Capacity != 0):
                        Remaining_Capacity = (Cal_Capacity/DesignCapacity_NH02)*100
                    elif(Cal_Capacity == None):
                        Remaining_Capacity = 0
                    else:
                        Remaining_Capacity = 0
                    SOH_Goop_Min['Cal'].append(Remaining_Capacity)
                    SOH_Goop_Min['SOH'].append(Cal_Capacity)
                    SOH_Goop_Min['Cal_E'].append(SOh_E)
                    
                    # for index, header in enumerate(headers2, start=shitty):
                    #     cell = sheet8.cell(row=4+31, column=index, value=header)
                    # cell = sheet8.cell(row=the_loop +31, column=shitty, value=first_min_v)
                    # cell = sheet8.cell(row=the_loop +31, column=shitty+1, value=last_max_v)
                    # cell = sheet8.cell(row=the_loop +31, column=shitty+2, value=SOC_Start)
                    # cell = sheet8.cell(row=the_loop +31, column=shitty+3, value=SOC_End)
                    # cell = sheet8.cell(row=the_loop +31, column=shitty+4, value=Cal_Capacity)
                    # cell = sheet8.cell(row=the_loop +31, column=shitty+5, value=Remaining_Capacity)    
                    # cell = sheet8.cell(row=the_loop +31, column=shitty+6, value=SOh_E)  
                    shitty += 7
                if first_min_v == None and last_max_v == None:
                    lost_data2.append(app+1)
                    lost_data4.append(app+1)
                # print(lost_data2)
                    # thread1 = threading.Thread(target=CCha)
                    # thread1.start()
                    # thread1.join()
                CCha()
            else:
                def CCha():
                    def SOH(input_voltage):
                        global closest_percentage
                        closest_percentage = 0
                        Volt = [ 2.7011, 2.7546, 2.7684, 2.7789, 2.7877, 2.7956, 2.8026, 2.8092, 2.8156, 2.8212, 2.827, 2.8327, 2.8378, 2.8432, 2.8482, 2.8533, 2.8584, 2.8632, 2.8678, 2.8727, 2.8775, 2.8823, 2.887, 2.8917, 2.8961, 2.9008, 2.9055, 2.9102, 2.9145, 2.9192, 2.9235, 2.9282, 2.9317, 2.9372, 2.9421, 2.9461, 2.9505, 2.9548, 2.9594, 2.9631, 2.9685, 2.9726, 2.9768, 2.9819, 2.9856, 2.9898, 2.9942, 2.998, 3.0025, 3.007, 3.0111, 3.0153, 3.0197, 3.0233, 3.0277, 3.0319, 3.0358, 3.0399, 3.044, 3.048, 3.052, 3.0559, 3.0598, 3.0637, 3.0675, 3.0715, 3.0754, 3.0789, 3.0828, 3.0868, 3.0903, 3.0942, 3.098, 3.1016, 3.1052, 3.1087, 3.1124, 3.116, 3.1195, 3.1231, 3.1266, 3.1299, 3.1336, 3.1373, 3.1405, 3.144, 3.1472, 3.1504, 3.1538, 3.1573, 3.1606, 3.1638, 3.167, 3.1703, 3.1739, 3.1766, 3.1797, 3.183, 3.1859, 3.1891, 3.1923, 3.1951, 3.1981, 3.2011, 3.2045, 3.2072, 3.21, 3.213, 3.2157, 3.2192, 3.2216, 3.2246, 3.2275, 3.2302, 3.233, 3.2359, 3.2385, 3.2408, 3.2436, 3.2463, 3.2489, 3.2515, 3.2541, 3.2566, 3.2592, 3.262, 3.2645, 3.2668, 3.2697, 3.2722, 3.2748, 3.2776, 3.2797, 3.2823, 3.2847, 3.2871, 3.2892, 3.292, 3.2946, 3.2965, 3.2989, 3.3014, 3.3036, 3.3056, 3.3085, 3.3106, 3.3129, 3.3153, 3.3177, 3.3199, 3.322, 3.3244, 3.3266, 3.3286, 3.3311, 3.3333, 3.3353, 3.3377, 3.3396, 3.3419, 3.3439, 3.346, 3.3483, 3.3503, 3.3521, 3.3544, 3.3566, 3.3585, 3.3606, 3.3625, 3.3645, 3.3667, 3.3684, 3.3705, 3.3725, 3.3744, 3.3763, 3.3784, 3.3802, 3.3822, 3.384, 3.3864, 3.3879, 3.3899, 3.3913, 3.3937, 3.3954, 3.3972, 3.399, 3.4009, 3.4027, 3.4046, 3.4064, 3.408, 3.4098, 3.4116, 3.4131, 3.4151, 3.4168, 3.4186, 3.4203, 3.4219, 3.4235, 3.4253, 3.4271, 3.4288, 3.4302, 3.432, 3.4339, 3.4355, 3.437, 3.4389, 3.4403, 3.442, 3.4436, 3.4451, 3.4469, 3.4483, 3.45, 3.4514, 3.4531, 3.4547, 3.4562, 3.4577, 3.4592, 3.4608, 3.4623, 3.4637, 3.4652, 3.4667, 3.4682, 3.4698, 3.4712, 3.4727, 3.4744, 3.4757, 3.4771, 3.4785, 3.48, 3.4814, 3.4829, 3.4844, 3.4857, 3.4871, 3.4884, 3.4899, 3.4912, 3.4926, 3.494, 3.4955, 3.4968, 3.498, 3.4994, 3.5009, 3.5021, 3.5036, 3.5046, 3.5058, 3.5076, 3.5086, 3.5101, 3.511, 3.5124, 3.5135, 3.5148, 3.5163, 3.5176, 3.5186, 3.52, 3.5213, 3.5225, 3.5236, 3.5248, 3.5257, 3.5271, 3.5284, 3.5296, 3.5309, 3.5318, 3.5327, 3.534, 3.5352, 3.5362, 3.5374, 3.5383, 3.539, 3.5401, 3.5412, 3.5422, 3.5431, 3.5435, 3.5446, 3.545, 3.5457, 3.5465, 3.547, 3.5475, 3.5484, 3.5489, 3.5495, 3.5499, 3.5503, 3.5506, 3.5513, 3.5516, 3.5524, 3.5527, 3.5529, 3.5533, 3.5537, 3.554, 3.5544, 3.5551, 3.5554, 3.5555, 3.5562, 3.5563, 3.5564, 3.5567, 3.5571, 3.5574, 3.5578, 3.5581, 3.5585, 3.5587, 3.5592, 3.5591, 3.5595, 3.5602, 3.5605, 3.5608, 3.5608, 3.561, 3.5616, 3.5617, 3.562, 3.5626, 3.5627, 3.5633, 3.5633, 3.5637, 3.5641, 3.5644, 3.5648, 3.5649, 3.5647, 3.5655, 3.5659, 3.5661, 3.5665, 3.5673, 3.5669, 3.5675, 3.5677, 3.5679, 3.5686, 3.569, 3.5689, 3.5698, 3.5701, 3.5703, 3.5707, 3.5711, 3.5712, 3.5716, 3.5723, 3.5724, 3.5727, 3.5732, 3.5737, 3.574, 3.5743, 3.5747, 3.5751, 3.5754, 3.576, 3.5763, 3.5767, 3.5771, 3.5773, 3.5781, 3.5784, 3.5789, 3.5792, 3.5793, 3.5802, 3.5805, 3.5808, 3.5813, 3.5817, 3.5821, 3.5827, 3.5832, 3.5837, 3.5841, 3.5847, 3.5852, 3.5857, 3.586, 3.5865, 3.5871, 3.5873, 3.5879, 3.5886, 3.5894, 3.5896, 3.59, 3.5906, 3.591, 3.5916, 3.5922, 3.5925, 3.5929, 3.594, 3.5942, 3.5943, 3.5953, 3.5955, 3.5965, 3.5968, 3.5974, 3.5977, 3.5986, 3.5986, 3.5996, 3.5995, 3.6005, 3.6008, 3.6018, 3.6019, 3.6026, 3.603, 3.6034, 3.604, 3.6048, 3.6051, 3.6057, 3.6062, 3.6066, 3.6074, 3.6078, 3.6084, 3.6089, 3.6095, 3.61, 3.6105, 3.6109, 3.6115, 3.612, 3.6127, 3.6133, 3.6136, 3.6142, 3.6146, 3.6152, 3.6157, 3.6163, 3.6169, 3.6173, 3.6178, 3.6184, 3.6189, 3.6195, 3.6199, 3.6204, 3.621, 3.6214, 3.6218, 3.6224, 3.6228, 3.6233, 3.6238, 3.6243, 3.6248, 3.6255, 3.6259, 3.6263, 3.627, 3.6273, 3.6278, 3.6284, 3.6287, 3.6292, 3.6296, 3.6301, 3.6306, 3.6311, 3.6316, 3.632, 3.6325, 3.633, 3.6334, 3.6339, 3.6343, 3.6348, 3.6352, 3.6358, 3.6361, 3.6364, 3.6369, 3.6374, 3.6379, 3.6382, 3.6389, 3.6392, 3.6393, 3.6401, 3.6405, 3.6407, 3.6413, 3.6416, 3.6421, 3.6425, 3.6432, 3.6434, 3.644, 3.6443, 3.6448, 3.645, 3.6454, 3.646, 3.6464, 3.647, 3.6475, 3.6478, 3.6478, 3.6483, 3.6492, 3.6494, 3.6498, 3.65, 3.6508, 3.6511, 3.6516, 3.652, 3.6524, 3.6527, 3.6532, 3.6537, 3.6542, 3.6548, 3.6551, 3.6554, 3.6559, 3.6564, 3.6568, 3.6573, 3.6578, 3.6581, 3.6586, 3.659, 3.6594, 3.6598, 3.66, 3.6606, 3.6612, 3.6613, 3.6621, 3.6626, 3.6629, 3.6634, 3.6637, 3.6641, 3.6645, 3.665, 3.6653, 3.6659, 3.666, 3.6664, 3.6669, 3.6675, 3.6677, 3.6682, 3.6686, 3.6688, 3.6694, 3.6698, 3.6701, 3.6706, 3.6709, 3.6714, 3.6717, 3.6719, 3.6726, 3.6727, 3.6732, 3.6735, 3.6739, 3.6745, 3.6748, 3.675, 3.6754, 3.6756, 3.676, 3.6765, 3.6769, 3.6771, 3.6775, 3.6777, 3.678, 3.6783, 3.6788, 3.6793, 3.6793, 3.6796, 3.68, 3.6804, 3.6805, 3.6811, 3.6812, 3.6817, 3.6819, 3.6822, 3.6825, 3.6827, 3.683, 3.6834, 3.6837, 3.6838, 3.6842, 3.6846, 3.6849, 3.685, 3.6852, 3.6858, 3.6861, 3.6862, 3.6864, 3.6867, 3.6873, 3.6874, 3.6875, 3.6878, 3.6882, 3.6884, 3.6886, 3.6889, 3.6892, 3.6893, 3.6897, 3.69, 3.6901, 3.6905, 3.6907, 3.6911, 3.6911, 3.6916, 3.6917, 3.6921, 3.6922, 3.6925, 3.6929, 3.6928, 3.6931, 3.6936, 3.6936, 3.6939, 3.6942, 3.6945, 3.6946, 3.6949, 3.6951, 3.6953, 3.6958, 3.6961, 3.6962, 3.6964, 3.6967, 3.6969, 3.6967, 3.6972, 3.6976, 3.6977, 3.6979, 3.6982, 3.6987, 3.6988, 3.6989, 3.6995, 3.6993, 3.6997, 3.7002, 3.7003, 3.7005, 3.7002, 3.7007, 3.7009, 3.7011, 3.7024, 3.7008, 3.7018, 3.7023, 3.7023, 3.7024, 3.7023, 3.7026, 3.7041, 3.7035, 3.7038, 3.7044, 3.7044, 3.7047, 3.705, 3.7062, 3.7053, 3.7056, 3.706, 3.7064, 3.7063, 3.7067, 3.7067, 3.707, 3.7073, 3.7076, 3.7081, 3.7078, 3.7081, 3.7086, 3.7088, 3.7089, 3.7094, 3.7096, 3.7098, 3.7101, 3.7102, 3.7104, 3.7105, 3.711, 3.7112, 3.7114, 3.7115, 3.712, 3.7123, 3.7127, 3.7129, 3.7131, 3.7134, 3.7135, 3.714, 3.7142, 3.7146, 3.7149, 3.7149, 3.7153, 3.7156, 3.7158, 3.7161, 3.7164, 3.7168, 3.7169, 3.7171, 3.7177, 3.7178, 3.718, 3.7183, 3.7188, 3.719, 3.7192, 3.7194, 3.7199, 3.7201, 3.7202, 3.7205, 3.721, 3.7213, 3.7217, 3.7219, 3.7222, 3.7225, 3.7227, 3.7231, 3.7233, 3.7235, 3.7239, 3.7242, 3.7245, 3.7249, 3.7252, 3.7254, 3.7259, 3.7261, 3.7263, 3.7268, 3.727, 3.7272, 3.7278, 3.7281, 3.7283, 3.7285, 3.7291, 3.7294, 3.7297, 3.7299, 3.7302, 3.7307, 3.7308, 3.7314, 3.7316, 3.7319, 3.7324, 3.7329, 3.7329, 3.7332, 3.7336, 3.7341, 3.7343, 3.7347, 3.735, 3.7353, 3.7358, 3.7361, 3.7365, 3.7368, 3.737, 3.7378, 3.738, 3.7382, 3.7387, 3.739, 3.7394, 3.7397, 3.7402, 3.7405, 3.7408, 3.7413, 3.7416, 3.7421, 3.7422, 3.7428, 3.7432, 3.7435, 3.7439, 3.7443, 3.7448, 3.7451, 3.7457, 3.7458, 3.7464, 3.7467, 3.7473, 3.7475, 3.7479, 3.7484, 3.7487, 3.749, 3.7495, 3.7498, 3.7504, 3.7505, 3.7513, 3.7518, 3.7519, 3.7524, 3.7529, 3.7534, 3.7537, 3.7542, 3.7545, 3.7548, 3.7554, 3.7559, 3.7562, 3.7566, 3.7572, 3.7576, 3.7581, 3.7585, 3.759, 3.7593, 3.7598, 3.7602, 3.7606, 3.7612, 3.7616, 3.7622, 3.7626, 3.763, 3.7634, 3.7638, 3.7645, 3.7648, 3.7653, 3.7658, 3.7662, 3.767, 3.7672, 3.7677, 3.7682, 3.7685, 3.769, 3.7695, 3.7701, 3.7706, 3.7711, 3.7716, 3.772, 3.7726, 3.7728, 3.7736, 3.774, 3.7744, 3.775, 3.7755, 3.776, 3.7766, 3.7769, 3.7774, 3.778, 3.7787, 3.779, 3.7797, 3.78, 3.7806, 3.7812, 3.7817, 3.7823, 3.783, 3.7834, 3.7839, 3.7844, 3.7849, 3.7854, 3.7859, 3.7865, 3.7872, 3.7876, 3.7882, 3.7887, 3.7894, 3.7898, 3.7907, 3.7914, 3.7919, 3.7924, 3.793, 3.7936, 3.7943, 3.7947, 3.7953, 3.7958, 3.7964, 3.7969, 3.7975, 3.7982, 3.7989, 3.7994, 3.7999, 3.8006, 3.801, 3.8018, 3.8023, 3.8028, 3.8035, 3.804, 3.8047, 3.8053, 3.8059, 3.8066, 3.807, 3.8078, 3.8083, 3.8089, 3.8096, 3.8102, 3.8109, 3.8114, 3.8121, 3.8127, 3.8133, 3.8141, 3.8148, 3.8153, 3.8159, 3.8164, 3.8172, 3.8178, 3.8186, 3.8193, 3.8197, 3.8205, 3.821, 3.8216, 3.8225, 3.8231, 3.8238, 3.8245, 3.8252, 3.8258, 3.8265, 3.8271, 3.8279, 3.8283, 3.8292, 3.8299, 3.8306, 3.8314, 3.8319, 3.8329, 3.8333, 3.8338, 3.8348, 3.8355, 3.8362, 3.8368, 3.8376, 3.8384, 3.8392, 3.8399, 3.8406, 3.8411, 3.842, 3.8428, 3.8432, 3.8441, 3.8449, 3.8458, 3.8464, 3.8471, 3.848, 3.8488, 3.8492, 3.8502, 3.8507, 3.8514, 3.8524, 3.8531, 3.8539, 3.8548, 3.8555, 3.8561, 3.8569, 3.858, 3.8587, 3.8593, 3.86, 3.8609, 3.8616, 3.8625, 3.8635, 3.8641, 3.8649, 3.8658, 3.8664, 3.8674, 3.8679, 3.869, 3.8698, 3.8705, 3.8712, 3.8721, 3.8725, 3.8735, 3.8743, 3.875, 3.8758, 3.8767, 3.8774, 3.8784, 3.8792, 3.8798, 3.8808, 3.8815, 3.8828, 3.8838, 3.8849, 3.885, 3.8856, 3.8864, 3.887, 3.888, 3.8888, 3.8897, 3.8904, 3.8914, 3.892, 3.8928, 3.8938, 3.8944, 3.8952, 3.8961, 3.8968, 3.8978, 3.8983, 3.8991, 3.8999, 3.9008, 3.9014, 3.9022, 3.9031, 3.9039, 3.9045, 3.9053, 3.906, 3.9069, 3.9076, 3.9084, 3.9092, 3.9099, 3.9108, 3.9114, 3.9121, 3.9129, 3.9136, 3.9145, 3.9153, 3.9159, 3.9167, 3.9175, 3.918, 3.9188, 3.9196, 3.9203, 3.921, 3.9218, 3.9224, 3.9233, 3.924, 3.9247, 3.9254, 3.9261, 3.9268, 3.9276, 3.9282, 3.929, 3.9297, 3.9305, 3.931, 3.9319, 3.9328, 3.9334, 3.934, 3.9349, 3.9355, 3.9361, 3.9369, 3.9376, 3.9383, 3.9391, 3.9398, 3.9405, 3.9413, 3.9418, 3.9425, 3.9435, 3.944, 3.9446, 3.9456, 3.9462, 3.9469, 3.9477, 3.9483, 3.9491, 3.9496, 3.9505, 3.9511, 3.9518, 3.9526, 3.9533, 3.954, 3.9547, 3.9554, 3.9561, 3.9566, 3.9575, 3.9582, 3.9588, 3.9595, 3.9603, 3.9611, 3.9618, 3.9624, 3.9631, 3.964, 3.9647, 3.9652, 3.966, 3.9667, 3.9676, 3.968, 3.9688, 3.9695, 3.9702, 3.9709, 3.9716, 3.9723, 3.973, 3.9737, 3.9745, 3.9751, 3.9759, 3.9767, 3.9773, 3.9779, 3.9788, 3.9794, 3.9801, 3.981, 3.9816, 3.9823, 3.983, 3.9837, 3.9846, 3.9851, 3.9856, 3.9864, 3.9871, 3.9879, 3.9885, 3.9892, 3.9901, 3.9908, 3.9914, 3.9921, 3.9929, 3.9935, 3.9943, 3.9949, 3.9958, 3.9964, 3.9972, 3.9979, 3.9985, 3.9994, 4, 4.0007, 4.0013, 4.0022, 4.0029, 4.0038, 4.0042, 4.005, 4.0058, 4.0065, 4.0072, 4.008, 4.0086, 4.0093, 4.0099, 4.0108, 4.0116, 4.0122, 4.0129, 4.0135, 4.0144, 4.0151, 4.0158, 4.0164, 4.0172, 4.0178, 4.0187, 4.0194, 4.0201, 4.021, 4.0215, 4.0223, 4.0231, 4.0239, 4.0244, 4.0252, 4.026, 4.0268, 4.0273, 4.028, 4.0288, 4.0295, 4.0304, 4.0311, 4.0318, 4.0325, 4.0333, 4.034, 4.0347, 4.0356, 4.0362, 4.0371, 4.0377, 4.0385, 4.0391, 4.0399, 4.0405, 4.0414, 4.042, 4.0428, 4.0436, 4.0443, 4.0449, 4.0458, 4.0465, 4.0473, 4.048, 4.0488, 4.0495, 4.0503, 4.051, 4.0517, 4.0525, 4.0531, 4.054, 4.0547, 4.0554, 4.0563, 4.057, 4.0578, 4.0585, 4.0592, 4.0601, 4.0607, 4.0616, 4.0621, 4.0631, 4.0637, 4.0645, 4.0652, 4.0659, 4.0667, 4.0675, 4.0683, 4.069, 4.0697, 4.0706, 4.0714, 4.0721, 4.0728, 4.0735, 4.0744, 4.0751, 4.0758, 4.0766, 4.0773, 4.0781, 4.079, 4.0797, 4.0805, 4.0812, 4.0821, 4.0827, 4.0834, 4.0843, 4.0852, 4.0858, 4.0867, 4.0872, 4.0882, 4.089, 4.0897, 4.0905, 4.0912, 4.0919, 4.0928, 4.0936, 4.0944, 4.0951, 4.0959, 4.0965, 4.0974, 4.0982, 4.099, 4.1, 4.1006, 4.1015, 4.1021, 4.1029, 4.1037, 4.1044, 4.1053, 4.1062, 4.1069, 4.1076, 4.1084, 4.1091, 4.1099, 4.1108, 4.1114, 4.1123, 4.113, 4.1139, 4.1146, 4.1154, 4.1162, 4.117, 4.1179, 4.1186, 4.1194, 4.1202, 4.1209, 4.1218, 4.1225, 4.1235, 4.1243, 4.1252, 4.1257, 4.1266, 4.1273, 4.1282, 4.129, 4.1297, 4.1306, 4.1312, 4.1321, 4.1329, 4.1337, 4.1346, 4.1354, 4.1361, 4.1369, 4.1379, 4.1386, 4.1392, 4.1402, 4.1409, 4.1417, 4.1427, 4.1433, 4.1442, 4.1451, 4.1458, 4.1467, 4.1474, 4.1483, 4.149, 4.1498, 4.1507, 4.1514, 4.1522, 4.1531, 4.1538, 4.1548, 4.1555, 4.1564, 4.1571, 4.1579, 4.1586, 4.1596, 4.1604, 4.1611, 4.1621, 4.1628, 4.1637, 4.1645, 4.1653, 4.1661, 4.1669, 4.1677, 4.1684, 4.1693, 4.1701, 4.1711, 4.1719, 4.1726, 4.1735, 4.1743, 4.1752, 4.1759, 4.1768, 4.1775, 4.1785, 4.1793, 4.1802, 4.181, 4.1817, 4.1825, 4.1834, 4.1843, 4.1849, 4.186, 4.1867, 4.1876, 4.1884, 4.1894, 4.1901, 4.1909, 4.1917, 4.1925, 4.1935, 4.1944, 4.1951, 4.1959, 4.1967, 4.1975, 4.1985, 4.1992, 4.2, 4.1992, 4.1992, 4.1991, 4.1991, 4.1993, 4.1991, 4.1993, 4.1993, 4.1991, 4.1993, 4.1991, 4.1992, 4.199, 4.1991, 4.1994, 4.1996, 4.2001, 4.2005, 4.2009, 4.2002, 4.2, 4.2001, 4.2002, 4.2001, 4.2001, 4.2, 4.2001, 4.2002, 4.2001, 4.2003, 4.2, 4.2001, 4.2002, 4.2002, 4.2002, 4.2002, 4.2001, 4.2, 4.2001, 4.2, 4.2001, 4.2, 4.2002, 4.2003, 4.2003, 4.2002, 4.2, 4.2001, 4.2001, 4.2002, 4.2002, 4.2001, 4.2003, 4.2001, 4.2002, 4.2001, 4.2001, 4.2004, 4.2003, 4.2003, 4.2006, 4.2006, 4.2008, 4.2004, 4.2005, 4.2004, 4.2004, 4.2003, 4.2004, 4.2005, 4.2004, 4.2001, 4.2004, 4.2004, 4.2005, 4.2004, 4.2004, 4.2004, 4.2004, 4.2004, 4.2003, 4.2003, 4.2004, 4.2001, 4.2003, 4.2003, 4.2003, 4.2002, 4.2003, 4.2002, 4.2001, 4.2004, 4.2002, 4.2003, 4.2004, 4.2003, 4.2006, 4.2003, 4.2002, 4.2002, 4.2002, 4.2003, 4.2003, 4.2003, 4.2002, 4.2003, 4.2002, 4.2003, 4.2002, 4.2003, 4.2003, 4.2005, 4.2006, 4.2005, 4.2006, 4.2007, 4.2005, 4.2005, 4.2005, 4.2007, 4.2006, 4.2005, 4.2007, 4.2007, 4.2007, 4.2007, 4.2006, 4.2004, 4.2005, 4.2006, 4.2006, 4.2006, 4.2005, 4.2006]

                        Percentage = [0.00,0.07,0.13,0.20,0.26,0.33,0.39,0.46,0.52,0.59,0.65,0.72,0.79,0.85,0.92,0.98,1.05,1.11,1.18,1.24,1.31,1.37,1.44,1.51,1.57,1.64,1.70,1.77,1.83,1.90,1.96,2.03,2.09,2.16,2.23,2.29,2.36,2.42,2.49,2.55,2.62,2.68,2.75,2.81,2.88,2.95,3.01,3.08,3.14,3.21,3.27,3.34,3.40,3.47,3.54,3.60,3.67,3.73,3.80,3.86,3.93,3.99,4.06,4.12,4.19,4.26,4.32,4.39,4.45,4.52,4.58,4.65,4.71,4.78,4.84,4.91,4.98,5.04,5.11,5.17,5.24,5.30,5.37,5.43,5.50,5.56,5.63,5.70,5.76,5.83,5.89,5.96,6.02,6.09,6.15,6.22,6.28,6.35,6.42,6.48,6.55,6.61,6.68,6.74,6.81,6.87,6.94,7.00,7.07,7.14,7.20,7.27,7.33,7.40,7.46,7.53,7.59,7.66,7.73,7.79,7.86,7.92,7.99,8.05,8.12,8.18,8.25,8.32,8.38,8.45,8.51,8.58,8.64,8.71,8.77,8.84,8.91,8.97,9.04,9.10,9.17,9.23,9.30,9.36,9.43,9.50,9.56,9.63,9.69,9.76,9.82,9.89,9.95,10.02,10.09,10.15,10.22,10.28,10.35,10.41,10.48,10.55,10.61,10.68,10.74,10.81,10.87,10.94,11.00,11.07,11.14,11.20,11.27,11.33,11.40,11.46,11.53,11.59,11.66,11.73,11.79,11.86,11.92,11.99,12.05,12.12,12.18,12.25,12.32,12.38,12.45,12.51,12.58,12.64,12.71,12.77,12.84,12.91,12.97,13.04,13.10,13.17,13.23,13.30,13.36,13.43,13.50,13.56,13.63,13.69,13.76,13.82,13.89,13.95,14.02,14.09,14.15,14.22,14.28,14.35,14.41,14.48,14.54,14.61,14.68,14.74,14.81,14.87,14.94,15.00,15.07,15.13,15.20,15.27,15.33,15.40,15.46,15.53,15.59,15.66,15.73,15.79,15.86,15.92,15.99,16.05,16.12,16.18,16.25,16.32,16.38,16.45,16.51,16.58,16.64,16.71,16.77,16.84,16.91,16.97,17.04,17.10,17.17,17.23,17.30,17.36,17.43,17.50,17.56,17.63,17.69,17.76,17.82,17.89,17.95,18.02,18.09,18.15,18.22,18.28,18.35,18.41,18.48,18.54,18.61,18.68,18.74,18.81,18.87,18.94,19.00,19.07,19.13,19.20,19.27,19.33,19.40,19.46,19.53,19.59,19.66,19.72,19.79,19.86,19.92,19.99,20.05,20.12,20.18,20.25,20.31,20.38,20.45,20.51,20.58,20.64,20.71,20.77,20.84,20.90,20.97,21.04,21.10,21.17,21.23,21.30,21.36,21.43,21.50,21.56,21.63,21.69,21.76,21.82,21.89,21.95,22.02,22.09,22.15,22.22,22.28,22.35,22.41,22.48,22.54,22.61,22.68,22.74,22.81,22.87,22.94,23.00,23.07,23.13,23.20,23.27,23.33,23.40,23.46,23.53,23.59,23.66,23.73,23.79,23.86,23.92,23.99,24.05,24.12,24.18,24.25,24.32,24.38,24.45,24.51,24.58,24.64,24.71,24.77,24.84,24.91,24.97,25.04,25.10,25.17,25.23,25.30,25.37,25.43,25.50,25.56,25.63,25.69,25.76,25.82,25.89,25.96,26.02,26.09,26.15,26.22,26.28,26.35,26.42,26.48,26.55,26.61,26.68,26.74,26.81,26.87,26.94,27.01,27.07,27.14,27.20,27.27,27.33,27.40,27.47,27.53,27.60,27.66,27.73,27.79,27.86,27.92,27.99,28.06,28.12,28.19,28.25,28.32,28.38,28.45,28.52,28.58,28.65,28.71,28.78,28.84,28.91,28.97,29.04,29.11,29.17,29.24,29.30,29.37,29.43,29.50,29.57,29.63,29.70,29.76,29.83,29.89,29.96,30.02,30.09,30.16,30.22,30.29,30.35,30.42,30.48,30.55,30.62,30.68,30.75,30.81,30.88,30.94,31.01,31.07,31.14,31.21,31.27,31.34,31.40,31.47,31.53,31.60,31.66,31.73,31.80,31.86,31.93,31.99,32.06,32.12,32.19,32.26,32.32,32.39,32.45,32.52,32.58,32.65,32.71,32.78,32.85,32.91,32.98,33.04,33.11,33.17,33.24,33.31,33.37,33.44,33.50,33.57,33.63,33.70,33.76,33.83,33.90,33.96,34.03,34.09,34.16,34.22,34.29,34.35,34.42,34.49,34.55,34.62,34.68,34.75,34.81,34.88,34.95,35.01,35.08,35.14,35.21,35.27,35.34,35.40,35.47,35.54,35.60,35.67,35.73,35.80,35.86,35.93,35.99,36.06,36.13,36.19,36.26,36.32,36.39,36.45,36.52,36.59,36.65,36.72,36.78,36.85,36.91,36.98,37.04,37.11,37.18,37.24,37.31,37.37,37.44,37.50,37.57,37.63,37.70,37.77,37.83,37.90,37.96,38.03,38.09,38.16,38.22,38.29,38.36,38.42,38.49,38.55,38.62,38.68,38.75,38.81,38.88,38.95,39.01,39.08,39.14,39.21,39.27,39.34,39.41,39.47,39.54,39.60,39.67,39.73,39.80,39.86,39.93,40.00,40.06,40.13,40.19,40.26,40.32,40.39,40.45,40.52,40.59,40.65,40.72,40.78,40.85,40.91,40.98,41.05,41.11,41.18,41.24,41.31,41.37,41.44,41.50,41.57,41.64,41.70,41.77,41.83,41.90,41.96,42.03,42.10,42.16,42.23,42.29,42.36,42.42,42.49,42.55,42.62,42.69,42.75,42.82,42.88,42.95,43.01,43.08,43.14,43.21,43.28,43.34,43.41,43.47,43.54,43.60,43.67,43.74,43.80,43.87,43.93,44.00,44.06,44.13,44.19,44.26,44.33,44.39,44.46,44.52,44.59,44.65,44.72,44.78,44.85,44.92,44.98,45.05,45.11,45.18,45.24,45.31,45.37,45.44,45.51,45.57,45.64,45.70,45.77,45.83,45.90,45.96,46.03,46.10,46.16,46.23,46.29,46.36,46.42,46.49,46.55,46.62,46.69,46.75,46.82,46.88,46.95,47.01,47.08,47.15,47.21,47.28,47.34,47.41,47.47,47.54,47.60,47.67,47.74,47.80,47.87,47.93,48.00,48.06,48.13,48.19,48.26,48.33,48.39,48.46,48.52,48.59,48.65,48.72,48.78,48.85,48.92,48.98,49.05,49.11,49.18,49.24,49.31,49.37,49.44,49.51,49.57,49.64,49.70,49.77,49.83,49.90,49.96,50.03,50.10,50.16,50.23,50.29,50.36,50.42,50.49,50.55,50.62,50.69,50.75,50.82,50.88,50.95,51.01,51.08,51.14,51.21,51.28,51.34,51.41,51.47,51.54,51.60,51.67,51.73,51.80,51.87,51.93,52.00,52.06,52.13,52.19,52.26,52.32,52.39,52.46,52.52,52.59,52.65,52.72,52.78,52.85,52.91,52.98,53.05,53.11,53.18,53.24,53.31,53.37,53.44,53.51,53.57,53.64,53.70,53.77,53.83,53.90,53.96,54.03,54.10,54.16,54.23,54.29,54.36,54.42,54.49,54.55,54.62,54.68,54.75,54.82,54.88,54.95,55.01,55.08,55.14,55.21,55.28,55.34,55.41,55.47,55.54,55.60,55.67,55.73,55.80,55.86,55.93,56.00,56.06,56.13,56.19,56.26,56.32,56.39,56.45,56.52,56.59,56.65,56.72,56.78,56.85,56.91,56.98,57.04,57.11,57.18,57.24,57.31,57.37,57.44,57.50,57.57,57.63,57.70,57.77,57.83,57.90,57.96,58.03,58.09,58.16,58.22,58.29,58.36,58.42,58.49,58.55,58.62,58.68,58.75,58.81,58.88,58.95,59.01,59.08,59.14,59.21,59.27,59.34,59.40,59.47,59.54,59.60,59.67,59.73,59.80,59.86,59.93,59.99,60.06,60.12,60.19,60.26,60.32,60.39,60.45,60.52,60.58,60.65,60.71,60.78,60.85,60.91,60.98,61.04,61.11,61.17,61.24,61.30,61.37,61.43,61.50,61.57,61.63,61.70,61.76,61.83,61.89,61.96,62.02,62.09,62.16,62.22,62.29,62.35,62.42,62.48,62.55,62.61,62.68,62.74,62.81,62.88,62.94,63.01,63.07,63.14,63.20,63.27,63.33,63.40,63.47,63.53,63.60,63.66,63.73,63.79,63.86,63.92,63.99,64.05,64.12,64.19,64.25,64.32,64.38,64.45,64.51,64.58,64.64,64.71,64.77,64.84,64.91,64.97,65.04,65.10,65.17,65.23,65.30,65.36,65.43,65.50,65.56,65.63,65.69,65.76,65.82,65.89,65.95,66.02,66.08,66.15,66.22,66.28,66.35,66.41,66.48,66.54,66.61,66.67,66.74,66.81,66.87,66.94,67.00,67.07,67.13,67.20,67.26,67.33,67.39,67.46,67.53,67.59,67.66,67.72,67.79,67.85,67.92,67.98,68.05,68.12,68.18,68.25,68.31,68.38,68.44,68.51,68.57,68.64,68.70,68.77,68.84,68.90,68.97,69.03,69.10,69.16,69.23,69.29,69.36,69.43,69.49,69.56,69.62,69.69,69.75,69.82,69.88,69.95,70.02,70.08,70.15,70.21,70.28,70.34,70.41,70.47,70.54,70.61,70.67,70.74,70.80,70.87,70.93,71.00,71.06,71.13,71.20,71.26,71.33,71.39,71.46,71.52,71.59,71.65,71.72,71.79,71.85,71.92,71.98,72.05,72.11,72.18,72.24,72.31,72.38,72.44,72.51,72.57,72.64,72.70,72.77,72.83,72.90,72.97,73.03,73.10,73.16,73.23,73.29,73.36,73.42,73.49,73.56,73.62,73.69,73.75,73.82,73.88,73.95,74.01,74.08,74.15,74.21,74.28,74.34,74.41,74.47,74.54,74.60,74.67,74.74,74.80,74.87,74.93,75.00,75.06,75.13,75.20,75.26,75.33,75.39,75.46,75.52,75.59,75.65,75.72,75.79,75.85,75.92,75.98,76.05,76.11,76.18,76.24,76.31,76.38,76.44,76.51,76.57,76.64,76.70,76.77,76.83,76.90,76.97,77.03,77.10,77.16,77.23,77.29,77.36,77.42,77.49,77.56,77.62,77.69,77.75,77.82,77.88,77.95,78.01,78.08,78.15,78.21,78.28,78.34,78.41,78.47,78.54,78.60,78.67,78.74,78.80,78.87,78.93,79.00,79.06,79.13,79.19,79.26,79.33,79.39,79.46,79.52,79.59,79.65,79.72,79.78,79.85,79.92,79.98,80.05,80.11,80.18,80.24,80.31,80.37,80.44,80.51,80.57,80.64,80.70,80.77,80.83,80.90,80.96,81.03,81.10,81.16,81.23,81.29,81.36,81.42,81.49,81.55,81.62,81.69,81.75,81.82,81.88,81.95,82.01,82.08,82.14,82.21,82.27,82.34,82.41,82.47,82.54,82.60,82.67,82.73,82.80,82.86,82.93,83.00,83.06,83.13,83.19,83.26,83.32,83.39,83.45,83.52,83.59,83.65,83.72,83.78,83.85,83.91,83.98,84.04,84.11,84.18,84.24,84.31,84.37,84.44,84.50,84.57,84.63,84.70,84.77,84.83,84.90,84.96,85.03,85.09,85.16,85.22,85.29,85.36,85.42,85.49,85.55,85.62,85.68,85.75,85.81,85.88,85.95,86.01,86.08,86.14,86.21,86.27,86.34,86.40,86.47,86.53,86.60,86.67,86.73,86.80,86.86,86.93,86.99,87.06,87.12,87.19,87.26,87.32,87.39,87.45,87.52,87.58,87.65,87.71,87.78,87.85,87.91,87.98,88.04,88.11,88.17,88.24,88.30,88.37,88.44,88.50,88.57,88.63,88.70,88.76,88.83,88.89,88.96,89.03,89.09,89.16,89.22,89.29,89.35,89.42,89.48,89.55,89.61,89.68,89.75,89.81,89.88,89.94,90.01,90.07,90.14,90.20,90.27,90.34,90.40,90.47,90.53,90.60,90.66,90.73,90.79,90.86,90.93,90.99,91.06,91.12,91.19,91.25,91.32,91.38,91.45,91.52,91.58,91.65,91.71,91.78,91.84,91.91,91.97,92.04,92.11,92.17,92.24,92.30,92.37,92.43,92.50,92.56,92.63,92.69,92.76,92.83,92.89,92.96,93.02,93.09,93.15,93.22,93.28,93.35,93.42,93.48,93.55,93.61,93.68,93.74,93.81,93.87,93.94,94.01,94.07,94.14,94.20,94.27,94.33,94.40,94.46,94.53,94.60,94.66,94.73,94.79,94.86,94.92,94.99,95.05,95.12,95.18,95.25,95.32,95.38,95.45,95.51,95.58,95.64,95.71,95.77,95.84,95.91,95.97,96.04,96.10,96.17,96.23,96.30,96.36,96.43,96.50,96.56,96.63,96.69,96.76,96.82,96.89,96.95,97.02,97.09,97.15,97.21,97.26,97.32,97.37,97.42,97.48,97.53,97.58,97.62,97.67,97.72,97.76,97.80,97.85,97.89,97.93,97.98,98.02,98.06,98.10,98.14,98.18,98.21,98.25,98.28,98.32,98.35,98.38,98.42,98.45,98.48,98.51,98.54,98.57,98.60,98.63,98.66,98.68,98.71,98.74,98.76,98.79,98.82,98.84,98.86,98.89,98.91,98.94,98.96,98.98,99.00,99.02,99.05,99.07,99.09,99.11,99.13,99.15,99.17,99.19,99.21,99.23,99.25,99.27,99.28,99.30,99.32,99.34,99.35,99.37,99.39,99.40,99.42,99.43,99.45,99.46,99.48,99.49,99.51,99.52,99.54,99.55,99.56,99.58,99.59,99.60,99.62,99.63,99.64,99.66,99.67,99.68,99.69,99.70,99.72,99.73,99.74,99.75,99.76,99.77,99.78,99.79,99.80,99.81,99.82,99.83,99.84,99.85,99.86,99.87,99.88,99.89,99.90,99.91,99.92,99.93,99.94,99.94,99.95,99.96,99.97,99.98,99.99,99.99,100.00,100.01,100.02,100.02,100.03,100.04,100.05,100.05,100.05,]

                        mapping = list(zip(Volt,Percentage))
                        # print(len(Volt))
                        # print(len(Percentage))
                        # print(len(mapping))

                        # Find the closest voltage in the list of Volt
                        closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                        # Find the index of the closest voltage
                        index = Volt.index(closest_voltage)
                        # print(index)
                        # print(Percentage[index])

                        # Check if the index is not the last index to avoid index out of range
                        if index < len(Volt) - 1:
                            if(input_voltage > closest_voltage) or index == 0  and input_voltage != 0:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")
                                # Calculate the slope using the closest_voltage and next_voltage
                                if next_voltage != closest_voltage :
                                    slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                    closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                else:
                                    next_voltage = Volt[index + 2]
                                    slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                    closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)

                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage < closest_voltage and input_voltage != 0:
                                next_voltage = Volt[index - 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage == 0:
                                closest_percentage = 0
                            else:
                                next_voltage = Volt[index + 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif index + 1 == len(Volt):

                            if input_voltage < closest_voltage and input_voltage != 0:
                                # print("This is herer 3")
                                next_voltage = Volt[index - 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage == 0:
                                closest_percentage = 0
                            elif input_voltage > Volt[index]:
                                # print(Volt[index])
                                closest_percentage = 0
                            else:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")                     
                        else:
                            print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                        return closest_percentage
                    def SOH2(input_voltage):
                        global closest_percentage2
                        closest_percentage2 = 0
                        # print("This is SOH2")
                        Volt = [ 2.7011, 2.7546, 2.7684, 2.7789, 2.7877, 2.7956, 2.8026, 2.8092, 2.8156, 2.8212, 2.827, 2.8327, 2.8378, 2.8432, 2.8482, 2.8533, 2.8584, 2.8632, 2.8678, 2.8727, 2.8775, 2.8823, 2.887, 2.8917, 2.8961, 2.9008, 2.9055, 2.9102, 2.9145, 2.9192, 2.9235, 2.9282, 2.9317, 2.9372, 2.9421, 2.9461, 2.9505, 2.9548, 2.9594, 2.9631, 2.9685, 2.9726, 2.9768, 2.9819, 2.9856, 2.9898, 2.9942, 2.998, 3.0025, 3.007, 3.0111, 3.0153, 3.0197, 3.0233, 3.0277, 3.0319, 3.0358, 3.0399, 3.044, 3.048, 3.052, 3.0559, 3.0598, 3.0637, 3.0675, 3.0715, 3.0754, 3.0789, 3.0828, 3.0868, 3.0903, 3.0942, 3.098, 3.1016, 3.1052, 3.1087, 3.1124, 3.116, 3.1195, 3.1231, 3.1266, 3.1299, 3.1336, 3.1373, 3.1405, 3.144, 3.1472, 3.1504, 3.1538, 3.1573, 3.1606, 3.1638, 3.167, 3.1703, 3.1739, 3.1766, 3.1797, 3.183, 3.1859, 3.1891, 3.1923, 3.1951, 3.1981, 3.2011, 3.2045, 3.2072, 3.21, 3.213, 3.2157, 3.2192, 3.2216, 3.2246, 3.2275, 3.2302, 3.233, 3.2359, 3.2385, 3.2408, 3.2436, 3.2463, 3.2489, 3.2515, 3.2541, 3.2566, 3.2592, 3.262, 3.2645, 3.2668, 3.2697, 3.2722, 3.2748, 3.2776, 3.2797, 3.2823, 3.2847, 3.2871, 3.2892, 3.292, 3.2946, 3.2965, 3.2989, 3.3014, 3.3036, 3.3056, 3.3085, 3.3106, 3.3129, 3.3153, 3.3177, 3.3199, 3.322, 3.3244, 3.3266, 3.3286, 3.3311, 3.3333, 3.3353, 3.3377, 3.3396, 3.3419, 3.3439, 3.346, 3.3483, 3.3503, 3.3521, 3.3544, 3.3566, 3.3585, 3.3606, 3.3625, 3.3645, 3.3667, 3.3684, 3.3705, 3.3725, 3.3744, 3.3763, 3.3784, 3.3802, 3.3822, 3.384, 3.3864, 3.3879, 3.3899, 3.3913, 3.3937, 3.3954, 3.3972, 3.399, 3.4009, 3.4027, 3.4046, 3.4064, 3.408, 3.4098, 3.4116, 3.4131, 3.4151, 3.4168, 3.4186, 3.4203, 3.4219, 3.4235, 3.4253, 3.4271, 3.4288, 3.4302, 3.432, 3.4339, 3.4355, 3.437, 3.4389, 3.4403, 3.442, 3.4436, 3.4451, 3.4469, 3.4483, 3.45, 3.4514, 3.4531, 3.4547, 3.4562, 3.4577, 3.4592, 3.4608, 3.4623, 3.4637, 3.4652, 3.4667, 3.4682, 3.4698, 3.4712, 3.4727, 3.4744, 3.4757, 3.4771, 3.4785, 3.48, 3.4814, 3.4829, 3.4844, 3.4857, 3.4871, 3.4884, 3.4899, 3.4912, 3.4926, 3.494, 3.4955, 3.4968, 3.498, 3.4994, 3.5009, 3.5021, 3.5036, 3.5046, 3.5058, 3.5076, 3.5086, 3.5101, 3.511, 3.5124, 3.5135, 3.5148, 3.5163, 3.5176, 3.5186, 3.52, 3.5213, 3.5225, 3.5236, 3.5248, 3.5257, 3.5271, 3.5284, 3.5296, 3.5309, 3.5318, 3.5327, 3.534, 3.5352, 3.5362, 3.5374, 3.5383, 3.539, 3.5401, 3.5412, 3.5422, 3.5431, 3.5435, 3.5446, 3.545, 3.5457, 3.5465, 3.547, 3.5475, 3.5484, 3.5489, 3.5495, 3.5499, 3.5503, 3.5506, 3.5513, 3.5516, 3.5524, 3.5527, 3.5529, 3.5533, 3.5537, 3.554, 3.5544, 3.5551, 3.5554, 3.5555, 3.5562, 3.5563, 3.5564, 3.5567, 3.5571, 3.5574, 3.5578, 3.5581, 3.5585, 3.5587, 3.5592, 3.5591, 3.5595, 3.5602, 3.5605, 3.5608, 3.5608, 3.561, 3.5616, 3.5617, 3.562, 3.5626, 3.5627, 3.5633, 3.5633, 3.5637, 3.5641, 3.5644, 3.5648, 3.5649, 3.5647, 3.5655, 3.5659, 3.5661, 3.5665, 3.5673, 3.5669, 3.5675, 3.5677, 3.5679, 3.5686, 3.569, 3.5689, 3.5698, 3.5701, 3.5703, 3.5707, 3.5711, 3.5712, 3.5716, 3.5723, 3.5724, 3.5727, 3.5732, 3.5737, 3.574, 3.5743, 3.5747, 3.5751, 3.5754, 3.576, 3.5763, 3.5767, 3.5771, 3.5773, 3.5781, 3.5784, 3.5789, 3.5792, 3.5793, 3.5802, 3.5805, 3.5808, 3.5813, 3.5817, 3.5821, 3.5827, 3.5832, 3.5837, 3.5841, 3.5847, 3.5852, 3.5857, 3.586, 3.5865, 3.5871, 3.5873, 3.5879, 3.5886, 3.5894, 3.5896, 3.59, 3.5906, 3.591, 3.5916, 3.5922, 3.5925, 3.5929, 3.594, 3.5942, 3.5943, 3.5953, 3.5955, 3.5965, 3.5968, 3.5974, 3.5977, 3.5986, 3.5986, 3.5996, 3.5995, 3.6005, 3.6008, 3.6018, 3.6019, 3.6026, 3.603, 3.6034, 3.604, 3.6048, 3.6051, 3.6057, 3.6062, 3.6066, 3.6074, 3.6078, 3.6084, 3.6089, 3.6095, 3.61, 3.6105, 3.6109, 3.6115, 3.612, 3.6127, 3.6133, 3.6136, 3.6142, 3.6146, 3.6152, 3.6157, 3.6163, 3.6169, 3.6173, 3.6178, 3.6184, 3.6189, 3.6195, 3.6199, 3.6204, 3.621, 3.6214, 3.6218, 3.6224, 3.6228, 3.6233, 3.6238, 3.6243, 3.6248, 3.6255, 3.6259, 3.6263, 3.627, 3.6273, 3.6278, 3.6284, 3.6287, 3.6292, 3.6296, 3.6301, 3.6306, 3.6311, 3.6316, 3.632, 3.6325, 3.633, 3.6334, 3.6339, 3.6343, 3.6348, 3.6352, 3.6358, 3.6361, 3.6364, 3.6369, 3.6374, 3.6379, 3.6382, 3.6389, 3.6392, 3.6393, 3.6401, 3.6405, 3.6407, 3.6413, 3.6416, 3.6421, 3.6425, 3.6432, 3.6434, 3.644, 3.6443, 3.6448, 3.645, 3.6454, 3.646, 3.6464, 3.647, 3.6475, 3.6478, 3.6478, 3.6483, 3.6492, 3.6494, 3.6498, 3.65, 3.6508, 3.6511, 3.6516, 3.652, 3.6524, 3.6527, 3.6532, 3.6537, 3.6542, 3.6548, 3.6551, 3.6554, 3.6559, 3.6564, 3.6568, 3.6573, 3.6578, 3.6581, 3.6586, 3.659, 3.6594, 3.6598, 3.66, 3.6606, 3.6612, 3.6613, 3.6621, 3.6626, 3.6629, 3.6634, 3.6637, 3.6641, 3.6645, 3.665, 3.6653, 3.6659, 3.666, 3.6664, 3.6669, 3.6675, 3.6677, 3.6682, 3.6686, 3.6688, 3.6694, 3.6698, 3.6701, 3.6706, 3.6709, 3.6714, 3.6717, 3.6719, 3.6726, 3.6727, 3.6732, 3.6735, 3.6739, 3.6745, 3.6748, 3.675, 3.6754, 3.6756, 3.676, 3.6765, 3.6769, 3.6771, 3.6775, 3.6777, 3.678, 3.6783, 3.6788, 3.6793, 3.6793, 3.6796, 3.68, 3.6804, 3.6805, 3.6811, 3.6812, 3.6817, 3.6819, 3.6822, 3.6825, 3.6827, 3.683, 3.6834, 3.6837, 3.6838, 3.6842, 3.6846, 3.6849, 3.685, 3.6852, 3.6858, 3.6861, 3.6862, 3.6864, 3.6867, 3.6873, 3.6874, 3.6875, 3.6878, 3.6882, 3.6884, 3.6886, 3.6889, 3.6892, 3.6893, 3.6897, 3.69, 3.6901, 3.6905, 3.6907, 3.6911, 3.6911, 3.6916, 3.6917, 3.6921, 3.6922, 3.6925, 3.6929, 3.6928, 3.6931, 3.6936, 3.6936, 3.6939, 3.6942, 3.6945, 3.6946, 3.6949, 3.6951, 3.6953, 3.6958, 3.6961, 3.6962, 3.6964, 3.6967, 3.6969, 3.6967, 3.6972, 3.6976, 3.6977, 3.6979, 3.6982, 3.6987, 3.6988, 3.6989, 3.6995, 3.6993, 3.6997, 3.7002, 3.7003, 3.7005, 3.7002, 3.7007, 3.7009, 3.7011, 3.7024, 3.7008, 3.7018, 3.7023, 3.7023, 3.7024, 3.7023, 3.7026, 3.7041, 3.7035, 3.7038, 3.7044, 3.7044, 3.7047, 3.705, 3.7062, 3.7053, 3.7056, 3.706, 3.7064, 3.7063, 3.7067, 3.7067, 3.707, 3.7073, 3.7076, 3.7081, 3.7078, 3.7081, 3.7086, 3.7088, 3.7089, 3.7094, 3.7096, 3.7098, 3.7101, 3.7102, 3.7104, 3.7105, 3.711, 3.7112, 3.7114, 3.7115, 3.712, 3.7123, 3.7127, 3.7129, 3.7131, 3.7134, 3.7135, 3.714, 3.7142, 3.7146, 3.7149, 3.7149, 3.7153, 3.7156, 3.7158, 3.7161, 3.7164, 3.7168, 3.7169, 3.7171, 3.7177, 3.7178, 3.718, 3.7183, 3.7188, 3.719, 3.7192, 3.7194, 3.7199, 3.7201, 3.7202, 3.7205, 3.721, 3.7213, 3.7217, 3.7219, 3.7222, 3.7225, 3.7227, 3.7231, 3.7233, 3.7235, 3.7239, 3.7242, 3.7245, 3.7249, 3.7252, 3.7254, 3.7259, 3.7261, 3.7263, 3.7268, 3.727, 3.7272, 3.7278, 3.7281, 3.7283, 3.7285, 3.7291, 3.7294, 3.7297, 3.7299, 3.7302, 3.7307, 3.7308, 3.7314, 3.7316, 3.7319, 3.7324, 3.7329, 3.7329, 3.7332, 3.7336, 3.7341, 3.7343, 3.7347, 3.735, 3.7353, 3.7358, 3.7361, 3.7365, 3.7368, 3.737, 3.7378, 3.738, 3.7382, 3.7387, 3.739, 3.7394, 3.7397, 3.7402, 3.7405, 3.7408, 3.7413, 3.7416, 3.7421, 3.7422, 3.7428, 3.7432, 3.7435, 3.7439, 3.7443, 3.7448, 3.7451, 3.7457, 3.7458, 3.7464, 3.7467, 3.7473, 3.7475, 3.7479, 3.7484, 3.7487, 3.749, 3.7495, 3.7498, 3.7504, 3.7505, 3.7513, 3.7518, 3.7519, 3.7524, 3.7529, 3.7534, 3.7537, 3.7542, 3.7545, 3.7548, 3.7554, 3.7559, 3.7562, 3.7566, 3.7572, 3.7576, 3.7581, 3.7585, 3.759, 3.7593, 3.7598, 3.7602, 3.7606, 3.7612, 3.7616, 3.7622, 3.7626, 3.763, 3.7634, 3.7638, 3.7645, 3.7648, 3.7653, 3.7658, 3.7662, 3.767, 3.7672, 3.7677, 3.7682, 3.7685, 3.769, 3.7695, 3.7701, 3.7706, 3.7711, 3.7716, 3.772, 3.7726, 3.7728, 3.7736, 3.774, 3.7744, 3.775, 3.7755, 3.776, 3.7766, 3.7769, 3.7774, 3.778, 3.7787, 3.779, 3.7797, 3.78, 3.7806, 3.7812, 3.7817, 3.7823, 3.783, 3.7834, 3.7839, 3.7844, 3.7849, 3.7854, 3.7859, 3.7865, 3.7872, 3.7876, 3.7882, 3.7887, 3.7894, 3.7898, 3.7907, 3.7914, 3.7919, 3.7924, 3.793, 3.7936, 3.7943, 3.7947, 3.7953, 3.7958, 3.7964, 3.7969, 3.7975, 3.7982, 3.7989, 3.7994, 3.7999, 3.8006, 3.801, 3.8018, 3.8023, 3.8028, 3.8035, 3.804, 3.8047, 3.8053, 3.8059, 3.8066, 3.807, 3.8078, 3.8083, 3.8089, 3.8096, 3.8102, 3.8109, 3.8114, 3.8121, 3.8127, 3.8133, 3.8141, 3.8148, 3.8153, 3.8159, 3.8164, 3.8172, 3.8178, 3.8186, 3.8193, 3.8197, 3.8205, 3.821, 3.8216, 3.8225, 3.8231, 3.8238, 3.8245, 3.8252, 3.8258, 3.8265, 3.8271, 3.8279, 3.8283, 3.8292, 3.8299, 3.8306, 3.8314, 3.8319, 3.8329, 3.8333, 3.8338, 3.8348, 3.8355, 3.8362, 3.8368, 3.8376, 3.8384, 3.8392, 3.8399, 3.8406, 3.8411, 3.842, 3.8428, 3.8432, 3.8441, 3.8449, 3.8458, 3.8464, 3.8471, 3.848, 3.8488, 3.8492, 3.8502, 3.8507, 3.8514, 3.8524, 3.8531, 3.8539, 3.8548, 3.8555, 3.8561, 3.8569, 3.858, 3.8587, 3.8593, 3.86, 3.8609, 3.8616, 3.8625, 3.8635, 3.8641, 3.8649, 3.8658, 3.8664, 3.8674, 3.8679, 3.869, 3.8698, 3.8705, 3.8712, 3.8721, 3.8725, 3.8735, 3.8743, 3.875, 3.8758, 3.8767, 3.8774, 3.8784, 3.8792, 3.8798, 3.8808, 3.8815, 3.8828, 3.8838, 3.8849, 3.885, 3.8856, 3.8864, 3.887, 3.888, 3.8888, 3.8897, 3.8904, 3.8914, 3.892, 3.8928, 3.8938, 3.8944, 3.8952, 3.8961, 3.8968, 3.8978, 3.8983, 3.8991, 3.8999, 3.9008, 3.9014, 3.9022, 3.9031, 3.9039, 3.9045, 3.9053, 3.906, 3.9069, 3.9076, 3.9084, 3.9092, 3.9099, 3.9108, 3.9114, 3.9121, 3.9129, 3.9136, 3.9145, 3.9153, 3.9159, 3.9167, 3.9175, 3.918, 3.9188, 3.9196, 3.9203, 3.921, 3.9218, 3.9224, 3.9233, 3.924, 3.9247, 3.9254, 3.9261, 3.9268, 3.9276, 3.9282, 3.929, 3.9297, 3.9305, 3.931, 3.9319, 3.9328, 3.9334, 3.934, 3.9349, 3.9355, 3.9361, 3.9369, 3.9376, 3.9383, 3.9391, 3.9398, 3.9405, 3.9413, 3.9418, 3.9425, 3.9435, 3.944, 3.9446, 3.9456, 3.9462, 3.9469, 3.9477, 3.9483, 3.9491, 3.9496, 3.9505, 3.9511, 3.9518, 3.9526, 3.9533, 3.954, 3.9547, 3.9554, 3.9561, 3.9566, 3.9575, 3.9582, 3.9588, 3.9595, 3.9603, 3.9611, 3.9618, 3.9624, 3.9631, 3.964, 3.9647, 3.9652, 3.966, 3.9667, 3.9676, 3.968, 3.9688, 3.9695, 3.9702, 3.9709, 3.9716, 3.9723, 3.973, 3.9737, 3.9745, 3.9751, 3.9759, 3.9767, 3.9773, 3.9779, 3.9788, 3.9794, 3.9801, 3.981, 3.9816, 3.9823, 3.983, 3.9837, 3.9846, 3.9851, 3.9856, 3.9864, 3.9871, 3.9879, 3.9885, 3.9892, 3.9901, 3.9908, 3.9914, 3.9921, 3.9929, 3.9935, 3.9943, 3.9949, 3.9958, 3.9964, 3.9972, 3.9979, 3.9985, 3.9994, 4, 4.0007, 4.0013, 4.0022, 4.0029, 4.0038, 4.0042, 4.005, 4.0058, 4.0065, 4.0072, 4.008, 4.0086, 4.0093, 4.0099, 4.0108, 4.0116, 4.0122, 4.0129, 4.0135, 4.0144, 4.0151, 4.0158, 4.0164, 4.0172, 4.0178, 4.0187, 4.0194, 4.0201, 4.021, 4.0215, 4.0223, 4.0231, 4.0239, 4.0244, 4.0252, 4.026, 4.0268, 4.0273, 4.028, 4.0288, 4.0295, 4.0304, 4.0311, 4.0318, 4.0325, 4.0333, 4.034, 4.0347, 4.0356, 4.0362, 4.0371, 4.0377, 4.0385, 4.0391, 4.0399, 4.0405, 4.0414, 4.042, 4.0428, 4.0436, 4.0443, 4.0449, 4.0458, 4.0465, 4.0473, 4.048, 4.0488, 4.0495, 4.0503, 4.051, 4.0517, 4.0525, 4.0531, 4.054, 4.0547, 4.0554, 4.0563, 4.057, 4.0578, 4.0585, 4.0592, 4.0601, 4.0607, 4.0616, 4.0621, 4.0631, 4.0637, 4.0645, 4.0652, 4.0659, 4.0667, 4.0675, 4.0683, 4.069, 4.0697, 4.0706, 4.0714, 4.0721, 4.0728, 4.0735, 4.0744, 4.0751, 4.0758, 4.0766, 4.0773, 4.0781, 4.079, 4.0797, 4.0805, 4.0812, 4.0821, 4.0827, 4.0834, 4.0843, 4.0852, 4.0858, 4.0867, 4.0872, 4.0882, 4.089, 4.0897, 4.0905, 4.0912, 4.0919, 4.0928, 4.0936, 4.0944, 4.0951, 4.0959, 4.0965, 4.0974, 4.0982, 4.099, 4.1, 4.1006, 4.1015, 4.1021, 4.1029, 4.1037, 4.1044, 4.1053, 4.1062, 4.1069, 4.1076, 4.1084, 4.1091, 4.1099, 4.1108, 4.1114, 4.1123, 4.113, 4.1139, 4.1146, 4.1154, 4.1162, 4.117, 4.1179, 4.1186, 4.1194, 4.1202, 4.1209, 4.1218, 4.1225, 4.1235, 4.1243, 4.1252, 4.1257, 4.1266, 4.1273, 4.1282, 4.129, 4.1297, 4.1306, 4.1312, 4.1321, 4.1329, 4.1337, 4.1346, 4.1354, 4.1361, 4.1369, 4.1379, 4.1386, 4.1392, 4.1402, 4.1409, 4.1417, 4.1427, 4.1433, 4.1442, 4.1451, 4.1458, 4.1467, 4.1474, 4.1483, 4.149, 4.1498, 4.1507, 4.1514, 4.1522, 4.1531, 4.1538, 4.1548, 4.1555, 4.1564, 4.1571, 4.1579, 4.1586, 4.1596, 4.1604, 4.1611, 4.1621, 4.1628, 4.1637, 4.1645, 4.1653, 4.1661, 4.1669, 4.1677, 4.1684, 4.1693, 4.1701, 4.1711, 4.1719, 4.1726, 4.1735, 4.1743, 4.1752, 4.1759, 4.1768, 4.1775, 4.1785, 4.1793, 4.1802, 4.181, 4.1817, 4.1825, 4.1834, 4.1843, 4.1849, 4.186, 4.1867, 4.1876, 4.1884, 4.1894, 4.1901, 4.1909, 4.1917, 4.1925, 4.1935, 4.1944, 4.1951, 4.1959, 4.1967, 4.1975, 4.1985, 4.1992, 4.2, 4.1992, 4.1992, 4.1991, 4.1991, 4.1993, 4.1991, 4.1993, 4.1993, 4.1991, 4.1993, 4.1991, 4.1992, 4.199, 4.1991, 4.1994, 4.1996, 4.2001, 4.2005, 4.2009, 4.2002, 4.2, 4.2001, 4.2002, 4.2001, 4.2001, 4.2, 4.2001, 4.2002, 4.2001, 4.2003, 4.2, 4.2001, 4.2002, 4.2002, 4.2002, 4.2002, 4.2001, 4.2, 4.2001, 4.2, 4.2001, 4.2, 4.2002, 4.2003, 4.2003, 4.2002, 4.2, 4.2001, 4.2001, 4.2002, 4.2002, 4.2001, 4.2003, 4.2001, 4.2002, 4.2001, 4.2001, 4.2004, 4.2003, 4.2003, 4.2006, 4.2006, 4.2008, 4.2004, 4.2005, 4.2004, 4.2004, 4.2003, 4.2004, 4.2005, 4.2004, 4.2001, 4.2004, 4.2004, 4.2005, 4.2004, 4.2004, 4.2004, 4.2004, 4.2004, 4.2003, 4.2003, 4.2004, 4.2001, 4.2003, 4.2003, 4.2003, 4.2002, 4.2003, 4.2002, 4.2001, 4.2004, 4.2002, 4.2003, 4.2004, 4.2003, 4.2006, 4.2003, 4.2002, 4.2002, 4.2002, 4.2003, 4.2003, 4.2003, 4.2002, 4.2003, 4.2002, 4.2003, 4.2002, 4.2003, 4.2003, 4.2005, 4.2006, 4.2005, 4.2006, 4.2007, 4.2005, 4.2005, 4.2005, 4.2007, 4.2006, 4.2005, 4.2007, 4.2007, 4.2007, 4.2007, 4.2006, 4.2004, 4.2005, 4.2006, 4.2006, 4.2006, 4.2005, 4.2006]

                        Percentage = [0.00,0.07,0.13,0.20,0.26,0.33,0.39,0.46,0.52,0.59,0.65,0.72,0.79,0.85,0.92,0.98,1.05,1.11,1.18,1.24,1.31,1.37,1.44,1.51,1.57,1.64,1.70,1.77,1.83,1.90,1.96,2.03,2.09,2.16,2.23,2.29,2.36,2.42,2.49,2.55,2.62,2.68,2.75,2.81,2.88,2.95,3.01,3.08,3.14,3.21,3.27,3.34,3.40,3.47,3.54,3.60,3.67,3.73,3.80,3.86,3.93,3.99,4.06,4.12,4.19,4.26,4.32,4.39,4.45,4.52,4.58,4.65,4.71,4.78,4.84,4.91,4.98,5.04,5.11,5.17,5.24,5.30,5.37,5.43,5.50,5.56,5.63,5.70,5.76,5.83,5.89,5.96,6.02,6.09,6.15,6.22,6.28,6.35,6.42,6.48,6.55,6.61,6.68,6.74,6.81,6.87,6.94,7.00,7.07,7.14,7.20,7.27,7.33,7.40,7.46,7.53,7.59,7.66,7.73,7.79,7.86,7.92,7.99,8.05,8.12,8.18,8.25,8.32,8.38,8.45,8.51,8.58,8.64,8.71,8.77,8.84,8.91,8.97,9.04,9.10,9.17,9.23,9.30,9.36,9.43,9.50,9.56,9.63,9.69,9.76,9.82,9.89,9.95,10.02,10.09,10.15,10.22,10.28,10.35,10.41,10.48,10.55,10.61,10.68,10.74,10.81,10.87,10.94,11.00,11.07,11.14,11.20,11.27,11.33,11.40,11.46,11.53,11.59,11.66,11.73,11.79,11.86,11.92,11.99,12.05,12.12,12.18,12.25,12.32,12.38,12.45,12.51,12.58,12.64,12.71,12.77,12.84,12.91,12.97,13.04,13.10,13.17,13.23,13.30,13.36,13.43,13.50,13.56,13.63,13.69,13.76,13.82,13.89,13.95,14.02,14.09,14.15,14.22,14.28,14.35,14.41,14.48,14.54,14.61,14.68,14.74,14.81,14.87,14.94,15.00,15.07,15.13,15.20,15.27,15.33,15.40,15.46,15.53,15.59,15.66,15.73,15.79,15.86,15.92,15.99,16.05,16.12,16.18,16.25,16.32,16.38,16.45,16.51,16.58,16.64,16.71,16.77,16.84,16.91,16.97,17.04,17.10,17.17,17.23,17.30,17.36,17.43,17.50,17.56,17.63,17.69,17.76,17.82,17.89,17.95,18.02,18.09,18.15,18.22,18.28,18.35,18.41,18.48,18.54,18.61,18.68,18.74,18.81,18.87,18.94,19.00,19.07,19.13,19.20,19.27,19.33,19.40,19.46,19.53,19.59,19.66,19.72,19.79,19.86,19.92,19.99,20.05,20.12,20.18,20.25,20.31,20.38,20.45,20.51,20.58,20.64,20.71,20.77,20.84,20.90,20.97,21.04,21.10,21.17,21.23,21.30,21.36,21.43,21.50,21.56,21.63,21.69,21.76,21.82,21.89,21.95,22.02,22.09,22.15,22.22,22.28,22.35,22.41,22.48,22.54,22.61,22.68,22.74,22.81,22.87,22.94,23.00,23.07,23.13,23.20,23.27,23.33,23.40,23.46,23.53,23.59,23.66,23.73,23.79,23.86,23.92,23.99,24.05,24.12,24.18,24.25,24.32,24.38,24.45,24.51,24.58,24.64,24.71,24.77,24.84,24.91,24.97,25.04,25.10,25.17,25.23,25.30,25.37,25.43,25.50,25.56,25.63,25.69,25.76,25.82,25.89,25.96,26.02,26.09,26.15,26.22,26.28,26.35,26.42,26.48,26.55,26.61,26.68,26.74,26.81,26.87,26.94,27.01,27.07,27.14,27.20,27.27,27.33,27.40,27.47,27.53,27.60,27.66,27.73,27.79,27.86,27.92,27.99,28.06,28.12,28.19,28.25,28.32,28.38,28.45,28.52,28.58,28.65,28.71,28.78,28.84,28.91,28.97,29.04,29.11,29.17,29.24,29.30,29.37,29.43,29.50,29.57,29.63,29.70,29.76,29.83,29.89,29.96,30.02,30.09,30.16,30.22,30.29,30.35,30.42,30.48,30.55,30.62,30.68,30.75,30.81,30.88,30.94,31.01,31.07,31.14,31.21,31.27,31.34,31.40,31.47,31.53,31.60,31.66,31.73,31.80,31.86,31.93,31.99,32.06,32.12,32.19,32.26,32.32,32.39,32.45,32.52,32.58,32.65,32.71,32.78,32.85,32.91,32.98,33.04,33.11,33.17,33.24,33.31,33.37,33.44,33.50,33.57,33.63,33.70,33.76,33.83,33.90,33.96,34.03,34.09,34.16,34.22,34.29,34.35,34.42,34.49,34.55,34.62,34.68,34.75,34.81,34.88,34.95,35.01,35.08,35.14,35.21,35.27,35.34,35.40,35.47,35.54,35.60,35.67,35.73,35.80,35.86,35.93,35.99,36.06,36.13,36.19,36.26,36.32,36.39,36.45,36.52,36.59,36.65,36.72,36.78,36.85,36.91,36.98,37.04,37.11,37.18,37.24,37.31,37.37,37.44,37.50,37.57,37.63,37.70,37.77,37.83,37.90,37.96,38.03,38.09,38.16,38.22,38.29,38.36,38.42,38.49,38.55,38.62,38.68,38.75,38.81,38.88,38.95,39.01,39.08,39.14,39.21,39.27,39.34,39.41,39.47,39.54,39.60,39.67,39.73,39.80,39.86,39.93,40.00,40.06,40.13,40.19,40.26,40.32,40.39,40.45,40.52,40.59,40.65,40.72,40.78,40.85,40.91,40.98,41.05,41.11,41.18,41.24,41.31,41.37,41.44,41.50,41.57,41.64,41.70,41.77,41.83,41.90,41.96,42.03,42.10,42.16,42.23,42.29,42.36,42.42,42.49,42.55,42.62,42.69,42.75,42.82,42.88,42.95,43.01,43.08,43.14,43.21,43.28,43.34,43.41,43.47,43.54,43.60,43.67,43.74,43.80,43.87,43.93,44.00,44.06,44.13,44.19,44.26,44.33,44.39,44.46,44.52,44.59,44.65,44.72,44.78,44.85,44.92,44.98,45.05,45.11,45.18,45.24,45.31,45.37,45.44,45.51,45.57,45.64,45.70,45.77,45.83,45.90,45.96,46.03,46.10,46.16,46.23,46.29,46.36,46.42,46.49,46.55,46.62,46.69,46.75,46.82,46.88,46.95,47.01,47.08,47.15,47.21,47.28,47.34,47.41,47.47,47.54,47.60,47.67,47.74,47.80,47.87,47.93,48.00,48.06,48.13,48.19,48.26,48.33,48.39,48.46,48.52,48.59,48.65,48.72,48.78,48.85,48.92,48.98,49.05,49.11,49.18,49.24,49.31,49.37,49.44,49.51,49.57,49.64,49.70,49.77,49.83,49.90,49.96,50.03,50.10,50.16,50.23,50.29,50.36,50.42,50.49,50.55,50.62,50.69,50.75,50.82,50.88,50.95,51.01,51.08,51.14,51.21,51.28,51.34,51.41,51.47,51.54,51.60,51.67,51.73,51.80,51.87,51.93,52.00,52.06,52.13,52.19,52.26,52.32,52.39,52.46,52.52,52.59,52.65,52.72,52.78,52.85,52.91,52.98,53.05,53.11,53.18,53.24,53.31,53.37,53.44,53.51,53.57,53.64,53.70,53.77,53.83,53.90,53.96,54.03,54.10,54.16,54.23,54.29,54.36,54.42,54.49,54.55,54.62,54.68,54.75,54.82,54.88,54.95,55.01,55.08,55.14,55.21,55.28,55.34,55.41,55.47,55.54,55.60,55.67,55.73,55.80,55.86,55.93,56.00,56.06,56.13,56.19,56.26,56.32,56.39,56.45,56.52,56.59,56.65,56.72,56.78,56.85,56.91,56.98,57.04,57.11,57.18,57.24,57.31,57.37,57.44,57.50,57.57,57.63,57.70,57.77,57.83,57.90,57.96,58.03,58.09,58.16,58.22,58.29,58.36,58.42,58.49,58.55,58.62,58.68,58.75,58.81,58.88,58.95,59.01,59.08,59.14,59.21,59.27,59.34,59.40,59.47,59.54,59.60,59.67,59.73,59.80,59.86,59.93,59.99,60.06,60.12,60.19,60.26,60.32,60.39,60.45,60.52,60.58,60.65,60.71,60.78,60.85,60.91,60.98,61.04,61.11,61.17,61.24,61.30,61.37,61.43,61.50,61.57,61.63,61.70,61.76,61.83,61.89,61.96,62.02,62.09,62.16,62.22,62.29,62.35,62.42,62.48,62.55,62.61,62.68,62.74,62.81,62.88,62.94,63.01,63.07,63.14,63.20,63.27,63.33,63.40,63.47,63.53,63.60,63.66,63.73,63.79,63.86,63.92,63.99,64.05,64.12,64.19,64.25,64.32,64.38,64.45,64.51,64.58,64.64,64.71,64.77,64.84,64.91,64.97,65.04,65.10,65.17,65.23,65.30,65.36,65.43,65.50,65.56,65.63,65.69,65.76,65.82,65.89,65.95,66.02,66.08,66.15,66.22,66.28,66.35,66.41,66.48,66.54,66.61,66.67,66.74,66.81,66.87,66.94,67.00,67.07,67.13,67.20,67.26,67.33,67.39,67.46,67.53,67.59,67.66,67.72,67.79,67.85,67.92,67.98,68.05,68.12,68.18,68.25,68.31,68.38,68.44,68.51,68.57,68.64,68.70,68.77,68.84,68.90,68.97,69.03,69.10,69.16,69.23,69.29,69.36,69.43,69.49,69.56,69.62,69.69,69.75,69.82,69.88,69.95,70.02,70.08,70.15,70.21,70.28,70.34,70.41,70.47,70.54,70.61,70.67,70.74,70.80,70.87,70.93,71.00,71.06,71.13,71.20,71.26,71.33,71.39,71.46,71.52,71.59,71.65,71.72,71.79,71.85,71.92,71.98,72.05,72.11,72.18,72.24,72.31,72.38,72.44,72.51,72.57,72.64,72.70,72.77,72.83,72.90,72.97,73.03,73.10,73.16,73.23,73.29,73.36,73.42,73.49,73.56,73.62,73.69,73.75,73.82,73.88,73.95,74.01,74.08,74.15,74.21,74.28,74.34,74.41,74.47,74.54,74.60,74.67,74.74,74.80,74.87,74.93,75.00,75.06,75.13,75.20,75.26,75.33,75.39,75.46,75.52,75.59,75.65,75.72,75.79,75.85,75.92,75.98,76.05,76.11,76.18,76.24,76.31,76.38,76.44,76.51,76.57,76.64,76.70,76.77,76.83,76.90,76.97,77.03,77.10,77.16,77.23,77.29,77.36,77.42,77.49,77.56,77.62,77.69,77.75,77.82,77.88,77.95,78.01,78.08,78.15,78.21,78.28,78.34,78.41,78.47,78.54,78.60,78.67,78.74,78.80,78.87,78.93,79.00,79.06,79.13,79.19,79.26,79.33,79.39,79.46,79.52,79.59,79.65,79.72,79.78,79.85,79.92,79.98,80.05,80.11,80.18,80.24,80.31,80.37,80.44,80.51,80.57,80.64,80.70,80.77,80.83,80.90,80.96,81.03,81.10,81.16,81.23,81.29,81.36,81.42,81.49,81.55,81.62,81.69,81.75,81.82,81.88,81.95,82.01,82.08,82.14,82.21,82.27,82.34,82.41,82.47,82.54,82.60,82.67,82.73,82.80,82.86,82.93,83.00,83.06,83.13,83.19,83.26,83.32,83.39,83.45,83.52,83.59,83.65,83.72,83.78,83.85,83.91,83.98,84.04,84.11,84.18,84.24,84.31,84.37,84.44,84.50,84.57,84.63,84.70,84.77,84.83,84.90,84.96,85.03,85.09,85.16,85.22,85.29,85.36,85.42,85.49,85.55,85.62,85.68,85.75,85.81,85.88,85.95,86.01,86.08,86.14,86.21,86.27,86.34,86.40,86.47,86.53,86.60,86.67,86.73,86.80,86.86,86.93,86.99,87.06,87.12,87.19,87.26,87.32,87.39,87.45,87.52,87.58,87.65,87.71,87.78,87.85,87.91,87.98,88.04,88.11,88.17,88.24,88.30,88.37,88.44,88.50,88.57,88.63,88.70,88.76,88.83,88.89,88.96,89.03,89.09,89.16,89.22,89.29,89.35,89.42,89.48,89.55,89.61,89.68,89.75,89.81,89.88,89.94,90.01,90.07,90.14,90.20,90.27,90.34,90.40,90.47,90.53,90.60,90.66,90.73,90.79,90.86,90.93,90.99,91.06,91.12,91.19,91.25,91.32,91.38,91.45,91.52,91.58,91.65,91.71,91.78,91.84,91.91,91.97,92.04,92.11,92.17,92.24,92.30,92.37,92.43,92.50,92.56,92.63,92.69,92.76,92.83,92.89,92.96,93.02,93.09,93.15,93.22,93.28,93.35,93.42,93.48,93.55,93.61,93.68,93.74,93.81,93.87,93.94,94.01,94.07,94.14,94.20,94.27,94.33,94.40,94.46,94.53,94.60,94.66,94.73,94.79,94.86,94.92,94.99,95.05,95.12,95.18,95.25,95.32,95.38,95.45,95.51,95.58,95.64,95.71,95.77,95.84,95.91,95.97,96.04,96.10,96.17,96.23,96.30,96.36,96.43,96.50,96.56,96.63,96.69,96.76,96.82,96.89,96.95,97.02,97.09,97.15,97.21,97.26,97.32,97.37,97.42,97.48,97.53,97.58,97.62,97.67,97.72,97.76,97.80,97.85,97.89,97.93,97.98,98.02,98.06,98.10,98.14,98.18,98.21,98.25,98.28,98.32,98.35,98.38,98.42,98.45,98.48,98.51,98.54,98.57,98.60,98.63,98.66,98.68,98.71,98.74,98.76,98.79,98.82,98.84,98.86,98.89,98.91,98.94,98.96,98.98,99.00,99.02,99.05,99.07,99.09,99.11,99.13,99.15,99.17,99.19,99.21,99.23,99.25,99.27,99.28,99.30,99.32,99.34,99.35,99.37,99.39,99.40,99.42,99.43,99.45,99.46,99.48,99.49,99.51,99.52,99.54,99.55,99.56,99.58,99.59,99.60,99.62,99.63,99.64,99.66,99.67,99.68,99.69,99.70,99.72,99.73,99.74,99.75,99.76,99.77,99.78,99.79,99.80,99.81,99.82,99.83,99.84,99.85,99.86,99.87,99.88,99.89,99.90,99.91,99.92,99.93,99.94,99.94,99.95,99.96,99.97,99.98,99.99,99.99,100.00,100.01,100.02,100.02,100.03,100.04,100.05,100.05,100.05,]

                        mapping = list(zip(Volt,Percentage))
                        # print(len(Volt))
                        # print(len(Percentage))
                        # print(len(mapping))

                        # Find the closest voltage in the list of Volt
                        closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                        # Find the index of the closest voltage
                        index = Volt.index(closest_voltage)
                        # print(index)
                        # print(Percentage[index])

                        # Check if the index is not the last index to avoid index out of range
                        if index < len(Volt) - 1:
                            # Get the next voltage value
                            if(input_voltage > closest_voltage) or index == 0 and input_voltage != 0:
                                next_voltage = Volt[index + 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                            elif input_voltage < closest_voltage and input_voltage != 0:
                                next_voltage = Volt[index - 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                                closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                            elif input_voltage == 0:
                                closest_percentage2 = 0
                            else:
                                next_voltage = Volt[index + 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")            
                        elif index + 1 == len(Volt):

                            if input_voltage < closest_voltage and input_voltage != 0:
                            #    print("This is herer 3")
                                next_voltage = Volt[index - 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                                closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                            elif input_voltage == 0:
                                closest_percentage2 = 0
                            elif input_voltage > Volt[index]:
                                # print(Volt[index])
                                closest_percentage2 = 0
                            else:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")          
                                            
                        else:
                            print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                    # Find indices of 'Group1' in the 'Group' list


                    # Sort the array
                    sorted_array = sorted(Charge_Set['Group'])

                    # Get unique numbers using set
                    unique_numbers = set(Charge_Set['Group'])

                    # Convert the unique numbers back to a sorted list if needed
                    sorted_unique_numbers = sorted(list(unique_numbers))
                    first_min_v = None
                    last_max_v = None
                    # Print the sorted unique numbers
                    # print(sorted_unique_numbers)
                    shitty = 1
                    for numbers in sorted_unique_numbers:
                        group_indices = [i for i, group in enumerate(Charge_Set['Group']) if group == numbers]

                        # Check if there are 'Group1' elements in the 'Group' list
                        if group_indices:
                            # Access the first and last 'Max_V' values for 'Group1'
                            first_min_v = Charge_Set['Min_V'][group_indices[0]]
                            last_max_v = Charge_Set['Max_V'][group_indices[-1]]

                            # print(f"First Max_V for 'Group1': {first_min_v}")
                            # print(f"Last Max_V for 'Group1': {last_max_v}")
                        else:
                            first_min_v = 0
                            last_max_v = 0
                        # print(f"First Max_V for 'Group1': {first_min_v}")
                        # print(f"Last Max_V for 'Group1': {last_max_v}")
                        SOH_Goop['Goop'].append(numbers)
                        SOH_Goop['Min_V'].append(first_min_v)
                        SOH_Goop['Max_V'].append(last_max_v)
                        thread1 = threading.Thread(target=SOH, args=(first_min_v,))
                        # print(f'This is last_m:{last_max_v}')
                        thread2 = threading.Thread(target=SOH2, args=(last_max_v,))
                        thread1.start()
                        thread2.start()
                        thread1.join()
                        thread2.join()
                        # print(f'This is SOH2:{closest_percentage2}')
                        # Calculate energy for 'Group1' where 'Group' is 0
                # Find indices of 'Group1' where 'Group' is 0
                        
                        energy_sum = 0
                        save_data_start = None
                        a= None
                        # Iterate over the numerical indices of Charge_Set['Group']
                        for index, group in enumerate(Charge_Set['Group']):
                            if group == numbers:
                                voltage = Charge_Set['Voltage'][index]
                                current = Charge_Set['Current'][index]
                                time_diff = Charge_Set['Time_Diff'][index]
                                # Split the original string by space to get the time portion
                                teiam = str( Charge_Set['Timestamps'][index])
                                # print(teiam)
                                split_string = teiam.split(" ")

                                # Check if there are at least two parts (date and time)
                                if len(split_string) >= 2:
                                    # Join the time portion and discard the date
                                    time_portion = " ".join(split_string[1:])
                                    # print(time_portion)

                                if save_data_start == None:
                                
                                    save_data_start = time_portion
                                    SOH_Goop['S_Time'].append(time_portion)
                                    
                                save_end = time_portion
                                # print(f'This is time diff { leg}')
                                # print(index)  # Assuming 'Diff' represents time intervals
                                # print(f'This is V {voltage}')
                                # print(f'This is C {current}')
                                # print(f'This is D {time_diff}')
                                if index != 0:
                                    energy =  abs(0.5*(current+previous_current))  * time_diff
                                    previous_current = current
                                    # print(f'This is Energy {energy}')
                                    # cell = sheetPackProcess.cell(row=gginp, column=28, value= energy)
                                    # gginp += 1
                                else :
                                    energy = 0
                                    previous_current = current
                                energy_sum += energy
                        SOH_Goop['E_Time'].append(save_end)
                        energy_sum = energy_sum* voltage
                        # print(f'Total energy for "Group1": {energy_sum} Joules')
                            

                        # print(f"Total energy for 'Group1' where 'Group' is 0: {energy_sum} Joules")
                        
                        SOC_Start = closest_percentage
                        SOC_End = closest_percentage2
                        Difference_SOC = SOC_End - SOC_Start
                        DesignCapacity_NH02 = 26.8
                        SOh_E = energy_sum/(3600*1000)
                        SOH_Goop['Start_SOC'].append(SOC_Start)
                        SOH_Goop['End_SOC'].append(SOC_End)
                        SOH_Goop['Charge'].append(SOh_E)
                        if Difference_SOC != 0:
                            Cal_Capacity = SOh_E/(Difference_SOC/100)
                        else:
                            Cal_Capacity = 0
                        if(Cal_Capacity != 0):
                            Remaining_Capacity = (Cal_Capacity/DesignCapacity_NH02)*100
                        elif(Cal_Capacity == None):
                            Remaining_Capacity = 0
                        else:
                            Remaining_Capacity = 0
                        SOH_Goop['Cal'].append(Remaining_Capacity)
                        SOH_Goop['SOH'].append(Cal_Capacity)
                        SOH_Goop['Cal_E'].append(SOh_E)
                        
                        # for index, header in enumerate(headers1, start=shitty):
                        #     cell = sheet8.cell(row=4, column=index, value=header)

                        # cell = sheet8.cell(row=the_loop, column=shitty, value=first_min_v)
                        # cell = sheet8.cell(row=the_loop, column=shitty+1, value=last_max_v)
                        # cell = sheet8.cell(row=the_loop, column=shitty+2, value=SOC_Start)
                        # cell = sheet8.cell(row=the_loop, column=shitty+3, value=SOC_End)
                        # cell = sheet8.cell(row=the_loop, column=shitty+4, value=Cal_Capacity)
                        # cell = sheet8.cell(row=the_loop, column=shitty+5, value=Remaining_Capacity)    
                        # cell = sheet8.cell(row=the_loop, column=shitty+6, value=SOh_E)  
                        shitty += 7
                    if first_min_v == None and last_max_v == None:
                        lost_data.append(app+1)
                        lost_data3.append(app+1)
                    # print(lost_data)
                def SOH(input_voltage):
                        global closest_percentage
                        closest_percentage = 0
                        Volt = [ 2.7011, 2.7546, 2.7684, 2.7789, 2.7877, 2.7956, 2.8026, 2.8092, 2.8156, 2.8212, 2.827, 2.8327, 2.8378, 2.8432, 2.8482, 2.8533, 2.8584, 2.8632, 2.8678, 2.8727, 2.8775, 2.8823, 2.887, 2.8917, 2.8961, 2.9008, 2.9055, 2.9102, 2.9145, 2.9192, 2.9235, 2.9282, 2.9317, 2.9372, 2.9421, 2.9461, 2.9505, 2.9548, 2.9594, 2.9631, 2.9685, 2.9726, 2.9768, 2.9819, 2.9856, 2.9898, 2.9942, 2.998, 3.0025, 3.007, 3.0111, 3.0153, 3.0197, 3.0233, 3.0277, 3.0319, 3.0358, 3.0399, 3.044, 3.048, 3.052, 3.0559, 3.0598, 3.0637, 3.0675, 3.0715, 3.0754, 3.0789, 3.0828, 3.0868, 3.0903, 3.0942, 3.098, 3.1016, 3.1052, 3.1087, 3.1124, 3.116, 3.1195, 3.1231, 3.1266, 3.1299, 3.1336, 3.1373, 3.1405, 3.144, 3.1472, 3.1504, 3.1538, 3.1573, 3.1606, 3.1638, 3.167, 3.1703, 3.1739, 3.1766, 3.1797, 3.183, 3.1859, 3.1891, 3.1923, 3.1951, 3.1981, 3.2011, 3.2045, 3.2072, 3.21, 3.213, 3.2157, 3.2192, 3.2216, 3.2246, 3.2275, 3.2302, 3.233, 3.2359, 3.2385, 3.2408, 3.2436, 3.2463, 3.2489, 3.2515, 3.2541, 3.2566, 3.2592, 3.262, 3.2645, 3.2668, 3.2697, 3.2722, 3.2748, 3.2776, 3.2797, 3.2823, 3.2847, 3.2871, 3.2892, 3.292, 3.2946, 3.2965, 3.2989, 3.3014, 3.3036, 3.3056, 3.3085, 3.3106, 3.3129, 3.3153, 3.3177, 3.3199, 3.322, 3.3244, 3.3266, 3.3286, 3.3311, 3.3333, 3.3353, 3.3377, 3.3396, 3.3419, 3.3439, 3.346, 3.3483, 3.3503, 3.3521, 3.3544, 3.3566, 3.3585, 3.3606, 3.3625, 3.3645, 3.3667, 3.3684, 3.3705, 3.3725, 3.3744, 3.3763, 3.3784, 3.3802, 3.3822, 3.384, 3.3864, 3.3879, 3.3899, 3.3913, 3.3937, 3.3954, 3.3972, 3.399, 3.4009, 3.4027, 3.4046, 3.4064, 3.408, 3.4098, 3.4116, 3.4131, 3.4151, 3.4168, 3.4186, 3.4203, 3.4219, 3.4235, 3.4253, 3.4271, 3.4288, 3.4302, 3.432, 3.4339, 3.4355, 3.437, 3.4389, 3.4403, 3.442, 3.4436, 3.4451, 3.4469, 3.4483, 3.45, 3.4514, 3.4531, 3.4547, 3.4562, 3.4577, 3.4592, 3.4608, 3.4623, 3.4637, 3.4652, 3.4667, 3.4682, 3.4698, 3.4712, 3.4727, 3.4744, 3.4757, 3.4771, 3.4785, 3.48, 3.4814, 3.4829, 3.4844, 3.4857, 3.4871, 3.4884, 3.4899, 3.4912, 3.4926, 3.494, 3.4955, 3.4968, 3.498, 3.4994, 3.5009, 3.5021, 3.5036, 3.5046, 3.5058, 3.5076, 3.5086, 3.5101, 3.511, 3.5124, 3.5135, 3.5148, 3.5163, 3.5176, 3.5186, 3.52, 3.5213, 3.5225, 3.5236, 3.5248, 3.5257, 3.5271, 3.5284, 3.5296, 3.5309, 3.5318, 3.5327, 3.534, 3.5352, 3.5362, 3.5374, 3.5383, 3.539, 3.5401, 3.5412, 3.5422, 3.5431, 3.5435, 3.5446, 3.545, 3.5457, 3.5465, 3.547, 3.5475, 3.5484, 3.5489, 3.5495, 3.5499, 3.5503, 3.5506, 3.5513, 3.5516, 3.5524, 3.5527, 3.5529, 3.5533, 3.5537, 3.554, 3.5544, 3.5551, 3.5554, 3.5555, 3.5562, 3.5563, 3.5564, 3.5567, 3.5571, 3.5574, 3.5578, 3.5581, 3.5585, 3.5587, 3.5592, 3.5591, 3.5595, 3.5602, 3.5605, 3.5608, 3.5608, 3.561, 3.5616, 3.5617, 3.562, 3.5626, 3.5627, 3.5633, 3.5633, 3.5637, 3.5641, 3.5644, 3.5648, 3.5649, 3.5647, 3.5655, 3.5659, 3.5661, 3.5665, 3.5673, 3.5669, 3.5675, 3.5677, 3.5679, 3.5686, 3.569, 3.5689, 3.5698, 3.5701, 3.5703, 3.5707, 3.5711, 3.5712, 3.5716, 3.5723, 3.5724, 3.5727, 3.5732, 3.5737, 3.574, 3.5743, 3.5747, 3.5751, 3.5754, 3.576, 3.5763, 3.5767, 3.5771, 3.5773, 3.5781, 3.5784, 3.5789, 3.5792, 3.5793, 3.5802, 3.5805, 3.5808, 3.5813, 3.5817, 3.5821, 3.5827, 3.5832, 3.5837, 3.5841, 3.5847, 3.5852, 3.5857, 3.586, 3.5865, 3.5871, 3.5873, 3.5879, 3.5886, 3.5894, 3.5896, 3.59, 3.5906, 3.591, 3.5916, 3.5922, 3.5925, 3.5929, 3.594, 3.5942, 3.5943, 3.5953, 3.5955, 3.5965, 3.5968, 3.5974, 3.5977, 3.5986, 3.5986, 3.5996, 3.5995, 3.6005, 3.6008, 3.6018, 3.6019, 3.6026, 3.603, 3.6034, 3.604, 3.6048, 3.6051, 3.6057, 3.6062, 3.6066, 3.6074, 3.6078, 3.6084, 3.6089, 3.6095, 3.61, 3.6105, 3.6109, 3.6115, 3.612, 3.6127, 3.6133, 3.6136, 3.6142, 3.6146, 3.6152, 3.6157, 3.6163, 3.6169, 3.6173, 3.6178, 3.6184, 3.6189, 3.6195, 3.6199, 3.6204, 3.621, 3.6214, 3.6218, 3.6224, 3.6228, 3.6233, 3.6238, 3.6243, 3.6248, 3.6255, 3.6259, 3.6263, 3.627, 3.6273, 3.6278, 3.6284, 3.6287, 3.6292, 3.6296, 3.6301, 3.6306, 3.6311, 3.6316, 3.632, 3.6325, 3.633, 3.6334, 3.6339, 3.6343, 3.6348, 3.6352, 3.6358, 3.6361, 3.6364, 3.6369, 3.6374, 3.6379, 3.6382, 3.6389, 3.6392, 3.6393, 3.6401, 3.6405, 3.6407, 3.6413, 3.6416, 3.6421, 3.6425, 3.6432, 3.6434, 3.644, 3.6443, 3.6448, 3.645, 3.6454, 3.646, 3.6464, 3.647, 3.6475, 3.6478, 3.6478, 3.6483, 3.6492, 3.6494, 3.6498, 3.65, 3.6508, 3.6511, 3.6516, 3.652, 3.6524, 3.6527, 3.6532, 3.6537, 3.6542, 3.6548, 3.6551, 3.6554, 3.6559, 3.6564, 3.6568, 3.6573, 3.6578, 3.6581, 3.6586, 3.659, 3.6594, 3.6598, 3.66, 3.6606, 3.6612, 3.6613, 3.6621, 3.6626, 3.6629, 3.6634, 3.6637, 3.6641, 3.6645, 3.665, 3.6653, 3.6659, 3.666, 3.6664, 3.6669, 3.6675, 3.6677, 3.6682, 3.6686, 3.6688, 3.6694, 3.6698, 3.6701, 3.6706, 3.6709, 3.6714, 3.6717, 3.6719, 3.6726, 3.6727, 3.6732, 3.6735, 3.6739, 3.6745, 3.6748, 3.675, 3.6754, 3.6756, 3.676, 3.6765, 3.6769, 3.6771, 3.6775, 3.6777, 3.678, 3.6783, 3.6788, 3.6793, 3.6793, 3.6796, 3.68, 3.6804, 3.6805, 3.6811, 3.6812, 3.6817, 3.6819, 3.6822, 3.6825, 3.6827, 3.683, 3.6834, 3.6837, 3.6838, 3.6842, 3.6846, 3.6849, 3.685, 3.6852, 3.6858, 3.6861, 3.6862, 3.6864, 3.6867, 3.6873, 3.6874, 3.6875, 3.6878, 3.6882, 3.6884, 3.6886, 3.6889, 3.6892, 3.6893, 3.6897, 3.69, 3.6901, 3.6905, 3.6907, 3.6911, 3.6911, 3.6916, 3.6917, 3.6921, 3.6922, 3.6925, 3.6929, 3.6928, 3.6931, 3.6936, 3.6936, 3.6939, 3.6942, 3.6945, 3.6946, 3.6949, 3.6951, 3.6953, 3.6958, 3.6961, 3.6962, 3.6964, 3.6967, 3.6969, 3.6967, 3.6972, 3.6976, 3.6977, 3.6979, 3.6982, 3.6987, 3.6988, 3.6989, 3.6995, 3.6993, 3.6997, 3.7002, 3.7003, 3.7005, 3.7002, 3.7007, 3.7009, 3.7011, 3.7024, 3.7008, 3.7018, 3.7023, 3.7023, 3.7024, 3.7023, 3.7026, 3.7041, 3.7035, 3.7038, 3.7044, 3.7044, 3.7047, 3.705, 3.7062, 3.7053, 3.7056, 3.706, 3.7064, 3.7063, 3.7067, 3.7067, 3.707, 3.7073, 3.7076, 3.7081, 3.7078, 3.7081, 3.7086, 3.7088, 3.7089, 3.7094, 3.7096, 3.7098, 3.7101, 3.7102, 3.7104, 3.7105, 3.711, 3.7112, 3.7114, 3.7115, 3.712, 3.7123, 3.7127, 3.7129, 3.7131, 3.7134, 3.7135, 3.714, 3.7142, 3.7146, 3.7149, 3.7149, 3.7153, 3.7156, 3.7158, 3.7161, 3.7164, 3.7168, 3.7169, 3.7171, 3.7177, 3.7178, 3.718, 3.7183, 3.7188, 3.719, 3.7192, 3.7194, 3.7199, 3.7201, 3.7202, 3.7205, 3.721, 3.7213, 3.7217, 3.7219, 3.7222, 3.7225, 3.7227, 3.7231, 3.7233, 3.7235, 3.7239, 3.7242, 3.7245, 3.7249, 3.7252, 3.7254, 3.7259, 3.7261, 3.7263, 3.7268, 3.727, 3.7272, 3.7278, 3.7281, 3.7283, 3.7285, 3.7291, 3.7294, 3.7297, 3.7299, 3.7302, 3.7307, 3.7308, 3.7314, 3.7316, 3.7319, 3.7324, 3.7329, 3.7329, 3.7332, 3.7336, 3.7341, 3.7343, 3.7347, 3.735, 3.7353, 3.7358, 3.7361, 3.7365, 3.7368, 3.737, 3.7378, 3.738, 3.7382, 3.7387, 3.739, 3.7394, 3.7397, 3.7402, 3.7405, 3.7408, 3.7413, 3.7416, 3.7421, 3.7422, 3.7428, 3.7432, 3.7435, 3.7439, 3.7443, 3.7448, 3.7451, 3.7457, 3.7458, 3.7464, 3.7467, 3.7473, 3.7475, 3.7479, 3.7484, 3.7487, 3.749, 3.7495, 3.7498, 3.7504, 3.7505, 3.7513, 3.7518, 3.7519, 3.7524, 3.7529, 3.7534, 3.7537, 3.7542, 3.7545, 3.7548, 3.7554, 3.7559, 3.7562, 3.7566, 3.7572, 3.7576, 3.7581, 3.7585, 3.759, 3.7593, 3.7598, 3.7602, 3.7606, 3.7612, 3.7616, 3.7622, 3.7626, 3.763, 3.7634, 3.7638, 3.7645, 3.7648, 3.7653, 3.7658, 3.7662, 3.767, 3.7672, 3.7677, 3.7682, 3.7685, 3.769, 3.7695, 3.7701, 3.7706, 3.7711, 3.7716, 3.772, 3.7726, 3.7728, 3.7736, 3.774, 3.7744, 3.775, 3.7755, 3.776, 3.7766, 3.7769, 3.7774, 3.778, 3.7787, 3.779, 3.7797, 3.78, 3.7806, 3.7812, 3.7817, 3.7823, 3.783, 3.7834, 3.7839, 3.7844, 3.7849, 3.7854, 3.7859, 3.7865, 3.7872, 3.7876, 3.7882, 3.7887, 3.7894, 3.7898, 3.7907, 3.7914, 3.7919, 3.7924, 3.793, 3.7936, 3.7943, 3.7947, 3.7953, 3.7958, 3.7964, 3.7969, 3.7975, 3.7982, 3.7989, 3.7994, 3.7999, 3.8006, 3.801, 3.8018, 3.8023, 3.8028, 3.8035, 3.804, 3.8047, 3.8053, 3.8059, 3.8066, 3.807, 3.8078, 3.8083, 3.8089, 3.8096, 3.8102, 3.8109, 3.8114, 3.8121, 3.8127, 3.8133, 3.8141, 3.8148, 3.8153, 3.8159, 3.8164, 3.8172, 3.8178, 3.8186, 3.8193, 3.8197, 3.8205, 3.821, 3.8216, 3.8225, 3.8231, 3.8238, 3.8245, 3.8252, 3.8258, 3.8265, 3.8271, 3.8279, 3.8283, 3.8292, 3.8299, 3.8306, 3.8314, 3.8319, 3.8329, 3.8333, 3.8338, 3.8348, 3.8355, 3.8362, 3.8368, 3.8376, 3.8384, 3.8392, 3.8399, 3.8406, 3.8411, 3.842, 3.8428, 3.8432, 3.8441, 3.8449, 3.8458, 3.8464, 3.8471, 3.848, 3.8488, 3.8492, 3.8502, 3.8507, 3.8514, 3.8524, 3.8531, 3.8539, 3.8548, 3.8555, 3.8561, 3.8569, 3.858, 3.8587, 3.8593, 3.86, 3.8609, 3.8616, 3.8625, 3.8635, 3.8641, 3.8649, 3.8658, 3.8664, 3.8674, 3.8679, 3.869, 3.8698, 3.8705, 3.8712, 3.8721, 3.8725, 3.8735, 3.8743, 3.875, 3.8758, 3.8767, 3.8774, 3.8784, 3.8792, 3.8798, 3.8808, 3.8815, 3.8828, 3.8838, 3.8849, 3.885, 3.8856, 3.8864, 3.887, 3.888, 3.8888, 3.8897, 3.8904, 3.8914, 3.892, 3.8928, 3.8938, 3.8944, 3.8952, 3.8961, 3.8968, 3.8978, 3.8983, 3.8991, 3.8999, 3.9008, 3.9014, 3.9022, 3.9031, 3.9039, 3.9045, 3.9053, 3.906, 3.9069, 3.9076, 3.9084, 3.9092, 3.9099, 3.9108, 3.9114, 3.9121, 3.9129, 3.9136, 3.9145, 3.9153, 3.9159, 3.9167, 3.9175, 3.918, 3.9188, 3.9196, 3.9203, 3.921, 3.9218, 3.9224, 3.9233, 3.924, 3.9247, 3.9254, 3.9261, 3.9268, 3.9276, 3.9282, 3.929, 3.9297, 3.9305, 3.931, 3.9319, 3.9328, 3.9334, 3.934, 3.9349, 3.9355, 3.9361, 3.9369, 3.9376, 3.9383, 3.9391, 3.9398, 3.9405, 3.9413, 3.9418, 3.9425, 3.9435, 3.944, 3.9446, 3.9456, 3.9462, 3.9469, 3.9477, 3.9483, 3.9491, 3.9496, 3.9505, 3.9511, 3.9518, 3.9526, 3.9533, 3.954, 3.9547, 3.9554, 3.9561, 3.9566, 3.9575, 3.9582, 3.9588, 3.9595, 3.9603, 3.9611, 3.9618, 3.9624, 3.9631, 3.964, 3.9647, 3.9652, 3.966, 3.9667, 3.9676, 3.968, 3.9688, 3.9695, 3.9702, 3.9709, 3.9716, 3.9723, 3.973, 3.9737, 3.9745, 3.9751, 3.9759, 3.9767, 3.9773, 3.9779, 3.9788, 3.9794, 3.9801, 3.981, 3.9816, 3.9823, 3.983, 3.9837, 3.9846, 3.9851, 3.9856, 3.9864, 3.9871, 3.9879, 3.9885, 3.9892, 3.9901, 3.9908, 3.9914, 3.9921, 3.9929, 3.9935, 3.9943, 3.9949, 3.9958, 3.9964, 3.9972, 3.9979, 3.9985, 3.9994, 4, 4.0007, 4.0013, 4.0022, 4.0029, 4.0038, 4.0042, 4.005, 4.0058, 4.0065, 4.0072, 4.008, 4.0086, 4.0093, 4.0099, 4.0108, 4.0116, 4.0122, 4.0129, 4.0135, 4.0144, 4.0151, 4.0158, 4.0164, 4.0172, 4.0178, 4.0187, 4.0194, 4.0201, 4.021, 4.0215, 4.0223, 4.0231, 4.0239, 4.0244, 4.0252, 4.026, 4.0268, 4.0273, 4.028, 4.0288, 4.0295, 4.0304, 4.0311, 4.0318, 4.0325, 4.0333, 4.034, 4.0347, 4.0356, 4.0362, 4.0371, 4.0377, 4.0385, 4.0391, 4.0399, 4.0405, 4.0414, 4.042, 4.0428, 4.0436, 4.0443, 4.0449, 4.0458, 4.0465, 4.0473, 4.048, 4.0488, 4.0495, 4.0503, 4.051, 4.0517, 4.0525, 4.0531, 4.054, 4.0547, 4.0554, 4.0563, 4.057, 4.0578, 4.0585, 4.0592, 4.0601, 4.0607, 4.0616, 4.0621, 4.0631, 4.0637, 4.0645, 4.0652, 4.0659, 4.0667, 4.0675, 4.0683, 4.069, 4.0697, 4.0706, 4.0714, 4.0721, 4.0728, 4.0735, 4.0744, 4.0751, 4.0758, 4.0766, 4.0773, 4.0781, 4.079, 4.0797, 4.0805, 4.0812, 4.0821, 4.0827, 4.0834, 4.0843, 4.0852, 4.0858, 4.0867, 4.0872, 4.0882, 4.089, 4.0897, 4.0905, 4.0912, 4.0919, 4.0928, 4.0936, 4.0944, 4.0951, 4.0959, 4.0965, 4.0974, 4.0982, 4.099, 4.1, 4.1006, 4.1015, 4.1021, 4.1029, 4.1037, 4.1044, 4.1053, 4.1062, 4.1069, 4.1076, 4.1084, 4.1091, 4.1099, 4.1108, 4.1114, 4.1123, 4.113, 4.1139, 4.1146, 4.1154, 4.1162, 4.117, 4.1179, 4.1186, 4.1194, 4.1202, 4.1209, 4.1218, 4.1225, 4.1235, 4.1243, 4.1252, 4.1257, 4.1266, 4.1273, 4.1282, 4.129, 4.1297, 4.1306, 4.1312, 4.1321, 4.1329, 4.1337, 4.1346, 4.1354, 4.1361, 4.1369, 4.1379, 4.1386, 4.1392, 4.1402, 4.1409, 4.1417, 4.1427, 4.1433, 4.1442, 4.1451, 4.1458, 4.1467, 4.1474, 4.1483, 4.149, 4.1498, 4.1507, 4.1514, 4.1522, 4.1531, 4.1538, 4.1548, 4.1555, 4.1564, 4.1571, 4.1579, 4.1586, 4.1596, 4.1604, 4.1611, 4.1621, 4.1628, 4.1637, 4.1645, 4.1653, 4.1661, 4.1669, 4.1677, 4.1684, 4.1693, 4.1701, 4.1711, 4.1719, 4.1726, 4.1735, 4.1743, 4.1752, 4.1759, 4.1768, 4.1775, 4.1785, 4.1793, 4.1802, 4.181, 4.1817, 4.1825, 4.1834, 4.1843, 4.1849, 4.186, 4.1867, 4.1876, 4.1884, 4.1894, 4.1901, 4.1909, 4.1917, 4.1925, 4.1935, 4.1944, 4.1951, 4.1959, 4.1967, 4.1975, 4.1985, 4.1992, 4.2, 4.1992, 4.1992, 4.1991, 4.1991, 4.1993, 4.1991, 4.1993, 4.1993, 4.1991, 4.1993, 4.1991, 4.1992, 4.199, 4.1991, 4.1994, 4.1996, 4.2001, 4.2005, 4.2009, 4.2002, 4.2, 4.2001, 4.2002, 4.2001, 4.2001, 4.2, 4.2001, 4.2002, 4.2001, 4.2003, 4.2, 4.2001, 4.2002, 4.2002, 4.2002, 4.2002, 4.2001, 4.2, 4.2001, 4.2, 4.2001, 4.2, 4.2002, 4.2003, 4.2003, 4.2002, 4.2, 4.2001, 4.2001, 4.2002, 4.2002, 4.2001, 4.2003, 4.2001, 4.2002, 4.2001, 4.2001, 4.2004, 4.2003, 4.2003, 4.2006, 4.2006, 4.2008, 4.2004, 4.2005, 4.2004, 4.2004, 4.2003, 4.2004, 4.2005, 4.2004, 4.2001, 4.2004, 4.2004, 4.2005, 4.2004, 4.2004, 4.2004, 4.2004, 4.2004, 4.2003, 4.2003, 4.2004, 4.2001, 4.2003, 4.2003, 4.2003, 4.2002, 4.2003, 4.2002, 4.2001, 4.2004, 4.2002, 4.2003, 4.2004, 4.2003, 4.2006, 4.2003, 4.2002, 4.2002, 4.2002, 4.2003, 4.2003, 4.2003, 4.2002, 4.2003, 4.2002, 4.2003, 4.2002, 4.2003, 4.2003, 4.2005, 4.2006, 4.2005, 4.2006, 4.2007, 4.2005, 4.2005, 4.2005, 4.2007, 4.2006, 4.2005, 4.2007, 4.2007, 4.2007, 4.2007, 4.2006, 4.2004, 4.2005, 4.2006, 4.2006, 4.2006, 4.2005, 4.2006]

                        Percentage = [0.00,0.07,0.13,0.20,0.26,0.33,0.39,0.46,0.52,0.59,0.65,0.72,0.79,0.85,0.92,0.98,1.05,1.11,1.18,1.24,1.31,1.37,1.44,1.51,1.57,1.64,1.70,1.77,1.83,1.90,1.96,2.03,2.09,2.16,2.23,2.29,2.36,2.42,2.49,2.55,2.62,2.68,2.75,2.81,2.88,2.95,3.01,3.08,3.14,3.21,3.27,3.34,3.40,3.47,3.54,3.60,3.67,3.73,3.80,3.86,3.93,3.99,4.06,4.12,4.19,4.26,4.32,4.39,4.45,4.52,4.58,4.65,4.71,4.78,4.84,4.91,4.98,5.04,5.11,5.17,5.24,5.30,5.37,5.43,5.50,5.56,5.63,5.70,5.76,5.83,5.89,5.96,6.02,6.09,6.15,6.22,6.28,6.35,6.42,6.48,6.55,6.61,6.68,6.74,6.81,6.87,6.94,7.00,7.07,7.14,7.20,7.27,7.33,7.40,7.46,7.53,7.59,7.66,7.73,7.79,7.86,7.92,7.99,8.05,8.12,8.18,8.25,8.32,8.38,8.45,8.51,8.58,8.64,8.71,8.77,8.84,8.91,8.97,9.04,9.10,9.17,9.23,9.30,9.36,9.43,9.50,9.56,9.63,9.69,9.76,9.82,9.89,9.95,10.02,10.09,10.15,10.22,10.28,10.35,10.41,10.48,10.55,10.61,10.68,10.74,10.81,10.87,10.94,11.00,11.07,11.14,11.20,11.27,11.33,11.40,11.46,11.53,11.59,11.66,11.73,11.79,11.86,11.92,11.99,12.05,12.12,12.18,12.25,12.32,12.38,12.45,12.51,12.58,12.64,12.71,12.77,12.84,12.91,12.97,13.04,13.10,13.17,13.23,13.30,13.36,13.43,13.50,13.56,13.63,13.69,13.76,13.82,13.89,13.95,14.02,14.09,14.15,14.22,14.28,14.35,14.41,14.48,14.54,14.61,14.68,14.74,14.81,14.87,14.94,15.00,15.07,15.13,15.20,15.27,15.33,15.40,15.46,15.53,15.59,15.66,15.73,15.79,15.86,15.92,15.99,16.05,16.12,16.18,16.25,16.32,16.38,16.45,16.51,16.58,16.64,16.71,16.77,16.84,16.91,16.97,17.04,17.10,17.17,17.23,17.30,17.36,17.43,17.50,17.56,17.63,17.69,17.76,17.82,17.89,17.95,18.02,18.09,18.15,18.22,18.28,18.35,18.41,18.48,18.54,18.61,18.68,18.74,18.81,18.87,18.94,19.00,19.07,19.13,19.20,19.27,19.33,19.40,19.46,19.53,19.59,19.66,19.72,19.79,19.86,19.92,19.99,20.05,20.12,20.18,20.25,20.31,20.38,20.45,20.51,20.58,20.64,20.71,20.77,20.84,20.90,20.97,21.04,21.10,21.17,21.23,21.30,21.36,21.43,21.50,21.56,21.63,21.69,21.76,21.82,21.89,21.95,22.02,22.09,22.15,22.22,22.28,22.35,22.41,22.48,22.54,22.61,22.68,22.74,22.81,22.87,22.94,23.00,23.07,23.13,23.20,23.27,23.33,23.40,23.46,23.53,23.59,23.66,23.73,23.79,23.86,23.92,23.99,24.05,24.12,24.18,24.25,24.32,24.38,24.45,24.51,24.58,24.64,24.71,24.77,24.84,24.91,24.97,25.04,25.10,25.17,25.23,25.30,25.37,25.43,25.50,25.56,25.63,25.69,25.76,25.82,25.89,25.96,26.02,26.09,26.15,26.22,26.28,26.35,26.42,26.48,26.55,26.61,26.68,26.74,26.81,26.87,26.94,27.01,27.07,27.14,27.20,27.27,27.33,27.40,27.47,27.53,27.60,27.66,27.73,27.79,27.86,27.92,27.99,28.06,28.12,28.19,28.25,28.32,28.38,28.45,28.52,28.58,28.65,28.71,28.78,28.84,28.91,28.97,29.04,29.11,29.17,29.24,29.30,29.37,29.43,29.50,29.57,29.63,29.70,29.76,29.83,29.89,29.96,30.02,30.09,30.16,30.22,30.29,30.35,30.42,30.48,30.55,30.62,30.68,30.75,30.81,30.88,30.94,31.01,31.07,31.14,31.21,31.27,31.34,31.40,31.47,31.53,31.60,31.66,31.73,31.80,31.86,31.93,31.99,32.06,32.12,32.19,32.26,32.32,32.39,32.45,32.52,32.58,32.65,32.71,32.78,32.85,32.91,32.98,33.04,33.11,33.17,33.24,33.31,33.37,33.44,33.50,33.57,33.63,33.70,33.76,33.83,33.90,33.96,34.03,34.09,34.16,34.22,34.29,34.35,34.42,34.49,34.55,34.62,34.68,34.75,34.81,34.88,34.95,35.01,35.08,35.14,35.21,35.27,35.34,35.40,35.47,35.54,35.60,35.67,35.73,35.80,35.86,35.93,35.99,36.06,36.13,36.19,36.26,36.32,36.39,36.45,36.52,36.59,36.65,36.72,36.78,36.85,36.91,36.98,37.04,37.11,37.18,37.24,37.31,37.37,37.44,37.50,37.57,37.63,37.70,37.77,37.83,37.90,37.96,38.03,38.09,38.16,38.22,38.29,38.36,38.42,38.49,38.55,38.62,38.68,38.75,38.81,38.88,38.95,39.01,39.08,39.14,39.21,39.27,39.34,39.41,39.47,39.54,39.60,39.67,39.73,39.80,39.86,39.93,40.00,40.06,40.13,40.19,40.26,40.32,40.39,40.45,40.52,40.59,40.65,40.72,40.78,40.85,40.91,40.98,41.05,41.11,41.18,41.24,41.31,41.37,41.44,41.50,41.57,41.64,41.70,41.77,41.83,41.90,41.96,42.03,42.10,42.16,42.23,42.29,42.36,42.42,42.49,42.55,42.62,42.69,42.75,42.82,42.88,42.95,43.01,43.08,43.14,43.21,43.28,43.34,43.41,43.47,43.54,43.60,43.67,43.74,43.80,43.87,43.93,44.00,44.06,44.13,44.19,44.26,44.33,44.39,44.46,44.52,44.59,44.65,44.72,44.78,44.85,44.92,44.98,45.05,45.11,45.18,45.24,45.31,45.37,45.44,45.51,45.57,45.64,45.70,45.77,45.83,45.90,45.96,46.03,46.10,46.16,46.23,46.29,46.36,46.42,46.49,46.55,46.62,46.69,46.75,46.82,46.88,46.95,47.01,47.08,47.15,47.21,47.28,47.34,47.41,47.47,47.54,47.60,47.67,47.74,47.80,47.87,47.93,48.00,48.06,48.13,48.19,48.26,48.33,48.39,48.46,48.52,48.59,48.65,48.72,48.78,48.85,48.92,48.98,49.05,49.11,49.18,49.24,49.31,49.37,49.44,49.51,49.57,49.64,49.70,49.77,49.83,49.90,49.96,50.03,50.10,50.16,50.23,50.29,50.36,50.42,50.49,50.55,50.62,50.69,50.75,50.82,50.88,50.95,51.01,51.08,51.14,51.21,51.28,51.34,51.41,51.47,51.54,51.60,51.67,51.73,51.80,51.87,51.93,52.00,52.06,52.13,52.19,52.26,52.32,52.39,52.46,52.52,52.59,52.65,52.72,52.78,52.85,52.91,52.98,53.05,53.11,53.18,53.24,53.31,53.37,53.44,53.51,53.57,53.64,53.70,53.77,53.83,53.90,53.96,54.03,54.10,54.16,54.23,54.29,54.36,54.42,54.49,54.55,54.62,54.68,54.75,54.82,54.88,54.95,55.01,55.08,55.14,55.21,55.28,55.34,55.41,55.47,55.54,55.60,55.67,55.73,55.80,55.86,55.93,56.00,56.06,56.13,56.19,56.26,56.32,56.39,56.45,56.52,56.59,56.65,56.72,56.78,56.85,56.91,56.98,57.04,57.11,57.18,57.24,57.31,57.37,57.44,57.50,57.57,57.63,57.70,57.77,57.83,57.90,57.96,58.03,58.09,58.16,58.22,58.29,58.36,58.42,58.49,58.55,58.62,58.68,58.75,58.81,58.88,58.95,59.01,59.08,59.14,59.21,59.27,59.34,59.40,59.47,59.54,59.60,59.67,59.73,59.80,59.86,59.93,59.99,60.06,60.12,60.19,60.26,60.32,60.39,60.45,60.52,60.58,60.65,60.71,60.78,60.85,60.91,60.98,61.04,61.11,61.17,61.24,61.30,61.37,61.43,61.50,61.57,61.63,61.70,61.76,61.83,61.89,61.96,62.02,62.09,62.16,62.22,62.29,62.35,62.42,62.48,62.55,62.61,62.68,62.74,62.81,62.88,62.94,63.01,63.07,63.14,63.20,63.27,63.33,63.40,63.47,63.53,63.60,63.66,63.73,63.79,63.86,63.92,63.99,64.05,64.12,64.19,64.25,64.32,64.38,64.45,64.51,64.58,64.64,64.71,64.77,64.84,64.91,64.97,65.04,65.10,65.17,65.23,65.30,65.36,65.43,65.50,65.56,65.63,65.69,65.76,65.82,65.89,65.95,66.02,66.08,66.15,66.22,66.28,66.35,66.41,66.48,66.54,66.61,66.67,66.74,66.81,66.87,66.94,67.00,67.07,67.13,67.20,67.26,67.33,67.39,67.46,67.53,67.59,67.66,67.72,67.79,67.85,67.92,67.98,68.05,68.12,68.18,68.25,68.31,68.38,68.44,68.51,68.57,68.64,68.70,68.77,68.84,68.90,68.97,69.03,69.10,69.16,69.23,69.29,69.36,69.43,69.49,69.56,69.62,69.69,69.75,69.82,69.88,69.95,70.02,70.08,70.15,70.21,70.28,70.34,70.41,70.47,70.54,70.61,70.67,70.74,70.80,70.87,70.93,71.00,71.06,71.13,71.20,71.26,71.33,71.39,71.46,71.52,71.59,71.65,71.72,71.79,71.85,71.92,71.98,72.05,72.11,72.18,72.24,72.31,72.38,72.44,72.51,72.57,72.64,72.70,72.77,72.83,72.90,72.97,73.03,73.10,73.16,73.23,73.29,73.36,73.42,73.49,73.56,73.62,73.69,73.75,73.82,73.88,73.95,74.01,74.08,74.15,74.21,74.28,74.34,74.41,74.47,74.54,74.60,74.67,74.74,74.80,74.87,74.93,75.00,75.06,75.13,75.20,75.26,75.33,75.39,75.46,75.52,75.59,75.65,75.72,75.79,75.85,75.92,75.98,76.05,76.11,76.18,76.24,76.31,76.38,76.44,76.51,76.57,76.64,76.70,76.77,76.83,76.90,76.97,77.03,77.10,77.16,77.23,77.29,77.36,77.42,77.49,77.56,77.62,77.69,77.75,77.82,77.88,77.95,78.01,78.08,78.15,78.21,78.28,78.34,78.41,78.47,78.54,78.60,78.67,78.74,78.80,78.87,78.93,79.00,79.06,79.13,79.19,79.26,79.33,79.39,79.46,79.52,79.59,79.65,79.72,79.78,79.85,79.92,79.98,80.05,80.11,80.18,80.24,80.31,80.37,80.44,80.51,80.57,80.64,80.70,80.77,80.83,80.90,80.96,81.03,81.10,81.16,81.23,81.29,81.36,81.42,81.49,81.55,81.62,81.69,81.75,81.82,81.88,81.95,82.01,82.08,82.14,82.21,82.27,82.34,82.41,82.47,82.54,82.60,82.67,82.73,82.80,82.86,82.93,83.00,83.06,83.13,83.19,83.26,83.32,83.39,83.45,83.52,83.59,83.65,83.72,83.78,83.85,83.91,83.98,84.04,84.11,84.18,84.24,84.31,84.37,84.44,84.50,84.57,84.63,84.70,84.77,84.83,84.90,84.96,85.03,85.09,85.16,85.22,85.29,85.36,85.42,85.49,85.55,85.62,85.68,85.75,85.81,85.88,85.95,86.01,86.08,86.14,86.21,86.27,86.34,86.40,86.47,86.53,86.60,86.67,86.73,86.80,86.86,86.93,86.99,87.06,87.12,87.19,87.26,87.32,87.39,87.45,87.52,87.58,87.65,87.71,87.78,87.85,87.91,87.98,88.04,88.11,88.17,88.24,88.30,88.37,88.44,88.50,88.57,88.63,88.70,88.76,88.83,88.89,88.96,89.03,89.09,89.16,89.22,89.29,89.35,89.42,89.48,89.55,89.61,89.68,89.75,89.81,89.88,89.94,90.01,90.07,90.14,90.20,90.27,90.34,90.40,90.47,90.53,90.60,90.66,90.73,90.79,90.86,90.93,90.99,91.06,91.12,91.19,91.25,91.32,91.38,91.45,91.52,91.58,91.65,91.71,91.78,91.84,91.91,91.97,92.04,92.11,92.17,92.24,92.30,92.37,92.43,92.50,92.56,92.63,92.69,92.76,92.83,92.89,92.96,93.02,93.09,93.15,93.22,93.28,93.35,93.42,93.48,93.55,93.61,93.68,93.74,93.81,93.87,93.94,94.01,94.07,94.14,94.20,94.27,94.33,94.40,94.46,94.53,94.60,94.66,94.73,94.79,94.86,94.92,94.99,95.05,95.12,95.18,95.25,95.32,95.38,95.45,95.51,95.58,95.64,95.71,95.77,95.84,95.91,95.97,96.04,96.10,96.17,96.23,96.30,96.36,96.43,96.50,96.56,96.63,96.69,96.76,96.82,96.89,96.95,97.02,97.09,97.15,97.21,97.26,97.32,97.37,97.42,97.48,97.53,97.58,97.62,97.67,97.72,97.76,97.80,97.85,97.89,97.93,97.98,98.02,98.06,98.10,98.14,98.18,98.21,98.25,98.28,98.32,98.35,98.38,98.42,98.45,98.48,98.51,98.54,98.57,98.60,98.63,98.66,98.68,98.71,98.74,98.76,98.79,98.82,98.84,98.86,98.89,98.91,98.94,98.96,98.98,99.00,99.02,99.05,99.07,99.09,99.11,99.13,99.15,99.17,99.19,99.21,99.23,99.25,99.27,99.28,99.30,99.32,99.34,99.35,99.37,99.39,99.40,99.42,99.43,99.45,99.46,99.48,99.49,99.51,99.52,99.54,99.55,99.56,99.58,99.59,99.60,99.62,99.63,99.64,99.66,99.67,99.68,99.69,99.70,99.72,99.73,99.74,99.75,99.76,99.77,99.78,99.79,99.80,99.81,99.82,99.83,99.84,99.85,99.86,99.87,99.88,99.89,99.90,99.91,99.92,99.93,99.94,99.94,99.95,99.96,99.97,99.98,99.99,99.99,100.00,100.01,100.02,100.02,100.03,100.04,100.05,100.05,100.05,]

                        mapping = list(zip(Volt,Percentage))
                        # print(len(Volt))
                        # print(len(Percentage))
                        # print(len(mapping))

                        # Find the closest voltage in the list of Volt
                        closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                        # Find the index of the closest voltage
                        index = Volt.index(closest_voltage)
                        # print(index)
                        # print(Percentage[index])

                        # Check if the index is not the last index to avoid index out of range
                        if index < len(Volt) - 1:
                            if(input_voltage > closest_voltage) or index == 0  and input_voltage != 0:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")
                                # Calculate the slope using the closest_voltage and next_voltage
                                if next_voltage != closest_voltage :
                                    slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                    closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                else:
                                    next_voltage = Volt[index + 2]
                                    slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                    closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)

                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage < closest_voltage and input_voltage != 0:
                                next_voltage = Volt[index - 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage == 0:
                                closest_percentage = 0
                            else:
                                next_voltage = Volt[index + 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif index + 1 == len(Volt):

                            if input_voltage < closest_voltage and input_voltage != 0:
                                # print("This is herer 3")
                                next_voltage = Volt[index - 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage == 0:
                                closest_percentage = 0
                            elif input_voltage > Volt[index]:
                                # print(Volt[index])
                                closest_percentage = 0
                            else:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")                     
                        else:
                            print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                        return closest_percentage
                def SOH2(input_voltage):
                    global closest_percentage2
                    closest_percentage2 = 0
                    # print("This is SOH2")
                    Volt = [ 2.7011, 2.7546, 2.7684, 2.7789, 2.7877, 2.7956, 2.8026, 2.8092, 2.8156, 2.8212, 2.827, 2.8327, 2.8378, 2.8432, 2.8482, 2.8533, 2.8584, 2.8632, 2.8678, 2.8727, 2.8775, 2.8823, 2.887, 2.8917, 2.8961, 2.9008, 2.9055, 2.9102, 2.9145, 2.9192, 2.9235, 2.9282, 2.9317, 2.9372, 2.9421, 2.9461, 2.9505, 2.9548, 2.9594, 2.9631, 2.9685, 2.9726, 2.9768, 2.9819, 2.9856, 2.9898, 2.9942, 2.998, 3.0025, 3.007, 3.0111, 3.0153, 3.0197, 3.0233, 3.0277, 3.0319, 3.0358, 3.0399, 3.044, 3.048, 3.052, 3.0559, 3.0598, 3.0637, 3.0675, 3.0715, 3.0754, 3.0789, 3.0828, 3.0868, 3.0903, 3.0942, 3.098, 3.1016, 3.1052, 3.1087, 3.1124, 3.116, 3.1195, 3.1231, 3.1266, 3.1299, 3.1336, 3.1373, 3.1405, 3.144, 3.1472, 3.1504, 3.1538, 3.1573, 3.1606, 3.1638, 3.167, 3.1703, 3.1739, 3.1766, 3.1797, 3.183, 3.1859, 3.1891, 3.1923, 3.1951, 3.1981, 3.2011, 3.2045, 3.2072, 3.21, 3.213, 3.2157, 3.2192, 3.2216, 3.2246, 3.2275, 3.2302, 3.233, 3.2359, 3.2385, 3.2408, 3.2436, 3.2463, 3.2489, 3.2515, 3.2541, 3.2566, 3.2592, 3.262, 3.2645, 3.2668, 3.2697, 3.2722, 3.2748, 3.2776, 3.2797, 3.2823, 3.2847, 3.2871, 3.2892, 3.292, 3.2946, 3.2965, 3.2989, 3.3014, 3.3036, 3.3056, 3.3085, 3.3106, 3.3129, 3.3153, 3.3177, 3.3199, 3.322, 3.3244, 3.3266, 3.3286, 3.3311, 3.3333, 3.3353, 3.3377, 3.3396, 3.3419, 3.3439, 3.346, 3.3483, 3.3503, 3.3521, 3.3544, 3.3566, 3.3585, 3.3606, 3.3625, 3.3645, 3.3667, 3.3684, 3.3705, 3.3725, 3.3744, 3.3763, 3.3784, 3.3802, 3.3822, 3.384, 3.3864, 3.3879, 3.3899, 3.3913, 3.3937, 3.3954, 3.3972, 3.399, 3.4009, 3.4027, 3.4046, 3.4064, 3.408, 3.4098, 3.4116, 3.4131, 3.4151, 3.4168, 3.4186, 3.4203, 3.4219, 3.4235, 3.4253, 3.4271, 3.4288, 3.4302, 3.432, 3.4339, 3.4355, 3.437, 3.4389, 3.4403, 3.442, 3.4436, 3.4451, 3.4469, 3.4483, 3.45, 3.4514, 3.4531, 3.4547, 3.4562, 3.4577, 3.4592, 3.4608, 3.4623, 3.4637, 3.4652, 3.4667, 3.4682, 3.4698, 3.4712, 3.4727, 3.4744, 3.4757, 3.4771, 3.4785, 3.48, 3.4814, 3.4829, 3.4844, 3.4857, 3.4871, 3.4884, 3.4899, 3.4912, 3.4926, 3.494, 3.4955, 3.4968, 3.498, 3.4994, 3.5009, 3.5021, 3.5036, 3.5046, 3.5058, 3.5076, 3.5086, 3.5101, 3.511, 3.5124, 3.5135, 3.5148, 3.5163, 3.5176, 3.5186, 3.52, 3.5213, 3.5225, 3.5236, 3.5248, 3.5257, 3.5271, 3.5284, 3.5296, 3.5309, 3.5318, 3.5327, 3.534, 3.5352, 3.5362, 3.5374, 3.5383, 3.539, 3.5401, 3.5412, 3.5422, 3.5431, 3.5435, 3.5446, 3.545, 3.5457, 3.5465, 3.547, 3.5475, 3.5484, 3.5489, 3.5495, 3.5499, 3.5503, 3.5506, 3.5513, 3.5516, 3.5524, 3.5527, 3.5529, 3.5533, 3.5537, 3.554, 3.5544, 3.5551, 3.5554, 3.5555, 3.5562, 3.5563, 3.5564, 3.5567, 3.5571, 3.5574, 3.5578, 3.5581, 3.5585, 3.5587, 3.5592, 3.5591, 3.5595, 3.5602, 3.5605, 3.5608, 3.5608, 3.561, 3.5616, 3.5617, 3.562, 3.5626, 3.5627, 3.5633, 3.5633, 3.5637, 3.5641, 3.5644, 3.5648, 3.5649, 3.5647, 3.5655, 3.5659, 3.5661, 3.5665, 3.5673, 3.5669, 3.5675, 3.5677, 3.5679, 3.5686, 3.569, 3.5689, 3.5698, 3.5701, 3.5703, 3.5707, 3.5711, 3.5712, 3.5716, 3.5723, 3.5724, 3.5727, 3.5732, 3.5737, 3.574, 3.5743, 3.5747, 3.5751, 3.5754, 3.576, 3.5763, 3.5767, 3.5771, 3.5773, 3.5781, 3.5784, 3.5789, 3.5792, 3.5793, 3.5802, 3.5805, 3.5808, 3.5813, 3.5817, 3.5821, 3.5827, 3.5832, 3.5837, 3.5841, 3.5847, 3.5852, 3.5857, 3.586, 3.5865, 3.5871, 3.5873, 3.5879, 3.5886, 3.5894, 3.5896, 3.59, 3.5906, 3.591, 3.5916, 3.5922, 3.5925, 3.5929, 3.594, 3.5942, 3.5943, 3.5953, 3.5955, 3.5965, 3.5968, 3.5974, 3.5977, 3.5986, 3.5986, 3.5996, 3.5995, 3.6005, 3.6008, 3.6018, 3.6019, 3.6026, 3.603, 3.6034, 3.604, 3.6048, 3.6051, 3.6057, 3.6062, 3.6066, 3.6074, 3.6078, 3.6084, 3.6089, 3.6095, 3.61, 3.6105, 3.6109, 3.6115, 3.612, 3.6127, 3.6133, 3.6136, 3.6142, 3.6146, 3.6152, 3.6157, 3.6163, 3.6169, 3.6173, 3.6178, 3.6184, 3.6189, 3.6195, 3.6199, 3.6204, 3.621, 3.6214, 3.6218, 3.6224, 3.6228, 3.6233, 3.6238, 3.6243, 3.6248, 3.6255, 3.6259, 3.6263, 3.627, 3.6273, 3.6278, 3.6284, 3.6287, 3.6292, 3.6296, 3.6301, 3.6306, 3.6311, 3.6316, 3.632, 3.6325, 3.633, 3.6334, 3.6339, 3.6343, 3.6348, 3.6352, 3.6358, 3.6361, 3.6364, 3.6369, 3.6374, 3.6379, 3.6382, 3.6389, 3.6392, 3.6393, 3.6401, 3.6405, 3.6407, 3.6413, 3.6416, 3.6421, 3.6425, 3.6432, 3.6434, 3.644, 3.6443, 3.6448, 3.645, 3.6454, 3.646, 3.6464, 3.647, 3.6475, 3.6478, 3.6478, 3.6483, 3.6492, 3.6494, 3.6498, 3.65, 3.6508, 3.6511, 3.6516, 3.652, 3.6524, 3.6527, 3.6532, 3.6537, 3.6542, 3.6548, 3.6551, 3.6554, 3.6559, 3.6564, 3.6568, 3.6573, 3.6578, 3.6581, 3.6586, 3.659, 3.6594, 3.6598, 3.66, 3.6606, 3.6612, 3.6613, 3.6621, 3.6626, 3.6629, 3.6634, 3.6637, 3.6641, 3.6645, 3.665, 3.6653, 3.6659, 3.666, 3.6664, 3.6669, 3.6675, 3.6677, 3.6682, 3.6686, 3.6688, 3.6694, 3.6698, 3.6701, 3.6706, 3.6709, 3.6714, 3.6717, 3.6719, 3.6726, 3.6727, 3.6732, 3.6735, 3.6739, 3.6745, 3.6748, 3.675, 3.6754, 3.6756, 3.676, 3.6765, 3.6769, 3.6771, 3.6775, 3.6777, 3.678, 3.6783, 3.6788, 3.6793, 3.6793, 3.6796, 3.68, 3.6804, 3.6805, 3.6811, 3.6812, 3.6817, 3.6819, 3.6822, 3.6825, 3.6827, 3.683, 3.6834, 3.6837, 3.6838, 3.6842, 3.6846, 3.6849, 3.685, 3.6852, 3.6858, 3.6861, 3.6862, 3.6864, 3.6867, 3.6873, 3.6874, 3.6875, 3.6878, 3.6882, 3.6884, 3.6886, 3.6889, 3.6892, 3.6893, 3.6897, 3.69, 3.6901, 3.6905, 3.6907, 3.6911, 3.6911, 3.6916, 3.6917, 3.6921, 3.6922, 3.6925, 3.6929, 3.6928, 3.6931, 3.6936, 3.6936, 3.6939, 3.6942, 3.6945, 3.6946, 3.6949, 3.6951, 3.6953, 3.6958, 3.6961, 3.6962, 3.6964, 3.6967, 3.6969, 3.6967, 3.6972, 3.6976, 3.6977, 3.6979, 3.6982, 3.6987, 3.6988, 3.6989, 3.6995, 3.6993, 3.6997, 3.7002, 3.7003, 3.7005, 3.7002, 3.7007, 3.7009, 3.7011, 3.7024, 3.7008, 3.7018, 3.7023, 3.7023, 3.7024, 3.7023, 3.7026, 3.7041, 3.7035, 3.7038, 3.7044, 3.7044, 3.7047, 3.705, 3.7062, 3.7053, 3.7056, 3.706, 3.7064, 3.7063, 3.7067, 3.7067, 3.707, 3.7073, 3.7076, 3.7081, 3.7078, 3.7081, 3.7086, 3.7088, 3.7089, 3.7094, 3.7096, 3.7098, 3.7101, 3.7102, 3.7104, 3.7105, 3.711, 3.7112, 3.7114, 3.7115, 3.712, 3.7123, 3.7127, 3.7129, 3.7131, 3.7134, 3.7135, 3.714, 3.7142, 3.7146, 3.7149, 3.7149, 3.7153, 3.7156, 3.7158, 3.7161, 3.7164, 3.7168, 3.7169, 3.7171, 3.7177, 3.7178, 3.718, 3.7183, 3.7188, 3.719, 3.7192, 3.7194, 3.7199, 3.7201, 3.7202, 3.7205, 3.721, 3.7213, 3.7217, 3.7219, 3.7222, 3.7225, 3.7227, 3.7231, 3.7233, 3.7235, 3.7239, 3.7242, 3.7245, 3.7249, 3.7252, 3.7254, 3.7259, 3.7261, 3.7263, 3.7268, 3.727, 3.7272, 3.7278, 3.7281, 3.7283, 3.7285, 3.7291, 3.7294, 3.7297, 3.7299, 3.7302, 3.7307, 3.7308, 3.7314, 3.7316, 3.7319, 3.7324, 3.7329, 3.7329, 3.7332, 3.7336, 3.7341, 3.7343, 3.7347, 3.735, 3.7353, 3.7358, 3.7361, 3.7365, 3.7368, 3.737, 3.7378, 3.738, 3.7382, 3.7387, 3.739, 3.7394, 3.7397, 3.7402, 3.7405, 3.7408, 3.7413, 3.7416, 3.7421, 3.7422, 3.7428, 3.7432, 3.7435, 3.7439, 3.7443, 3.7448, 3.7451, 3.7457, 3.7458, 3.7464, 3.7467, 3.7473, 3.7475, 3.7479, 3.7484, 3.7487, 3.749, 3.7495, 3.7498, 3.7504, 3.7505, 3.7513, 3.7518, 3.7519, 3.7524, 3.7529, 3.7534, 3.7537, 3.7542, 3.7545, 3.7548, 3.7554, 3.7559, 3.7562, 3.7566, 3.7572, 3.7576, 3.7581, 3.7585, 3.759, 3.7593, 3.7598, 3.7602, 3.7606, 3.7612, 3.7616, 3.7622, 3.7626, 3.763, 3.7634, 3.7638, 3.7645, 3.7648, 3.7653, 3.7658, 3.7662, 3.767, 3.7672, 3.7677, 3.7682, 3.7685, 3.769, 3.7695, 3.7701, 3.7706, 3.7711, 3.7716, 3.772, 3.7726, 3.7728, 3.7736, 3.774, 3.7744, 3.775, 3.7755, 3.776, 3.7766, 3.7769, 3.7774, 3.778, 3.7787, 3.779, 3.7797, 3.78, 3.7806, 3.7812, 3.7817, 3.7823, 3.783, 3.7834, 3.7839, 3.7844, 3.7849, 3.7854, 3.7859, 3.7865, 3.7872, 3.7876, 3.7882, 3.7887, 3.7894, 3.7898, 3.7907, 3.7914, 3.7919, 3.7924, 3.793, 3.7936, 3.7943, 3.7947, 3.7953, 3.7958, 3.7964, 3.7969, 3.7975, 3.7982, 3.7989, 3.7994, 3.7999, 3.8006, 3.801, 3.8018, 3.8023, 3.8028, 3.8035, 3.804, 3.8047, 3.8053, 3.8059, 3.8066, 3.807, 3.8078, 3.8083, 3.8089, 3.8096, 3.8102, 3.8109, 3.8114, 3.8121, 3.8127, 3.8133, 3.8141, 3.8148, 3.8153, 3.8159, 3.8164, 3.8172, 3.8178, 3.8186, 3.8193, 3.8197, 3.8205, 3.821, 3.8216, 3.8225, 3.8231, 3.8238, 3.8245, 3.8252, 3.8258, 3.8265, 3.8271, 3.8279, 3.8283, 3.8292, 3.8299, 3.8306, 3.8314, 3.8319, 3.8329, 3.8333, 3.8338, 3.8348, 3.8355, 3.8362, 3.8368, 3.8376, 3.8384, 3.8392, 3.8399, 3.8406, 3.8411, 3.842, 3.8428, 3.8432, 3.8441, 3.8449, 3.8458, 3.8464, 3.8471, 3.848, 3.8488, 3.8492, 3.8502, 3.8507, 3.8514, 3.8524, 3.8531, 3.8539, 3.8548, 3.8555, 3.8561, 3.8569, 3.858, 3.8587, 3.8593, 3.86, 3.8609, 3.8616, 3.8625, 3.8635, 3.8641, 3.8649, 3.8658, 3.8664, 3.8674, 3.8679, 3.869, 3.8698, 3.8705, 3.8712, 3.8721, 3.8725, 3.8735, 3.8743, 3.875, 3.8758, 3.8767, 3.8774, 3.8784, 3.8792, 3.8798, 3.8808, 3.8815, 3.8828, 3.8838, 3.8849, 3.885, 3.8856, 3.8864, 3.887, 3.888, 3.8888, 3.8897, 3.8904, 3.8914, 3.892, 3.8928, 3.8938, 3.8944, 3.8952, 3.8961, 3.8968, 3.8978, 3.8983, 3.8991, 3.8999, 3.9008, 3.9014, 3.9022, 3.9031, 3.9039, 3.9045, 3.9053, 3.906, 3.9069, 3.9076, 3.9084, 3.9092, 3.9099, 3.9108, 3.9114, 3.9121, 3.9129, 3.9136, 3.9145, 3.9153, 3.9159, 3.9167, 3.9175, 3.918, 3.9188, 3.9196, 3.9203, 3.921, 3.9218, 3.9224, 3.9233, 3.924, 3.9247, 3.9254, 3.9261, 3.9268, 3.9276, 3.9282, 3.929, 3.9297, 3.9305, 3.931, 3.9319, 3.9328, 3.9334, 3.934, 3.9349, 3.9355, 3.9361, 3.9369, 3.9376, 3.9383, 3.9391, 3.9398, 3.9405, 3.9413, 3.9418, 3.9425, 3.9435, 3.944, 3.9446, 3.9456, 3.9462, 3.9469, 3.9477, 3.9483, 3.9491, 3.9496, 3.9505, 3.9511, 3.9518, 3.9526, 3.9533, 3.954, 3.9547, 3.9554, 3.9561, 3.9566, 3.9575, 3.9582, 3.9588, 3.9595, 3.9603, 3.9611, 3.9618, 3.9624, 3.9631, 3.964, 3.9647, 3.9652, 3.966, 3.9667, 3.9676, 3.968, 3.9688, 3.9695, 3.9702, 3.9709, 3.9716, 3.9723, 3.973, 3.9737, 3.9745, 3.9751, 3.9759, 3.9767, 3.9773, 3.9779, 3.9788, 3.9794, 3.9801, 3.981, 3.9816, 3.9823, 3.983, 3.9837, 3.9846, 3.9851, 3.9856, 3.9864, 3.9871, 3.9879, 3.9885, 3.9892, 3.9901, 3.9908, 3.9914, 3.9921, 3.9929, 3.9935, 3.9943, 3.9949, 3.9958, 3.9964, 3.9972, 3.9979, 3.9985, 3.9994, 4, 4.0007, 4.0013, 4.0022, 4.0029, 4.0038, 4.0042, 4.005, 4.0058, 4.0065, 4.0072, 4.008, 4.0086, 4.0093, 4.0099, 4.0108, 4.0116, 4.0122, 4.0129, 4.0135, 4.0144, 4.0151, 4.0158, 4.0164, 4.0172, 4.0178, 4.0187, 4.0194, 4.0201, 4.021, 4.0215, 4.0223, 4.0231, 4.0239, 4.0244, 4.0252, 4.026, 4.0268, 4.0273, 4.028, 4.0288, 4.0295, 4.0304, 4.0311, 4.0318, 4.0325, 4.0333, 4.034, 4.0347, 4.0356, 4.0362, 4.0371, 4.0377, 4.0385, 4.0391, 4.0399, 4.0405, 4.0414, 4.042, 4.0428, 4.0436, 4.0443, 4.0449, 4.0458, 4.0465, 4.0473, 4.048, 4.0488, 4.0495, 4.0503, 4.051, 4.0517, 4.0525, 4.0531, 4.054, 4.0547, 4.0554, 4.0563, 4.057, 4.0578, 4.0585, 4.0592, 4.0601, 4.0607, 4.0616, 4.0621, 4.0631, 4.0637, 4.0645, 4.0652, 4.0659, 4.0667, 4.0675, 4.0683, 4.069, 4.0697, 4.0706, 4.0714, 4.0721, 4.0728, 4.0735, 4.0744, 4.0751, 4.0758, 4.0766, 4.0773, 4.0781, 4.079, 4.0797, 4.0805, 4.0812, 4.0821, 4.0827, 4.0834, 4.0843, 4.0852, 4.0858, 4.0867, 4.0872, 4.0882, 4.089, 4.0897, 4.0905, 4.0912, 4.0919, 4.0928, 4.0936, 4.0944, 4.0951, 4.0959, 4.0965, 4.0974, 4.0982, 4.099, 4.1, 4.1006, 4.1015, 4.1021, 4.1029, 4.1037, 4.1044, 4.1053, 4.1062, 4.1069, 4.1076, 4.1084, 4.1091, 4.1099, 4.1108, 4.1114, 4.1123, 4.113, 4.1139, 4.1146, 4.1154, 4.1162, 4.117, 4.1179, 4.1186, 4.1194, 4.1202, 4.1209, 4.1218, 4.1225, 4.1235, 4.1243, 4.1252, 4.1257, 4.1266, 4.1273, 4.1282, 4.129, 4.1297, 4.1306, 4.1312, 4.1321, 4.1329, 4.1337, 4.1346, 4.1354, 4.1361, 4.1369, 4.1379, 4.1386, 4.1392, 4.1402, 4.1409, 4.1417, 4.1427, 4.1433, 4.1442, 4.1451, 4.1458, 4.1467, 4.1474, 4.1483, 4.149, 4.1498, 4.1507, 4.1514, 4.1522, 4.1531, 4.1538, 4.1548, 4.1555, 4.1564, 4.1571, 4.1579, 4.1586, 4.1596, 4.1604, 4.1611, 4.1621, 4.1628, 4.1637, 4.1645, 4.1653, 4.1661, 4.1669, 4.1677, 4.1684, 4.1693, 4.1701, 4.1711, 4.1719, 4.1726, 4.1735, 4.1743, 4.1752, 4.1759, 4.1768, 4.1775, 4.1785, 4.1793, 4.1802, 4.181, 4.1817, 4.1825, 4.1834, 4.1843, 4.1849, 4.186, 4.1867, 4.1876, 4.1884, 4.1894, 4.1901, 4.1909, 4.1917, 4.1925, 4.1935, 4.1944, 4.1951, 4.1959, 4.1967, 4.1975, 4.1985, 4.1992, 4.2, 4.1992, 4.1992, 4.1991, 4.1991, 4.1993, 4.1991, 4.1993, 4.1993, 4.1991, 4.1993, 4.1991, 4.1992, 4.199, 4.1991, 4.1994, 4.1996, 4.2001, 4.2005, 4.2009, 4.2002, 4.2, 4.2001, 4.2002, 4.2001, 4.2001, 4.2, 4.2001, 4.2002, 4.2001, 4.2003, 4.2, 4.2001, 4.2002, 4.2002, 4.2002, 4.2002, 4.2001, 4.2, 4.2001, 4.2, 4.2001, 4.2, 4.2002, 4.2003, 4.2003, 4.2002, 4.2, 4.2001, 4.2001, 4.2002, 4.2002, 4.2001, 4.2003, 4.2001, 4.2002, 4.2001, 4.2001, 4.2004, 4.2003, 4.2003, 4.2006, 4.2006, 4.2008, 4.2004, 4.2005, 4.2004, 4.2004, 4.2003, 4.2004, 4.2005, 4.2004, 4.2001, 4.2004, 4.2004, 4.2005, 4.2004, 4.2004, 4.2004, 4.2004, 4.2004, 4.2003, 4.2003, 4.2004, 4.2001, 4.2003, 4.2003, 4.2003, 4.2002, 4.2003, 4.2002, 4.2001, 4.2004, 4.2002, 4.2003, 4.2004, 4.2003, 4.2006, 4.2003, 4.2002, 4.2002, 4.2002, 4.2003, 4.2003, 4.2003, 4.2002, 4.2003, 4.2002, 4.2003, 4.2002, 4.2003, 4.2003, 4.2005, 4.2006, 4.2005, 4.2006, 4.2007, 4.2005, 4.2005, 4.2005, 4.2007, 4.2006, 4.2005, 4.2007, 4.2007, 4.2007, 4.2007, 4.2006, 4.2004, 4.2005, 4.2006, 4.2006, 4.2006, 4.2005, 4.2006]

                    Percentage = [0.00,0.07,0.13,0.20,0.26,0.33,0.39,0.46,0.52,0.59,0.65,0.72,0.79,0.85,0.92,0.98,1.05,1.11,1.18,1.24,1.31,1.37,1.44,1.51,1.57,1.64,1.70,1.77,1.83,1.90,1.96,2.03,2.09,2.16,2.23,2.29,2.36,2.42,2.49,2.55,2.62,2.68,2.75,2.81,2.88,2.95,3.01,3.08,3.14,3.21,3.27,3.34,3.40,3.47,3.54,3.60,3.67,3.73,3.80,3.86,3.93,3.99,4.06,4.12,4.19,4.26,4.32,4.39,4.45,4.52,4.58,4.65,4.71,4.78,4.84,4.91,4.98,5.04,5.11,5.17,5.24,5.30,5.37,5.43,5.50,5.56,5.63,5.70,5.76,5.83,5.89,5.96,6.02,6.09,6.15,6.22,6.28,6.35,6.42,6.48,6.55,6.61,6.68,6.74,6.81,6.87,6.94,7.00,7.07,7.14,7.20,7.27,7.33,7.40,7.46,7.53,7.59,7.66,7.73,7.79,7.86,7.92,7.99,8.05,8.12,8.18,8.25,8.32,8.38,8.45,8.51,8.58,8.64,8.71,8.77,8.84,8.91,8.97,9.04,9.10,9.17,9.23,9.30,9.36,9.43,9.50,9.56,9.63,9.69,9.76,9.82,9.89,9.95,10.02,10.09,10.15,10.22,10.28,10.35,10.41,10.48,10.55,10.61,10.68,10.74,10.81,10.87,10.94,11.00,11.07,11.14,11.20,11.27,11.33,11.40,11.46,11.53,11.59,11.66,11.73,11.79,11.86,11.92,11.99,12.05,12.12,12.18,12.25,12.32,12.38,12.45,12.51,12.58,12.64,12.71,12.77,12.84,12.91,12.97,13.04,13.10,13.17,13.23,13.30,13.36,13.43,13.50,13.56,13.63,13.69,13.76,13.82,13.89,13.95,14.02,14.09,14.15,14.22,14.28,14.35,14.41,14.48,14.54,14.61,14.68,14.74,14.81,14.87,14.94,15.00,15.07,15.13,15.20,15.27,15.33,15.40,15.46,15.53,15.59,15.66,15.73,15.79,15.86,15.92,15.99,16.05,16.12,16.18,16.25,16.32,16.38,16.45,16.51,16.58,16.64,16.71,16.77,16.84,16.91,16.97,17.04,17.10,17.17,17.23,17.30,17.36,17.43,17.50,17.56,17.63,17.69,17.76,17.82,17.89,17.95,18.02,18.09,18.15,18.22,18.28,18.35,18.41,18.48,18.54,18.61,18.68,18.74,18.81,18.87,18.94,19.00,19.07,19.13,19.20,19.27,19.33,19.40,19.46,19.53,19.59,19.66,19.72,19.79,19.86,19.92,19.99,20.05,20.12,20.18,20.25,20.31,20.38,20.45,20.51,20.58,20.64,20.71,20.77,20.84,20.90,20.97,21.04,21.10,21.17,21.23,21.30,21.36,21.43,21.50,21.56,21.63,21.69,21.76,21.82,21.89,21.95,22.02,22.09,22.15,22.22,22.28,22.35,22.41,22.48,22.54,22.61,22.68,22.74,22.81,22.87,22.94,23.00,23.07,23.13,23.20,23.27,23.33,23.40,23.46,23.53,23.59,23.66,23.73,23.79,23.86,23.92,23.99,24.05,24.12,24.18,24.25,24.32,24.38,24.45,24.51,24.58,24.64,24.71,24.77,24.84,24.91,24.97,25.04,25.10,25.17,25.23,25.30,25.37,25.43,25.50,25.56,25.63,25.69,25.76,25.82,25.89,25.96,26.02,26.09,26.15,26.22,26.28,26.35,26.42,26.48,26.55,26.61,26.68,26.74,26.81,26.87,26.94,27.01,27.07,27.14,27.20,27.27,27.33,27.40,27.47,27.53,27.60,27.66,27.73,27.79,27.86,27.92,27.99,28.06,28.12,28.19,28.25,28.32,28.38,28.45,28.52,28.58,28.65,28.71,28.78,28.84,28.91,28.97,29.04,29.11,29.17,29.24,29.30,29.37,29.43,29.50,29.57,29.63,29.70,29.76,29.83,29.89,29.96,30.02,30.09,30.16,30.22,30.29,30.35,30.42,30.48,30.55,30.62,30.68,30.75,30.81,30.88,30.94,31.01,31.07,31.14,31.21,31.27,31.34,31.40,31.47,31.53,31.60,31.66,31.73,31.80,31.86,31.93,31.99,32.06,32.12,32.19,32.26,32.32,32.39,32.45,32.52,32.58,32.65,32.71,32.78,32.85,32.91,32.98,33.04,33.11,33.17,33.24,33.31,33.37,33.44,33.50,33.57,33.63,33.70,33.76,33.83,33.90,33.96,34.03,34.09,34.16,34.22,34.29,34.35,34.42,34.49,34.55,34.62,34.68,34.75,34.81,34.88,34.95,35.01,35.08,35.14,35.21,35.27,35.34,35.40,35.47,35.54,35.60,35.67,35.73,35.80,35.86,35.93,35.99,36.06,36.13,36.19,36.26,36.32,36.39,36.45,36.52,36.59,36.65,36.72,36.78,36.85,36.91,36.98,37.04,37.11,37.18,37.24,37.31,37.37,37.44,37.50,37.57,37.63,37.70,37.77,37.83,37.90,37.96,38.03,38.09,38.16,38.22,38.29,38.36,38.42,38.49,38.55,38.62,38.68,38.75,38.81,38.88,38.95,39.01,39.08,39.14,39.21,39.27,39.34,39.41,39.47,39.54,39.60,39.67,39.73,39.80,39.86,39.93,40.00,40.06,40.13,40.19,40.26,40.32,40.39,40.45,40.52,40.59,40.65,40.72,40.78,40.85,40.91,40.98,41.05,41.11,41.18,41.24,41.31,41.37,41.44,41.50,41.57,41.64,41.70,41.77,41.83,41.90,41.96,42.03,42.10,42.16,42.23,42.29,42.36,42.42,42.49,42.55,42.62,42.69,42.75,42.82,42.88,42.95,43.01,43.08,43.14,43.21,43.28,43.34,43.41,43.47,43.54,43.60,43.67,43.74,43.80,43.87,43.93,44.00,44.06,44.13,44.19,44.26,44.33,44.39,44.46,44.52,44.59,44.65,44.72,44.78,44.85,44.92,44.98,45.05,45.11,45.18,45.24,45.31,45.37,45.44,45.51,45.57,45.64,45.70,45.77,45.83,45.90,45.96,46.03,46.10,46.16,46.23,46.29,46.36,46.42,46.49,46.55,46.62,46.69,46.75,46.82,46.88,46.95,47.01,47.08,47.15,47.21,47.28,47.34,47.41,47.47,47.54,47.60,47.67,47.74,47.80,47.87,47.93,48.00,48.06,48.13,48.19,48.26,48.33,48.39,48.46,48.52,48.59,48.65,48.72,48.78,48.85,48.92,48.98,49.05,49.11,49.18,49.24,49.31,49.37,49.44,49.51,49.57,49.64,49.70,49.77,49.83,49.90,49.96,50.03,50.10,50.16,50.23,50.29,50.36,50.42,50.49,50.55,50.62,50.69,50.75,50.82,50.88,50.95,51.01,51.08,51.14,51.21,51.28,51.34,51.41,51.47,51.54,51.60,51.67,51.73,51.80,51.87,51.93,52.00,52.06,52.13,52.19,52.26,52.32,52.39,52.46,52.52,52.59,52.65,52.72,52.78,52.85,52.91,52.98,53.05,53.11,53.18,53.24,53.31,53.37,53.44,53.51,53.57,53.64,53.70,53.77,53.83,53.90,53.96,54.03,54.10,54.16,54.23,54.29,54.36,54.42,54.49,54.55,54.62,54.68,54.75,54.82,54.88,54.95,55.01,55.08,55.14,55.21,55.28,55.34,55.41,55.47,55.54,55.60,55.67,55.73,55.80,55.86,55.93,56.00,56.06,56.13,56.19,56.26,56.32,56.39,56.45,56.52,56.59,56.65,56.72,56.78,56.85,56.91,56.98,57.04,57.11,57.18,57.24,57.31,57.37,57.44,57.50,57.57,57.63,57.70,57.77,57.83,57.90,57.96,58.03,58.09,58.16,58.22,58.29,58.36,58.42,58.49,58.55,58.62,58.68,58.75,58.81,58.88,58.95,59.01,59.08,59.14,59.21,59.27,59.34,59.40,59.47,59.54,59.60,59.67,59.73,59.80,59.86,59.93,59.99,60.06,60.12,60.19,60.26,60.32,60.39,60.45,60.52,60.58,60.65,60.71,60.78,60.85,60.91,60.98,61.04,61.11,61.17,61.24,61.30,61.37,61.43,61.50,61.57,61.63,61.70,61.76,61.83,61.89,61.96,62.02,62.09,62.16,62.22,62.29,62.35,62.42,62.48,62.55,62.61,62.68,62.74,62.81,62.88,62.94,63.01,63.07,63.14,63.20,63.27,63.33,63.40,63.47,63.53,63.60,63.66,63.73,63.79,63.86,63.92,63.99,64.05,64.12,64.19,64.25,64.32,64.38,64.45,64.51,64.58,64.64,64.71,64.77,64.84,64.91,64.97,65.04,65.10,65.17,65.23,65.30,65.36,65.43,65.50,65.56,65.63,65.69,65.76,65.82,65.89,65.95,66.02,66.08,66.15,66.22,66.28,66.35,66.41,66.48,66.54,66.61,66.67,66.74,66.81,66.87,66.94,67.00,67.07,67.13,67.20,67.26,67.33,67.39,67.46,67.53,67.59,67.66,67.72,67.79,67.85,67.92,67.98,68.05,68.12,68.18,68.25,68.31,68.38,68.44,68.51,68.57,68.64,68.70,68.77,68.84,68.90,68.97,69.03,69.10,69.16,69.23,69.29,69.36,69.43,69.49,69.56,69.62,69.69,69.75,69.82,69.88,69.95,70.02,70.08,70.15,70.21,70.28,70.34,70.41,70.47,70.54,70.61,70.67,70.74,70.80,70.87,70.93,71.00,71.06,71.13,71.20,71.26,71.33,71.39,71.46,71.52,71.59,71.65,71.72,71.79,71.85,71.92,71.98,72.05,72.11,72.18,72.24,72.31,72.38,72.44,72.51,72.57,72.64,72.70,72.77,72.83,72.90,72.97,73.03,73.10,73.16,73.23,73.29,73.36,73.42,73.49,73.56,73.62,73.69,73.75,73.82,73.88,73.95,74.01,74.08,74.15,74.21,74.28,74.34,74.41,74.47,74.54,74.60,74.67,74.74,74.80,74.87,74.93,75.00,75.06,75.13,75.20,75.26,75.33,75.39,75.46,75.52,75.59,75.65,75.72,75.79,75.85,75.92,75.98,76.05,76.11,76.18,76.24,76.31,76.38,76.44,76.51,76.57,76.64,76.70,76.77,76.83,76.90,76.97,77.03,77.10,77.16,77.23,77.29,77.36,77.42,77.49,77.56,77.62,77.69,77.75,77.82,77.88,77.95,78.01,78.08,78.15,78.21,78.28,78.34,78.41,78.47,78.54,78.60,78.67,78.74,78.80,78.87,78.93,79.00,79.06,79.13,79.19,79.26,79.33,79.39,79.46,79.52,79.59,79.65,79.72,79.78,79.85,79.92,79.98,80.05,80.11,80.18,80.24,80.31,80.37,80.44,80.51,80.57,80.64,80.70,80.77,80.83,80.90,80.96,81.03,81.10,81.16,81.23,81.29,81.36,81.42,81.49,81.55,81.62,81.69,81.75,81.82,81.88,81.95,82.01,82.08,82.14,82.21,82.27,82.34,82.41,82.47,82.54,82.60,82.67,82.73,82.80,82.86,82.93,83.00,83.06,83.13,83.19,83.26,83.32,83.39,83.45,83.52,83.59,83.65,83.72,83.78,83.85,83.91,83.98,84.04,84.11,84.18,84.24,84.31,84.37,84.44,84.50,84.57,84.63,84.70,84.77,84.83,84.90,84.96,85.03,85.09,85.16,85.22,85.29,85.36,85.42,85.49,85.55,85.62,85.68,85.75,85.81,85.88,85.95,86.01,86.08,86.14,86.21,86.27,86.34,86.40,86.47,86.53,86.60,86.67,86.73,86.80,86.86,86.93,86.99,87.06,87.12,87.19,87.26,87.32,87.39,87.45,87.52,87.58,87.65,87.71,87.78,87.85,87.91,87.98,88.04,88.11,88.17,88.24,88.30,88.37,88.44,88.50,88.57,88.63,88.70,88.76,88.83,88.89,88.96,89.03,89.09,89.16,89.22,89.29,89.35,89.42,89.48,89.55,89.61,89.68,89.75,89.81,89.88,89.94,90.01,90.07,90.14,90.20,90.27,90.34,90.40,90.47,90.53,90.60,90.66,90.73,90.79,90.86,90.93,90.99,91.06,91.12,91.19,91.25,91.32,91.38,91.45,91.52,91.58,91.65,91.71,91.78,91.84,91.91,91.97,92.04,92.11,92.17,92.24,92.30,92.37,92.43,92.50,92.56,92.63,92.69,92.76,92.83,92.89,92.96,93.02,93.09,93.15,93.22,93.28,93.35,93.42,93.48,93.55,93.61,93.68,93.74,93.81,93.87,93.94,94.01,94.07,94.14,94.20,94.27,94.33,94.40,94.46,94.53,94.60,94.66,94.73,94.79,94.86,94.92,94.99,95.05,95.12,95.18,95.25,95.32,95.38,95.45,95.51,95.58,95.64,95.71,95.77,95.84,95.91,95.97,96.04,96.10,96.17,96.23,96.30,96.36,96.43,96.50,96.56,96.63,96.69,96.76,96.82,96.89,96.95,97.02,97.09,97.15,97.21,97.26,97.32,97.37,97.42,97.48,97.53,97.58,97.62,97.67,97.72,97.76,97.80,97.85,97.89,97.93,97.98,98.02,98.06,98.10,98.14,98.18,98.21,98.25,98.28,98.32,98.35,98.38,98.42,98.45,98.48,98.51,98.54,98.57,98.60,98.63,98.66,98.68,98.71,98.74,98.76,98.79,98.82,98.84,98.86,98.89,98.91,98.94,98.96,98.98,99.00,99.02,99.05,99.07,99.09,99.11,99.13,99.15,99.17,99.19,99.21,99.23,99.25,99.27,99.28,99.30,99.32,99.34,99.35,99.37,99.39,99.40,99.42,99.43,99.45,99.46,99.48,99.49,99.51,99.52,99.54,99.55,99.56,99.58,99.59,99.60,99.62,99.63,99.64,99.66,99.67,99.68,99.69,99.70,99.72,99.73,99.74,99.75,99.76,99.77,99.78,99.79,99.80,99.81,99.82,99.83,99.84,99.85,99.86,99.87,99.88,99.89,99.90,99.91,99.92,99.93,99.94,99.94,99.95,99.96,99.97,99.98,99.99,99.99,100.00,100.01,100.02,100.02,100.03,100.04,100.05,100.05,100.05,]

                    mapping = list(zip(Volt,Percentage))
                    # print(len(Volt))
                    # print(len(Percentage))
                    # print(len(mapping))

                    # Find the closest voltage in the list of Volt
                    closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                    # Find the index of the closest voltage
                    index = Volt.index(closest_voltage)
                    # print(index)
                    # print(Percentage[index])

                    # Check if the index is not the last index to avoid index out of range
                    if index < len(Volt) - 1:
                        # Get the next voltage value
                        if(input_voltage > closest_voltage) or index == 0 and input_voltage != 0:
                            next_voltage = Volt[index + 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        elif input_voltage < closest_voltage and input_voltage != 0:
                            next_voltage = Volt[index - 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                            closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        elif input_voltage == 0:
                            closest_percentage2 = 0
                        else:
                            next_voltage = Volt[index + 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")            
                    elif index + 1 == len(Volt):

                        if input_voltage < closest_voltage and input_voltage != 0:
                        #    print("This is herer 3")
                            next_voltage = Volt[index - 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                            closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        elif input_voltage == 0:
                            closest_percentage2 = 0
                        elif input_voltage > Volt[index]:
                            # print(Volt[index])
                            closest_percentage2 = 0
                        else:
                            next_voltage = Volt[index + 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")          
                                        
                    else:
                        print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                # Find indices of 'Group1' in the 'Group' list


                # Sort the array
                sorted_array = sorted(Charge_Set['Group'])

                # Get unique numbers using set
                unique_numbers = set(Charge_Set['Group'])

                # Convert the unique numbers back to a sorted list if needed
                sorted_unique_numbers = sorted(list(unique_numbers))
                first_min_v = None
                last_max_v = None
                # Print the sorted unique numbers
                # print(sorted_unique_numbers)
                shitty = 1
                for numbers in sorted_unique_numbers:
                    group_indices = [i for i, group in enumerate(Charge_Set['Group']) if group == numbers]

                    # Check if there are 'Group1' elements in the 'Group' list
                    if group_indices:
                        # Access the first and last 'Max_V' values for 'Group1'
                        first_min_v = Charge_Set['Min_V'][group_indices[0]]
                        last_max_v = Charge_Set['Max_V'][group_indices[-1]]

                        # print(f"First Max_V for 'Group1': {first_min_v}")
                        # print(f"Last Max_V for 'Group1': {last_max_v}")
                    else:
                        first_min_v = 0
                        last_max_v = 0
                    # print(f"First Max_V for 'Group1': {first_min_v}")
                    # print(f"Last Max_V for 'Group1': {last_max_v}")
                    SOH_Goop['Goop'].append(numbers)
                    SOH_Goop['Min_V'].append(first_min_v)
                    SOH_Goop['Max_V'].append(last_max_v)
                    thread1 = threading.Thread(target=SOH, args=(first_min_v,))
                    # print(f'This is last_m:{last_max_v}')
                    thread2 = threading.Thread(target=SOH2, args=(last_max_v,))
                    thread1.start()
                    thread2.start()
                    thread1.join()
                    thread2.join()
                    # print(f'This is SOH2:{closest_percentage2}')
                    # Calculate energy for 'Group1' where 'Group' is 0
            # Find indices of 'Group1' where 'Group' is 0
                    
                    energy_sum = 0
                    save_data_start = None
                    a= None
                    # Iterate over the numerical indices of Charge_Set['Group']
                    for index, group in enumerate(Charge_Set['Group']):
                        if group == numbers:
                            voltage = Charge_Set['Voltage'][index]
                            current = Charge_Set['Current'][index]
                            time_diff = Charge_Set['Time_Diff'][index]
                            # Split the original string by space to get the time portion
                            teiam = str( Charge_Set['Timestamps'][index])
                            # print(teiam)
                            split_string = teiam.split(" ")

                            # Check if there are at least two parts (date and time)
                            if len(split_string) >= 2:
                                # Join the time portion and discard the date
                                time_portion = " ".join(split_string[1:])
                                # print(time_portion)

                            if save_data_start == None:
                            
                                save_data_start = time_portion
                                SOH_Goop['S_Time'].append(time_portion)
                                
                            save_end = time_portion
                            # print(f'This is time diff { leg}')
                            # print(index)  # Assuming 'Diff' represents time intervals
                            # print(f'This is V {voltage}')
                            # print(f'This is C {current}')
                            # print(f'This is D {time_diff}')
                            if index != 0:
                                energy =  abs(0.5*(current+previous_current))  * time_diff
                                previous_current = current
                                # print(f'This is Energy {energy}')
                                # cell = sheetPackProcess.cell(row=gginp, column=28, value= energy)
                                # gginp += 1
                            else :
                                energy = 0
                                previous_current = current
                            energy_sum += energy
                    SOH_Goop['E_Time'].append(save_end)
                    energy_sum = energy_sum* voltage
                    # print(f'Total energy for "Group1": {energy_sum} Joules')
                        

                    # print(f"Total energy for 'Group1' where 'Group' is 0: {energy_sum} Joules")
                    
                    SOC_Start = closest_percentage
                    SOC_End = closest_percentage2
                    Difference_SOC = SOC_End - SOC_Start
                    DesignCapacity_NH02 = 26.8
                    SOh_E = energy_sum/(3600*1000)
                    SOH_Goop['Start_SOC'].append(SOC_Start)
                    SOH_Goop['End_SOC'].append(SOC_End)
                    SOH_Goop['Charge'].append(SOh_E)
                    if Difference_SOC != 0:
                        Cal_Capacity = SOh_E/(Difference_SOC/100)
                    else:
                        Cal_Capacity = 0
                    if(Cal_Capacity != 0):
                        Remaining_Capacity = (Cal_Capacity/DesignCapacity_NH02)*100
                    elif(Cal_Capacity == None):
                        Remaining_Capacity = 0
                    else:
                        Remaining_Capacity = 0
                    SOH_Goop['Cal'].append(Remaining_Capacity)
                    SOH_Goop['SOH'].append(Cal_Capacity)
                    SOH_Goop['Cal_E'].append(SOh_E)
                    
                    # for index, header in enumerate(headers1, start=shitty):
                    #     cell = sheet8.cell(row=4, column=index, value=header)

                    # cell = sheet8.cell(row=the_loop, column=shitty, value=first_min_v)
                    # cell = sheet8.cell(row=the_loop, column=shitty+1, value=last_max_v)
                    # cell = sheet8.cell(row=the_loop, column=shitty+2, value=SOC_Start)
                    # cell = sheet8.cell(row=the_loop, column=shitty+3, value=SOC_End)
                    # cell = sheet8.cell(row=the_loop, column=shitty+4, value=Cal_Capacity)
                    # cell = sheet8.cell(row=the_loop, column=shitty+5, value=Remaining_Capacity)    
                    # cell = sheet8.cell(row=the_loop, column=shitty+6, value=SOh_E)  
                    shitty += 7
                if first_min_v == None and last_max_v == None:
                    lost_data.append(app+1)
                    lost_data3.append(app+1)
                # print(lost_data)
            def SOH_Min(input_voltage):
                    global closest_percentage
                    closest_percentage = 0
                    Volt = [2.500,2.951	,3.202,3.364,3.477,3.515,3.549,3.584,3.614,3.646,3.674,3.702,3.736,3.779,3.832,3.897,3.953,4.007,4.064,4.125,4.200]
                    Percentage = [0,0.05,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]
                    
                
                    mapping = list(zip(Volt,Percentage))
                    # print(len(Volt))
                    # print(len(Percentage))
                    # print(len(mapping))

                    # Find the closest voltage in the list of Volt
                    closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                    # Find the index of the closest voltage
                    index = Volt.index(closest_voltage)
                    # print(index)
                    # print(Percentage[index])

                    # Check if the index is not the last index to avoid index out of range
                    if index < len(Volt) - 1:
                        if(input_voltage > closest_voltage) or index == 0  and input_voltage != 0:
                            next_voltage = Volt[index + 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")
                            # Calculate the slope using the closest_voltage and next_voltage
                            if next_voltage != closest_voltage :
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            else:
                                next_voltage = Volt[index + 2]
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                        elif (closest_voltage - input_voltage) < 0.1 :    
                            next_voltage = Volt[index + 1]
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif input_voltage < closest_voltage and input_voltage != 0:
                            next_voltage = Volt[index - 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / ( closest_voltage - next_voltage)
                            closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif input_voltage == 0:
                            closest_percentage = 0
                        else:
                            next_voltage = Volt[index + 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                    elif index + 1 == len(Volt):

                        if input_voltage < closest_voltage and input_voltage != 0:
                            # print("This is herer 3")
                            next_voltage = Volt[index - 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                            closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif input_voltage == 0:
                            closest_percentage = 0
                        elif input_voltage > Volt[index]:
                            # print(Volt[index])
                            closest_percentage = 0
                        else:
                            next_voltage = Volt[index + 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                                
                    else:
                        print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                    return closest_percentage
        
            def SOH2_Min(input_voltage):
                global closest_percentage2
                closest_percentage2 = 0
                # print("This is SOH2")
                Volt = [2.500,2.951	,3.202,3.364,3.477,3.515,3.549,3.584,3.614,3.646,3.674,3.702,3.736,3.779,3.832,3.897,3.953,4.007,4.064,4.125,4.200]
            
                Percentage = [0,0.05,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]

                mapping = list(zip(Volt,Percentage))
                # print(len(Volt))
                # print(len(Percentage))
                # print(len(mapping))

                # Find the closest voltage in the list of Volt
                closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                # Find the index of the closest voltage
                index = Volt.index(closest_voltage)
                # print(index)
                # print(Percentage[index])

                # Check if the index is not the last index to avoid index out of range
                if index < len(Volt) - 1:
                    # Get the next voltage value
                    if(input_voltage > closest_voltage) or index == 0 and input_voltage != 0:
                        next_voltage = Volt[index + 1]
                        
                        # Calculate the slope using the closest_voltage and next_voltage
                        slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                        closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                    elif (closest_voltage - input_voltage) < 0.1 :    
                        next_voltage = Volt[index + 1]
                        slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                        closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                                
                        # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                    elif input_voltage < closest_voltage and input_voltage != 0:
                        next_voltage = Volt[index - 1]
                        
                        # Calculate the slope using the closest_voltage and next_voltage
                        slope = (Percentage[index + 1] - Percentage[index]) / ( closest_voltage - next_voltage)
                        closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                        
                        # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                    elif input_voltage == 0:
                        closest_percentage2 = 0
                    else:
                        next_voltage = Volt[index + 1]
                        
                        # Calculate the slope using the closest_voltage and next_voltage
                        slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                        closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                        
                        # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")            
                elif index + 1 == len(Volt):

                    if input_voltage < closest_voltage and input_voltage != 0:
                        # print("This is herer 3")
                        next_voltage = Volt[index - 1]
                        # print(f"This is next{next_voltage}")
                        # print(f"This is next{closest_voltage}")        
                        # Calculate the slope using the closest_voltage and next_voltage
                        slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                        closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                        
                        # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                    elif input_voltage == 0:
                        closest_percentage2 = 0
                    elif input_voltage > Volt[index]:
                        # print(Volt[index])
                        closest_percentage2 = 0
                    else:
                        next_voltage = Volt[index + 1]
                        # print(f"This is next{next_voltage}")
                        # print(f"This is next{closest_voltage}")        
                        # Calculate the slope using the closest_voltage and next_voltage
                        slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                        closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                        
                        # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")          
                    
                else:
                    print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
            # Find indices of 'Group1' in the 'Group' list


            # Sort the array
            sorted_array = sorted(Discharge_Set['Group'])

            # Get unique numbers using set
            unique_numbers = set(Discharge_Set['Group'])

            # Convert the unique numbers back to a sorted list if needed
            sorted_unique_numbers = sorted(list(unique_numbers))
            first_min_v = None
            last_max_v = None
            # Print the sorted unique numbers
            # print(sorted_unique_numbers)
            shitty = 1
            for numbers in sorted_unique_numbers:
                group_indices = [i for i, group in enumerate(Discharge_Set['Group']) if group == numbers]

                # Check if there are 'Group1' elements in the 'Group' list
                if group_indices:
                    # Access the first and last 'Max_V' values for 'Group1'
                    first_min_v = Discharge_Set['Max_V'][group_indices[0]]
                    last_max_v = Discharge_Set['Min_V'][group_indices[-1]]

                    # print(f"First Max_V for 'Group1': {first_min_v}")
                    # print(f"Last Max_V for 'Group1': {last_max_v}")
                else:
                    first_min_v = 0
                    last_max_v = 0
                # print(f"First Max_V for 'Group1': {first_min_v}")
                # print(f"Last Max_V for 'Group1': {last_max_v}")
                SOH_Goop_Min['Goop'].append(numbers)
                SOH_Goop_Min['Min_V'].append(first_min_v)
                SOH_Goop_Min['Max_V'].append(last_max_v)
                thread1 = threading.Thread(target=SOH_Min, args=(first_min_v,))
                # print(f'This is last_m:{last_max_v}')
                thread2 = threading.Thread(target=SOH2_Min, args=(last_max_v,))
                thread1.start()
                thread2.start()
                thread1.join()
                thread2.join()
                # print(f'This is SOH2:{closest_percentage2}')
                # Calculate energy for 'Group1' where 'Group' is 0
        # Find indices of 'Group1' where 'Group' is 0
                
                energy_sum = 0
                save_data_start = None
                a= None
                # Iterate over the numerical indices of Discharge_Set['Group']
                for index, group in enumerate(Discharge_Set['Group']):
                    if group == numbers:
                        voltage = Discharge_Set['Voltage'][index]
                        current = Discharge_Set['Current'][index]
                        time_diff = Discharge_Set['Time_Diff'][index]
                        # Split the original string by space to get the time portion
                        teiam = str( Discharge_Set['Timestamps'][index])
                        # print(teiam)
                        split_string = teiam.split(" ")

                        # Check if there are at least two parts (date and time)
                        if len(split_string) >= 2:
                            # Join the time portion and discard the date
                            time_portion = " ".join(split_string[1:])
                            # print(time_portion)

                        if save_data_start == None:
                        
                            save_data_start = time_portion
                            SOH_Goop_Min['S_Time'].append(time_portion)
                            
                        save_end = time_portion

                        # print(f'This is time diff { leg}')
                        # print(index)  # Assuming 'Diff' represents time intervals
                        # print(f'This is V {voltage}')
                        # print(f'This is C {current}')
                        # print(f'This is D {time_diff}')
                        if index != 0:
                            energy =  abs(0.5*(current+previous_current))  * time_diff
                            previous_current = current
                            # print(f'This is Energy {energy}')
                            # cell = sheetPackProcess.cell(row=gginp, column=27, value= energy)
                            # gginp += 1

                        else :
                            energy = 0
                            previous_current = current
                        energy_sum += energy
                SOH_Goop_Min['E_Time'].append(save_end)
                energy_sum = energy_sum* voltage

                # print(f'Total energy for "Group1": {energy_sum} Joules')
                    

                # print(f"Total energy for 'Group1' where 'Group' is 0: {energy_sum} Joules")
                
                SOC_Start = closest_percentage
                SOC_End = closest_percentage2
                Difference_SOC = SOC_Start - SOC_End 
                DesignCapacity_NH02 = 26.8
                SOh_E = energy_sum/(3600*1000)
                SOH_Goop_Min['Start_SOC'].append(SOC_Start*100)
                SOH_Goop_Min['End_SOC'].append(SOC_End*100)
                SOH_Goop_Min['Discharge'].append(SOh_E)
                if Difference_SOC != 0:
                    Cal_Capacity = abs(SOh_E)/(Difference_SOC)
                else:
                    Cal_Capacity = 0
                if(Cal_Capacity != 0):
                    Remaining_Capacity = (Cal_Capacity/DesignCapacity_NH02)*100
                elif(Cal_Capacity == None):
                    Remaining_Capacity = 0
                else:
                    Remaining_Capacity = 0
                SOH_Goop_Min['Cal'].append(Remaining_Capacity)
                SOH_Goop_Min['SOH'].append(Cal_Capacity)
                SOH_Goop_Min['Cal_E'].append(SOh_E)
                
                # for index, header in enumerate(headers2, start=shitty):
                #     cell = sheet8.cell(row=4+31, column=index, value=header)
                # cell = sheet8.cell(row=the_loop +31, column=shitty, value=first_min_v)
                # cell = sheet8.cell(row=the_loop +31, column=shitty+1, value=last_max_v)
                # cell = sheet8.cell(row=the_loop +31, column=shitty+2, value=SOC_Start)
                # cell = sheet8.cell(row=the_loop +31, column=shitty+3, value=SOC_End)
                # cell = sheet8.cell(row=the_loop +31, column=shitty+4, value=Cal_Capacity)
                # cell = sheet8.cell(row=the_loop +31, column=shitty+5, value=Remaining_Capacity)    
                # cell = sheet8.cell(row=the_loop +31, column=shitty+6, value=SOh_E)  
                shitty += 7

                if first_min_v == None and last_max_v == None:
                    lost_data2.append(app+1)
                    lost_data4.append(app+1)
                # print(lost_data2)
                    # thread1 = threading.Thread(target=CCha)
                    # thread1.start()
                    # thread1.join()
                CCha()
        else:
                
                def CCha():
                    def SOH(input_voltage):
                        global closest_percentage
                        closest_percentage = 0
                        Volt = [ 2.7011, 2.7546, 2.7684, 2.7789, 2.7877, 2.7956, 2.8026, 2.8092, 2.8156, 2.8212, 2.827, 2.8327, 2.8378, 2.8432, 2.8482, 2.8533, 2.8584, 2.8632, 2.8678, 2.8727, 2.8775, 2.8823, 2.887, 2.8917, 2.8961, 2.9008, 2.9055, 2.9102, 2.9145, 2.9192, 2.9235, 2.9282, 2.9317, 2.9372, 2.9421, 2.9461, 2.9505, 2.9548, 2.9594, 2.9631, 2.9685, 2.9726, 2.9768, 2.9819, 2.9856, 2.9898, 2.9942, 2.998, 3.0025, 3.007, 3.0111, 3.0153, 3.0197, 3.0233, 3.0277, 3.0319, 3.0358, 3.0399, 3.044, 3.048, 3.052, 3.0559, 3.0598, 3.0637, 3.0675, 3.0715, 3.0754, 3.0789, 3.0828, 3.0868, 3.0903, 3.0942, 3.098, 3.1016, 3.1052, 3.1087, 3.1124, 3.116, 3.1195, 3.1231, 3.1266, 3.1299, 3.1336, 3.1373, 3.1405, 3.144, 3.1472, 3.1504, 3.1538, 3.1573, 3.1606, 3.1638, 3.167, 3.1703, 3.1739, 3.1766, 3.1797, 3.183, 3.1859, 3.1891, 3.1923, 3.1951, 3.1981, 3.2011, 3.2045, 3.2072, 3.21, 3.213, 3.2157, 3.2192, 3.2216, 3.2246, 3.2275, 3.2302, 3.233, 3.2359, 3.2385, 3.2408, 3.2436, 3.2463, 3.2489, 3.2515, 3.2541, 3.2566, 3.2592, 3.262, 3.2645, 3.2668, 3.2697, 3.2722, 3.2748, 3.2776, 3.2797, 3.2823, 3.2847, 3.2871, 3.2892, 3.292, 3.2946, 3.2965, 3.2989, 3.3014, 3.3036, 3.3056, 3.3085, 3.3106, 3.3129, 3.3153, 3.3177, 3.3199, 3.322, 3.3244, 3.3266, 3.3286, 3.3311, 3.3333, 3.3353, 3.3377, 3.3396, 3.3419, 3.3439, 3.346, 3.3483, 3.3503, 3.3521, 3.3544, 3.3566, 3.3585, 3.3606, 3.3625, 3.3645, 3.3667, 3.3684, 3.3705, 3.3725, 3.3744, 3.3763, 3.3784, 3.3802, 3.3822, 3.384, 3.3864, 3.3879, 3.3899, 3.3913, 3.3937, 3.3954, 3.3972, 3.399, 3.4009, 3.4027, 3.4046, 3.4064, 3.408, 3.4098, 3.4116, 3.4131, 3.4151, 3.4168, 3.4186, 3.4203, 3.4219, 3.4235, 3.4253, 3.4271, 3.4288, 3.4302, 3.432, 3.4339, 3.4355, 3.437, 3.4389, 3.4403, 3.442, 3.4436, 3.4451, 3.4469, 3.4483, 3.45, 3.4514, 3.4531, 3.4547, 3.4562, 3.4577, 3.4592, 3.4608, 3.4623, 3.4637, 3.4652, 3.4667, 3.4682, 3.4698, 3.4712, 3.4727, 3.4744, 3.4757, 3.4771, 3.4785, 3.48, 3.4814, 3.4829, 3.4844, 3.4857, 3.4871, 3.4884, 3.4899, 3.4912, 3.4926, 3.494, 3.4955, 3.4968, 3.498, 3.4994, 3.5009, 3.5021, 3.5036, 3.5046, 3.5058, 3.5076, 3.5086, 3.5101, 3.511, 3.5124, 3.5135, 3.5148, 3.5163, 3.5176, 3.5186, 3.52, 3.5213, 3.5225, 3.5236, 3.5248, 3.5257, 3.5271, 3.5284, 3.5296, 3.5309, 3.5318, 3.5327, 3.534, 3.5352, 3.5362, 3.5374, 3.5383, 3.539, 3.5401, 3.5412, 3.5422, 3.5431, 3.5435, 3.5446, 3.545, 3.5457, 3.5465, 3.547, 3.5475, 3.5484, 3.5489, 3.5495, 3.5499, 3.5503, 3.5506, 3.5513, 3.5516, 3.5524, 3.5527, 3.5529, 3.5533, 3.5537, 3.554, 3.5544, 3.5551, 3.5554, 3.5555, 3.5562, 3.5563, 3.5564, 3.5567, 3.5571, 3.5574, 3.5578, 3.5581, 3.5585, 3.5587, 3.5592, 3.5591, 3.5595, 3.5602, 3.5605, 3.5608, 3.5608, 3.561, 3.5616, 3.5617, 3.562, 3.5626, 3.5627, 3.5633, 3.5633, 3.5637, 3.5641, 3.5644, 3.5648, 3.5649, 3.5647, 3.5655, 3.5659, 3.5661, 3.5665, 3.5673, 3.5669, 3.5675, 3.5677, 3.5679, 3.5686, 3.569, 3.5689, 3.5698, 3.5701, 3.5703, 3.5707, 3.5711, 3.5712, 3.5716, 3.5723, 3.5724, 3.5727, 3.5732, 3.5737, 3.574, 3.5743, 3.5747, 3.5751, 3.5754, 3.576, 3.5763, 3.5767, 3.5771, 3.5773, 3.5781, 3.5784, 3.5789, 3.5792, 3.5793, 3.5802, 3.5805, 3.5808, 3.5813, 3.5817, 3.5821, 3.5827, 3.5832, 3.5837, 3.5841, 3.5847, 3.5852, 3.5857, 3.586, 3.5865, 3.5871, 3.5873, 3.5879, 3.5886, 3.5894, 3.5896, 3.59, 3.5906, 3.591, 3.5916, 3.5922, 3.5925, 3.5929, 3.594, 3.5942, 3.5943, 3.5953, 3.5955, 3.5965, 3.5968, 3.5974, 3.5977, 3.5986, 3.5986, 3.5996, 3.5995, 3.6005, 3.6008, 3.6018, 3.6019, 3.6026, 3.603, 3.6034, 3.604, 3.6048, 3.6051, 3.6057, 3.6062, 3.6066, 3.6074, 3.6078, 3.6084, 3.6089, 3.6095, 3.61, 3.6105, 3.6109, 3.6115, 3.612, 3.6127, 3.6133, 3.6136, 3.6142, 3.6146, 3.6152, 3.6157, 3.6163, 3.6169, 3.6173, 3.6178, 3.6184, 3.6189, 3.6195, 3.6199, 3.6204, 3.621, 3.6214, 3.6218, 3.6224, 3.6228, 3.6233, 3.6238, 3.6243, 3.6248, 3.6255, 3.6259, 3.6263, 3.627, 3.6273, 3.6278, 3.6284, 3.6287, 3.6292, 3.6296, 3.6301, 3.6306, 3.6311, 3.6316, 3.632, 3.6325, 3.633, 3.6334, 3.6339, 3.6343, 3.6348, 3.6352, 3.6358, 3.6361, 3.6364, 3.6369, 3.6374, 3.6379, 3.6382, 3.6389, 3.6392, 3.6393, 3.6401, 3.6405, 3.6407, 3.6413, 3.6416, 3.6421, 3.6425, 3.6432, 3.6434, 3.644, 3.6443, 3.6448, 3.645, 3.6454, 3.646, 3.6464, 3.647, 3.6475, 3.6478, 3.6478, 3.6483, 3.6492, 3.6494, 3.6498, 3.65, 3.6508, 3.6511, 3.6516, 3.652, 3.6524, 3.6527, 3.6532, 3.6537, 3.6542, 3.6548, 3.6551, 3.6554, 3.6559, 3.6564, 3.6568, 3.6573, 3.6578, 3.6581, 3.6586, 3.659, 3.6594, 3.6598, 3.66, 3.6606, 3.6612, 3.6613, 3.6621, 3.6626, 3.6629, 3.6634, 3.6637, 3.6641, 3.6645, 3.665, 3.6653, 3.6659, 3.666, 3.6664, 3.6669, 3.6675, 3.6677, 3.6682, 3.6686, 3.6688, 3.6694, 3.6698, 3.6701, 3.6706, 3.6709, 3.6714, 3.6717, 3.6719, 3.6726, 3.6727, 3.6732, 3.6735, 3.6739, 3.6745, 3.6748, 3.675, 3.6754, 3.6756, 3.676, 3.6765, 3.6769, 3.6771, 3.6775, 3.6777, 3.678, 3.6783, 3.6788, 3.6793, 3.6793, 3.6796, 3.68, 3.6804, 3.6805, 3.6811, 3.6812, 3.6817, 3.6819, 3.6822, 3.6825, 3.6827, 3.683, 3.6834, 3.6837, 3.6838, 3.6842, 3.6846, 3.6849, 3.685, 3.6852, 3.6858, 3.6861, 3.6862, 3.6864, 3.6867, 3.6873, 3.6874, 3.6875, 3.6878, 3.6882, 3.6884, 3.6886, 3.6889, 3.6892, 3.6893, 3.6897, 3.69, 3.6901, 3.6905, 3.6907, 3.6911, 3.6911, 3.6916, 3.6917, 3.6921, 3.6922, 3.6925, 3.6929, 3.6928, 3.6931, 3.6936, 3.6936, 3.6939, 3.6942, 3.6945, 3.6946, 3.6949, 3.6951, 3.6953, 3.6958, 3.6961, 3.6962, 3.6964, 3.6967, 3.6969, 3.6967, 3.6972, 3.6976, 3.6977, 3.6979, 3.6982, 3.6987, 3.6988, 3.6989, 3.6995, 3.6993, 3.6997, 3.7002, 3.7003, 3.7005, 3.7002, 3.7007, 3.7009, 3.7011, 3.7024, 3.7008, 3.7018, 3.7023, 3.7023, 3.7024, 3.7023, 3.7026, 3.7041, 3.7035, 3.7038, 3.7044, 3.7044, 3.7047, 3.705, 3.7062, 3.7053, 3.7056, 3.706, 3.7064, 3.7063, 3.7067, 3.7067, 3.707, 3.7073, 3.7076, 3.7081, 3.7078, 3.7081, 3.7086, 3.7088, 3.7089, 3.7094, 3.7096, 3.7098, 3.7101, 3.7102, 3.7104, 3.7105, 3.711, 3.7112, 3.7114, 3.7115, 3.712, 3.7123, 3.7127, 3.7129, 3.7131, 3.7134, 3.7135, 3.714, 3.7142, 3.7146, 3.7149, 3.7149, 3.7153, 3.7156, 3.7158, 3.7161, 3.7164, 3.7168, 3.7169, 3.7171, 3.7177, 3.7178, 3.718, 3.7183, 3.7188, 3.719, 3.7192, 3.7194, 3.7199, 3.7201, 3.7202, 3.7205, 3.721, 3.7213, 3.7217, 3.7219, 3.7222, 3.7225, 3.7227, 3.7231, 3.7233, 3.7235, 3.7239, 3.7242, 3.7245, 3.7249, 3.7252, 3.7254, 3.7259, 3.7261, 3.7263, 3.7268, 3.727, 3.7272, 3.7278, 3.7281, 3.7283, 3.7285, 3.7291, 3.7294, 3.7297, 3.7299, 3.7302, 3.7307, 3.7308, 3.7314, 3.7316, 3.7319, 3.7324, 3.7329, 3.7329, 3.7332, 3.7336, 3.7341, 3.7343, 3.7347, 3.735, 3.7353, 3.7358, 3.7361, 3.7365, 3.7368, 3.737, 3.7378, 3.738, 3.7382, 3.7387, 3.739, 3.7394, 3.7397, 3.7402, 3.7405, 3.7408, 3.7413, 3.7416, 3.7421, 3.7422, 3.7428, 3.7432, 3.7435, 3.7439, 3.7443, 3.7448, 3.7451, 3.7457, 3.7458, 3.7464, 3.7467, 3.7473, 3.7475, 3.7479, 3.7484, 3.7487, 3.749, 3.7495, 3.7498, 3.7504, 3.7505, 3.7513, 3.7518, 3.7519, 3.7524, 3.7529, 3.7534, 3.7537, 3.7542, 3.7545, 3.7548, 3.7554, 3.7559, 3.7562, 3.7566, 3.7572, 3.7576, 3.7581, 3.7585, 3.759, 3.7593, 3.7598, 3.7602, 3.7606, 3.7612, 3.7616, 3.7622, 3.7626, 3.763, 3.7634, 3.7638, 3.7645, 3.7648, 3.7653, 3.7658, 3.7662, 3.767, 3.7672, 3.7677, 3.7682, 3.7685, 3.769, 3.7695, 3.7701, 3.7706, 3.7711, 3.7716, 3.772, 3.7726, 3.7728, 3.7736, 3.774, 3.7744, 3.775, 3.7755, 3.776, 3.7766, 3.7769, 3.7774, 3.778, 3.7787, 3.779, 3.7797, 3.78, 3.7806, 3.7812, 3.7817, 3.7823, 3.783, 3.7834, 3.7839, 3.7844, 3.7849, 3.7854, 3.7859, 3.7865, 3.7872, 3.7876, 3.7882, 3.7887, 3.7894, 3.7898, 3.7907, 3.7914, 3.7919, 3.7924, 3.793, 3.7936, 3.7943, 3.7947, 3.7953, 3.7958, 3.7964, 3.7969, 3.7975, 3.7982, 3.7989, 3.7994, 3.7999, 3.8006, 3.801, 3.8018, 3.8023, 3.8028, 3.8035, 3.804, 3.8047, 3.8053, 3.8059, 3.8066, 3.807, 3.8078, 3.8083, 3.8089, 3.8096, 3.8102, 3.8109, 3.8114, 3.8121, 3.8127, 3.8133, 3.8141, 3.8148, 3.8153, 3.8159, 3.8164, 3.8172, 3.8178, 3.8186, 3.8193, 3.8197, 3.8205, 3.821, 3.8216, 3.8225, 3.8231, 3.8238, 3.8245, 3.8252, 3.8258, 3.8265, 3.8271, 3.8279, 3.8283, 3.8292, 3.8299, 3.8306, 3.8314, 3.8319, 3.8329, 3.8333, 3.8338, 3.8348, 3.8355, 3.8362, 3.8368, 3.8376, 3.8384, 3.8392, 3.8399, 3.8406, 3.8411, 3.842, 3.8428, 3.8432, 3.8441, 3.8449, 3.8458, 3.8464, 3.8471, 3.848, 3.8488, 3.8492, 3.8502, 3.8507, 3.8514, 3.8524, 3.8531, 3.8539, 3.8548, 3.8555, 3.8561, 3.8569, 3.858, 3.8587, 3.8593, 3.86, 3.8609, 3.8616, 3.8625, 3.8635, 3.8641, 3.8649, 3.8658, 3.8664, 3.8674, 3.8679, 3.869, 3.8698, 3.8705, 3.8712, 3.8721, 3.8725, 3.8735, 3.8743, 3.875, 3.8758, 3.8767, 3.8774, 3.8784, 3.8792, 3.8798, 3.8808, 3.8815, 3.8828, 3.8838, 3.8849, 3.885, 3.8856, 3.8864, 3.887, 3.888, 3.8888, 3.8897, 3.8904, 3.8914, 3.892, 3.8928, 3.8938, 3.8944, 3.8952, 3.8961, 3.8968, 3.8978, 3.8983, 3.8991, 3.8999, 3.9008, 3.9014, 3.9022, 3.9031, 3.9039, 3.9045, 3.9053, 3.906, 3.9069, 3.9076, 3.9084, 3.9092, 3.9099, 3.9108, 3.9114, 3.9121, 3.9129, 3.9136, 3.9145, 3.9153, 3.9159, 3.9167, 3.9175, 3.918, 3.9188, 3.9196, 3.9203, 3.921, 3.9218, 3.9224, 3.9233, 3.924, 3.9247, 3.9254, 3.9261, 3.9268, 3.9276, 3.9282, 3.929, 3.9297, 3.9305, 3.931, 3.9319, 3.9328, 3.9334, 3.934, 3.9349, 3.9355, 3.9361, 3.9369, 3.9376, 3.9383, 3.9391, 3.9398, 3.9405, 3.9413, 3.9418, 3.9425, 3.9435, 3.944, 3.9446, 3.9456, 3.9462, 3.9469, 3.9477, 3.9483, 3.9491, 3.9496, 3.9505, 3.9511, 3.9518, 3.9526, 3.9533, 3.954, 3.9547, 3.9554, 3.9561, 3.9566, 3.9575, 3.9582, 3.9588, 3.9595, 3.9603, 3.9611, 3.9618, 3.9624, 3.9631, 3.964, 3.9647, 3.9652, 3.966, 3.9667, 3.9676, 3.968, 3.9688, 3.9695, 3.9702, 3.9709, 3.9716, 3.9723, 3.973, 3.9737, 3.9745, 3.9751, 3.9759, 3.9767, 3.9773, 3.9779, 3.9788, 3.9794, 3.9801, 3.981, 3.9816, 3.9823, 3.983, 3.9837, 3.9846, 3.9851, 3.9856, 3.9864, 3.9871, 3.9879, 3.9885, 3.9892, 3.9901, 3.9908, 3.9914, 3.9921, 3.9929, 3.9935, 3.9943, 3.9949, 3.9958, 3.9964, 3.9972, 3.9979, 3.9985, 3.9994, 4, 4.0007, 4.0013, 4.0022, 4.0029, 4.0038, 4.0042, 4.005, 4.0058, 4.0065, 4.0072, 4.008, 4.0086, 4.0093, 4.0099, 4.0108, 4.0116, 4.0122, 4.0129, 4.0135, 4.0144, 4.0151, 4.0158, 4.0164, 4.0172, 4.0178, 4.0187, 4.0194, 4.0201, 4.021, 4.0215, 4.0223, 4.0231, 4.0239, 4.0244, 4.0252, 4.026, 4.0268, 4.0273, 4.028, 4.0288, 4.0295, 4.0304, 4.0311, 4.0318, 4.0325, 4.0333, 4.034, 4.0347, 4.0356, 4.0362, 4.0371, 4.0377, 4.0385, 4.0391, 4.0399, 4.0405, 4.0414, 4.042, 4.0428, 4.0436, 4.0443, 4.0449, 4.0458, 4.0465, 4.0473, 4.048, 4.0488, 4.0495, 4.0503, 4.051, 4.0517, 4.0525, 4.0531, 4.054, 4.0547, 4.0554, 4.0563, 4.057, 4.0578, 4.0585, 4.0592, 4.0601, 4.0607, 4.0616, 4.0621, 4.0631, 4.0637, 4.0645, 4.0652, 4.0659, 4.0667, 4.0675, 4.0683, 4.069, 4.0697, 4.0706, 4.0714, 4.0721, 4.0728, 4.0735, 4.0744, 4.0751, 4.0758, 4.0766, 4.0773, 4.0781, 4.079, 4.0797, 4.0805, 4.0812, 4.0821, 4.0827, 4.0834, 4.0843, 4.0852, 4.0858, 4.0867, 4.0872, 4.0882, 4.089, 4.0897, 4.0905, 4.0912, 4.0919, 4.0928, 4.0936, 4.0944, 4.0951, 4.0959, 4.0965, 4.0974, 4.0982, 4.099, 4.1, 4.1006, 4.1015, 4.1021, 4.1029, 4.1037, 4.1044, 4.1053, 4.1062, 4.1069, 4.1076, 4.1084, 4.1091, 4.1099, 4.1108, 4.1114, 4.1123, 4.113, 4.1139, 4.1146, 4.1154, 4.1162, 4.117, 4.1179, 4.1186, 4.1194, 4.1202, 4.1209, 4.1218, 4.1225, 4.1235, 4.1243, 4.1252, 4.1257, 4.1266, 4.1273, 4.1282, 4.129, 4.1297, 4.1306, 4.1312, 4.1321, 4.1329, 4.1337, 4.1346, 4.1354, 4.1361, 4.1369, 4.1379, 4.1386, 4.1392, 4.1402, 4.1409, 4.1417, 4.1427, 4.1433, 4.1442, 4.1451, 4.1458, 4.1467, 4.1474, 4.1483, 4.149, 4.1498, 4.1507, 4.1514, 4.1522, 4.1531, 4.1538, 4.1548, 4.1555, 4.1564, 4.1571, 4.1579, 4.1586, 4.1596, 4.1604, 4.1611, 4.1621, 4.1628, 4.1637, 4.1645, 4.1653, 4.1661, 4.1669, 4.1677, 4.1684, 4.1693, 4.1701, 4.1711, 4.1719, 4.1726, 4.1735, 4.1743, 4.1752, 4.1759, 4.1768, 4.1775, 4.1785, 4.1793, 4.1802, 4.181, 4.1817, 4.1825, 4.1834, 4.1843, 4.1849, 4.186, 4.1867, 4.1876, 4.1884, 4.1894, 4.1901, 4.1909, 4.1917, 4.1925, 4.1935, 4.1944, 4.1951, 4.1959, 4.1967, 4.1975, 4.1985, 4.1992, 4.2, 4.1992, 4.1992, 4.1991, 4.1991, 4.1993, 4.1991, 4.1993, 4.1993, 4.1991, 4.1993, 4.1991, 4.1992, 4.199, 4.1991, 4.1994, 4.1996, 4.2001, 4.2005, 4.2009, 4.2002, 4.2, 4.2001, 4.2002, 4.2001, 4.2001, 4.2, 4.2001, 4.2002, 4.2001, 4.2003, 4.2, 4.2001, 4.2002, 4.2002, 4.2002, 4.2002, 4.2001, 4.2, 4.2001, 4.2, 4.2001, 4.2, 4.2002, 4.2003, 4.2003, 4.2002, 4.2, 4.2001, 4.2001, 4.2002, 4.2002, 4.2001, 4.2003, 4.2001, 4.2002, 4.2001, 4.2001, 4.2004, 4.2003, 4.2003, 4.2006, 4.2006, 4.2008, 4.2004, 4.2005, 4.2004, 4.2004, 4.2003, 4.2004, 4.2005, 4.2004, 4.2001, 4.2004, 4.2004, 4.2005, 4.2004, 4.2004, 4.2004, 4.2004, 4.2004, 4.2003, 4.2003, 4.2004, 4.2001, 4.2003, 4.2003, 4.2003, 4.2002, 4.2003, 4.2002, 4.2001, 4.2004, 4.2002, 4.2003, 4.2004, 4.2003, 4.2006, 4.2003, 4.2002, 4.2002, 4.2002, 4.2003, 4.2003, 4.2003, 4.2002, 4.2003, 4.2002, 4.2003, 4.2002, 4.2003, 4.2003, 4.2005, 4.2006, 4.2005, 4.2006, 4.2007, 4.2005, 4.2005, 4.2005, 4.2007, 4.2006, 4.2005, 4.2007, 4.2007, 4.2007, 4.2007, 4.2006, 4.2004, 4.2005, 4.2006, 4.2006, 4.2006, 4.2005, 4.2006]

                        Percentage = [0.00,0.07,0.13,0.20,0.26,0.33,0.39,0.46,0.52,0.59,0.65,0.72,0.79,0.85,0.92,0.98,1.05,1.11,1.18,1.24,1.31,1.37,1.44,1.51,1.57,1.64,1.70,1.77,1.83,1.90,1.96,2.03,2.09,2.16,2.23,2.29,2.36,2.42,2.49,2.55,2.62,2.68,2.75,2.81,2.88,2.95,3.01,3.08,3.14,3.21,3.27,3.34,3.40,3.47,3.54,3.60,3.67,3.73,3.80,3.86,3.93,3.99,4.06,4.12,4.19,4.26,4.32,4.39,4.45,4.52,4.58,4.65,4.71,4.78,4.84,4.91,4.98,5.04,5.11,5.17,5.24,5.30,5.37,5.43,5.50,5.56,5.63,5.70,5.76,5.83,5.89,5.96,6.02,6.09,6.15,6.22,6.28,6.35,6.42,6.48,6.55,6.61,6.68,6.74,6.81,6.87,6.94,7.00,7.07,7.14,7.20,7.27,7.33,7.40,7.46,7.53,7.59,7.66,7.73,7.79,7.86,7.92,7.99,8.05,8.12,8.18,8.25,8.32,8.38,8.45,8.51,8.58,8.64,8.71,8.77,8.84,8.91,8.97,9.04,9.10,9.17,9.23,9.30,9.36,9.43,9.50,9.56,9.63,9.69,9.76,9.82,9.89,9.95,10.02,10.09,10.15,10.22,10.28,10.35,10.41,10.48,10.55,10.61,10.68,10.74,10.81,10.87,10.94,11.00,11.07,11.14,11.20,11.27,11.33,11.40,11.46,11.53,11.59,11.66,11.73,11.79,11.86,11.92,11.99,12.05,12.12,12.18,12.25,12.32,12.38,12.45,12.51,12.58,12.64,12.71,12.77,12.84,12.91,12.97,13.04,13.10,13.17,13.23,13.30,13.36,13.43,13.50,13.56,13.63,13.69,13.76,13.82,13.89,13.95,14.02,14.09,14.15,14.22,14.28,14.35,14.41,14.48,14.54,14.61,14.68,14.74,14.81,14.87,14.94,15.00,15.07,15.13,15.20,15.27,15.33,15.40,15.46,15.53,15.59,15.66,15.73,15.79,15.86,15.92,15.99,16.05,16.12,16.18,16.25,16.32,16.38,16.45,16.51,16.58,16.64,16.71,16.77,16.84,16.91,16.97,17.04,17.10,17.17,17.23,17.30,17.36,17.43,17.50,17.56,17.63,17.69,17.76,17.82,17.89,17.95,18.02,18.09,18.15,18.22,18.28,18.35,18.41,18.48,18.54,18.61,18.68,18.74,18.81,18.87,18.94,19.00,19.07,19.13,19.20,19.27,19.33,19.40,19.46,19.53,19.59,19.66,19.72,19.79,19.86,19.92,19.99,20.05,20.12,20.18,20.25,20.31,20.38,20.45,20.51,20.58,20.64,20.71,20.77,20.84,20.90,20.97,21.04,21.10,21.17,21.23,21.30,21.36,21.43,21.50,21.56,21.63,21.69,21.76,21.82,21.89,21.95,22.02,22.09,22.15,22.22,22.28,22.35,22.41,22.48,22.54,22.61,22.68,22.74,22.81,22.87,22.94,23.00,23.07,23.13,23.20,23.27,23.33,23.40,23.46,23.53,23.59,23.66,23.73,23.79,23.86,23.92,23.99,24.05,24.12,24.18,24.25,24.32,24.38,24.45,24.51,24.58,24.64,24.71,24.77,24.84,24.91,24.97,25.04,25.10,25.17,25.23,25.30,25.37,25.43,25.50,25.56,25.63,25.69,25.76,25.82,25.89,25.96,26.02,26.09,26.15,26.22,26.28,26.35,26.42,26.48,26.55,26.61,26.68,26.74,26.81,26.87,26.94,27.01,27.07,27.14,27.20,27.27,27.33,27.40,27.47,27.53,27.60,27.66,27.73,27.79,27.86,27.92,27.99,28.06,28.12,28.19,28.25,28.32,28.38,28.45,28.52,28.58,28.65,28.71,28.78,28.84,28.91,28.97,29.04,29.11,29.17,29.24,29.30,29.37,29.43,29.50,29.57,29.63,29.70,29.76,29.83,29.89,29.96,30.02,30.09,30.16,30.22,30.29,30.35,30.42,30.48,30.55,30.62,30.68,30.75,30.81,30.88,30.94,31.01,31.07,31.14,31.21,31.27,31.34,31.40,31.47,31.53,31.60,31.66,31.73,31.80,31.86,31.93,31.99,32.06,32.12,32.19,32.26,32.32,32.39,32.45,32.52,32.58,32.65,32.71,32.78,32.85,32.91,32.98,33.04,33.11,33.17,33.24,33.31,33.37,33.44,33.50,33.57,33.63,33.70,33.76,33.83,33.90,33.96,34.03,34.09,34.16,34.22,34.29,34.35,34.42,34.49,34.55,34.62,34.68,34.75,34.81,34.88,34.95,35.01,35.08,35.14,35.21,35.27,35.34,35.40,35.47,35.54,35.60,35.67,35.73,35.80,35.86,35.93,35.99,36.06,36.13,36.19,36.26,36.32,36.39,36.45,36.52,36.59,36.65,36.72,36.78,36.85,36.91,36.98,37.04,37.11,37.18,37.24,37.31,37.37,37.44,37.50,37.57,37.63,37.70,37.77,37.83,37.90,37.96,38.03,38.09,38.16,38.22,38.29,38.36,38.42,38.49,38.55,38.62,38.68,38.75,38.81,38.88,38.95,39.01,39.08,39.14,39.21,39.27,39.34,39.41,39.47,39.54,39.60,39.67,39.73,39.80,39.86,39.93,40.00,40.06,40.13,40.19,40.26,40.32,40.39,40.45,40.52,40.59,40.65,40.72,40.78,40.85,40.91,40.98,41.05,41.11,41.18,41.24,41.31,41.37,41.44,41.50,41.57,41.64,41.70,41.77,41.83,41.90,41.96,42.03,42.10,42.16,42.23,42.29,42.36,42.42,42.49,42.55,42.62,42.69,42.75,42.82,42.88,42.95,43.01,43.08,43.14,43.21,43.28,43.34,43.41,43.47,43.54,43.60,43.67,43.74,43.80,43.87,43.93,44.00,44.06,44.13,44.19,44.26,44.33,44.39,44.46,44.52,44.59,44.65,44.72,44.78,44.85,44.92,44.98,45.05,45.11,45.18,45.24,45.31,45.37,45.44,45.51,45.57,45.64,45.70,45.77,45.83,45.90,45.96,46.03,46.10,46.16,46.23,46.29,46.36,46.42,46.49,46.55,46.62,46.69,46.75,46.82,46.88,46.95,47.01,47.08,47.15,47.21,47.28,47.34,47.41,47.47,47.54,47.60,47.67,47.74,47.80,47.87,47.93,48.00,48.06,48.13,48.19,48.26,48.33,48.39,48.46,48.52,48.59,48.65,48.72,48.78,48.85,48.92,48.98,49.05,49.11,49.18,49.24,49.31,49.37,49.44,49.51,49.57,49.64,49.70,49.77,49.83,49.90,49.96,50.03,50.10,50.16,50.23,50.29,50.36,50.42,50.49,50.55,50.62,50.69,50.75,50.82,50.88,50.95,51.01,51.08,51.14,51.21,51.28,51.34,51.41,51.47,51.54,51.60,51.67,51.73,51.80,51.87,51.93,52.00,52.06,52.13,52.19,52.26,52.32,52.39,52.46,52.52,52.59,52.65,52.72,52.78,52.85,52.91,52.98,53.05,53.11,53.18,53.24,53.31,53.37,53.44,53.51,53.57,53.64,53.70,53.77,53.83,53.90,53.96,54.03,54.10,54.16,54.23,54.29,54.36,54.42,54.49,54.55,54.62,54.68,54.75,54.82,54.88,54.95,55.01,55.08,55.14,55.21,55.28,55.34,55.41,55.47,55.54,55.60,55.67,55.73,55.80,55.86,55.93,56.00,56.06,56.13,56.19,56.26,56.32,56.39,56.45,56.52,56.59,56.65,56.72,56.78,56.85,56.91,56.98,57.04,57.11,57.18,57.24,57.31,57.37,57.44,57.50,57.57,57.63,57.70,57.77,57.83,57.90,57.96,58.03,58.09,58.16,58.22,58.29,58.36,58.42,58.49,58.55,58.62,58.68,58.75,58.81,58.88,58.95,59.01,59.08,59.14,59.21,59.27,59.34,59.40,59.47,59.54,59.60,59.67,59.73,59.80,59.86,59.93,59.99,60.06,60.12,60.19,60.26,60.32,60.39,60.45,60.52,60.58,60.65,60.71,60.78,60.85,60.91,60.98,61.04,61.11,61.17,61.24,61.30,61.37,61.43,61.50,61.57,61.63,61.70,61.76,61.83,61.89,61.96,62.02,62.09,62.16,62.22,62.29,62.35,62.42,62.48,62.55,62.61,62.68,62.74,62.81,62.88,62.94,63.01,63.07,63.14,63.20,63.27,63.33,63.40,63.47,63.53,63.60,63.66,63.73,63.79,63.86,63.92,63.99,64.05,64.12,64.19,64.25,64.32,64.38,64.45,64.51,64.58,64.64,64.71,64.77,64.84,64.91,64.97,65.04,65.10,65.17,65.23,65.30,65.36,65.43,65.50,65.56,65.63,65.69,65.76,65.82,65.89,65.95,66.02,66.08,66.15,66.22,66.28,66.35,66.41,66.48,66.54,66.61,66.67,66.74,66.81,66.87,66.94,67.00,67.07,67.13,67.20,67.26,67.33,67.39,67.46,67.53,67.59,67.66,67.72,67.79,67.85,67.92,67.98,68.05,68.12,68.18,68.25,68.31,68.38,68.44,68.51,68.57,68.64,68.70,68.77,68.84,68.90,68.97,69.03,69.10,69.16,69.23,69.29,69.36,69.43,69.49,69.56,69.62,69.69,69.75,69.82,69.88,69.95,70.02,70.08,70.15,70.21,70.28,70.34,70.41,70.47,70.54,70.61,70.67,70.74,70.80,70.87,70.93,71.00,71.06,71.13,71.20,71.26,71.33,71.39,71.46,71.52,71.59,71.65,71.72,71.79,71.85,71.92,71.98,72.05,72.11,72.18,72.24,72.31,72.38,72.44,72.51,72.57,72.64,72.70,72.77,72.83,72.90,72.97,73.03,73.10,73.16,73.23,73.29,73.36,73.42,73.49,73.56,73.62,73.69,73.75,73.82,73.88,73.95,74.01,74.08,74.15,74.21,74.28,74.34,74.41,74.47,74.54,74.60,74.67,74.74,74.80,74.87,74.93,75.00,75.06,75.13,75.20,75.26,75.33,75.39,75.46,75.52,75.59,75.65,75.72,75.79,75.85,75.92,75.98,76.05,76.11,76.18,76.24,76.31,76.38,76.44,76.51,76.57,76.64,76.70,76.77,76.83,76.90,76.97,77.03,77.10,77.16,77.23,77.29,77.36,77.42,77.49,77.56,77.62,77.69,77.75,77.82,77.88,77.95,78.01,78.08,78.15,78.21,78.28,78.34,78.41,78.47,78.54,78.60,78.67,78.74,78.80,78.87,78.93,79.00,79.06,79.13,79.19,79.26,79.33,79.39,79.46,79.52,79.59,79.65,79.72,79.78,79.85,79.92,79.98,80.05,80.11,80.18,80.24,80.31,80.37,80.44,80.51,80.57,80.64,80.70,80.77,80.83,80.90,80.96,81.03,81.10,81.16,81.23,81.29,81.36,81.42,81.49,81.55,81.62,81.69,81.75,81.82,81.88,81.95,82.01,82.08,82.14,82.21,82.27,82.34,82.41,82.47,82.54,82.60,82.67,82.73,82.80,82.86,82.93,83.00,83.06,83.13,83.19,83.26,83.32,83.39,83.45,83.52,83.59,83.65,83.72,83.78,83.85,83.91,83.98,84.04,84.11,84.18,84.24,84.31,84.37,84.44,84.50,84.57,84.63,84.70,84.77,84.83,84.90,84.96,85.03,85.09,85.16,85.22,85.29,85.36,85.42,85.49,85.55,85.62,85.68,85.75,85.81,85.88,85.95,86.01,86.08,86.14,86.21,86.27,86.34,86.40,86.47,86.53,86.60,86.67,86.73,86.80,86.86,86.93,86.99,87.06,87.12,87.19,87.26,87.32,87.39,87.45,87.52,87.58,87.65,87.71,87.78,87.85,87.91,87.98,88.04,88.11,88.17,88.24,88.30,88.37,88.44,88.50,88.57,88.63,88.70,88.76,88.83,88.89,88.96,89.03,89.09,89.16,89.22,89.29,89.35,89.42,89.48,89.55,89.61,89.68,89.75,89.81,89.88,89.94,90.01,90.07,90.14,90.20,90.27,90.34,90.40,90.47,90.53,90.60,90.66,90.73,90.79,90.86,90.93,90.99,91.06,91.12,91.19,91.25,91.32,91.38,91.45,91.52,91.58,91.65,91.71,91.78,91.84,91.91,91.97,92.04,92.11,92.17,92.24,92.30,92.37,92.43,92.50,92.56,92.63,92.69,92.76,92.83,92.89,92.96,93.02,93.09,93.15,93.22,93.28,93.35,93.42,93.48,93.55,93.61,93.68,93.74,93.81,93.87,93.94,94.01,94.07,94.14,94.20,94.27,94.33,94.40,94.46,94.53,94.60,94.66,94.73,94.79,94.86,94.92,94.99,95.05,95.12,95.18,95.25,95.32,95.38,95.45,95.51,95.58,95.64,95.71,95.77,95.84,95.91,95.97,96.04,96.10,96.17,96.23,96.30,96.36,96.43,96.50,96.56,96.63,96.69,96.76,96.82,96.89,96.95,97.02,97.09,97.15,97.21,97.26,97.32,97.37,97.42,97.48,97.53,97.58,97.62,97.67,97.72,97.76,97.80,97.85,97.89,97.93,97.98,98.02,98.06,98.10,98.14,98.18,98.21,98.25,98.28,98.32,98.35,98.38,98.42,98.45,98.48,98.51,98.54,98.57,98.60,98.63,98.66,98.68,98.71,98.74,98.76,98.79,98.82,98.84,98.86,98.89,98.91,98.94,98.96,98.98,99.00,99.02,99.05,99.07,99.09,99.11,99.13,99.15,99.17,99.19,99.21,99.23,99.25,99.27,99.28,99.30,99.32,99.34,99.35,99.37,99.39,99.40,99.42,99.43,99.45,99.46,99.48,99.49,99.51,99.52,99.54,99.55,99.56,99.58,99.59,99.60,99.62,99.63,99.64,99.66,99.67,99.68,99.69,99.70,99.72,99.73,99.74,99.75,99.76,99.77,99.78,99.79,99.80,99.81,99.82,99.83,99.84,99.85,99.86,99.87,99.88,99.89,99.90,99.91,99.92,99.93,99.94,99.94,99.95,99.96,99.97,99.98,99.99,99.99,100.00,100.01,100.02,100.02,100.03,100.04,100.05,100.05,100.05,]

                        mapping = list(zip(Volt,Percentage))
                        # print(len(Volt))
                        # print(len(Percentage))
                        # print(len(mapping))

                        # Find the closest voltage in the list of Volt
                        closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                        # Find the index of the closest voltage
                        index = Volt.index(closest_voltage)
                        # print(index)
                        # print(Percentage[index])

                        # Check if the index is not the last index to avoid index out of range
                        if index < len(Volt) - 1:
                            if(input_voltage > closest_voltage) or index == 0  and input_voltage != 0:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")
                                # Calculate the slope using the closest_voltage and next_voltage
                                if next_voltage != closest_voltage :
                                    slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                    closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                else:
                                    next_voltage = Volt[index + 2]
                                    slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                    closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)

                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage < closest_voltage and input_voltage != 0:
                                next_voltage = Volt[index - 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage == 0:
                                closest_percentage = 0
                            else:
                                next_voltage = Volt[index + 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif index + 1 == len(Volt):

                            if input_voltage < closest_voltage and input_voltage != 0:
                                # print("This is herer 3")
                                next_voltage = Volt[index - 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage == 0:
                                closest_percentage = 0
                            elif input_voltage > Volt[index]:
                                # print(Volt[index])
                                closest_percentage = 0
                            else:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")                     
                        else:
                            print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                        return closest_percentage
                    def SOH2(input_voltage):
                        global closest_percentage2
                        closest_percentage2 = 0
                        # print("This is SOH2")
                        Volt = [ 2.7011, 2.7546, 2.7684, 2.7789, 2.7877, 2.7956, 2.8026, 2.8092, 2.8156, 2.8212, 2.827, 2.8327, 2.8378, 2.8432, 2.8482, 2.8533, 2.8584, 2.8632, 2.8678, 2.8727, 2.8775, 2.8823, 2.887, 2.8917, 2.8961, 2.9008, 2.9055, 2.9102, 2.9145, 2.9192, 2.9235, 2.9282, 2.9317, 2.9372, 2.9421, 2.9461, 2.9505, 2.9548, 2.9594, 2.9631, 2.9685, 2.9726, 2.9768, 2.9819, 2.9856, 2.9898, 2.9942, 2.998, 3.0025, 3.007, 3.0111, 3.0153, 3.0197, 3.0233, 3.0277, 3.0319, 3.0358, 3.0399, 3.044, 3.048, 3.052, 3.0559, 3.0598, 3.0637, 3.0675, 3.0715, 3.0754, 3.0789, 3.0828, 3.0868, 3.0903, 3.0942, 3.098, 3.1016, 3.1052, 3.1087, 3.1124, 3.116, 3.1195, 3.1231, 3.1266, 3.1299, 3.1336, 3.1373, 3.1405, 3.144, 3.1472, 3.1504, 3.1538, 3.1573, 3.1606, 3.1638, 3.167, 3.1703, 3.1739, 3.1766, 3.1797, 3.183, 3.1859, 3.1891, 3.1923, 3.1951, 3.1981, 3.2011, 3.2045, 3.2072, 3.21, 3.213, 3.2157, 3.2192, 3.2216, 3.2246, 3.2275, 3.2302, 3.233, 3.2359, 3.2385, 3.2408, 3.2436, 3.2463, 3.2489, 3.2515, 3.2541, 3.2566, 3.2592, 3.262, 3.2645, 3.2668, 3.2697, 3.2722, 3.2748, 3.2776, 3.2797, 3.2823, 3.2847, 3.2871, 3.2892, 3.292, 3.2946, 3.2965, 3.2989, 3.3014, 3.3036, 3.3056, 3.3085, 3.3106, 3.3129, 3.3153, 3.3177, 3.3199, 3.322, 3.3244, 3.3266, 3.3286, 3.3311, 3.3333, 3.3353, 3.3377, 3.3396, 3.3419, 3.3439, 3.346, 3.3483, 3.3503, 3.3521, 3.3544, 3.3566, 3.3585, 3.3606, 3.3625, 3.3645, 3.3667, 3.3684, 3.3705, 3.3725, 3.3744, 3.3763, 3.3784, 3.3802, 3.3822, 3.384, 3.3864, 3.3879, 3.3899, 3.3913, 3.3937, 3.3954, 3.3972, 3.399, 3.4009, 3.4027, 3.4046, 3.4064, 3.408, 3.4098, 3.4116, 3.4131, 3.4151, 3.4168, 3.4186, 3.4203, 3.4219, 3.4235, 3.4253, 3.4271, 3.4288, 3.4302, 3.432, 3.4339, 3.4355, 3.437, 3.4389, 3.4403, 3.442, 3.4436, 3.4451, 3.4469, 3.4483, 3.45, 3.4514, 3.4531, 3.4547, 3.4562, 3.4577, 3.4592, 3.4608, 3.4623, 3.4637, 3.4652, 3.4667, 3.4682, 3.4698, 3.4712, 3.4727, 3.4744, 3.4757, 3.4771, 3.4785, 3.48, 3.4814, 3.4829, 3.4844, 3.4857, 3.4871, 3.4884, 3.4899, 3.4912, 3.4926, 3.494, 3.4955, 3.4968, 3.498, 3.4994, 3.5009, 3.5021, 3.5036, 3.5046, 3.5058, 3.5076, 3.5086, 3.5101, 3.511, 3.5124, 3.5135, 3.5148, 3.5163, 3.5176, 3.5186, 3.52, 3.5213, 3.5225, 3.5236, 3.5248, 3.5257, 3.5271, 3.5284, 3.5296, 3.5309, 3.5318, 3.5327, 3.534, 3.5352, 3.5362, 3.5374, 3.5383, 3.539, 3.5401, 3.5412, 3.5422, 3.5431, 3.5435, 3.5446, 3.545, 3.5457, 3.5465, 3.547, 3.5475, 3.5484, 3.5489, 3.5495, 3.5499, 3.5503, 3.5506, 3.5513, 3.5516, 3.5524, 3.5527, 3.5529, 3.5533, 3.5537, 3.554, 3.5544, 3.5551, 3.5554, 3.5555, 3.5562, 3.5563, 3.5564, 3.5567, 3.5571, 3.5574, 3.5578, 3.5581, 3.5585, 3.5587, 3.5592, 3.5591, 3.5595, 3.5602, 3.5605, 3.5608, 3.5608, 3.561, 3.5616, 3.5617, 3.562, 3.5626, 3.5627, 3.5633, 3.5633, 3.5637, 3.5641, 3.5644, 3.5648, 3.5649, 3.5647, 3.5655, 3.5659, 3.5661, 3.5665, 3.5673, 3.5669, 3.5675, 3.5677, 3.5679, 3.5686, 3.569, 3.5689, 3.5698, 3.5701, 3.5703, 3.5707, 3.5711, 3.5712, 3.5716, 3.5723, 3.5724, 3.5727, 3.5732, 3.5737, 3.574, 3.5743, 3.5747, 3.5751, 3.5754, 3.576, 3.5763, 3.5767, 3.5771, 3.5773, 3.5781, 3.5784, 3.5789, 3.5792, 3.5793, 3.5802, 3.5805, 3.5808, 3.5813, 3.5817, 3.5821, 3.5827, 3.5832, 3.5837, 3.5841, 3.5847, 3.5852, 3.5857, 3.586, 3.5865, 3.5871, 3.5873, 3.5879, 3.5886, 3.5894, 3.5896, 3.59, 3.5906, 3.591, 3.5916, 3.5922, 3.5925, 3.5929, 3.594, 3.5942, 3.5943, 3.5953, 3.5955, 3.5965, 3.5968, 3.5974, 3.5977, 3.5986, 3.5986, 3.5996, 3.5995, 3.6005, 3.6008, 3.6018, 3.6019, 3.6026, 3.603, 3.6034, 3.604, 3.6048, 3.6051, 3.6057, 3.6062, 3.6066, 3.6074, 3.6078, 3.6084, 3.6089, 3.6095, 3.61, 3.6105, 3.6109, 3.6115, 3.612, 3.6127, 3.6133, 3.6136, 3.6142, 3.6146, 3.6152, 3.6157, 3.6163, 3.6169, 3.6173, 3.6178, 3.6184, 3.6189, 3.6195, 3.6199, 3.6204, 3.621, 3.6214, 3.6218, 3.6224, 3.6228, 3.6233, 3.6238, 3.6243, 3.6248, 3.6255, 3.6259, 3.6263, 3.627, 3.6273, 3.6278, 3.6284, 3.6287, 3.6292, 3.6296, 3.6301, 3.6306, 3.6311, 3.6316, 3.632, 3.6325, 3.633, 3.6334, 3.6339, 3.6343, 3.6348, 3.6352, 3.6358, 3.6361, 3.6364, 3.6369, 3.6374, 3.6379, 3.6382, 3.6389, 3.6392, 3.6393, 3.6401, 3.6405, 3.6407, 3.6413, 3.6416, 3.6421, 3.6425, 3.6432, 3.6434, 3.644, 3.6443, 3.6448, 3.645, 3.6454, 3.646, 3.6464, 3.647, 3.6475, 3.6478, 3.6478, 3.6483, 3.6492, 3.6494, 3.6498, 3.65, 3.6508, 3.6511, 3.6516, 3.652, 3.6524, 3.6527, 3.6532, 3.6537, 3.6542, 3.6548, 3.6551, 3.6554, 3.6559, 3.6564, 3.6568, 3.6573, 3.6578, 3.6581, 3.6586, 3.659, 3.6594, 3.6598, 3.66, 3.6606, 3.6612, 3.6613, 3.6621, 3.6626, 3.6629, 3.6634, 3.6637, 3.6641, 3.6645, 3.665, 3.6653, 3.6659, 3.666, 3.6664, 3.6669, 3.6675, 3.6677, 3.6682, 3.6686, 3.6688, 3.6694, 3.6698, 3.6701, 3.6706, 3.6709, 3.6714, 3.6717, 3.6719, 3.6726, 3.6727, 3.6732, 3.6735, 3.6739, 3.6745, 3.6748, 3.675, 3.6754, 3.6756, 3.676, 3.6765, 3.6769, 3.6771, 3.6775, 3.6777, 3.678, 3.6783, 3.6788, 3.6793, 3.6793, 3.6796, 3.68, 3.6804, 3.6805, 3.6811, 3.6812, 3.6817, 3.6819, 3.6822, 3.6825, 3.6827, 3.683, 3.6834, 3.6837, 3.6838, 3.6842, 3.6846, 3.6849, 3.685, 3.6852, 3.6858, 3.6861, 3.6862, 3.6864, 3.6867, 3.6873, 3.6874, 3.6875, 3.6878, 3.6882, 3.6884, 3.6886, 3.6889, 3.6892, 3.6893, 3.6897, 3.69, 3.6901, 3.6905, 3.6907, 3.6911, 3.6911, 3.6916, 3.6917, 3.6921, 3.6922, 3.6925, 3.6929, 3.6928, 3.6931, 3.6936, 3.6936, 3.6939, 3.6942, 3.6945, 3.6946, 3.6949, 3.6951, 3.6953, 3.6958, 3.6961, 3.6962, 3.6964, 3.6967, 3.6969, 3.6967, 3.6972, 3.6976, 3.6977, 3.6979, 3.6982, 3.6987, 3.6988, 3.6989, 3.6995, 3.6993, 3.6997, 3.7002, 3.7003, 3.7005, 3.7002, 3.7007, 3.7009, 3.7011, 3.7024, 3.7008, 3.7018, 3.7023, 3.7023, 3.7024, 3.7023, 3.7026, 3.7041, 3.7035, 3.7038, 3.7044, 3.7044, 3.7047, 3.705, 3.7062, 3.7053, 3.7056, 3.706, 3.7064, 3.7063, 3.7067, 3.7067, 3.707, 3.7073, 3.7076, 3.7081, 3.7078, 3.7081, 3.7086, 3.7088, 3.7089, 3.7094, 3.7096, 3.7098, 3.7101, 3.7102, 3.7104, 3.7105, 3.711, 3.7112, 3.7114, 3.7115, 3.712, 3.7123, 3.7127, 3.7129, 3.7131, 3.7134, 3.7135, 3.714, 3.7142, 3.7146, 3.7149, 3.7149, 3.7153, 3.7156, 3.7158, 3.7161, 3.7164, 3.7168, 3.7169, 3.7171, 3.7177, 3.7178, 3.718, 3.7183, 3.7188, 3.719, 3.7192, 3.7194, 3.7199, 3.7201, 3.7202, 3.7205, 3.721, 3.7213, 3.7217, 3.7219, 3.7222, 3.7225, 3.7227, 3.7231, 3.7233, 3.7235, 3.7239, 3.7242, 3.7245, 3.7249, 3.7252, 3.7254, 3.7259, 3.7261, 3.7263, 3.7268, 3.727, 3.7272, 3.7278, 3.7281, 3.7283, 3.7285, 3.7291, 3.7294, 3.7297, 3.7299, 3.7302, 3.7307, 3.7308, 3.7314, 3.7316, 3.7319, 3.7324, 3.7329, 3.7329, 3.7332, 3.7336, 3.7341, 3.7343, 3.7347, 3.735, 3.7353, 3.7358, 3.7361, 3.7365, 3.7368, 3.737, 3.7378, 3.738, 3.7382, 3.7387, 3.739, 3.7394, 3.7397, 3.7402, 3.7405, 3.7408, 3.7413, 3.7416, 3.7421, 3.7422, 3.7428, 3.7432, 3.7435, 3.7439, 3.7443, 3.7448, 3.7451, 3.7457, 3.7458, 3.7464, 3.7467, 3.7473, 3.7475, 3.7479, 3.7484, 3.7487, 3.749, 3.7495, 3.7498, 3.7504, 3.7505, 3.7513, 3.7518, 3.7519, 3.7524, 3.7529, 3.7534, 3.7537, 3.7542, 3.7545, 3.7548, 3.7554, 3.7559, 3.7562, 3.7566, 3.7572, 3.7576, 3.7581, 3.7585, 3.759, 3.7593, 3.7598, 3.7602, 3.7606, 3.7612, 3.7616, 3.7622, 3.7626, 3.763, 3.7634, 3.7638, 3.7645, 3.7648, 3.7653, 3.7658, 3.7662, 3.767, 3.7672, 3.7677, 3.7682, 3.7685, 3.769, 3.7695, 3.7701, 3.7706, 3.7711, 3.7716, 3.772, 3.7726, 3.7728, 3.7736, 3.774, 3.7744, 3.775, 3.7755, 3.776, 3.7766, 3.7769, 3.7774, 3.778, 3.7787, 3.779, 3.7797, 3.78, 3.7806, 3.7812, 3.7817, 3.7823, 3.783, 3.7834, 3.7839, 3.7844, 3.7849, 3.7854, 3.7859, 3.7865, 3.7872, 3.7876, 3.7882, 3.7887, 3.7894, 3.7898, 3.7907, 3.7914, 3.7919, 3.7924, 3.793, 3.7936, 3.7943, 3.7947, 3.7953, 3.7958, 3.7964, 3.7969, 3.7975, 3.7982, 3.7989, 3.7994, 3.7999, 3.8006, 3.801, 3.8018, 3.8023, 3.8028, 3.8035, 3.804, 3.8047, 3.8053, 3.8059, 3.8066, 3.807, 3.8078, 3.8083, 3.8089, 3.8096, 3.8102, 3.8109, 3.8114, 3.8121, 3.8127, 3.8133, 3.8141, 3.8148, 3.8153, 3.8159, 3.8164, 3.8172, 3.8178, 3.8186, 3.8193, 3.8197, 3.8205, 3.821, 3.8216, 3.8225, 3.8231, 3.8238, 3.8245, 3.8252, 3.8258, 3.8265, 3.8271, 3.8279, 3.8283, 3.8292, 3.8299, 3.8306, 3.8314, 3.8319, 3.8329, 3.8333, 3.8338, 3.8348, 3.8355, 3.8362, 3.8368, 3.8376, 3.8384, 3.8392, 3.8399, 3.8406, 3.8411, 3.842, 3.8428, 3.8432, 3.8441, 3.8449, 3.8458, 3.8464, 3.8471, 3.848, 3.8488, 3.8492, 3.8502, 3.8507, 3.8514, 3.8524, 3.8531, 3.8539, 3.8548, 3.8555, 3.8561, 3.8569, 3.858, 3.8587, 3.8593, 3.86, 3.8609, 3.8616, 3.8625, 3.8635, 3.8641, 3.8649, 3.8658, 3.8664, 3.8674, 3.8679, 3.869, 3.8698, 3.8705, 3.8712, 3.8721, 3.8725, 3.8735, 3.8743, 3.875, 3.8758, 3.8767, 3.8774, 3.8784, 3.8792, 3.8798, 3.8808, 3.8815, 3.8828, 3.8838, 3.8849, 3.885, 3.8856, 3.8864, 3.887, 3.888, 3.8888, 3.8897, 3.8904, 3.8914, 3.892, 3.8928, 3.8938, 3.8944, 3.8952, 3.8961, 3.8968, 3.8978, 3.8983, 3.8991, 3.8999, 3.9008, 3.9014, 3.9022, 3.9031, 3.9039, 3.9045, 3.9053, 3.906, 3.9069, 3.9076, 3.9084, 3.9092, 3.9099, 3.9108, 3.9114, 3.9121, 3.9129, 3.9136, 3.9145, 3.9153, 3.9159, 3.9167, 3.9175, 3.918, 3.9188, 3.9196, 3.9203, 3.921, 3.9218, 3.9224, 3.9233, 3.924, 3.9247, 3.9254, 3.9261, 3.9268, 3.9276, 3.9282, 3.929, 3.9297, 3.9305, 3.931, 3.9319, 3.9328, 3.9334, 3.934, 3.9349, 3.9355, 3.9361, 3.9369, 3.9376, 3.9383, 3.9391, 3.9398, 3.9405, 3.9413, 3.9418, 3.9425, 3.9435, 3.944, 3.9446, 3.9456, 3.9462, 3.9469, 3.9477, 3.9483, 3.9491, 3.9496, 3.9505, 3.9511, 3.9518, 3.9526, 3.9533, 3.954, 3.9547, 3.9554, 3.9561, 3.9566, 3.9575, 3.9582, 3.9588, 3.9595, 3.9603, 3.9611, 3.9618, 3.9624, 3.9631, 3.964, 3.9647, 3.9652, 3.966, 3.9667, 3.9676, 3.968, 3.9688, 3.9695, 3.9702, 3.9709, 3.9716, 3.9723, 3.973, 3.9737, 3.9745, 3.9751, 3.9759, 3.9767, 3.9773, 3.9779, 3.9788, 3.9794, 3.9801, 3.981, 3.9816, 3.9823, 3.983, 3.9837, 3.9846, 3.9851, 3.9856, 3.9864, 3.9871, 3.9879, 3.9885, 3.9892, 3.9901, 3.9908, 3.9914, 3.9921, 3.9929, 3.9935, 3.9943, 3.9949, 3.9958, 3.9964, 3.9972, 3.9979, 3.9985, 3.9994, 4, 4.0007, 4.0013, 4.0022, 4.0029, 4.0038, 4.0042, 4.005, 4.0058, 4.0065, 4.0072, 4.008, 4.0086, 4.0093, 4.0099, 4.0108, 4.0116, 4.0122, 4.0129, 4.0135, 4.0144, 4.0151, 4.0158, 4.0164, 4.0172, 4.0178, 4.0187, 4.0194, 4.0201, 4.021, 4.0215, 4.0223, 4.0231, 4.0239, 4.0244, 4.0252, 4.026, 4.0268, 4.0273, 4.028, 4.0288, 4.0295, 4.0304, 4.0311, 4.0318, 4.0325, 4.0333, 4.034, 4.0347, 4.0356, 4.0362, 4.0371, 4.0377, 4.0385, 4.0391, 4.0399, 4.0405, 4.0414, 4.042, 4.0428, 4.0436, 4.0443, 4.0449, 4.0458, 4.0465, 4.0473, 4.048, 4.0488, 4.0495, 4.0503, 4.051, 4.0517, 4.0525, 4.0531, 4.054, 4.0547, 4.0554, 4.0563, 4.057, 4.0578, 4.0585, 4.0592, 4.0601, 4.0607, 4.0616, 4.0621, 4.0631, 4.0637, 4.0645, 4.0652, 4.0659, 4.0667, 4.0675, 4.0683, 4.069, 4.0697, 4.0706, 4.0714, 4.0721, 4.0728, 4.0735, 4.0744, 4.0751, 4.0758, 4.0766, 4.0773, 4.0781, 4.079, 4.0797, 4.0805, 4.0812, 4.0821, 4.0827, 4.0834, 4.0843, 4.0852, 4.0858, 4.0867, 4.0872, 4.0882, 4.089, 4.0897, 4.0905, 4.0912, 4.0919, 4.0928, 4.0936, 4.0944, 4.0951, 4.0959, 4.0965, 4.0974, 4.0982, 4.099, 4.1, 4.1006, 4.1015, 4.1021, 4.1029, 4.1037, 4.1044, 4.1053, 4.1062, 4.1069, 4.1076, 4.1084, 4.1091, 4.1099, 4.1108, 4.1114, 4.1123, 4.113, 4.1139, 4.1146, 4.1154, 4.1162, 4.117, 4.1179, 4.1186, 4.1194, 4.1202, 4.1209, 4.1218, 4.1225, 4.1235, 4.1243, 4.1252, 4.1257, 4.1266, 4.1273, 4.1282, 4.129, 4.1297, 4.1306, 4.1312, 4.1321, 4.1329, 4.1337, 4.1346, 4.1354, 4.1361, 4.1369, 4.1379, 4.1386, 4.1392, 4.1402, 4.1409, 4.1417, 4.1427, 4.1433, 4.1442, 4.1451, 4.1458, 4.1467, 4.1474, 4.1483, 4.149, 4.1498, 4.1507, 4.1514, 4.1522, 4.1531, 4.1538, 4.1548, 4.1555, 4.1564, 4.1571, 4.1579, 4.1586, 4.1596, 4.1604, 4.1611, 4.1621, 4.1628, 4.1637, 4.1645, 4.1653, 4.1661, 4.1669, 4.1677, 4.1684, 4.1693, 4.1701, 4.1711, 4.1719, 4.1726, 4.1735, 4.1743, 4.1752, 4.1759, 4.1768, 4.1775, 4.1785, 4.1793, 4.1802, 4.181, 4.1817, 4.1825, 4.1834, 4.1843, 4.1849, 4.186, 4.1867, 4.1876, 4.1884, 4.1894, 4.1901, 4.1909, 4.1917, 4.1925, 4.1935, 4.1944, 4.1951, 4.1959, 4.1967, 4.1975, 4.1985, 4.1992, 4.2, 4.1992, 4.1992, 4.1991, 4.1991, 4.1993, 4.1991, 4.1993, 4.1993, 4.1991, 4.1993, 4.1991, 4.1992, 4.199, 4.1991, 4.1994, 4.1996, 4.2001, 4.2005, 4.2009, 4.2002, 4.2, 4.2001, 4.2002, 4.2001, 4.2001, 4.2, 4.2001, 4.2002, 4.2001, 4.2003, 4.2, 4.2001, 4.2002, 4.2002, 4.2002, 4.2002, 4.2001, 4.2, 4.2001, 4.2, 4.2001, 4.2, 4.2002, 4.2003, 4.2003, 4.2002, 4.2, 4.2001, 4.2001, 4.2002, 4.2002, 4.2001, 4.2003, 4.2001, 4.2002, 4.2001, 4.2001, 4.2004, 4.2003, 4.2003, 4.2006, 4.2006, 4.2008, 4.2004, 4.2005, 4.2004, 4.2004, 4.2003, 4.2004, 4.2005, 4.2004, 4.2001, 4.2004, 4.2004, 4.2005, 4.2004, 4.2004, 4.2004, 4.2004, 4.2004, 4.2003, 4.2003, 4.2004, 4.2001, 4.2003, 4.2003, 4.2003, 4.2002, 4.2003, 4.2002, 4.2001, 4.2004, 4.2002, 4.2003, 4.2004, 4.2003, 4.2006, 4.2003, 4.2002, 4.2002, 4.2002, 4.2003, 4.2003, 4.2003, 4.2002, 4.2003, 4.2002, 4.2003, 4.2002, 4.2003, 4.2003, 4.2005, 4.2006, 4.2005, 4.2006, 4.2007, 4.2005, 4.2005, 4.2005, 4.2007, 4.2006, 4.2005, 4.2007, 4.2007, 4.2007, 4.2007, 4.2006, 4.2004, 4.2005, 4.2006, 4.2006, 4.2006, 4.2005, 4.2006]

                        Percentage = [0.00,0.07,0.13,0.20,0.26,0.33,0.39,0.46,0.52,0.59,0.65,0.72,0.79,0.85,0.92,0.98,1.05,1.11,1.18,1.24,1.31,1.37,1.44,1.51,1.57,1.64,1.70,1.77,1.83,1.90,1.96,2.03,2.09,2.16,2.23,2.29,2.36,2.42,2.49,2.55,2.62,2.68,2.75,2.81,2.88,2.95,3.01,3.08,3.14,3.21,3.27,3.34,3.40,3.47,3.54,3.60,3.67,3.73,3.80,3.86,3.93,3.99,4.06,4.12,4.19,4.26,4.32,4.39,4.45,4.52,4.58,4.65,4.71,4.78,4.84,4.91,4.98,5.04,5.11,5.17,5.24,5.30,5.37,5.43,5.50,5.56,5.63,5.70,5.76,5.83,5.89,5.96,6.02,6.09,6.15,6.22,6.28,6.35,6.42,6.48,6.55,6.61,6.68,6.74,6.81,6.87,6.94,7.00,7.07,7.14,7.20,7.27,7.33,7.40,7.46,7.53,7.59,7.66,7.73,7.79,7.86,7.92,7.99,8.05,8.12,8.18,8.25,8.32,8.38,8.45,8.51,8.58,8.64,8.71,8.77,8.84,8.91,8.97,9.04,9.10,9.17,9.23,9.30,9.36,9.43,9.50,9.56,9.63,9.69,9.76,9.82,9.89,9.95,10.02,10.09,10.15,10.22,10.28,10.35,10.41,10.48,10.55,10.61,10.68,10.74,10.81,10.87,10.94,11.00,11.07,11.14,11.20,11.27,11.33,11.40,11.46,11.53,11.59,11.66,11.73,11.79,11.86,11.92,11.99,12.05,12.12,12.18,12.25,12.32,12.38,12.45,12.51,12.58,12.64,12.71,12.77,12.84,12.91,12.97,13.04,13.10,13.17,13.23,13.30,13.36,13.43,13.50,13.56,13.63,13.69,13.76,13.82,13.89,13.95,14.02,14.09,14.15,14.22,14.28,14.35,14.41,14.48,14.54,14.61,14.68,14.74,14.81,14.87,14.94,15.00,15.07,15.13,15.20,15.27,15.33,15.40,15.46,15.53,15.59,15.66,15.73,15.79,15.86,15.92,15.99,16.05,16.12,16.18,16.25,16.32,16.38,16.45,16.51,16.58,16.64,16.71,16.77,16.84,16.91,16.97,17.04,17.10,17.17,17.23,17.30,17.36,17.43,17.50,17.56,17.63,17.69,17.76,17.82,17.89,17.95,18.02,18.09,18.15,18.22,18.28,18.35,18.41,18.48,18.54,18.61,18.68,18.74,18.81,18.87,18.94,19.00,19.07,19.13,19.20,19.27,19.33,19.40,19.46,19.53,19.59,19.66,19.72,19.79,19.86,19.92,19.99,20.05,20.12,20.18,20.25,20.31,20.38,20.45,20.51,20.58,20.64,20.71,20.77,20.84,20.90,20.97,21.04,21.10,21.17,21.23,21.30,21.36,21.43,21.50,21.56,21.63,21.69,21.76,21.82,21.89,21.95,22.02,22.09,22.15,22.22,22.28,22.35,22.41,22.48,22.54,22.61,22.68,22.74,22.81,22.87,22.94,23.00,23.07,23.13,23.20,23.27,23.33,23.40,23.46,23.53,23.59,23.66,23.73,23.79,23.86,23.92,23.99,24.05,24.12,24.18,24.25,24.32,24.38,24.45,24.51,24.58,24.64,24.71,24.77,24.84,24.91,24.97,25.04,25.10,25.17,25.23,25.30,25.37,25.43,25.50,25.56,25.63,25.69,25.76,25.82,25.89,25.96,26.02,26.09,26.15,26.22,26.28,26.35,26.42,26.48,26.55,26.61,26.68,26.74,26.81,26.87,26.94,27.01,27.07,27.14,27.20,27.27,27.33,27.40,27.47,27.53,27.60,27.66,27.73,27.79,27.86,27.92,27.99,28.06,28.12,28.19,28.25,28.32,28.38,28.45,28.52,28.58,28.65,28.71,28.78,28.84,28.91,28.97,29.04,29.11,29.17,29.24,29.30,29.37,29.43,29.50,29.57,29.63,29.70,29.76,29.83,29.89,29.96,30.02,30.09,30.16,30.22,30.29,30.35,30.42,30.48,30.55,30.62,30.68,30.75,30.81,30.88,30.94,31.01,31.07,31.14,31.21,31.27,31.34,31.40,31.47,31.53,31.60,31.66,31.73,31.80,31.86,31.93,31.99,32.06,32.12,32.19,32.26,32.32,32.39,32.45,32.52,32.58,32.65,32.71,32.78,32.85,32.91,32.98,33.04,33.11,33.17,33.24,33.31,33.37,33.44,33.50,33.57,33.63,33.70,33.76,33.83,33.90,33.96,34.03,34.09,34.16,34.22,34.29,34.35,34.42,34.49,34.55,34.62,34.68,34.75,34.81,34.88,34.95,35.01,35.08,35.14,35.21,35.27,35.34,35.40,35.47,35.54,35.60,35.67,35.73,35.80,35.86,35.93,35.99,36.06,36.13,36.19,36.26,36.32,36.39,36.45,36.52,36.59,36.65,36.72,36.78,36.85,36.91,36.98,37.04,37.11,37.18,37.24,37.31,37.37,37.44,37.50,37.57,37.63,37.70,37.77,37.83,37.90,37.96,38.03,38.09,38.16,38.22,38.29,38.36,38.42,38.49,38.55,38.62,38.68,38.75,38.81,38.88,38.95,39.01,39.08,39.14,39.21,39.27,39.34,39.41,39.47,39.54,39.60,39.67,39.73,39.80,39.86,39.93,40.00,40.06,40.13,40.19,40.26,40.32,40.39,40.45,40.52,40.59,40.65,40.72,40.78,40.85,40.91,40.98,41.05,41.11,41.18,41.24,41.31,41.37,41.44,41.50,41.57,41.64,41.70,41.77,41.83,41.90,41.96,42.03,42.10,42.16,42.23,42.29,42.36,42.42,42.49,42.55,42.62,42.69,42.75,42.82,42.88,42.95,43.01,43.08,43.14,43.21,43.28,43.34,43.41,43.47,43.54,43.60,43.67,43.74,43.80,43.87,43.93,44.00,44.06,44.13,44.19,44.26,44.33,44.39,44.46,44.52,44.59,44.65,44.72,44.78,44.85,44.92,44.98,45.05,45.11,45.18,45.24,45.31,45.37,45.44,45.51,45.57,45.64,45.70,45.77,45.83,45.90,45.96,46.03,46.10,46.16,46.23,46.29,46.36,46.42,46.49,46.55,46.62,46.69,46.75,46.82,46.88,46.95,47.01,47.08,47.15,47.21,47.28,47.34,47.41,47.47,47.54,47.60,47.67,47.74,47.80,47.87,47.93,48.00,48.06,48.13,48.19,48.26,48.33,48.39,48.46,48.52,48.59,48.65,48.72,48.78,48.85,48.92,48.98,49.05,49.11,49.18,49.24,49.31,49.37,49.44,49.51,49.57,49.64,49.70,49.77,49.83,49.90,49.96,50.03,50.10,50.16,50.23,50.29,50.36,50.42,50.49,50.55,50.62,50.69,50.75,50.82,50.88,50.95,51.01,51.08,51.14,51.21,51.28,51.34,51.41,51.47,51.54,51.60,51.67,51.73,51.80,51.87,51.93,52.00,52.06,52.13,52.19,52.26,52.32,52.39,52.46,52.52,52.59,52.65,52.72,52.78,52.85,52.91,52.98,53.05,53.11,53.18,53.24,53.31,53.37,53.44,53.51,53.57,53.64,53.70,53.77,53.83,53.90,53.96,54.03,54.10,54.16,54.23,54.29,54.36,54.42,54.49,54.55,54.62,54.68,54.75,54.82,54.88,54.95,55.01,55.08,55.14,55.21,55.28,55.34,55.41,55.47,55.54,55.60,55.67,55.73,55.80,55.86,55.93,56.00,56.06,56.13,56.19,56.26,56.32,56.39,56.45,56.52,56.59,56.65,56.72,56.78,56.85,56.91,56.98,57.04,57.11,57.18,57.24,57.31,57.37,57.44,57.50,57.57,57.63,57.70,57.77,57.83,57.90,57.96,58.03,58.09,58.16,58.22,58.29,58.36,58.42,58.49,58.55,58.62,58.68,58.75,58.81,58.88,58.95,59.01,59.08,59.14,59.21,59.27,59.34,59.40,59.47,59.54,59.60,59.67,59.73,59.80,59.86,59.93,59.99,60.06,60.12,60.19,60.26,60.32,60.39,60.45,60.52,60.58,60.65,60.71,60.78,60.85,60.91,60.98,61.04,61.11,61.17,61.24,61.30,61.37,61.43,61.50,61.57,61.63,61.70,61.76,61.83,61.89,61.96,62.02,62.09,62.16,62.22,62.29,62.35,62.42,62.48,62.55,62.61,62.68,62.74,62.81,62.88,62.94,63.01,63.07,63.14,63.20,63.27,63.33,63.40,63.47,63.53,63.60,63.66,63.73,63.79,63.86,63.92,63.99,64.05,64.12,64.19,64.25,64.32,64.38,64.45,64.51,64.58,64.64,64.71,64.77,64.84,64.91,64.97,65.04,65.10,65.17,65.23,65.30,65.36,65.43,65.50,65.56,65.63,65.69,65.76,65.82,65.89,65.95,66.02,66.08,66.15,66.22,66.28,66.35,66.41,66.48,66.54,66.61,66.67,66.74,66.81,66.87,66.94,67.00,67.07,67.13,67.20,67.26,67.33,67.39,67.46,67.53,67.59,67.66,67.72,67.79,67.85,67.92,67.98,68.05,68.12,68.18,68.25,68.31,68.38,68.44,68.51,68.57,68.64,68.70,68.77,68.84,68.90,68.97,69.03,69.10,69.16,69.23,69.29,69.36,69.43,69.49,69.56,69.62,69.69,69.75,69.82,69.88,69.95,70.02,70.08,70.15,70.21,70.28,70.34,70.41,70.47,70.54,70.61,70.67,70.74,70.80,70.87,70.93,71.00,71.06,71.13,71.20,71.26,71.33,71.39,71.46,71.52,71.59,71.65,71.72,71.79,71.85,71.92,71.98,72.05,72.11,72.18,72.24,72.31,72.38,72.44,72.51,72.57,72.64,72.70,72.77,72.83,72.90,72.97,73.03,73.10,73.16,73.23,73.29,73.36,73.42,73.49,73.56,73.62,73.69,73.75,73.82,73.88,73.95,74.01,74.08,74.15,74.21,74.28,74.34,74.41,74.47,74.54,74.60,74.67,74.74,74.80,74.87,74.93,75.00,75.06,75.13,75.20,75.26,75.33,75.39,75.46,75.52,75.59,75.65,75.72,75.79,75.85,75.92,75.98,76.05,76.11,76.18,76.24,76.31,76.38,76.44,76.51,76.57,76.64,76.70,76.77,76.83,76.90,76.97,77.03,77.10,77.16,77.23,77.29,77.36,77.42,77.49,77.56,77.62,77.69,77.75,77.82,77.88,77.95,78.01,78.08,78.15,78.21,78.28,78.34,78.41,78.47,78.54,78.60,78.67,78.74,78.80,78.87,78.93,79.00,79.06,79.13,79.19,79.26,79.33,79.39,79.46,79.52,79.59,79.65,79.72,79.78,79.85,79.92,79.98,80.05,80.11,80.18,80.24,80.31,80.37,80.44,80.51,80.57,80.64,80.70,80.77,80.83,80.90,80.96,81.03,81.10,81.16,81.23,81.29,81.36,81.42,81.49,81.55,81.62,81.69,81.75,81.82,81.88,81.95,82.01,82.08,82.14,82.21,82.27,82.34,82.41,82.47,82.54,82.60,82.67,82.73,82.80,82.86,82.93,83.00,83.06,83.13,83.19,83.26,83.32,83.39,83.45,83.52,83.59,83.65,83.72,83.78,83.85,83.91,83.98,84.04,84.11,84.18,84.24,84.31,84.37,84.44,84.50,84.57,84.63,84.70,84.77,84.83,84.90,84.96,85.03,85.09,85.16,85.22,85.29,85.36,85.42,85.49,85.55,85.62,85.68,85.75,85.81,85.88,85.95,86.01,86.08,86.14,86.21,86.27,86.34,86.40,86.47,86.53,86.60,86.67,86.73,86.80,86.86,86.93,86.99,87.06,87.12,87.19,87.26,87.32,87.39,87.45,87.52,87.58,87.65,87.71,87.78,87.85,87.91,87.98,88.04,88.11,88.17,88.24,88.30,88.37,88.44,88.50,88.57,88.63,88.70,88.76,88.83,88.89,88.96,89.03,89.09,89.16,89.22,89.29,89.35,89.42,89.48,89.55,89.61,89.68,89.75,89.81,89.88,89.94,90.01,90.07,90.14,90.20,90.27,90.34,90.40,90.47,90.53,90.60,90.66,90.73,90.79,90.86,90.93,90.99,91.06,91.12,91.19,91.25,91.32,91.38,91.45,91.52,91.58,91.65,91.71,91.78,91.84,91.91,91.97,92.04,92.11,92.17,92.24,92.30,92.37,92.43,92.50,92.56,92.63,92.69,92.76,92.83,92.89,92.96,93.02,93.09,93.15,93.22,93.28,93.35,93.42,93.48,93.55,93.61,93.68,93.74,93.81,93.87,93.94,94.01,94.07,94.14,94.20,94.27,94.33,94.40,94.46,94.53,94.60,94.66,94.73,94.79,94.86,94.92,94.99,95.05,95.12,95.18,95.25,95.32,95.38,95.45,95.51,95.58,95.64,95.71,95.77,95.84,95.91,95.97,96.04,96.10,96.17,96.23,96.30,96.36,96.43,96.50,96.56,96.63,96.69,96.76,96.82,96.89,96.95,97.02,97.09,97.15,97.21,97.26,97.32,97.37,97.42,97.48,97.53,97.58,97.62,97.67,97.72,97.76,97.80,97.85,97.89,97.93,97.98,98.02,98.06,98.10,98.14,98.18,98.21,98.25,98.28,98.32,98.35,98.38,98.42,98.45,98.48,98.51,98.54,98.57,98.60,98.63,98.66,98.68,98.71,98.74,98.76,98.79,98.82,98.84,98.86,98.89,98.91,98.94,98.96,98.98,99.00,99.02,99.05,99.07,99.09,99.11,99.13,99.15,99.17,99.19,99.21,99.23,99.25,99.27,99.28,99.30,99.32,99.34,99.35,99.37,99.39,99.40,99.42,99.43,99.45,99.46,99.48,99.49,99.51,99.52,99.54,99.55,99.56,99.58,99.59,99.60,99.62,99.63,99.64,99.66,99.67,99.68,99.69,99.70,99.72,99.73,99.74,99.75,99.76,99.77,99.78,99.79,99.80,99.81,99.82,99.83,99.84,99.85,99.86,99.87,99.88,99.89,99.90,99.91,99.92,99.93,99.94,99.94,99.95,99.96,99.97,99.98,99.99,99.99,100.00,100.01,100.02,100.02,100.03,100.04,100.05,100.05,100.05,]

                        mapping = list(zip(Volt,Percentage))
                        # print(len(Volt))
                        # print(len(Percentage))
                        # print(len(mapping))

                        # Find the closest voltage in the list of Volt
                        closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                        # Find the index of the closest voltage
                        index = Volt.index(closest_voltage)
                        # print(index)
                        # print(Percentage[index])

                        # Check if the index is not the last index to avoid index out of range
                        if index < len(Volt) - 1:
                            # Get the next voltage value
                            if(input_voltage > closest_voltage) or index == 0 and input_voltage != 0:
                                next_voltage = Volt[index + 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                            elif input_voltage < closest_voltage and input_voltage != 0:
                                next_voltage = Volt[index - 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index ] - Percentage[index-1]) / ( closest_voltage - next_voltage)
                                closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                            elif input_voltage == 0:
                                closest_percentage2 = 0
                            else:
                                next_voltage = Volt[index + 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")            
                        elif index + 1 == len(Volt):

                            if input_voltage < closest_voltage and input_voltage != 0:
                            #    print("This is herer 3")
                                next_voltage = Volt[index - 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                                closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                            elif input_voltage == 0:
                                closest_percentage2 = 0
                            elif input_voltage > Volt[index]:
                                # print(Volt[index])
                                closest_percentage2 = 0
                            else:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")          
                                            
                        else:
                            print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                    # Find indices of 'Group1' in the 'Group' list


                    # Sort the array
                    sorted_array = sorted(Charge_Set['Group'])

                    # Get unique numbers using set
                    unique_numbers = set(Charge_Set['Group'])

                    # Convert the unique numbers back to a sorted list if needed
                    sorted_unique_numbers = sorted(list(unique_numbers))
                    first_min_v = None
                    last_max_v = None
                    # Print the sorted unique numbers
                    # print(sorted_unique_numbers)
                    shitty = 1
                    for numbers in sorted_unique_numbers:
                        group_indices = [i for i, group in enumerate(Charge_Set['Group']) if group == numbers]

                        # Check if there are 'Group1' elements in the 'Group' list
                        if group_indices:
                            # Access the first and last 'Max_V' values for 'Group1'
                            first_min_v = Charge_Set['Min_V'][group_indices[0]]
                            last_max_v = Charge_Set['Max_V'][group_indices[-1]]

                            # print(f"First Max_V for 'Group1': {first_min_v}")
                            # print(f"Last Max_V for 'Group1': {last_max_v}")
                        else:
                            first_min_v = 0
                            last_max_v = 0
                        # print(f"First Max_V for 'Group1': {first_min_v}")
                        # print(f"Last Max_V for 'Group1': {last_max_v}")
                        SOH_Goop['Goop'].append(numbers)
                        SOH_Goop['Min_V'].append(first_min_v)
                        SOH_Goop['Max_V'].append(last_max_v)
                        thread1 = threading.Thread(target=SOH, args=(first_min_v,))
                        # print(f'This is last_m:{last_max_v}')
                        thread2 = threading.Thread(target=SOH2, args=(last_max_v,))
                        thread1.start()
                        thread2.start()
                        thread1.join()
                        thread2.join()
                        # print(f'This is SOH2:{closest_percentage2}')
                        # Calculate energy for 'Group1' where 'Group' is 0
                # Find indices of 'Group1' where 'Group' is 0
                        
                        energy_sum = 0
                        save_data_start = None
                        a= None
                        # Iterate over the numerical indices of Charge_Set['Group']
                        for index, group in enumerate(Charge_Set['Group']):
                            if group == numbers:
                                voltage = Charge_Set['Voltage'][index]
                                current = Charge_Set['Current'][index]
                                time_diff = Charge_Set['Time_Diff'][index]
                                # Split the original string by space to get the time portion
                                teiam = str( Charge_Set['Timestamps'][index])
                                # print(teiam)
                                split_string = teiam.split(" ")

                                # Check if there are at least two parts (date and time)
                                if len(split_string) >= 2:
                                    # Join the time portion and discard the date
                                    time_portion = " ".join(split_string[1:])
                                    # print(time_portion)

                                if save_data_start == None:
                                
                                    save_data_start = time_portion
                                    SOH_Goop['S_Time'].append(time_portion)
                                    
                                save_end = time_portion
                                # print(f'This is time diff { leg}')
                                # print(index)  # Assuming 'Diff' represents time intervals
                                # print(f'This is V {voltage}')
                                # print(f'This is C {current}')
                                # print(f'This is D {time_diff}')
                                if index != 0:
                                    energy =  abs(0.5*(current+previous_current))  * time_diff
                                    previous_current = current
                                    # print(f'This is Energy {energy}')
                                    # cell = sheetPackProcess.cell(row=gginp, column=28, value= energy)
                                    # gginp += 1
                                else :
                                    energy = 0
                                    previous_current = current
                                energy_sum += energy
                        SOH_Goop['E_Time'].append(save_end)
                        energy_sum = energy_sum* voltage
                        # print(f'Total energy for "Group1": {energy_sum} Joules')
                            

                        # print(f"Total energy for 'Group1' where 'Group' is 0: {energy_sum} Joules")
                        
                        SOC_Start = closest_percentage
                        SOC_End = closest_percentage2
                        Difference_SOC = SOC_End - SOC_Start
                        DesignCapacity_NH02 = 26.8
                        SOh_E = energy_sum/(3600*1000)
                        SOH_Goop['Start_SOC'].append(SOC_Start)
                        SOH_Goop['End_SOC'].append(SOC_End)
                        SOH_Goop['Charge'].append(SOh_E)
                        if Difference_SOC != 0:
                            Cal_Capacity = SOh_E/(Difference_SOC/100)
                        else:
                            Cal_Capacity = 0
                        if(Cal_Capacity != 0):
                            Remaining_Capacity = (Cal_Capacity/DesignCapacity_NH02)*100
                        elif(Cal_Capacity == None):
                            Remaining_Capacity = 0
                        else:
                            Remaining_Capacity = 0
                        SOH_Goop['Cal'].append(Remaining_Capacity)
                        SOH_Goop['SOH'].append(Cal_Capacity)
                        SOH_Goop['Cal_E'].append(SOh_E)
                        
                        # for index, header in enumerate(headers1, start=shitty):
                        #     cell = sheet8.cell(row=4, column=index, value=header)

                        # cell = sheet8.cell(row=the_loop, column=shitty, value=first_min_v)
                        # cell = sheet8.cell(row=the_loop, column=shitty+1, value=last_max_v)
                        # cell = sheet8.cell(row=the_loop, column=shitty+2, value=SOC_Start)
                        # cell = sheet8.cell(row=the_loop, column=shitty+3, value=SOC_End)
                        # cell = sheet8.cell(row=the_loop, column=shitty+4, value=Cal_Capacity)
                        # cell = sheet8.cell(row=the_loop, column=shitty+5, value=Remaining_Capacity)    
                        # cell = sheet8.cell(row=the_loop, column=shitty+6, value=SOh_E)  
                        shitty += 7
                    if first_min_v == None and last_max_v == None:
                        lost_data.append(app+1)
                        lost_data3.append(app+1)
                    # print(lost_data)
                def SOH_Min(input_voltage):
                        global closest_percentage
                        closest_percentage = 0
                        Volt = [2.500,2.951	,3.202,3.364,3.477,3.515,3.549,3.584,3.614,3.646,3.674,3.702,3.736,3.779,3.832,3.897,3.953,4.007,4.064,4.125,4.200]
                        Percentage = [0,0.05,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]
                        
                    
                        mapping = list(zip(Volt,Percentage))
                        # print(len(Volt))
                        # print(len(Percentage))
                        # print(len(mapping))

                        # Find the closest voltage in the list of Volt
                        closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                        # Find the index of the closest voltage
                        index = Volt.index(closest_voltage)
                        # print(index)
                        # print(Percentage[index])

                        # Check if the index is not the last index to avoid index out of range
                        if index < len(Volt) - 1:
                            if(input_voltage > closest_voltage) or index == 0  and input_voltage != 0:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")
                                # Calculate the slope using the closest_voltage and next_voltage
                                if next_voltage != closest_voltage :
                                    slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                    closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                else:
                                    next_voltage = Volt[index + 2]
                                    slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                    closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                            elif (closest_voltage - input_voltage) < 0.1 :    
                                next_voltage = Volt[index + 1]
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage < closest_voltage and input_voltage != 0:
                                next_voltage = Volt[index - 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage == 0:
                                closest_percentage = 0
                            else:
                                next_voltage = Volt[index + 1]
                                
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                        elif index + 1 == len(Volt):

                            if input_voltage < closest_voltage and input_voltage != 0:
                                # print("This is herer 3")
                                next_voltage = Volt[index - 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                                closest_percentage = Percentage[index-1] + slope*(input_voltage - next_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                            elif input_voltage == 0:
                                closest_percentage = 0
                            elif input_voltage > Volt[index]:
                                # print(Volt[index])
                                closest_percentage = 0
                            else:
                                next_voltage = Volt[index + 1]
                                # print(f"This is next{next_voltage}")
                                # print(f"This is next{closest_voltage}")        
                                # Calculate the slope using the closest_voltage and next_voltage
                                slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                                closest_percentage = Percentage[index] + slope*(input_voltage - closest_voltage)
                                
                                # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage}")
                                    
                        else:
                            print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                        return closest_percentage
            
                def SOH2_Min(input_voltage):
                    global closest_percentage2
                    closest_percentage2 = 0
                    # print("This is SOH2")
                    Volt = [2.500,2.951	,3.202,3.364,3.477,3.515,3.549,3.584,3.614,3.646,3.674,3.702,3.736,3.779,3.832,3.897,3.953,4.007,4.064,4.125,4.200]
                
                    Percentage = [0,0.05,0.1,0.15,0.2,0.25,0.3,0.35,0.4,0.45,0.5,0.55,0.6,0.65,0.7,0.75,0.8,0.85,0.9,0.95,1]

                    mapping = list(zip(Volt,Percentage))
                    # print(len(Volt))
                    # print(len(Percentage))
                    # print(len(mapping))

                    # Find the closest voltage in the list of Volt
                    closest_voltage = min(Volt, key=lambda x: abs(x - input_voltage))
                    # Find the index of the closest voltage
                    index = Volt.index(closest_voltage)
                    # print(index)
                    # print(Percentage[index])

                    # Check if the index is not the last index to avoid index out of range
                    if index < len(Volt) - 1:
                        # Get the next voltage value
                        if(input_voltage > closest_voltage) or index == 0 and input_voltage != 0:
                            next_voltage = Volt[index + 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                        elif (closest_voltage - input_voltage) < 0.1 :    
                            next_voltage = Volt[index + 1]
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                                                    
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        elif input_voltage < closest_voltage and input_voltage != 0:
                            next_voltage = Volt[index - 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / ( closest_voltage - next_voltage)
                            closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        elif input_voltage == 0:
                            closest_percentage2 = 0
                        else:
                            next_voltage = Volt[index + 1]
                            
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")            
                    elif index + 1 == len(Volt):

                        if input_voltage < closest_voltage and input_voltage != 0:
                            # print("This is herer 3")
                            next_voltage = Volt[index - 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index] - Percentage[index - 1]) / ( closest_voltage - next_voltage)
                            closest_percentage2 = Percentage[index-1] + slope*(input_voltage - next_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")
                        elif input_voltage == 0:
                            closest_percentage2 = 0
                        elif input_voltage > Volt[index]:
                            # print(Volt[index])
                            closest_percentage2 = 0
                        else:
                            next_voltage = Volt[index + 1]
                            # print(f"This is next{next_voltage}")
                            # print(f"This is next{closest_voltage}")        
                            # Calculate the slope using the closest_voltage and next_voltage
                            slope = (Percentage[index + 1] - Percentage[index]) / (next_voltage - closest_voltage)
                            closest_percentage2 = Percentage[index] + slope*(input_voltage - closest_voltage)
                            
                            # print(f"The slope between {closest_voltage}V and {next_voltage}V is {closest_percentage2}")          
                     
                    else:
                        print("The closest voltage is the last one in the dataset, so there's no next voltage for slope calculation.")
                # Find indices of 'Group1' in the 'Group' list


                # Sort the array
                sorted_array = sorted(Discharge_Set['Group'])

                # Get unique numbers using set
                unique_numbers = set(Discharge_Set['Group'])

                # Convert the unique numbers back to a sorted list if needed
                sorted_unique_numbers = sorted(list(unique_numbers))
                first_min_v = None
                last_max_v = None
                # Print the sorted unique numbers
                # print(sorted_unique_numbers)
                shitty = 1
                for numbers in sorted_unique_numbers:
                    group_indices = [i for i, group in enumerate(Discharge_Set['Group']) if group == numbers]

                    # Check if there are 'Group1' elements in the 'Group' list
                    if group_indices:
                        # Access the first and last 'Max_V' values for 'Group1'
                        first_min_v = Discharge_Set['Max_V'][group_indices[0]]
                        last_max_v = Discharge_Set['Min_V'][group_indices[-1]]

                        # print(f"First Max_V for 'Group1': {first_min_v}")
                        # print(f"Last Max_V for 'Group1': {last_max_v}")
                    else:
                        first_min_v = 0
                        last_max_v = 0
                    # print(f"First Max_V for 'Group1': {first_min_v}")
                    # print(f"Last Max_V for 'Group1': {last_max_v}")
                    SOH_Goop_Min['Goop'].append(numbers)
                    SOH_Goop_Min['Min_V'].append(first_min_v)
                    SOH_Goop_Min['Max_V'].append(last_max_v)
                    thread1 = threading.Thread(target=SOH_Min, args=(first_min_v,))
                    # print(f'This is last_m:{last_max_v}')
                    thread2 = threading.Thread(target=SOH2_Min, args=(last_max_v,))
                    thread1.start()
                    thread2.start()
                    thread1.join()
                    thread2.join()
                    # print(f'This is SOH2:{closest_percentage2}')
                    # Calculate energy for 'Group1' where 'Group' is 0
            # Find indices of 'Group1' where 'Group' is 0
                    
                    energy_sum = 0
                    save_data_start = None
                    a= None
                    # Iterate over the numerical indices of Discharge_Set['Group']
                    for index, group in enumerate(Discharge_Set['Group']):
                        if group == numbers:
                            voltage = Discharge_Set['Voltage'][index]
                            current = Discharge_Set['Current'][index]
                            time_diff = Discharge_Set['Time_Diff'][index]
                            # Split the original string by space to get the time portion
                            teiam = str( Discharge_Set['Timestamps'][index])
                            # print(teiam)
                            split_string = teiam.split(" ")

                            # Check if there are at least two parts (date and time)
                            if len(split_string) >= 2:
                                # Join the time portion and discard the date
                                time_portion = " ".join(split_string[1:])
                                # print(time_portion)

                            if save_data_start == None:
                            
                                save_data_start = time_portion
                                SOH_Goop_Min['S_Time'].append(time_portion)
                                
                            save_end = time_portion

                            # print(f'This is time diff { leg}')
                            # print(index)  # Assuming 'Diff' represents time intervals
                            # print(f'This is V {voltage}')
                            # print(f'This is C {current}')
                            # print(f'This is D {time_diff}')
                            if index != 0:
                                energy =  abs(0.5*(current+previous_current))  * time_diff
                                previous_current = current
                                # print(f'This is Energy {energy}')
                                # cell = sheetPackProcess.cell(row=gginp, column=27, value= energy)
                                # gginp += 1

                            else :
                                energy = 0
                                previous_current = current
                            energy_sum += energy
                    SOH_Goop_Min['E_Time'].append(save_end)
                    energy_sum = energy_sum* voltage

                    # print(f'Total energy for "Group1": {energy_sum} Joules')
                        

                    # print(f"Total energy for 'Group1' where 'Group' is 0: {energy_sum} Joules")
                    
                    SOC_Start = closest_percentage
                    SOC_End = closest_percentage2
                    Difference_SOC = SOC_Start - SOC_End 
                    DesignCapacity_NH02 = 26.8
                    SOh_E = energy_sum/(3600*1000)
                    SOH_Goop_Min['Start_SOC'].append(SOC_Start*100)
                    SOH_Goop_Min['End_SOC'].append(SOC_End*100)
                    SOH_Goop_Min['Discharge'].append(SOh_E)
                    if Difference_SOC != 0:
                        Cal_Capacity = abs(SOh_E)/(Difference_SOC)
                    else:
                        Cal_Capacity = 0
                    if(Cal_Capacity != 0):
                        Remaining_Capacity = (Cal_Capacity/DesignCapacity_NH02)*100
                    elif(Cal_Capacity == None):
                        Remaining_Capacity = 0
                    else:
                        Remaining_Capacity = 0
                    SOH_Goop_Min['Cal'].append(Remaining_Capacity)
                    SOH_Goop_Min['SOH'].append(Cal_Capacity)
                    SOH_Goop_Min['Cal_E'].append(SOh_E)
                    
                    # for index, header in enumerate(headers2, start=shitty):
                    #     cell = sheet8.cell(row=4+31, column=index, value=header)
                    # cell = sheet8.cell(row=the_loop +31, column=shitty, value=first_min_v)
                    # cell = sheet8.cell(row=the_loop +31, column=shitty+1, value=last_max_v)
                    # cell = sheet8.cell(row=the_loop +31, column=shitty+2, value=SOC_Start)
                    # cell = sheet8.cell(row=the_loop +31, column=shitty+3, value=SOC_End)
                    # cell = sheet8.cell(row=the_loop +31, column=shitty+4, value=Cal_Capacity)
                    # cell = sheet8.cell(row=the_loop +31, column=shitty+5, value=Remaining_Capacity)    
                    # cell = sheet8.cell(row=the_loop +31, column=shitty+6, value=SOh_E)  
                    shitty += 7
                if first_min_v == None and last_max_v == None:
                    lost_data2.append(app+1)
                    lost_data4.append(app+1)
                # print(lost_data2)
                    # thread1 = threading.Thread(target=CCha)
                    # thread1.start()
                    # thread1.join()
                CCha()

    if app == 25:
        # Set the base filename
        date = str(date)
        modified_string = date.replace("-", "")
        base_filename = f'{modified_string}_MSF{ferry_ided}_Energy.xlsx'
        counter = 0
        for dex,group in enumerate(SOH_Goop['Goop']):
            if SOH_Goop['Charge'][dex] < 3 :
                SOH_Goop['SOH'][dex] = 9999
                
        

        sorted_data = sorted(zip(SOH_Goop['Goop'], SOH_Goop['SOH'], SOH_Goop['Cal_E']))
        sorted_dat = sorted(zip(SOH_Goop['Goop'],SOH_Goop['SOH']))
        # Group the data by 'Goop' value
        grouped = groupby(sorted_data, lambda x: x[0])
        groupe = groupby(sorted_dat, lambda x: x[0])
        # Initialize variables to store the minimum 'SOH' and its corresponding group
        min_soh = float('inf')  # Initialize with a large value
        min_soh_group = None
        # Define a tolerance value for floating-point comparisons
        tolerance = 1e-6
        for (group, group_data), (group2, group_dat) in zip(grouped, groupe):
            # print(f"Group: {group}")
            
            group_cal_e_values = [x[2] for x in group_data]
            group_soh_values = [x[1] for x in group_dat]
            # print(f"Group Cal_E Values: {group_cal_e_values}")
            # print(f"Group SOH Values: {group_soh_values}")

            
            # Check if any 'Cal_E' values are greater than 12.0 in this group
            if any(float(cal_e) > 12.0 for cal_e in group_cal_e_values):
                
                group_min_soh = min(group_soh_values)
                # print(f"Minimum SOH: {group_soh_values} (Group {group})")
                # print(f"Hi I am not stupid")
                
                # Update the minimum 'SOH' and its group if needed
                if group_min_soh < min_soh:
                    min_soh = group_min_soh
                    min_soh_group = group
                # Iterate over the groups and find the minimum 'SOH' value and its corresponding group

        # Print the minimum 'SOH' value and its corresponding group
        # print(f"Minimum SOH: {min_soh} (Group {min_soh_group})")
        # Print the minimum 'SOH' value and its corresponding group
        # print(f"Minimum SOH: {min_soh} (Group {min_soh_group})")
    #     lick = 5
    #     hick = 1
    #     for dex,group in enumerate(SOH_Goop['Goop']):
    #         print(min_soh_group)
    #         print(group)
    #         lgg = 0
    #         if group == min_soh_group:
    #             for i,value in enumerate(lost_data): #2
                    
    #                 print(f"This is hick{hick}")
    #                 print(f"This is Value{value}")

    #                 if hick != value and lgg == 0 :
    #                     cell = sheet.cell(row=lick, column=2, value=SOH_Goop['S_Time'][dex])
    #                     cell = sheet.cell(row=lick, column=3, value=SOH_Goop['E_Time'][dex])
    #                     cell = sheet.cell(row=lick, column=9, value=SOH_Goop['Max_V'][dex])
    #                     cell = sheet.cell(row=lick, column=8, value=SOH_Goop['Min_V'][dex])
    #                     cell = sheet.cell(row=lick, column=10, value=SOH_Goop['Start_SOC'][dex])
    #                     cell = sheet.cell(row=lick, column=11, value=SOH_Goop['End_SOC'][dex])
    #                     cell = sheet.cell(row=lick, column=6, value=SOH_Goop['Charge'][dex])
    #                     cell = sheet.cell(row=lick, column=7, value=0)
    #                     if SOH_Goop['Charge'][dex] < 3 :
    #                         cell = sheet.cell(row=lick, column=13, value=0)
    #                         cell = sheet.cell(row=lick, column=12, value=0)   
    #                     else:
    #                         cell = sheet.cell(row=lick, column=13, value=SOH_Goop['Cal'][dex])
    #                         cell = sheet.cell(row=lick, column=12, value=SOH_Goop['SOH'][dex])    
    #                     lick += 1
    #                     hick += 1
    #                     lgg += 1
    #                     if hick == value :
    #                         print(f"I am Nyi Nyi DEX {dex}")
    #                         print("I am Nyi Nyi")
    #                         break
    #                 elif hick == value: 
    #                     cell = sheet.cell(row=lick, column=2, value=SOH_Goop['S_Time'][dex])
    #                     cell = sheet.cell(row=lick, column=3, value=SOH_Goop['E_Time'][dex])
    #                     cell = sheet.cell(row=lick, column=9, value=0)
    #                     cell = sheet.cell(row=lick, column=8, value=0)
    #                     cell = sheet.cell(row=lick, column=10, value=0)
    #                     cell = sheet.cell(row=lick, column=11, value=0)
    #                     cell = sheet.cell(row=lick, column=13, value=0)
    #                     cell = sheet.cell(row=lick, column=12, value=0)
    #                     cell = sheet.cell(row=lick, column=7, value=0)
    #                     cell = sheet.cell(row=lick, column=6, value=0)
    #                     cell = sheet.cell(row=lick+1, column=9, value=SOH_Goop['Max_V'][dex])
    #                     cell = sheet.cell(row=lick+1, column=8, value=SOH_Goop['Min_V'][dex])
    #                     cell = sheet.cell(row=lick+1, column=10, value=SOH_Goop['Start_SOC'][dex])
    #                     cell = sheet.cell(row=lick+1, column=11, value=SOH_Goop['End_SOC'][dex])
    #                     cell = sheet.cell(row=lick+1, column=2, value=SOH_Goop['S_Time'][dex])
    #                     cell = sheet.cell(row=lick+1, column=3, value=SOH_Goop['E_Time'][dex])
    #                     cell = sheet.cell(row=lick+1, column=6, value=SOH_Goop['Charge'][dex])
    #                     cell = sheet.cell(row=lick+1, column=7, value=0)
    #                     if SOH_Goop['Charge'][dex] < 3 :
    #                         cell = sheet.cell(row=lick+1, column=13, value=0)
    #                         cell = sheet.cell(row=lick+1, column=12, value=0)   
    #                     else:
    #                         cell = sheet.cell(row=lick+1, column=13, value=SOH_Goop['Cal'][dex])
    #                         cell = sheet.cell(row=lick+1, column=12, value=SOH_Goop['SOH'][dex]) 
    #                     if len(lost_data) > 2:
    #                         hick += len(lost_data) - 1
    #                     else:
    #                         hick += 1

    #                     lick += 2
    #                     if len(lost_data) > 1:
    #                         lost_data.pop(0)
    #                     break
    #             if not lost_data   :
    #                     cell = sheet.cell(row=lick, column=2, value=SOH_Goop['S_Time'][dex])
    #                     cell = sheet.cell(row=lick, column=3, value=SOH_Goop['E_Time'][dex])
    #                     cell = sheet.cell(row=lick, column=9, value=SOH_Goop['Max_V'][dex])
    #                     cell = sheet.cell(row=lick, column=8, value=SOH_Goop['Min_V'][dex])
    #                     cell = sheet.cell(row=lick, column=10, value=SOH_Goop['Start_SOC'][dex])
    #                     cell = sheet.cell(row=lick, column=11, value=SOH_Goop['End_SOC'][dex])
    #                     cell = sheet.cell(row=lick, column=6, value=SOH_Goop['Charge'][dex])
    #                     cell = sheet.cell(row=lick, column=7, value=0)
    #                     if SOH_Goop['Charge'][dex] < 3 :
    #                         cell = sheet.cell(row=lick, column=13, value=0)
    #                         cell = sheet.cell(row=lick, column=12, value=0)   
    #                     else:
    #                         cell = sheet.cell(row=lick, column=13, value=SOH_Goop['Cal'][dex])
    #                         cell = sheet.cell(row=lick, column=12, value=SOH_Goop['SOH'][dex])    
    #                     lick += 1
    #     for dex,group in enumerate(SOH_Goop_Min['Goop']):
    #         if SOH_Goop_Min['Discharge'][dex] > -3 :
    #             SOH_Goop_Min['SOH'][dex] = 9999
                
        

        sorted_data_2 = sorted(zip(SOH_Goop_Min['Goop'], SOH_Goop_Min['SOH'], SOH_Goop_Min['Cal_E']))
        sorted_dat_2 = sorted(zip(SOH_Goop_Min['Goop'],SOH_Goop_Min['SOH']))
        # Group the data by 'Goop' value
        grouped2 = groupby(sorted_data_2, lambda x: x[0])
        groupe2 = groupby(sorted_dat_2, lambda x: x[0])
        # Initialize variables to store the minimum 'SOH' and its corresponding group
        min_soh2 = float('inf')  # Initialize with a large value
        min_soh_group2 = None
        # Define a tolerance value for floating-point comparisons
        tolerance = 1e-6
        for (group, group_data), (group2, group_dat) in zip(grouped2, groupe2):
            # print(f"Group: {group}")
            
            group_cal_e_values2 = [x[2] for x in group_data]
            group_soh_values2 = [x[1] for x in group_dat]
            # print(f"Group Cal_E Values: {group_cal_e_values2}")
            # print(f"Group SOH Values: {group_soh_values2}")

            
            # Check if any 'Cal_E' values are greater than 12.0 in this group
            if any(float(cal_e) < -12.0 for cal_e in group_cal_e_values2):
                
                group_min_soh = min(group_soh_values2)
                # print(f"Minimum SOH: {group_soh_values2} (Group {group})")
                # print(f"Hi I am not stupid")
                
                # Update the minimum 'SOH' and its group if needed
                if group_min_soh < min_soh2:
                    min_soh2 = group_min_soh
                    min_soh_group2 = group
                # Iterate over the groups and find the minimum 'SOH' value and its corresponding group




        # Print the minimum 'SOH' value and its corresponding group
        # print(f"Minimum SOH: {min_soh2} (Group {min_soh_group2})")
        # Print the minimum 'SOH' value and its corresponding group
        # print(f"Minimum SOH: {min_soh2} (Group {min_soh_group2})")
    #     lick = 5
    #     hick = 1
    #     for dex,group in enumerate(SOH_Goop_Min['Goop']):
    #         print(min_soh_group2)
    #         print(group)
    #         lgg = 0
    #         if group == min_soh_group2:
    #             for i,value in enumerate(lost_data2): #2
                    
    #                 print(f"This is hick{hick}")
    #                 print(f"This is Value{value}")

    #                 if hick != value and lgg == 0 :
    #                     cell = sheet200.cell(row=lick, column=2, value=SOH_Goop_Min['S_Time'][dex])
    #                     cell = sheet200.cell(row=lick, column=3, value=SOH_Goop_Min['E_Time'][dex])
    #                     cell = sheet200.cell(row=lick, column=9, value=SOH_Goop_Min['Max_V'][dex])
    #                     cell = sheet200.cell(row=lick, column=8, value=SOH_Goop_Min['Min_V'][dex])
    #                     cell = sheet200.cell(row=lick, column=10, value=SOH_Goop_Min['Start_SOC'][dex])
    #                     cell = sheet200.cell(row=lick, column=11, value=SOH_Goop_Min['End_SOC'][dex])
    #                     cell = sheet200.cell(row=lick, column=7, value=SOH_Goop_Min['Discharge'][dex])
    #                     cell = sheet200.cell(row=lick, column=6, value=0)
    #                     if SOH_Goop_Min['Discharge'][dex] > -3 :
    #                         cell = sheet200.cell(row=lick, column=13, value=0)
    #                         cell = sheet200.cell(row=lick, column=12, value=0)   
    #                     else:
    #                         cell = sheet200.cell(row=lick, column=13, value=SOH_Goop_Min['Cal'][dex])
    #                         cell = sheet200.cell(row=lick, column=12, value=SOH_Goop_Min['SOH'][dex])    
    #                     lick += 1
    #                     hick += 1
    #                     lgg += 1
    #                     if hick == value :
    #                         print(f"I am Nyi Nyi DEX {dex}")
    #                         print("I am Nyi Nyi")
    #                         break
    #                 elif hick == value: 
    #                     cell = sheet200.cell(row=lick, column=2, value=SOH_Goop_Min['S_Time'][dex])
    #                     cell = sheet200.cell(row=lick, column=3, value=SOH_Goop_Min['E_Time'][dex])
    #                     cell = sheet200.cell(row=lick, column=9, value=0)
    #                     cell = sheet200.cell(row=lick, column=8, value=0)
    #                     cell = sheet200.cell(row=lick, column=10, value=0)
    #                     cell = sheet200.cell(row=lick, column=11, value=0)
    #                     cell = sheet200.cell(row=lick, column=13, value=0)
    #                     cell = sheet200.cell(row=lick, column=12, value=0)
    #                     cell = sheet200.cell(row=lick, column=6, value=0)
    #                     cell = sheet200.cell(row=lick, column=7, value=0)
    #                     cell = sheet200.cell(row=lick+1, column=9, value=SOH_Goop_Min['Max_V'][dex])
    #                     cell = sheet200.cell(row=lick+1, column=8, value=SOH_Goop_Min['Min_V'][dex])
    #                     cell = sheet200.cell(row=lick+1, column=10, value=SOH_Goop_Min['Start_SOC'][dex])
    #                     cell = sheet200.cell(row=lick+1, column=11, value=SOH_Goop_Min['End_SOC'][dex])
    #                     cell = sheet200.cell(row=lick+1, column=2, value=SOH_Goop_Min['S_Time'][dex])
    #                     cell = sheet200.cell(row=lick+1, column=3, value=SOH_Goop_Min['E_Time'][dex])
    #                     cell = sheet200.cell(row=lick+1, column=7, value=SOH_Goop_Min['Discharge'][dex])
    #                     cell = sheet200.cell(row=lick+1, column=6, value=0)
    #                     if SOH_Goop_Min['Discharge'][dex] > -3 :
    #                         cell = sheet200.cell(row=lick+1, column=13, value=0)
    #                         cell = sheet200.cell(row=lick+1, column=12, value=0)   
    #                     else:
    #                         cell = sheet200.cell(row=lick+1, column=13, value=SOH_Goop_Min['Cal'][dex])
    #                         cell = sheet200.cell(row=lick+1, column=12, value=SOH_Goop_Min['SOH'][dex]) 
    #                     if len(lost_data2) > 2:
    #                         hick += len(lost_data2) - 1
    #                     else:
    #                         hick += 1

    #                     lick += 2
    #                     if len(lost_data2) > 1:
    #                         lost_data2.pop(0)
    #                     break
    #             if not lost_data2   :
    #                     cell = sheet200.cell(row=lick, column=2, value=SOH_Goop_Min['S_Time'][dex])
    #                     cell = sheet200.cell(row=lick, column=3, value=SOH_Goop_Min['E_Time'][dex])
    #                     cell = sheet200.cell(row=lick, column=9, value=SOH_Goop_Min['Max_V'][dex])
    #                     cell = sheet200.cell(row=lick, column=8, value=SOH_Goop_Min['Min_V'][dex])
    #                     cell = sheet200.cell(row=lick, column=10, value=SOH_Goop_Min['Start_SOC'][dex])
    #                     cell = sheet200.cell(row=lick, column=11, value=SOH_Goop_Min['End_SOC'][dex])
    #                     cell = sheet200.cell(row=lick, column=7, value=SOH_Goop_Min['Discharge'][dex])
    #                     cell = sheet200.cell(row=lick, column=6, value=0)
    #                     if SOH_Goop_Min['Discharge'][dex] > -3 :
    #                         cell = sheet200.cell(row=lick, column=13, value=0)
    #                         cell = sheet200.cell(row=lick, column=12, value=0)   
    #                     else:
    #                         cell = sheet200.cell(row=lick, column=13, value=SOH_Goop_Min['Cal'][dex])
    #                         cell = sheet200.cell(row=lick, column=12, value=SOH_Goop_Min['SOH'][dex])    
    #                     lick += 1
        while os.path.exists(base_filename):
            counter += 1
            base_filename = f'{modified_string}_MSF{ferry_ided}_SOH(Debugg)_{counter}.xlsx'
        # Save the workbook to a file
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']
        workbook.save(base_filename)

     # Print the filename
        print(f'Saved as: {base_filename}')  
      
        headers = [
            "Date","Start Time","Stop Time","Ferry","Pack No.",
            "Charged Energy (kWh)",
            "Discharged Energy (kWh)","Start_Min","End_Max","Start_SOC", "End_SOC", "Remaining_Capacity(kWh)","Remaining_Capacity(%)"
        ]
        wordbook2 = openpyxl.Workbook()
        worksheet = wordbook2.active  
        sheet3000 = wordbook2.create_sheet(title=f'Charge or DC {j}')
        for index, header in enumerate(headers, start=1):
            cell = sheet3000.cell(row=4, column=index, value=header)
        Header = 'StartTime'
        Header2 = 'StopTime'
                # Set the base filename

        cell = sheet3000.cell(row=1, column = 1, value = 'Ferry Number')
        cell = sheet3000.cell(row=2, column = 1, value = 'Start-Time')
        cell = sheet3000.cell(row=3, column = 1, value = 'Stop-Time')
        cell = sheet3000.cell(row=2, column = 3, value = '<-GMT +7 ')
        cell = sheet3000.cell(row=3, column = 3, value = '<-GMT +7 ')
        cell = sheet3000.cell(row=2, column = 1, value = Header)
        cell = sheet3000.cell(row=3, column = 1, value = Header2)
        cell = sheet3000.cell(row=2, column=2, value=o_start_datetime)
        cell = sheet3000.cell(row=3, column=2, value=o_end_datetime)
        cell = sheet3000.cell(row=1, column=2, value= ferry_ided)

        date = str(date)
        modified_string = date.replace("-", "/")
        i = 0
        for i in range(0,26):
            cell = sheet3000.cell(row=i+5, column=5, value=i + 1)
            cell = sheet3000.cell(row=i+5, column=5, value=i + 1)
            cell = sheet3000.cell(row=i+5, column=1, value= modified_string)
            cell = sheet3000.cell(row=i+5, column=4, value=ferry_ided)
        # print(f"This is Min SOH2 :{min_soh2}")
        # print(f"This is Min SOH :{min_soh}")
        if min_soh2 < min_soh:
            lick = 5
            hick = 1
            for dex,group in enumerate(SOH_Goop_Min['Goop']):
                # print(min_soh_group2)
                # print(group)

                lgg = 0
                if group == min_soh_group2:
                    for i,value in enumerate(lost_data4): #2
                        
                        # print(f"This is hick{hick}")
                        # print(f"This is Value{value}")

                        if hick != value and lgg == 0 :
                            cell = sheet3000.cell(row=lick, column=2, value=SOH_Goop_Min['S_Time'][dex])
                            cell = sheet3000.cell(row=lick, column=3, value=SOH_Goop_Min['E_Time'][dex])
                            cell = sheet3000.cell(row=lick, column=9, value=SOH_Goop_Min['Max_V'][dex])
                            cell = sheet3000.cell(row=lick, column=8, value=SOH_Goop_Min['Min_V'][dex])
                            cell = sheet3000.cell(row=lick, column=10, value=SOH_Goop_Min['Start_SOC'][dex])
                            cell = sheet3000.cell(row=lick, column=11, value=SOH_Goop_Min['End_SOC'][dex])
                            cell = sheet3000.cell(row=lick, column=7, value=SOH_Goop_Min['Discharge'][dex])
                            cell = sheet3000.cell(row=lick, column=6, value=0)
                            if SOH_Goop_Min['Discharge'][dex] > -3 :
                                cell = sheet3000.cell(row=lick, column=13, value=0)
                                cell = sheet3000.cell(row=lick, column=12, value=0)   
                            else:
                                cell = sheet3000.cell(row=lick, column=12, value=SOH_Goop_Min['SOH'][dex])
                                if float(SOH_Goop_Min['Cal'][dex]) > 100.0 and float(SOH_Goop_Min['Cal'][dex]) <130.0:
                                    cell = sheet3000.cell(row=lick, column=13, value=100) 
                                else:
                                    cell = sheet3000.cell(row=lick, column=13, value=SOH_Goop_Min['Cal'][dex]) 
                            lick += 1
                            hick += 1
                            lgg += 1
                            if hick == value :
                                # print(f"I am Nyi Nyi DEX {dex}")
                                # print("I am Nyi Nyi")
                                break
                        elif hick == value: 
                            cell = sheet3000.cell(row=lick, column=2, value=SOH_Goop_Min['S_Time'][dex])
                            cell = sheet3000.cell(row=lick, column=3, value=SOH_Goop_Min['E_Time'][dex])
                            cell = sheet3000.cell(row=lick, column=9, value=0)
                            cell = sheet3000.cell(row=lick, column=8, value=0)
                            cell = sheet3000.cell(row=lick, column=10, value=0)
                            cell = sheet3000.cell(row=lick, column=11, value=0)
                            cell = sheet3000.cell(row=lick, column=13, value=0)
                            cell = sheet3000.cell(row=lick, column=12, value=0)
                            cell = sheet3000.cell(row=lick, column=6, value=0)
                            cell = sheet3000.cell(row=lick, column=7, value=0)
                            cell = sheet3000.cell(row=lick+1, column=9, value=SOH_Goop_Min['Max_V'][dex])
                            cell = sheet3000.cell(row=lick+1, column=8, value=SOH_Goop_Min['Min_V'][dex])
                            cell = sheet3000.cell(row=lick+1, column=10, value=SOH_Goop_Min['Start_SOC'][dex])
                            cell = sheet3000.cell(row=lick+1, column=11, value=SOH_Goop_Min['End_SOC'][dex])
                            cell = sheet3000.cell(row=lick+1, column=2, value=SOH_Goop_Min['S_Time'][dex])
                            cell = sheet3000.cell(row=lick+1, column=3, value=SOH_Goop_Min['E_Time'][dex])
                            cell = sheet3000.cell(row=lick+1, column=7, value=SOH_Goop_Min['Discharge'][dex])
                            cell = sheet3000.cell(row=lick+1, column=6, value=0)
                            if SOH_Goop_Min['Discharge'][dex] > -3 :
                                cell = sheet3000.cell(row=lick+1, column=13, value=0)
                                cell = sheet3000.cell(row=lick+1, column=12, value=0)   
                            else:
                                
                                cell = sheet3000.cell(row=lick+1, column=12, value=SOH_Goop_Min['SOH'][dex])
                                if float(SOH_Goop_Min['Cal'][dex]) > 100.0 and float(SOH_Goop_Min['Cal'][dex]) <130.0:
                                    cell = sheet3000.cell(row=lick+1, column=13, value=100) 
                                else:
                                    cell = sheet3000.cell(row=lick+1, column=13, value=SOH_Goop_Min['Cal'][dex]) 
                            if len(lost_data4) > 2:
                                hick += len(lost_data4) - 1
                            else:
                                hick += 1

                            lick += 2
                            if len(lost_data4) > 1:
                                lost_data4.pop(0)
                            break
                    if not lost_data4   :
                            cell = sheet3000.cell(row=lick, column=2, value=SOH_Goop_Min['S_Time'][dex])
                            cell = sheet3000.cell(row=lick, column=3, value=SOH_Goop_Min['E_Time'][dex])
                            cell = sheet3000.cell(row=lick, column=9, value=SOH_Goop_Min['Max_V'][dex])
                            cell = sheet3000.cell(row=lick, column=8, value=SOH_Goop_Min['Min_V'][dex])
                            cell = sheet3000.cell(row=lick, column=10, value=SOH_Goop_Min['Start_SOC'][dex])
                            cell = sheet3000.cell(row=lick, column=11, value=SOH_Goop_Min['End_SOC'][dex])
                            cell = sheet3000.cell(row=lick, column=7, value=SOH_Goop_Min['Discharge'][dex])
                            cell = sheet3000.cell(row=lick, column=6, value=0)
                            if SOH_Goop_Min['Discharge'][dex] > -3 :
                                cell = sheet3000.cell(row=lick, column=13, value=0)
                                cell = sheet3000.cell(row=lick, column=12, value=0)   
                            else:
                                cell = sheet3000.cell(row=lick, column=12, value=SOH_Goop_Min['SOH'][dex])
                                if float(SOH_Goop_Min['Cal'][dex]) > 100.0 and float(SOH_Goop_Min['Cal'][dex]) <130.0:
                                    cell = sheet3000.cell(row=lick, column=13, value=100) 
                                else:
                                    cell = sheet3000.cell(row=lick, column=13, value=SOH_Goop_Min['Cal'][dex]) 
                                   
                            lick += 1
        else:
            lick = 5
            hick = 1
            for dex,group in enumerate(SOH_Goop['Goop']):
                # print(min_soh_group)
                # print(group)
                lgg = 0
                if group == min_soh_group:
                    for i,value in enumerate(lost_data3): #2
                        
                        # print(f"This is hick{hick}")
                        # print(f"This is Value{value}")

                        if hick != value and lgg == 0 :
                            cell = sheet3000.cell(row=lick, column=2, value=SOH_Goop['S_Time'][dex])
                            cell = sheet3000.cell(row=lick, column=3, value=SOH_Goop['E_Time'][dex])
                            cell = sheet3000.cell(row=lick, column=9, value=SOH_Goop['Max_V'][dex])
                            cell = sheet3000.cell(row=lick, column=8, value=SOH_Goop['Min_V'][dex])
                            cell = sheet3000.cell(row=lick, column=10, value=SOH_Goop['Start_SOC'][dex])
                            cell = sheet3000.cell(row=lick, column=11, value=SOH_Goop['End_SOC'][dex])
                            cell = sheet3000.cell(row=lick, column=6, value=SOH_Goop['Charge'][dex])
                            cell = sheet3000.cell(row=lick, column=7, value=0)
                            if SOH_Goop['Charge'][dex] < 3 :
                                cell = sheet3000.cell(row=lick, column=13, value=0)
                                cell = sheet3000.cell(row=lick, column=12, value=0)   
                            else:
                                cell = sheet3000.cell(row=lick, column=12, value=SOH_Goop['SOH'][dex])
                                if float(SOH_Goop['Cal'][dex]) > 100.0 and float(SOH_Goop['Cal'][dex]) <130.0:
                                    cell = sheet3000.cell(row=lick, column=13, value=100) 
                                else:
                                    cell = sheet3000.cell(row=lick, column=13, value=SOH_Goop['Cal'][dex])      
                            lick += 1
                            hick += 1
                            lgg += 1
                            if hick == value :
                                # print(f"I am Nyi Nyi DEX {dex}")
                                # print("I am Nyi Nyi")
                                break
                        elif hick == value: 
                            cell = sheet3000.cell(row=lick, column=2, value=SOH_Goop['S_Time'][dex])
                            cell = sheet3000.cell(row=lick, column=3, value=SOH_Goop['E_Time'][dex])
                            cell = sheet3000.cell(row=lick, column=9, value=0)
                            cell = sheet3000.cell(row=lick, column=8, value=0)
                            cell = sheet3000.cell(row=lick, column=10, value=0)
                            cell = sheet3000.cell(row=lick, column=11, value=0)
                            cell = sheet3000.cell(row=lick, column=13, value=0)
                            cell = sheet3000.cell(row=lick, column=12, value=0)
                            cell = sheet3000.cell(row=lick, column=7, value=0)
                            cell = sheet3000.cell(row=lick, column=6, value=0)
                            cell = sheet3000.cell(row=lick+1, column=9, value=SOH_Goop['Max_V'][dex])
                            cell = sheet3000.cell(row=lick+1, column=8, value=SOH_Goop['Min_V'][dex])
                            cell = sheet3000.cell(row=lick+1, column=10, value=SOH_Goop['Start_SOC'][dex])
                            cell = sheet3000.cell(row=lick+1, column=11, value=SOH_Goop['End_SOC'][dex])
                            cell = sheet3000.cell(row=lick+1, column=2, value=SOH_Goop['S_Time'][dex])
                            cell = sheet3000.cell(row=lick+1, column=3, value=SOH_Goop['E_Time'][dex])
                            cell = sheet3000.cell(row=lick+1, column=6, value=SOH_Goop['Charge'][dex])
                            cell = sheet3000.cell(row=lick+1, column=7, value=0)
                            if SOH_Goop['Charge'][dex] < 3 :
                                cell = sheet3000.cell(row=lick+1, column=13, value=0)
                                cell = sheet3000.cell(row=lick+1, column=12, value=0)   
                            else:
                                cell = sheet3000.cell(row=lick+1, column=12, value=SOH_Goop['SOH'][dex])
                                if float(SOH_Goop['Cal'][dex]) > 100.0 and float(SOH_Goop['Cal'][dex]) <130.0:
                                    cell = sheet3000.cell(row=lick+1, column=13, value=100) 
                                else:
                                    cell = sheet3000.cell(row=lick+1, column=13, value=SOH_Goop['Cal'][dex])  
                            if len(lost_data3) > 2:
                                hick += len(lost_data3) - 1
                            else:
                                hick += 1
                            lick += 2
                            if len(lost_data3) > 1:
                                lost_data3.pop(0)
                            break
                    if not lost_data3   :
                            cell = sheet3000.cell(row=lick, column=2, value=SOH_Goop['S_Time'][dex])
                            cell = sheet3000.cell(row=lick, column=3, value=SOH_Goop['E_Time'][dex])
                            cell = sheet3000.cell(row=lick, column=9, value=SOH_Goop['Max_V'][dex])
                            cell = sheet3000.cell(row=lick, column=8, value=SOH_Goop['Min_V'][dex])
                            cell = sheet3000.cell(row=lick, column=10, value=SOH_Goop['Start_SOC'][dex])
                            cell = sheet3000.cell(row=lick, column=11, value=SOH_Goop['End_SOC'][dex])
                            cell = sheet3000.cell(row=lick, column=6, value=SOH_Goop['Charge'][dex])
                            cell = sheet3000.cell(row=lick, column=7, value=0)
                            if SOH_Goop['Charge'][dex] < 3 :
                                cell = sheet3000.cell(row=lick, column=13, value=0)
                                cell = sheet3000.cell(row=lick, column=12, value=0)   
                            else:
                                cell = sheet3000.cell(row=lick, column=12, value=SOH_Goop['SOH'][dex])
                                if float(SOH_Goop['Cal'][dex]) > 100.0 and float(SOH_Goop['Cal'][dex]) <130.0:
                                    cell = sheet3000.cell(row=lick, column=13, value=100) 
                                else:
                                    cell = sheet3000.cell(row=lick, column=13, value=SOH_Goop['Cal'][dex])   
                            lick += 1

        date = str(date)
        modified_string = date.replace("-", "")
        base_filename2 = f'{modified_string}_MSF{ferry_ided}_SOH.xlsx'
        print(f'Saved as: hell {base_filename2}')
        counter = 0
        while os.path.exists(base_filename2):
            counter += 1
            base_filename2 = f'{modified_string}_MSF{ferry_ided}_SOH_{counter}.xlsx'
        # Save the wordbook2 to a file
        del wordbook2['Sheet']
        wordbook2.save(base_filename2)
        print(f'Saved as: {base_filename2}')
    print(app)

j = 1
def ProcessData(ferry_ided):
    start_time = time.time()  # Record the start time

    date_str = str(desired_time1)
    date = date_str.replace("22:00:00+07:00", "")

    # Create a new Excel workbook and add a sheet
    # workbook = openpyxl.Workbook()
    # worksheet = workbook.active
    global soop
    soop = 0
    global DataBP
    global lost_data
    global lost_data2
    global lost_data3
    global lost_data4
    lost_data = []
    lost_data2 = []
    lost_data3 = []
    lost_data4 = []
    global SOH_Goop
    global SOH_Goop_Min
    sampling = '30s'
    SOH_Goop = {
            'Goop': [],
            'SOH': [],
            'Max_V' : [],
            'Min_V':[],
            'Charge': [],
            'Start_SOC':[],
            'End_SOC':[],
            'SOH':[],
            'Cal':[],'S_Time':[],'E_Time':[],'Cal_E':[]
            } 
    SOH_Goop_Min = {
            'Goop': [],
            'SOH': [],
            'Max_V' : [],
            'Min_V':[],
            'Discharge': [],
            'Start_SOC':[],
            'End_SOC':[],
            'SOH':[],
            'Cal':[],'S_Time':[],'E_Time':[],'Cal_E':[]
            }     
    DataBP = True
    hex_values =[]  
    fields_c = []
    fields_v = []
    fields_max = []
    fields_min = []
    print("Yee Haw")
    global sheet
    global j 
    # sheet = workbook.create_sheet(title=f'Charge {j}')
    # sheet200 = workbook.create_sheet(title=f'Discharge {j}')
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    global sheetProcess
    sheetProcess = workbook.create_sheet(title=f'Fleet2 Energy')

    headers = [
        "Date","Start Time","Stop Time","Ferry","Pack No.",
        "Charged Energy (kWh)",
        "Discharged Energy (kWh)"
    ]
    for index, header in enumerate(headers, start=1):
            cell = sheetProcess.cell(row=4, column=index, value=header)
    Header = 'StartTime'
    Header2 = 'StopTime'
	        # Set the base filename

    cell = sheetProcess.cell(row=1, column = 1, value = 'Ferry Number')
    cell = sheetProcess.cell(row=2, column = 1, value = 'Start-Time')
    cell = sheetProcess.cell(row=3, column = 1, value = 'Stop-Time')
    cell = sheetProcess.cell(row=2, column = 3, value = '<-GMT +7 ')
    cell = sheetProcess.cell(row=3, column = 3, value = '<-GMT +7 ')
    cell = sheetProcess.cell(row=2, column = 1, value = Header)
    cell = sheetProcess.cell(row=3, column = 1, value = Header2)
    cell = sheetProcess.cell(row=2, column=2, value=o_start_datetime)
    cell = sheetProcess.cell(row=3, column=2, value=o_end_datetime)
    cell = sheetProcess.cell(row=1, column=2, value=ferry_ided)    
    # for index, header in enumerate(headers, start=1):
    #     cell = sheet.cell(row=4, column=index, value=header)
    # Header = 'StartTime'
    # Header2 = 'StopTime'
	#         # Set the base filename

    # cell = sheet.cell(row=1, column = 1, value = 'Ferry Number')
    # cell = sheet.cell(row=2, column = 1, value = 'Start-Time')
    # cell = sheet.cell(row=3, column = 1, value = 'Stop-Time')
    # cell = sheet.cell(row=2, column = 3, value = '<-GMT +7 ')
    # cell = sheet.cell(row=3, column = 3, value = '<-GMT +7 ')
    # cell = sheet.cell(row=2, column = 1, value = Header)
    # cell = sheet.cell(row=3, column = 1, value = Header2)
    # cell = sheet.cell(row=2, column=2, value=o_start_datetime)
    # cell = sheet.cell(row=3, column=2, value=o_end_datetime)
    # cell = sheet.cell(row=1, column=2, value= ferry_ided)

    # for index, header in enumerate(headers, start=1):
    #     cell = sheet200.cell(row=4, column=index, value=header)
    # Header = 'StartTime'
    # Header2 = 'StopTime'
	#         # Set the base filename

    # cell = sheet200.cell(row=1, column = 1, value = 'Ferry Number')
    # cell = sheet200.cell(row=2, column = 1, value = 'Start-Time')
    # cell = sheet200.cell(row=3, column = 1, value = 'Stop-Time')
    # cell = sheet200.cell(row=2, column = 3, value = '<-GMT +7 ')
    # cell = sheet200.cell(row=3, column = 3, value = '<-GMT +7 ')
    # cell = sheet200.cell(row=2, column = 1, value = Header)
    # cell = sheet200.cell(row=3, column = 1, value = Header2)
    # cell = sheet200.cell(row=2, column=2, value=o_start_datetime)
    # cell = sheet200.cell(row=3, column=2, value=o_end_datetime)
    # cell = sheet200.cell(row=1, column=2, value= ferry_ided)

    total_DC= []
    total_C = []
    j  += 1
    for num in range(1, 27):
        hex_value = hex(num)[2:].upper()  # Convert integer to hexadecimal and remove the '0x' prefix
        hex_values.append(hex_value)
        # print(hex_value)  
        fields = []
    for hex_val in hex_values:
        hex_val_int = int(hex_val, 16)
        if hex_val_int <= 0xF:
            field = f"0x180a000{hex_val.lower()}_S{hex_val}_BatPack_Current"
            fields_c.append(field)  
            field_max = f"0x180c000{hex_val.lower()}_S{hex_val}_MaxCell_Voltage"
            fields_max.append(field_max)
        else :
            field = f"0x180a00{hex_val.lower()}_S{hex_val}_BatPack_Current"
            fields_c.append(field)  
            field_max = f"0x180c00{hex_val.lower()}_S{hex_val}_MaxCell_Voltage"
            fields_max.append(field_max) 

    for hex_val in hex_values:
        hex_val_int = int(hex_val, 16)
        if hex_val_int <= 0xF:
            field = f"0x180a000{hex_val.lower()}_S{hex_val}_BatPack_Voltage"
            fields_v.append(field)  
            field_min = f"0x180c000{hex_val.lower()}_S{hex_val}_MinCell_Voltage"
            fields_min.append(field_min)            
        else :
            field = f"0x180a00{hex_val.lower()}_S{hex_val}_BatPack_Voltage"
            fields_v.append(field)  
            field_min = f"0x180c00{hex_val.lower()}_S{hex_val}_MinCell_Voltage"
            fields_min.append(field_min)   
    # print(fields_max)
    # print(fields_min)
    # print(fields_v)
    # print(fields_c)
    # print(len(fields_v))

    for num in range(1, 27):
        hex_value = hex(num)[2:].upper()  # Convert integer to hexadecimal and remove the '0x' prefix
        hex_values.append(hex_value)
        # print(hex_value)

    def Threads1():
        global result_totalC
        query_totalC = f' from(bucket:"datalogger")\
        |> range(start:{start_t}, stop:{end_t})\
        |> filter(fn:(r) => r._measurement == "mbcu")\
        |> filter(fn:(r) => r._field == "0x1801d0f3_BatPack_TotCurrent" )\
        |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
        |> aggregateWindow(every: {sampling}, fn: last, createEmpty: false)\
        |> yield(name: "last")'
        


        # print(query_totalC) 
        client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
        query_api = client.query_api()
        # Write a query and execute it

        result_totalC = query_api.query(org=org, query=query_totalC)
    def Threads2():
        global result_totalV
        query_totalV = f' from(bucket:"datalogger")\
        |> range(start:{start_t}, stop:{end_t})\
        |> filter(fn:(r) => r._measurement == "mbcu")\
        |> filter(fn:(r) => r._field == "0x1801d0f3_BatPack_TotVoltage" )\
        |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
        |> aggregateWindow(every: {sampling}, fn: last, createEmpty: false)\
        |> yield(name: "last")'
        
        


        # print(query_totalV) 
        client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
        query_api = client.query_api()
    # Write a query and execute it

        result_totalV = query_api.query(org=org, query=query_totalV)
    threads1 = threading.Thread(target=Threads1)
    threads2 = threading.Thread(target=Threads2)

    threads1.start()
    threads2.start()

        # Wait for both threads to finish
    threads1.join()
    threads2.join()
    # sheet2 = workbook.create_sheet(title=f'Volt & Current')

    # Start row and column
    # cell =sheet2.cell(row=1, column=1, value="Time_Stamp")
    # cell =sheet2.cell(row=1, column=2, value="Total_Voltage(V)")
    # cell =sheet2.cell(row=1, column=3, value="Total_Current(A)")
    start_row = 1  # You can adjust this to start from a different row
    column_d = 4  # Column D

    # Loop to write "Pack 1" to "Pack 26"
    for pack_number in range(1, 27):
        value = f"Pack_Current_{pack_number}"
        # sheet2.cell(row=start_row, column=column_d, value=value)
        column_d += 1
    loopy = 24
    column = 4

    global SystemCC
    global system_V_Timee
    SystemCC = []
    system_V_Timee = [] 
    for table in result_totalC:
        for record in table.records:   
            value = record.get_value()
            SystemCC.append(value)
            timestamp = record.get_time()
            converted_times = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
            # print(converted_times)
            system_V_Timee.append(converted_times)
            # cell = sheet2.cell(row=rows, column = 1, value = converted_times)
            # cell = sheet2.cell(row=rows, column = 3, value = value)
    # global sheet8
    # sheet8 = workbook.create_sheet(title=f'C_1')
    if SystemCC == []:
        return
    for i in range(26):
        
        def Thread1():
            global global_result2  # Access the global variables
            query2 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{fields_v[i]}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: true)'
            


            # print(query2) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result2 = query_api.query(org=org, query=query2)
            # print("processing")
            results2 = []

            for table in result2:
                for record in table.records:
                    results2.append((record.get_field(), record.get_value()))
            
            global_result2 = result2
        def Thread2():
            global global_result3  # Access the global variables


            query3 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{fields_c[i]}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: false)'



            # print(query3) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result3 = query_api.query(org=org, query=query3)
            # print("processing")
            results3 = []
            for table in result3:
                for record in table.records:
                    results3.append((record.get_field(), record.get_value()))
            
            
            global_result3 = result3
        def Thread3():
            global result4

            query4 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{fields_max[i]}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: false)'



            # print(query4) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result4 = query_api.query(org=org, query=query4)
            # print("processing")
            results4 = []
            for table in result4:
                for record in table.records:
                    results4.append((record.get_field(), record.get_value()))
            result_maxs = results4
            # print(results4)
        def Thread4():
            global result5
            query5 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{fields_min[i]}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: false)'

            # print(query5) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result5 = query_api.query(org=org, query=query5)
            # print("processing")
            results5 = []
            for table in result5:
                for record in table.records:
                    results5.append((record.get_field(), record.get_value()))
        threading_Timer_s = time.time()
        thread1 = threading.Thread(target=Thread1)
        thread2 = threading.Thread(target=Thread2)
        thread3 = threading.Thread(target=Thread3)
        thread4 = threading.Thread(target=Thread4)

        # print(results3)


        # Start the threads
        thread1.start()
        thread2.start()
        thread3.start()
        thread4.start()
        # Wait for both threads to finish
        thread1.join()
        thread2.join()
        thread3.join()
        thread4.join()
        result2 = global_result2
        result3 = global_result3
        threading_Timer_e = time.time()
        execution_time_thread = threading_Timer_e - threading_Timer_s
        print(f"Process_Data took {execution_time_thread:.6f} seconds to execute.")
        # print(query5) 
        # client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
        # query_api = client.query_api()
        # # Write a query and execute it

        # result5 = query_api.query(org=org, query=query5)
        # print("processing")
        # results5 = []
        # for table in result5:
        #     for record in table.records:
        #         results5.append((record.get_field(), record.get_value()))
        # # print(results5)
        the_loop = i + 5
        app = i 
        if i == 25:
            print("finished processing")
            DataBP = False
        column = i + 4

        show_result_window(result2,result3, the_loop,app,total_DC,total_C,column,soop,workbook,ferry_ided,sheetProcess)
    end_time = time.time()  # Record the end time
    execution_time = end_time - start_time  # Calculate the execution time
    print(f"Process_Data took {execution_time:.6f} seconds to execute.")


for i,ferry in enumerate(ferries):
    # ferrish = int(ferry)
    # if ferrish != 17:
    ProcessData(ferry)



date_str = str(desired_time1)
date = date_str.replace("22:00:00+07:00", "")
date = str(date)
modified_string = date.replace("-", "/")
modified_date_save1 = date.replace("-", "")
modified_date_save = modified_date_save1.replace(" ", "")

def OfflineCheck(ferry_ided,sheet_PackStat_Sum):
    global ferry_ide
    ferry_ide = ferry_ided
    start_time = time.time()  # Record the start time

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    date_str = str(desired_time1)
    date = date_str.replace("22:00:00+07:00", "")

    sampling = '30s'
    # Create a new Excel workbook and add a sheet
    # workbook = openpyxl.Workbook()
    # worksheet = workbook.active

    hex_values =[]  
    fields_c = []
    fields_v = []
    fields_max = []
    fields_min = []
    field_online_stat = []
    fields_MaxTemp = []
    fields_MinTemp = []



    for num in range(1, 27):
        hex_value = hex(num)[2:].upper()  # Convert integer to hexadecimal and remove the '0x' prefix
        hex_values.append(hex_value)
        # print(hex_value)  
        fields = []
    for hex_val in hex_values:
        hex_val_int = int(hex_val, 16)
        if hex_val_int <= 0xF:
            field = f"0x180a000{hex_val.lower()}_S{hex_val}_BatPack_Current"
            fields_c.append(field)  
            field = f"0x180b000{hex_val.lower()}_S{hex_val}_MinTemp"
            fields_MinTemp.append(field)  
            field_max = f"0x180c000{hex_val.lower()}_S{hex_val}_MaxCell_Voltage"
            fields_max.append(field_max)
        else :
            field = f"0x180a00{hex_val.lower()}_S{hex_val}_BatPack_Current"
            fields_c.append(field)  
            field = f"0x180b00{hex_val.lower()}_S{hex_val}_MinTemp"
            fields_MinTemp.append(field)  
            field_max = f"0x180c00{hex_val.lower()}_S{hex_val}_MaxCell_Voltage"
            fields_max.append(field_max) 

    for hex_val in hex_values:
        hex_val_int = int(hex_val, 16)
        if hex_val_int <= 0xF:
            field = f"0x180a000{hex_val.lower()}_S{hex_val}_BatPack_Voltage"
            fields_v.append(field)  
            field = f"0x180b000{hex_val.lower()}_S{hex_val}_MaxTemp"
            fields_MaxTemp.append(field)  
            field_min = f"0x180c000{hex_val.lower()}_S{hex_val}_MinCell_Voltage"
            fields_min.append(field_min)            
        else :
            field = f"0x180a00{hex_val.lower()}_S{hex_val}_BatPack_Voltage"
            fields_v.append(field)  
            field = f"0x180b00{hex_val.lower()}_S{hex_val}_MaxTemp"
            fields_MaxTemp.append(field)  
            field_min = f"0x180c00{hex_val.lower()}_S{hex_val}_MinCell_Voltage"
            fields_min.append(field_min)   

    for num in range(1, 27):
            field = f"0x1805d0f3_Pack{num}_Status"
            field_online_stat.append(field)


    headers = [
        "Date","Start Time","Stop Time","Ferry","Pack No.",
        "Fault_Pack","Possible_Cause",
        "Time_Disconnected","Time_Disconnected_Until","System_V","Pack_V","MaxCell_V","MinCell_V","MaxTemp","MinTemp,", "MaxCell_Over","MinCell_Under","Temp_Over","Temp_Under"
    ]

    for index, header in enumerate(headers, start=1):
            cell = sheet_PackStat_Sum.cell(row=1, column=index, value=header)    
    g = 1
    h = 2

    for num in range(1, 27):
        hex_value = hex(num)[2:].upper()  # Convert integer to hexadecimal and remove the '0x' prefix
        hex_values.append(hex_value)
        # print(hex_value)

        field_MaxTemp = fields_MaxTemp[num-1]
        field_MinTemp = fields_MinTemp[num-1]
        field_max = fields_max[num-1]
        field_min = fields_min[num-1]

        cell = sheet_PackStat_Sum.cell(row=num+1, column=5, value= num)
        cell = sheet_PackStat_Sum.cell(row=num+1, column=1, value= modified_string)
        cell = sheet_PackStat_Sum.cell(row=num+1, column=4, value=ferry_ided)
        cell = sheet_PackStat_Sum.cell(row=num+1, column=2, value= "5:00:00")
        cell = sheet_PackStat_Sum.cell(row=num+1, column=3, value="22:00:00")

        def Thread1():
            global result_MaxTemp # Access the global variables
            query2 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{field_MaxTemp}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: true)'
            


            # print(query2) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result_MaxTemp = query_api.query(org=org, query=query2)
            # print("processing")
            results_MaxTemps = []



        def Thread2():
            global result_MinTemp  # Access the global variables


            query3 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{field_MinTemp}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: false)'



            # print(query3) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result_MinTemp = query_api.query(org=org, query=query3)
            # print("processing")


        def Thread3():
            global result_MaxCell

            query4 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{field_max}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: false)'



            # print(query4) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result_MaxCell = query_api.query(org=org, query=query4)
            # print("processing")


        def Thread4():
            global result_MinCell
            query5 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{field_min}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: false)'

            # print(query5) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result_MinCell = query_api.query(org=org, query=query5)






        def Threads5():
            global result_totalV
            query_totalV = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "mbcu")\
            |> filter(fn:(r) => r._field == "0x1801d0f3_BatPack_TotVoltage" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: false)\
            |> yield(name: "last")'
            
            


            # print(query_totalV) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
        # Write a query and execute it

            result_totalV = query_api.query(org=org, query=query_totalV)

        def Thread6():
            global result_PackV  # Access the global variables
            query2 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{fields_v[num-1]}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: true)'
            


            # print(query2) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result2 = query_api.query(org=org, query=query2)
            result_PackV = result2

        def Thread7():
            global result_PackOnline  # Access the global variables
            queryPackOnline = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "mbcu")\
            |> filter(fn:(r) => r._field == "0x1801d0f3_BatPack_Fault_Strings_Number" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: true)'
            


            # print(queryPackOnline) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            resultPOnline = query_api.query(org=org, query=queryPackOnline)
            result_PackOnline = resultPOnline

        threading_Timer_s = time.time()
        thread1 = threading.Thread(target=Thread1)
        thread2 = threading.Thread(target=Thread2)
        thread3 = threading.Thread(target=Thread3)
        thread4 = threading.Thread(target=Thread4)
        thread5 = threading.Thread(target=Threads5)
        thread6 = threading.Thread(target=Thread6)
        thread7 = threading.Thread(target=Thread7)
        # print(results3)


        # Start the threads
        thread1.start()
        thread2.start()
        thread3.start()
        thread4.start()
        thread5.start()
        thread6.start()    
        thread7.start()    

        # Wait for both threads to finish
        thread1.join()
        thread2.join()
        thread3.join()
        thread4.join()
        thread5.join()
        thread6.join()
        thread7.join()

        threading_Timer_e = time.time()
        execution_time_thread = threading_Timer_e - threading_Timer_s
        print(f"Process_Data took {execution_time_thread:.6f} seconds to execute.")        
        Value_SystemV = []
        Value_PackV = []
        Value_SystemTime = []
        PackOffline = 0
        smooth = 2
        for table in result_totalV:
            for record in table.records:
                timestamp = record.get_time()
                converted_time1 = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                value = record.get_value()
                # if converted_time1 == timestamp_PackStat_Off[0]:
                #     Value_total_V = value
                Value_SystemV.append(value)
                Value_SystemTime.append(converted_time1)
                # cell = sheet_PackStat_V.cell(row=smooth, column = 2, value = value) 
                # cell = sheet_PackStat_V.cell(row=smooth, column = 1, value = converted_time1)
                smooth += 1

        for table in result_PackOnline:
            for record in table.records:
                timestamp = record.get_time()
                converted_time1 = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                value = record.get_value()
                # if converted_time1 == timestamp_PackStat_Off[0]:
                #     Value_total_V = value
                if value is not None:

                    if value < 26 or value > 0:
                        PackOffline = 1

        smooth = 2
        iish = 0
        saved = False
        saved_SV = 0
        saved_PV = 0
        saved_time = 0
        saved_2 = False

        for table in result_PackV:
            for record in table.records:
                timestamp = record.get_time()
                converted_time1 = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                value = record.get_value()
                # if converted_time1 == timestamp_PackStat_Off[0]:
                #     Value_PackV = value
                if value is not None:
                    Value_PackV.append(value)
                    difference_for_t = abs(value - Value_SystemV[iish])
                    if difference_for_t > 5.0 and saved == False:
                        saved = True
                        saved_time = Value_SystemTime[iish]
                        saved_SV = Value_SystemV[iish]
                        saved_PV = value
                    if difference_for_t > 5.0:
                        saved_2 = False
                    if difference_for_t < 1.0 and saved_2 == False:
                        saved_end_time = converted_time1
                        saved_2 = True
                    # cell = sheet_PackStat_V.cell(row=smooth, column = num + 4, value = value) 
                    # cell = sheet_PackStat_V.cell(row=smooth, column = 3, value = converted_time1) 
                    smooth += 1
                    iish += 1
            if saved_2 == False:
                saved_end_time = converted_time1

        if result_PackV == []:
            saved_time = Value_SystemTime[0]
            saved_end_time = Value_SystemTime[-1]
            print(saved_time)
      

        
        if Value_PackV != []:
            Pack_avg = sum(Value_PackV)/len(Value_PackV)
        else:
            Pack_avg = 0
        System_avg  = sum(Value_SystemV)/len(Value_SystemV)

        # cell = sheet_PackStat_V.cell(row=smooth+2, column = num + 4, value =Pack_avg) 
        # cell = sheet_PackStat_V.cell(row=smooth+2, column = 3, value = System_avg)         
        Voltage_difference = abs(System_avg-Pack_avg)
        last_voltage = abs(Value_SystemV[-1] - Value_PackV[-1])  
        if (Voltage_difference > 5.0 and PackOffline == 1)  or last_voltage > 5.0:

            # cell = sheet_PackStat_V.cell(row=smooth+3, column = num + 4, value = "Bad") 



            MaxCell_V = []
            MinCell_V = []
            Max_Temp = []
            Min_Temp = []
            Value_maxV = 0
            Value_minV = 0
            Value_maxtemp = 0
            Value_mintemp = 0
            Value_total_V = 0
            Value_PackV = 0

            for table in result_MaxTemp:
                for record in table.records:
                    timestamp = record.get_time()
                    converted_time1 = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                    
                    field = record.get_field()
                    value = record.get_value()

                    Max_Temp.append(value)
                    

                    if converted_time1 == saved_time:
                        Value_maxtemp = int(value)


            for table in result_MinTemp:
                for record in table.records:
                    timestamp = record.get_time()
                    converted_time1 = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                    value = record.get_value()

                    Min_Temp.append(value)

                    if converted_time1 == saved_time:
                        Value_mintemp = int(value)

            for table in result_MaxCell:

                for record in table.records:
                    timestamp = record.get_time()
                    converted_time1 = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                    value = record.get_value()/1000

                    MaxCell_V.append(value)

                    if converted_time1 == saved_time:
                        Value_maxV = float(value)

            for table in result_MinCell:
                for record in table.records:
                    timestamp = record.get_time()
                    converted_time1 = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                    value = record.get_value()/1000

                    MinCell_V.append(value)

                    if converted_time1 == saved_time:
                        Value_minV = float(value)

            BCU_con_p = 0
            Cell_Max_p = 0
            Cell_Min_p = 0
            Temp_Max_p = 0
            Temp_Min_p = 0 
            Volt_Diff_p = 0
            Fuse_DC = 0

            Total_Cell_V = Value_maxV + Value_minV
            if Total_Cell_V > 4.9 and Total_Cell_V <= 5.1:
                BCU_con_p = 1
            if Value_maxV > 4.19 :
                Cell_Max_p = 1
            if Value_minV < 3.05 and Value_minV > 0:
                Cell_Min_p = 1
            if Value_maxtemp >= 55 :
                Temp_Max_p = 1
            if Value_mintemp < 0 :
                Temp_Min_p = 1
            if BCU_con_p == 0 and Cell_Max_p == 0 and Cell_Min_p == 0 and Temp_Max_p == 0 and Temp_Min_p == 0:
                Volt_Diff = abs(saved_SV-saved_PV)
                if Volt_Diff > 5.0 or Volt_Diff < -5.0:
                    Volt_Diff_p = 1
            if Value_maxV == 0 and Value_minV == 0 and Value_maxtemp == 0 and Value_mintemp == 0 and saved_PV == 0:
                Fuse_DC = 1
            Fault_Massage = ["BCU_Connector_Problem","MaxCellV_Over","MinCellV_Under","MaxTemp_Over","MinTemp_Under","Fuse Disconnected","Volt Diff(Cause Undetected)","Lost_Data"]
            if BCU_con_p == 1:
                cell = sheet_PackStat_Sum.cell(row=num+1, column = 7, value = Fault_Massage[0])
            elif BCU_con_p == 0 and Cell_Max_p == 1:
                cell = sheet_PackStat_Sum.cell(row=num+1, column = 7, value = Fault_Massage[1])
            elif BCU_con_p == 0 and Cell_Max_p == 0 and Cell_Min_p == 1:
                cell = sheet_PackStat_Sum.cell(row=num+1, column = 7, value = Fault_Massage[2])
            elif BCU_con_p == 0 and Cell_Max_p == 0 and Cell_Min_p == 0 and Temp_Max_p == 1:
                cell = sheet_PackStat_Sum.cell(row=num+1, column = 7, value = Fault_Massage[3])
            elif BCU_con_p == 0 and Cell_Max_p == 0 and Cell_Min_p == 0 and Temp_Max_p == 0 and Temp_Min_p == 1:
                cell = sheet_PackStat_Sum.cell(row=num+1, column = 7, value = Fault_Massage[4])
            elif BCU_con_p == 0 and Cell_Max_p == 0 and Cell_Min_p == 0 and Temp_Max_p == 0 and Temp_Min_p == 0 and Fuse_DC == 1:
                cell = sheet_PackStat_Sum.cell(row=num+1, column = 7, value = Fault_Massage[5])
            elif BCU_con_p == 0 and Cell_Max_p == 0 and Cell_Min_p == 0 and Temp_Max_p == 0 and Temp_Min_p == 0 and Fuse_DC == 0 and Volt_Diff_p:
                cell = sheet_PackStat_Sum.cell(row=num+1, column = 7, value = Fault_Massage[6])
            else:
                cell = sheet_PackStat_Sum.cell(row=num+1, column = 7, value = Fault_Massage[7])

            Total_V_Col = 10
            Value_PackV_Col = 11
            Value_maxV_Col = 12
            Value_minV_Col = 13
            Value_maxtemp_Col = 14
            Value_mintemp_Col = 15
            Value_stime = 8
            Value_etime = 9




            cell = sheet_PackStat_Sum.cell(row=num+1, column = 6, value = 1) 
            cell = sheet_PackStat_Sum.cell(row=num+1, column = Total_V_Col, value = saved_SV) 
            cell = sheet_PackStat_Sum.cell(row=num+1, column = Value_PackV_Col, value = saved_PV)
            cell = sheet_PackStat_Sum.cell(row=num+1, column = Value_maxV_Col, value = Value_maxV)
            cell = sheet_PackStat_Sum.cell(row=num+1, column = Value_minV_Col, value = Value_minV)
            cell = sheet_PackStat_Sum.cell(row=num+1, column = Value_maxtemp_Col, value = Value_maxtemp)
            cell = sheet_PackStat_Sum.cell(row=num+1, column = Value_mintemp_Col, value = Value_mintemp)
            cell = sheet_PackStat_Sum.cell(row=num+1, column = Value_stime, value = saved_time)
            cell = sheet_PackStat_Sum.cell(row=num+1, column = Value_etime, value = saved_end_time)

    # Define the directory where you want to save the files


        else:
            # cell = sheet_PackStat_V.cell(row=smooth+3, column = num + 4, value = "Good")
            cell = sheet_PackStat_Sum.cell(row=num+1, column = 6, value = 0) 
        MaxCell_V = []
        MinCell_V = []
        Max_Temp = []
        Min_Temp = []


        for table in result_MaxTemp:
            for record in table.records:
                timestamp = record.get_time()
                converted_time1 = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                
                field = record.get_field()
                value = record.get_value()
                if value is not None:
                    Max_Temp.append(value)
                



        for table in result_MinTemp:    
            for record in table.records:
                timestamp = record.get_time()
                converted_time1 = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                value = record.get_value()
                if value is not None:
                    Min_Temp.append(value)

        for table in result_MaxCell:

            for record in table.records:
                timestamp = record.get_time()
                converted_time1 = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                value = record.get_value()/1000
                if value is not None:   
                    MaxCell_V.append(value)


        for table in result_MinCell:
            for record in table.records:
                timestamp = record.get_time()
                converted_time1 = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                value = record.get_value()/1000
                if value is not None:
                    MinCell_V.append(value)

        # Cell_Max_V_today = 0s
        # Cell_Min_V_today = 0
        # Temp_Max_today = 0
        # Temp_Min_today = 0

        if MaxCell_V != []:
            Cell_Max_V_today = max(MaxCell_V)
        if MinCell_V != []:
            Cell_Min_V_today = min(MinCell_V)
        if Max_Temp != []:
            Temp_Max_today = max(Max_Temp)
        if Min_Temp != []:
            Temp_Min_today = min(Min_Temp)
        
        Cell_Max_Over = 0
        Cell_Min_Under = 0
        Temp_Max_Over = 0
        Temp_Min_Under = 0
        if ferries == 2 or ferries == 18 or ferries ==  22 or ferries == 21:
            if Cell_Max_V_today > 4.12 :
                Cell_Max_Over = 1 
        else :           
            if Cell_Max_V_today > 4.16 :
                Cell_Max_Over = 1
        if Cell_Min_V_today < 3.4 :
            Cell_Min_Under = 1
        if Temp_Max_today > 40 :
            Temp_Max_Over = 1
        if Temp_Min_today <15:
            Temp_Min_Under = 0        

        Fault_Massage = ["Max Cell Volt Operating Higher than Recommended","Min Cell Volt Operating Higher than Recommended","Max Temp Operating Higher than Recommended","Min Temp Operating Lower than Recommeded"]

        MaxCLvl1 = 0
        MinCLvl1 = 0
        MaxTLvl1 = 0
        MinTLvl1 = 0
        
        Value_maxV_Col= 12
        Value_minV_Col = 13
        Value_maxtemp_Col = 14
        Value_mintemp_Col = 15

        Falut_Message_LvL1 = ""
        if Cell_Max_Over == 1:
            cell = sheet_PackStat_Sum.cell(row=num+1, column = 16, value = 1)
            cell = sheet_PackStat_Sum.cell(row=num+1, column=6, value=1)
            Falut_Message_LvL1 = Falut_Message_LvL1 +" "+ "Max_Cell_LvL 1"
            cell = sheet_PackStat_Sum.cell(row=num+1, column = Value_maxV_Col, value = Cell_Max_V_today)

        else:
            cell = sheet_PackStat_Sum.cell(row=num+1, column = 16, value = 0)

        if Cell_Min_Under == 1:
            cell = sheet_PackStat_Sum.cell(row=num+1, column = 17, value = 1)
            cell = sheet_PackStat_Sum.cell(row=num+1, column=6, value=1)
            Falut_Message_LvL1 = Falut_Message_LvL1+" " + "Min_Cell_LvL 1"
            cell = sheet_PackStat_Sum.cell(row=num+1, column = Value_minV_Col, value = Cell_Min_V_today)
            
        else:
            cell = sheet_PackStat_Sum.cell(row=num+1, column = 17, value = 0)
        if Temp_Max_Over == 1:
            cell = sheet_PackStat_Sum.cell(row=num+1, column = 18, value = 1)
            cell = sheet_PackStat_Sum.cell(row=num+1, column=6, value=1)
            Falut_Message_LvL1 = Falut_Message_LvL1+" " + "Max_T_LvL 1"
            cell = sheet_PackStat_Sum.cell(row=num+1, column = Value_maxV_Col, value = Temp_Max_today)

        else:
            cell = sheet_PackStat_Sum.cell(row=num+1, column = 18, value = 0)
        if Temp_Min_Under == 1:
            cell = sheet_PackStat_Sum.cell(row=num+1, column = 19, value = 1)
            cell = sheet_PackStat_Sum.cell(row=num+1, column=6, value=1)
            Falut_Message_LvL1 = Falut_Message_LvL1+" " + "Min_T_LvL 1"
            cell = sheet_PackStat_Sum.cell(row=num+1, column = Value_mintemp_Col, value = Temp_Min_today )


        else:
            cell = sheet_PackStat_Sum.cell(row=num+1, column = 19, value = 0)


        cell = sheet_PackStat_Sum.cell(row=num+1, column = 7, value = Falut_Message_LvL1)


    # Define the directory where you want to save the files
    
    # folder_path = os.path.join(os.path.expanduser('~'), 'Documents', f'PackStat_Check')

    # # If the directory does not exist, create it
    # if not os.path.exists(folder_path):
    #     os.makedirs(folder_path)

#     base_filename = f'{modified_date_save}_PackStat_MSF{ferry_ide}.xlsx'
#     counter = 0
#     while os.path.exists( base_filename):
#         counter += 1
#         base_filename = f'{modified_date_save}_PackStat_MSF{ferry_ide}_{counter}.xlsx'
#     # Save the wordbook2 to a file
#     del workbook['Sheet']
#     workbook.save( base_filename)

#     print(f'Saved as: {base_filename}')
# ferris = Get_ferry_id()
# for i in ferris:
#     OfflineCheck(i)


def ProcessRange(ferries):
    ferry_ided = ferries
    sampling = '30s'


    fields_MaxTemp = []
    fields_MinTemp = []
    fields_max = []
    fields_min = []



    date_str = str(desired_time1)
    date = date_str.replace("22:00:00+07:00", "")
    date = str(date)
    modified_string = date.replace("-", "/")

    modified_date_save1 = date.replace("-", "")
    modified_date_save = modified_date_save1.replace(" ", "")
    hex_values = []

    workbook = openpyxl.Workbook()
    sheet_Process = workbook.create_sheet(title = f'MaxMins')
    workbook2 = openpyxl.Workbook()
    sheet_Process_Vdiff = workbook2.create_sheet(title = f'MaxMins')
    sheet_PackStat_Sum = workbook2.create_sheet(title=f'PackFault')
    headers2 = [
        "Date","Start Time","Stop Time","MSF","Pack No.","Cell_V_Diff_Time","MaxCellV","MinCellV","Cell_V_Diff"

    ]   
    headers = [
        "Date","Start Time","Stop Time","MSF","Pack No.","MinCellV","MaxCellV","MaxTemp","MinTemp""Max_Cell_Diff","Time_C_Diff"

    ]  

    for index, header in enumerate(headers, start=1):
            cell = sheet_Process.cell(row=1, column=index, value=header)
    for index, header in enumerate(headers2, start=1):
            cell = sheet_Process_Vdiff.cell(row=1, column=index, value=header)
    for num in range(1, 27):
        hex_value = hex(num)[2:].upper()  # Convert integer to hexadecimal and remove the '0x' prefix
        hex_values.append(hex_value)
        # print(hex_value)  
        fields = []
    for hex_val in hex_values:
        hex_val_int = int(hex_val, 16)
        if hex_val_int <= 0xF:
            field = f"0x180b000{hex_val.lower()}_S{hex_val}_MaxTemp"
            fields_MaxTemp.append(field)  
            field_max = f"0x180c000{hex_val.lower()}_S{hex_val}_MaxCell_Voltage"
            fields_max.append(field_max)
        else :
            field = f"0x180b00{hex_val.lower()}_S{hex_val}_MaxTemp"
            fields_MaxTemp.append(field)  
            field_max = f"0x180c00{hex_val.lower()}_S{hex_val}_MaxCell_Voltage"
            fields_max.append(field_max) 

    for hex_val in hex_values:
        hex_val_int = int(hex_val, 16)
        if hex_val_int <= 0xF:
            field = f"0x180b000{hex_val.lower()}_S{hex_val}_MinTemp"
            fields_MinTemp.append(field)  
            field_min = f"0x180c000{hex_val.lower()}_S{hex_val}_MinCell_Voltage"
            fields_min.append(field_min)            
        else :
            field = f"0x180b00{hex_val.lower()}_S{hex_val}_MinTemp"
            fields_MinTemp.append(field)  
            field_min = f"0x180c00{hex_val.lower()}_S{hex_val}_MinCell_Voltage"
            fields_min.append(field_min)   
    # print(fields_max)

    for i in range(26):
        
        def Thread1():
            global result_MaxTemp # Access the global variables
            query2 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{fields_MaxTemp[i]}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: true)'
            


            # print(query2) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result_MaxTemp = query_api.query(org=org, query=query2)
            # print("processing")
            results_MaxTemps = []

            for table in result_MaxTemp:
                for record in table.records:
                    results_MaxTemps.append((record.get_field(), record.get_value()))
            
            global_result2 = result_MaxTemp
        def Thread2():
            global result_MinTemp  # Access the global variables


            query3 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{fields_MinTemp[i]}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: false)'



            # print(query3) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result_MinTemp = query_api.query(org=org, query=query3)
            # print("processing")
            results_MinTemps = []
            for table in result_MinTemp:
                for record in table.records:
                    results_MinTemps.append((record.get_field(), record.get_value()))
            
            
            global_result3 = result_MinTemp
        def Thread3():
            global result_MaxCell

            query4 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{fields_max[i]}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: false)'



            # print(query4) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result_MaxCell = query_api.query(org=org, query=query4)
            # print("processing")
            results_MaxCells = []
            for table in result_MaxCell:
                for record in table.records:
                    results_MaxCells.append((record.get_field(), record.get_value()))
            result_maxs = results_MaxCells
            # print(results_MaxCells)
        def Thread4():
            global result_MinCell
            query5 = f' from(bucket:"datalogger")\
            |> range(start:{start_t}, stop:{end_t})\
            |> filter(fn:(r) => r._measurement == "sbcu")\
            |> filter(fn:(r) => r._field == "{fields_min[i]}" )\
            |> filter(fn:(r) => r.ferry_id == "{ferry_ided}" )\
            |> aggregateWindow(every: {sampling}, fn: last, createEmpty: false)'

            # print(query5) 
            client = InfluxDBClient(url="http://mmr-influx.energyabsolute.co.th:8086/", token=token)
            query_api = client.query_api()
            # Write a query and execute it

            result_MinCell = query_api.query(org=org, query=query5)
            # print("processing")
            results_MinCells = []
            for table in result_MinCell:
                for record in table.records:
                    results_MinCells.append((record.get_field(), record.get_value()))
        threading_Timer_s = time.time()
        thread1 = threading.Thread(target=Thread1)
        thread2 = threading.Thread(target=Thread2)
        thread3 = threading.Thread(target=Thread3)
        thread4 = threading.Thread(target=Thread4)

        # print(results3)


        # Start the threads
        thread1.start()
        thread2.start()
        thread3.start()
        thread4.start()
        # Wait for both threads to finish
        thread1.join()
        thread2.join()
        thread3.join()
        thread4.join()


        threading_Timer_e = time.time()
        execution_time_thread = threading_Timer_e - threading_Timer_s
        print(f"Process_Data took {execution_time_thread:.6f} seconds to execute.")

        Value_maxtemp = []
        Value_mintemp = []
        Value_maxV = []
        Value_minV = []
        timestamps = []
        for table in result_MaxTemp:
            for record in table.records:
                timestamp = record.get_time()
                converted_time1 = timestamp.astimezone(target_time_zone).replace(tzinfo=None)
                timestamps.append(converted_time1)
                field = record.get_field()
                value = record.get_value()
                if value is not None:
                    Value_maxtemp.append(value)

        for table in result_MinTemp:
            for record in table.records:
                value = record.get_value()
                if value is not None:
                    Value_mintemp.append(value)

        for table in result_MaxCell:
            for record in table.records:
                value = record.get_value()/1000
                if value is not None:
                    Value_maxV.append(value)

        for table in result_MinCell:
            for record in table.records:
                value = record.get_value()/1000
                if value is not None:
                    Value_minV.append(value)


        if Value_maxtemp != []:
            MaxTemp = max(Value_maxtemp)
        elif Value_maxtemp is not None:
            MaxTemp = 0


        else:
            MaxTemp = 0
        if  Value_mintemp != []:    
            MinTemp = min(Value_mintemp)
        else:
            MinTemp = 0
        if  Value_maxV != []:
            MaxVolt = max(Value_maxV)
        else:
            MaxVolt = 0
        if  Value_minV != []:
            MinVolt = min(Value_minV)
        else:
            MinVolt = 0

        Value_CellDiff = []
        PValue = 0
        if Value_maxV != []:
            for k,maxV in enumerate(Value_maxV ):
                Value = maxV - Value_minV[k]
                Value_CellDiff.append(Value)
                if Value >  PValue:
                    PValue = Value
                    timy = k
        if timestamps != []:
            timmy = timestamps[timy]
            timmy_max = Value_maxV[timy]
            timmy_min = Value_minV[timy]
        else:
            timmy = 0
            timmy_max =0
            timmy_min = 0  
        if Value_CellDiff != []:
            Cell_Diff = max(Value_CellDiff)
        else:
            Cell_Diff = 0
        
        cell = sheet_Process.cell(row=i + 2, column = 4, value = ferries)
        cell = sheet_Process.cell(row=i + 2, column = 5, value = i+1)
        cell = sheet_Process.cell(row=i + 2, column = 1, value = modified_string)
        cell = sheet_Process.cell(row=i + 2, column=2, value= "5:00:00")
        cell = sheet_Process.cell(row=i + 2, column=3, value="22:00:00")  
        cell = sheet_Process.cell(row=i + 2, column=6, value=MinVolt)  
        cell = sheet_Process.cell(row=i + 2, column=7, value=MaxVolt)  
        cell = sheet_Process.cell(row=i + 2, column=8, value=MaxTemp)  
        cell = sheet_Process.cell(row=i + 2, column=9, value=MinTemp) 
        


        cell = sheet_Process_Vdiff.cell(row=i + 2, column = 4, value = ferries)
        cell = sheet_Process_Vdiff.cell(row=i + 2, column = 5, value = i+1)
        cell = sheet_Process_Vdiff.cell(row=i + 2, column = 1, value = modified_string)
        cell = sheet_Process_Vdiff.cell(row=i + 2, column=2, value= "5:00:00")
        cell = sheet_Process_Vdiff.cell(row=i + 2, column=3, value="22:00:00") 
        cell = sheet_Process_Vdiff.cell(row=i+2, column=7, value=timmy_max)  
        cell = sheet_Process_Vdiff.cell(row=i+2, column=8, value=timmy_min)  
        cell = sheet_Process_Vdiff.cell(row=i+2, column=6, value=timmy)  
        cell = sheet_Process_Vdiff.cell(row=i+2, column=9, value=Cell_Diff) 
    # Define the directory where you want to save the files

    OfflineCheck(ferries,sheet_PackStat_Sum)

    base_filename = f'{modified_date_save}_MSF{ferry_ided}_Usage.xlsx'
    counter = 0
    while os.path.exists( base_filename):
        counter += 1
        base_filename = f'{modified_date_save}_MSF{ferry_ided}_Usage_{counter}.xlsx'
    # Save the wordbook2 to a file
    del workbook['Sheet']
    workbook.save( base_filename)
    print(f'Saved as: {base_filename}')
    base_filename2 = f'{modified_date_save}_MSF{ferry_ided}_Diagnose.xlsx'
    counter = 0
    while os.path.exists( base_filename2):
        counter += 1
        base_filename2 = f'{modified_date_save}_MSF{ferry_ided}_Diagnose_{counter}.xlsx'
    # Save the wordbook2 to a file
    del workbook2['Sheet']
    workbook2.save(base_filename2)
    print(f'Saved as: {base_filename2}')    

ferris = Get_ferry_id()
for i in ferris:
    ProcessRange(i)
# ProcessRange(17)

# Sleep for 2 minutes (120 seconds) after the program is complete
sleep_duration = 120  # 2 minutes
end_time = time.time() + sleep_duration

while time.time() < end_time:
    time.sleep(1)  # Sleep for 1 second